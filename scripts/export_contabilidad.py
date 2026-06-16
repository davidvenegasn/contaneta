"""Generate a single Excel workbook with all CFDI data for accounting declarations.

Users included:
  - Diego (issuer_id=9)    — RESICO PF (régimen 626)
  - Perla (issuer_id=9103) — RESICO PF (régimen 626)
  - Manuel (issuer_id=11)  — PF con Actividad Empresarial (régimen 612)

Output: contabilidad_2026.xlsx in project root.
"""
from __future__ import annotations

import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add project root so we can import services
import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from services.sat.cfdi_relacion_labels import label_for_received, signed_amount

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "invoicing.db"
PERIOD = "2026-05"  # YYYY-MM — periodo de la declaración
OUT = ROOT / f"contabilidad_{PERIOD}.xlsx"

USERS = [
    {"id": 9,    "nombre": "Diego",  "rfc_label": "GAZD970429MKA", "regimen": "626 — RESICO Personas Físicas"},
    {"id": 9103, "nombre": "Perla",  "rfc_label": "CAJP980715DG6", "regimen": "626 — RESICO Personas Físicas"},
    {"id": 11,   "nombre": "Manuel", "rfc_label": "MOBJ970402176", "regimen": "612 — PF Actividad Empresarial"},
]

# ── Styles ───────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F2937")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11, name="Inter")
SUBHEAD_FILL  = PatternFill("solid", fgColor="E5E7EB")
SUBHEAD_FONT  = Font(bold=True, size=10, name="Inter")
TITLE_FONT    = Font(bold=True, size=16, name="Inter")
SUBTITLE_FONT = Font(italic=True, size=10, color="6B7280", name="Inter")
BORDER_THIN   = Border(
    left=Side(style="thin",  color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin",   color="E5E7EB"),
    bottom=Side(style="thin",color="E5E7EB"),
)
TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")
TOTAL_FONT = Font(bold=True, size=10, name="Inter")
MONEY_FMT  = '"$"#,##0.00'


def conn() -> sqlite3.Connection:
    c = sqlite3.connect(str(DB))
    c.row_factory = sqlite3.Row
    return c


def fetch_cfdi(c: sqlite3.Connection, issuer_id: int, direction: str) -> list[sqlite3.Row]:
    """Fetch all non-cancelled CFDI for a user/direction, sorted by fecha."""
    rows = c.execute(
        """
        SELECT uuid, fecha_emision, serie, folio,
               rfc_emisor, nombre_emisor, rfc_receptor, nombre_receptor,
               tipo_comprobante, tipo_relacion, metodo_pago, forma_pago, uso_cfdi,
               concepto, moneda,
               COALESCE(subtotal, 0)    AS subtotal,
               COALESCE(descuento, 0)   AS descuento,
               COALESCE(impuestos, 0)   AS iva,
               COALESCE(retenciones, 0) AS retenciones,
               COALESCE(total, 0)       AS total,
               status, xml_status
          FROM sat_cfdi
         WHERE issuer_id = ?
           AND direction = ?
           AND COALESCE(status, '') NOT IN ('0', 'C', 'cancelled', 'Cancelado')
           AND substr(fecha_emision, 1, 7) = ?
         ORDER BY fecha_emision ASC
        """,
        (issuer_id, direction, PERIOD),
    ).fetchall()
    return list(rows)


# ── Helpers ──────────────────────────────────────────────────────────────
def parse_date(s: str | None) -> str:
    if not s:
        return ""
    s = s.replace("T", " ").split(".")[0]
    return s[:10]


def month_key(s: str | None) -> str:
    return (s or "")[:7] if s else "????-??"


def autosize(ws, min_w: int = 10, max_w: int = 42):
    for col_cells in ws.columns:
        col = col_cells[0].column
        max_len = min_w
        for cell in col_cells:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, max_w)


def style_header_row(ws, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    ws.row_dimensions[row_idx].height = 28


# ── Build sheets ─────────────────────────────────────────────────────────
def write_user_movements(ws, rows: list[sqlite3.Row], titulo: str, kind: str):
    """kind: 'ingresos' (issued) or 'gastos' (received)."""
    ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"{len(rows)} CFDI vigentes (no cancelados)").font = SUBTITLE_FONT

    headers = [
        "Fecha", "UUID", "Serie", "Folio",
        "RFC contraparte", "Nombre contraparte",
        "Tipo", "Tipo (label)", "Método", "Forma", "Uso CFDI",
        "Concepto", "Moneda",
        "Subtotal", "Descuento", "IVA", "Retenciones", "Total",
    ]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers))

    total_sub = total_iva = total_ret = total_tot = 0.0
    for ridx, r in enumerate(rows, start=HEADER_ROW + 1):
        rfc_cp = r["rfc_receptor"] if kind == "ingresos" else r["rfc_emisor"]
        nom_cp = r["nombre_receptor"] if kind == "ingresos" else r["nombre_emisor"]
        sub = float(r["subtotal"] or 0)
        desc = float(r["descuento"] or 0)
        iva = float(r["iva"] or 0)
        ret = float(r["retenciones"] or 0)
        tot = float(r["total"] or 0)
        total_sub += sub
        total_iva += iva
        total_ret += ret
        total_tot += tot
        tc = r["tipo_comprobante"] or ""
        tr = r["tipo_relacion"] if "tipo_relacion" in r.keys() else None
        tipo_lbl = label_for_received(tc, tr) if kind == "gastos" else tc
        vals = [
            parse_date(r["fecha_emision"]),
            r["uuid"], r["serie"] or "", r["folio"] or "",
            rfc_cp or "", nom_cp or "",
            tc, tipo_lbl, r["metodo_pago"] or "",
            r["forma_pago"] or "", r["uso_cfdi"] or "",
            (r["concepto"] or "")[:60], r["moneda"] or "MXN",
            sub, desc, iva, ret, tot,
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=ridx, column=ci, value=v)
            cell.border = BORDER_THIN
            if ci >= 14:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")

    # Totals row
    if rows:
        trow = HEADER_ROW + 1 + len(rows)
        ws.cell(row=trow, column=13, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=trow, column=13).alignment = Alignment(horizontal="right")
        for ci, v in [(14, total_sub), (16, total_iva), (17, total_ret), (18, total_tot)]:
            cell = ws.cell(row=trow, column=ci, value=v)
            cell.number_format = MONEY_FMT
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.alignment = Alignment(horizontal="right")
            cell.border = BORDER_THIN

    ws.freeze_panes = "A5"
    autosize(ws)


def write_monthly_summary(ws, issued: list[sqlite3.Row], received: list[sqlite3.Row], titulo: str):
    ws.cell(row=1, column=1, value=titulo).font = TITLE_FONT
    ws.cell(row=2, column=1, value="Resumen mensual — ingresos, IVA trasladado, retenciones, gastos deducibles").font = SUBTITLE_FONT

    # Aggregate per month
    months: dict[str, dict[str, float]] = defaultdict(lambda: {
        "ing_sub": 0.0, "ing_iva": 0.0, "ing_ret": 0.0, "ing_tot": 0.0, "ing_n": 0,
        "gas_sub": 0.0, "gas_iva": 0.0, "gas_ret": 0.0, "gas_tot": 0.0, "gas_n": 0,
        "nc_tot": 0.0, "nc_n": 0,
        "antic_tot": 0.0, "antic_n": 0,
    })
    for r in issued:
        m = month_key(r["fecha_emision"])
        months[m]["ing_sub"] += float(r["subtotal"] or 0)
        months[m]["ing_iva"] += float(r["impuestos"] if "impuestos" in r.keys() else r["iva"] or 0)
        months[m]["ing_ret"] += float(r["retenciones"] or 0)
        months[m]["ing_tot"] += float(r["total"] or 0)
        months[m]["ing_n"] += 1
    for r in received:
        tc = (r["tipo_comprobante"] or "").upper()
        tr = r["tipo_relacion"] if "tipo_relacion" in r.keys() else None
        # Exclude nómina (N) and pago (P) from gastos deducibles
        if tc in ("N", "P"):
            continue
        m = month_key(r["fecha_emision"])
        tot = float(r["total"] or 0)
        # Classify Egresos by TipoRelacion
        if tc == "E" and (tr or "") in ("01", "03"):
            months[m]["nc_tot"] += tot
            months[m]["nc_n"] += 1
        elif tc == "E" and tr == "07":
            months[m]["antic_tot"] += tot
            months[m]["antic_n"] += 1
        elif tc == "E" and (tr or "") in ("04", "05", "06"):
            pass  # Neutral — sustitución/traslado, skip
        else:
            # Regular gastos (Ingreso type or unknown Egreso)
            months[m]["gas_sub"] += float(r["subtotal"] or 0)
            months[m]["gas_iva"] += float(r["impuestos"] if "impuestos" in r.keys() else r["iva"] or 0)
            months[m]["gas_ret"] += float(r["retenciones"] or 0)
            months[m]["gas_tot"] += tot
            months[m]["gas_n"] += 1

    headers = [
        "Mes",
        "# Ingresos", "Subtotal ingresos", "IVA trasladado", "Retenciones a favor", "Total ingresos",
        "# Gastos", "Subtotal gastos", "IVA acreditable", "Retenciones (en gasto)", "Total gastos",
        "# NC/Dev", "NC + Devoluciones", "# Anticipos", "Anticipos aplicados",
        "Gasto neto deducible", "Utilidad fiscal", "IVA neto",
    ]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers))

    money_cols = (3, 4, 5, 6, 8, 9, 10, 11, 13, 15, 16, 17, 18)
    total = defaultdict(float)
    sorted_months = sorted(months.keys())
    for ridx, m in enumerate(sorted_months, start=HEADER_ROW + 1):
        d = months[m]
        gasto_neto = d["gas_tot"] - d["nc_tot"] - d["antic_tot"]
        util = d["ing_sub"] - gasto_neto
        iva_neto = d["ing_iva"] - d["gas_iva"]
        vals = [
            m,
            d["ing_n"], d["ing_sub"], d["ing_iva"], d["ing_ret"], d["ing_tot"],
            d["gas_n"], d["gas_sub"], d["gas_iva"], d["gas_ret"], d["gas_tot"],
            d["nc_n"], d["nc_tot"], d["antic_n"], d["antic_tot"],
            gasto_neto, util, iva_neto,
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=ridx, column=ci, value=v)
            cell.border = BORDER_THIN
            if ci in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
        for k in ("ing_n", "ing_sub", "ing_iva", "ing_ret", "ing_tot",
                   "gas_n", "gas_sub", "gas_iva", "gas_ret", "gas_tot",
                   "nc_n", "nc_tot", "antic_n", "antic_tot"):
            total[k] += d[k]

    if sorted_months:
        trow = HEADER_ROW + 1 + len(sorted_months)
        t_gasto_neto = total["gas_tot"] - total["nc_tot"] - total["antic_tot"]
        ws.cell(row=trow, column=1, value="TOTAL")
        for ci, v in [
            (2, total["ing_n"]), (3, total["ing_sub"]), (4, total["ing_iva"]),
            (5, total["ing_ret"]), (6, total["ing_tot"]),
            (7, total["gas_n"]), (8, total["gas_sub"]), (9, total["gas_iva"]),
            (10, total["gas_ret"]), (11, total["gas_tot"]),
            (12, total["nc_n"]), (13, total["nc_tot"]),
            (14, total["antic_n"]), (15, total["antic_tot"]),
            (16, t_gasto_neto),
            (17, total["ing_sub"] - t_gasto_neto),
            (18, total["ing_iva"] - total["gas_iva"]),
        ]:
            cell = ws.cell(row=trow, column=ci, value=v)
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = BORDER_THIN
            if ci in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
        ws.cell(row=trow, column=1).font = TOTAL_FONT
        ws.cell(row=trow, column=1).fill = TOTAL_FILL

    ws.freeze_panes = "B5"
    autosize(ws)


def write_resumen_general(ws, users_data: list[dict]):
    ws.cell(row=1, column=1, value=f"ContaNeta — Declaración mensual {PERIOD}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generado el {datetime.now().strftime('%Y-%m-%d %H:%M')} · solo CFDI vigentes del periodo {PERIOD}").font = SUBTITLE_FONT

    headers = [
        "Usuario", "RFC", "Régimen fiscal",
        "# Ingresos", "Subtotal ingresos", "IVA trasladado", "Retenciones a favor",
        "# Gastos", "Subtotal gastos deducibles", "IVA acreditable",
        "Utilidad fiscal estimada", "IVA neto",
    ]
    HEADER_ROW = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=i, value=h)
    style_header_row(ws, HEADER_ROW, len(headers))

    for ridx, u in enumerate(users_data, start=HEADER_ROW + 1):
        i_sub = i_iva = i_ret = 0.0
        for r in u["issued"]:
            i_sub += float(r["subtotal"] or 0)
            i_iva += float(r["impuestos"] if "impuestos" in r.keys() else 0)
            i_ret += float(r["retenciones"] or 0)
        g_sub = g_iva = 0.0
        for r in u["received"]:
            if (r["tipo_comprobante"] or "").upper() in ("N", "P"):
                continue
            g_sub += float(r["subtotal"] or 0)
            g_iva += float(r["impuestos"] if "impuestos" in r.keys() else 0)
        utilidad = i_sub - g_sub
        iva_neto = i_iva - g_iva
        vals = [
            u["nombre"], u["rfc_label"], u["regimen"],
            len(u["issued"]), i_sub, i_iva, i_ret,
            sum(1 for r in u["received"] if (r["tipo_comprobante"] or "").upper() not in ("N", "P")),
            g_sub, g_iva, utilidad, iva_neto,
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=ridx, column=ci, value=v)
            cell.border = BORDER_THIN
            if ci in (5, 6, 7, 9, 10, 11, 12):
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")

    # Notes section
    notes_row = HEADER_ROW + 1 + len(users_data) + 2
    ws.cell(row=notes_row, column=1, value="Notas para el contador").font = TITLE_FONT
    notes = [
        "• Solo se incluyen CFDI con estado vigente (status 1/V). Cancelados quedan fuera.",
        "• Para gastos se excluyen tipo P (complementos de pago) y N (nómina) porque no son gastos deducibles.",
        "• Diego y Perla: régimen 626 (RESICO PF). ISR sobre ingresos según tabla del régimen; IVA 16% trasladado.",
        "• Manuel: régimen 612 (PF Actividad Empresarial). ISR sobre utilidad (ingresos − gastos); IVA 16%.",
        "• Las retenciones registradas en CFDI emitidos son ISR/IVA retenidos POR EL CLIENTE — acreditables contra el impuesto a cargo.",
        "• Para la declaración mensual revisa la hoja '<Nombre> — Mensual' de cada usuario.",
        "• Detalle CFDI por CFDI en las hojas '<Nombre> — Ingresos' y '<Nombre> — Gastos'.",
    ]
    for i, n in enumerate(notes, start=1):
        cell = ws.cell(row=notes_row + i, column=1, value=n)
        cell.font = Font(size=10, name="Inter", color="374151")
        ws.merge_cells(start_row=notes_row + i, start_column=1, end_row=notes_row + i, end_column=8)

    ws.freeze_panes = "A5"
    autosize(ws)


def main() -> None:
    c = conn()
    wb = Workbook()
    wb.remove(wb.active)

    users_data = []
    for u in USERS:
        issued = fetch_cfdi(c, u["id"], "issued")
        received = fetch_cfdi(c, u["id"], "received")
        users_data.append({**u, "issued": issued, "received": received})

    # Resumen General first
    ws = wb.create_sheet("Resumen General")
    write_resumen_general(ws, users_data)

    # Per-user sheets
    for u in users_data:
        nom = u["nombre"]
        ws_m = wb.create_sheet(f"{nom} — Mensual")
        write_monthly_summary(ws_m, u["issued"], u["received"],
                              f"{nom} ({u['rfc_label']}) · {u['regimen']}")

        ws_i = wb.create_sheet(f"{nom} — Ingresos")
        write_user_movements(ws_i, u["issued"],
                             f"{nom} · Ingresos (CFDI emitidos)", "ingresos")

        ws_g = wb.create_sheet(f"{nom} — Gastos")
        write_user_movements(ws_g, u["received"],
                             f"{nom} · Gastos (CFDI recibidos)", "gastos")

    wb.save(OUT)
    print(f"OK → {OUT}")
    for u in users_data:
        print(f"  {u['nombre']:8s}  ingresos={len(u['issued']):4d}  gastos={len(u['received']):4d}")


if __name__ == "__main__":
    main()
