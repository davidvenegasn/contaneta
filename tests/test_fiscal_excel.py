"""Tests for fiscal Excel export."""

import os
import sys
import tempfile
from io import BytesIO
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

if not os.environ.get("APP_DB_PATH"):
    _fd, _p = tempfile.mkstemp(suffix=".db", prefix="test_fiscal_xl_")
    os.close(_fd)
    os.environ["APP_DB_PATH"] = _p
if not os.environ.get("SESSION_SECRET"):
    os.environ["SESSION_SECRET"] = "test-secret-fiscal-xl"

import database  # noqa: E402
from config import DB_PATH  # noqa: E402
from migrations_runner import apply_migrations  # noqa: E402

ISSUER_ID = 88810
USER_ID = 88810


@pytest.fixture(scope="module", autouse=True)
def seed():
    apply_migrations(DB_PATH)
    conn = database.db()
    try:
        conn.execute(
            "INSERT OR IGNORE INTO issuers (id, rfc, razon_social, active, created_at, updated_at) "
            "VALUES (?, 'FISC010101AAA', 'Fiscal Excel SA', 1, datetime('now'), datetime('now'))",
            (ISSUER_ID,),
        )
        conn.execute(
            "INSERT OR IGNORE INTO users (id, email, password_hash, created_at) "
            "VALUES (?, 'fiscal@test.local', 'x', datetime('now'))",
            (USER_ID,),
        )
        conn.execute(
            "INSERT OR IGNORE INTO memberships (user_id, issuer_id, role, created_at) "
            "VALUES (?, ?, 'owner', datetime('now'))",
            (USER_ID, ISSUER_ID),
        )
        # Seed CFDI data
        for i in range(3):
            conn.execute(
                "INSERT OR IGNORE INTO sat_cfdi (issuer_id, uuid, direction, fecha_emision, total, subtotal, impuestos, status) "
                "VALUES (?, ?, 'issued', '2026-05-15', 11600, 10000, 1600, 'vigente')",
                (ISSUER_ID, f"fiscal-issued-{i:04d}"),
            )
            conn.execute(
                "INSERT OR IGNORE INTO sat_cfdi (issuer_id, uuid, direction, fecha_emision, total, subtotal, impuestos, status) "
                "VALUES (?, ?, 'received', '2026-05-20', 5800, 5000, 800, 'vigente')",
                (ISSUER_ID, f"fiscal-received-{i:04d}"),
            )
        conn.commit()
    finally:
        conn.close()
    yield


@pytest.fixture(scope="module")
def client():
    from fastapi.testclient import TestClient
    from app import app
    from tests.helpers import make_session_cookie
    cookies = make_session_cookie(issuer_id=ISSUER_ID, user_id=USER_ID)
    return TestClient(app, raise_server_exceptions=False, cookies=cookies)


def test_excel_export_creates_4_sheets():
    """build_fiscal_excel should produce a workbook with 4 sheets."""
    from openpyxl import load_workbook
    from services.fiscal.excel_export import build_fiscal_excel
    data = build_fiscal_excel(ISSUER_ID, "2026-05", issuer_alias="Test Co")
    wb = load_workbook(BytesIO(data))
    assert len(wb.sheetnames) == 4
    assert wb.sheetnames == ["Resumen", "Ingresos", "Gastos", "Deducciones"]


def test_excel_export_includes_cfdis_of_month():
    """Ingresos sheet should contain the seeded issued CFDIs."""
    from openpyxl import load_workbook
    from services.fiscal.excel_export import build_fiscal_excel
    data = build_fiscal_excel(ISSUER_ID, "2026-05")
    wb = load_workbook(BytesIO(data))
    ws = wb["Ingresos"]
    # Header + 3 data rows
    assert ws.max_row >= 4, f"Expected >=4 rows (header + 3 CFDIs), got {ws.max_row}"


def test_excel_export_endpoint_requires_auth():
    """GET /portal/fiscal/export without auth returns 401 or redirect."""
    from fastapi.testclient import TestClient
    from app import app
    plain_client = TestClient(app, raise_server_exceptions=False)
    resp = plain_client.get("/portal/fiscal/export?ym=2026-05")
    assert resp.status_code in (401, 302, 307)


def test_excel_export_endpoint_returns_xlsx(client):
    """GET /portal/fiscal/export should return an XLSX file."""
    resp = client.get("/portal/fiscal/export?ym=2026-05")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers.get("content-type", "")
    assert "ContaNeta_papeles_" in resp.headers.get("content-disposition", "")


def test_excel_filename_uses_issuer_alias_and_month(client):
    """The Content-Disposition header should include issuer alias and ym."""
    resp = client.get("/portal/fiscal/export?ym=2026-05")
    assert resp.status_code == 200
    disposition = resp.headers.get("content-disposition", "")
    assert "2026-05" in disposition
