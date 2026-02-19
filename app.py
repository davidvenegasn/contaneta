import base64
import hashlib
import hmac
import logging
import os
import re
import secrets
import sqlite3
import time
from datetime import datetime, date, timedelta
from typing import Optional, List, Union
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

from dotenv import load_dotenv
from migrations_runner import apply_migrations
from fastapi import FastAPI, Request, Form, HTTPException, Query, Body, Depends
from fastapi.responses import HTMLResponse, Response, RedirectResponse
from io import BytesIO
from fastapi.templating import Jinja2Templates

from facturapi_client import create_invoice, download_invoice, FacturapiError
from validators import validate_customer, validate_product

from fastapi.staticfiles import StaticFiles

app = FastAPI()

# Absolute paths (avoid issues if uvicorn is started from another working directory)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=TEMPLATES_DIR)

load_dotenv()

# DB path (default to local file inside this project folder)
DB_PATH = os.getenv("APP_DB_PATH") or os.path.join(BASE_DIR, "invoicing.db")

# SAT catalogs DB (from phpcfdi/resources-sat-catalogs)
CATALOGS_DB = os.path.join(BASE_DIR, "catalogs", "catalogs.db")

# Dev-mode token bypass (ONLY for local development)
DEV_MODE = os.getenv("DEV_MODE", "1") == "1"
DEV_TOKEN = os.getenv("DEV_TOKEN", "demo")

# Session cookie (auth sin token en URL)
SESSION_SECRET = os.getenv("SESSION_SECRET", secrets.token_hex(32))
SESSION_COOKIE_NAME = "portal_session"
SESSION_TTL_DAYS = int(os.getenv("SESSION_TTL_DAYS", "7"))
COOKIE_SECURE = os.getenv("COOKIE_SECURE", "0") == "1"  # True en prod (HTTPS)


# ----------------------------
# DB helpers (invoicing.db) - deben estar antes de auth helpers
# ----------------------------

def db() -> sqlite3.Connection:
    """Conexión a invoicing.db con timeout y pragmas para reducir I/O y locks (ver migrations_runner)."""
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute("PRAGMA busy_timeout = 5000;")
    conn.execute("PRAGMA journal_mode = WAL;")
    return conn

def db_rows(sql: str, params: tuple = ()) -> list[dict]:
    """Run a SELECT and return rows as list[dict] (safe for templates)."""
    conn = db()
    try:
        cur = conn.execute(sql, params)
        rows = cur.fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ----------------------------
# Auth helpers (deben estar antes de las rutas que usan Depends)
# ----------------------------

def get_issuer_by_token(token: str):
    # DEV bypass: allow opening the portal in Safari without a real token
    if DEV_MODE and (not token or token == DEV_TOKEN):
        return {
            "id": -1,
            "alias": "Contaneta",
            "rfc": "XIA190128J61",
            "regimen_fiscal": None,
            "facturapi_org_id": None,
            "active": 1,
        }

    conn = db()
    row = conn.execute(
        """
        SELECT
            i.id,
            i.rfc,
            i.razon_social,
            i.regimen_fiscal,
            i.active,
            i.facturapi_org_id,
            t.token
        FROM issuer_tokens t
        JOIN issuers i ON i.id = t.issuer_id
        WHERE t.token = ? AND t.active = 1 AND i.active = 1
        LIMIT 1
        """,
        (token,),
    ).fetchone()
    conn.close()

    if not row:
        raise ValueError("Token inválido o inactivo.")

    # Normaliza el issuer a dict y crea campos "compatibles" con el portal
    d = dict(row)
    regimen = (d.get("regimen_fiscal") or "").strip().upper()
    return {
        "id": d["id"],
        "rfc": d.get("rfc") or "",
        "alias": d.get("razon_social") or d.get("rfc") or "Emisor",
        "regimen_fiscal": regimen or None,
        "facturapi_org_id": d.get("facturapi_org_id"),  # puede no existir (None)
        "active": d.get("active", 1),
    }


def get_issuer_by_id(issuer_id: int):
    """Carga emisor por id (para sesión por cookie). Dev bypass si issuer_id == -1."""
    if DEV_MODE and issuer_id == -1:
        return {
            "id": -1,
            "alias": "Contaneta",
            "rfc": "XIA190128J61",
            "regimen_fiscal": None,
            "facturapi_org_id": None,
            "active": 1,
        }
    conn = db()
    row = conn.execute(
        "SELECT id, rfc, razon_social, regimen_fiscal, active, facturapi_org_id FROM issuers WHERE id = ? AND active = 1 LIMIT 1",
        (issuer_id,),
    ).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    regimen = (d.get("regimen_fiscal") or "").strip().upper()
    return {
        "id": d["id"],
        "rfc": d.get("rfc") or "",
        "alias": d.get("razon_social") or d.get("rfc") or "Emisor",
        "regimen_fiscal": regimen or None,
        "facturapi_org_id": d.get("facturapi_org_id"),
        "active": d.get("active", 1),
    }


def _sign_session(issuer_id: int) -> str:
    """Firma payload issuer_id|expiry_ts con HMAC. TTL = SESSION_TTL_DAYS."""
    expiry = int(time.time()) + SESSION_TTL_DAYS * 86400
    payload = f"{issuer_id}|{expiry}"
    sig = hmac.new(SESSION_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode(f"{payload}.{sig}".encode()).decode().rstrip("=")


def _verify_session(cookie_val: Optional[str]) -> Optional[int]:
    """Verifica cookie de sesión; devuelve issuer_id o None."""
    if not cookie_val or not cookie_val.strip():
        return None
    try:
        raw = base64.urlsafe_b64decode(cookie_val + "==")
        s = raw.decode()
    except Exception:
        return None
    if "." not in s:
        return None
    payload, sig = s.rsplit(".", 1)
    expected = hmac.new(SESSION_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, sig):
        return None
    parts = payload.split("|")
    if len(parts) != 2:
        return None
    issuer_id, expiry = int(parts[0]), int(parts[1])
    if time.time() > expiry:
        return None
    return issuer_id


def _session_cookie_params(request: Optional[Request] = None) -> dict:
    """
    Parámetros para set/clear cookie (HttpOnly, SameSite=Lax, Secure).
    Secure=True solo cuando el cliente usa HTTPS (evita que en HTTP la cookie no se guarde y no puedas abrir el portal).
    Se usa X-Forwarded-Proto si existe (proxy inverso) o request.url.scheme.
    """
    secure = COOKIE_SECURE
    if request is not None:
        proto = (request.headers.get("x-forwarded-proto") or "").strip().lower()
        if proto == "https":
            secure = True
        elif request.url.scheme == "https":
            secure = True
        elif not COOKIE_SECURE:
            secure = False  # Cliente por HTTP: no marcar Secure para que el navegador guarde la cookie
    return {
        "httponly": True,
        "samesite": "lax",
        "secure": secure,
        "max_age": SESSION_TTL_DAYS * 86400,
        "path": "/",
    }


def get_portal_issuer(request: Request) -> dict:
    """
    Dependency: autentica por cookie o por ?token= (legacy).
    Si hay token en query y es válido: setea cookie y devuelve issuer (la redirección sin token se maneja en middleware).
    Si hay cookie válida: setea request.state.issuer_id y request.state.issuer y devuelve issuer.
    Si no: lanza HTTPException (API) o HTTPException con código 302 para HTML (manejado por middleware).
    
    IMPORTANTE: Esta dependencia SIEMPRE devuelve dict o lanza excepción. Nunca devuelve RedirectResponse
    directamente porque FastAPI puede pasarlo como parámetro a la ruta en lugar de interceptarlo.
    """
    token_query = request.query_params.get("token", "").strip()
    cookie_val = request.cookies.get(SESSION_COOKIE_NAME)
    is_api = request.url.path.startswith("/api/") or request.url.path.startswith("/download/")

    # Legacy: ?token= válido → iniciar sesión (cookie) y devolver issuer
    # La redirección sin token se maneja en el middleware redirect_token_middleware
    if token_query:
        try:
            issuer = get_issuer_by_token(token_query)
        except ValueError:
            pass
        else:
            request.state.issuer_id = issuer["id"]
            request.state.issuer = issuer
            return issuer

    # Sesión por cookie
    issuer_id = _verify_session(cookie_val)
    if issuer_id is not None:
        issuer = get_issuer_by_id(issuer_id)
        if issuer:
            request.state.issuer_id = issuer["id"]
            request.state.issuer = issuer
            return issuer

    # DEV_MODE: permitir acceso sin login con issuer demo (opcional)
    if DEV_MODE and not is_api:
        demo = get_issuer_by_token(DEV_TOKEN)
        request.state.issuer_id = demo["id"]
        request.state.issuer = demo
        return demo

    # No autorizado: para HTML, usar HTTPException con código 302 que será manejado por middleware
    # Para API, usar HTTPException normal
    if is_api:
        raise HTTPException(status_code=401, detail="No autorizado")
    # Para rutas HTML sin autenticación, lanzar excepción que será capturada y redirigida
    raise HTTPException(status_code=401, detail="No autorizado - redirigir a /login")


# Middleware: redirección por token (legacy) y redirección a /login si no hay sesión (evitar JSON 401)
@app.middleware("http")
async def redirect_token_middleware(request: Request, call_next):
    """
    1. Si hay ?token= en URL del portal (no API), establecer cookie y redirigir sin token.
    2. Para rutas HTML del portal sin sesión ni token: redirigir a /login ANTES de la ruta
       (así el navegador recibe 302 y no JSON "No autorizado").
    """
    token_query = request.query_params.get("token", "").strip()
    is_api = request.url.path.startswith("/api/") or request.url.path.startswith("/download/")
    is_public = request.url.path.startswith("/q/") or request.url.path.startswith("/public/")
    is_portal_html = request.url.path.startswith("/portal/") and not is_api and not is_public

    # 1. Manejar ?token= en URL del portal (legacy)
    if token_query and is_portal_html:
        try:
            issuer = get_issuer_by_token(token_query)
            parsed = urlparse(str(request.url))
            qs = parse_qs(parsed.query, keep_blank_values=False)
            qs.pop("token", None)
            if qs:
                new_query = urlencode(qs, doseq=True)
                new_url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))
            else:
                new_url = parsed.path
            response = RedirectResponse(url=new_url, status_code=302)
            response.set_cookie(SESSION_COOKIE_NAME, _sign_session(issuer["id"]), **_session_cookie_params(request))
            return response
        except ValueError:
            pass  # Token inválido, continuar normalmente

    # 2. Portal HTML sin token: si no hay sesión válida, redirigir a /login (sino FastAPI devuelve JSON 401)
    if is_portal_html and not token_query:
        cookie_val = request.cookies.get(SESSION_COOKIE_NAME)
        if _verify_session(cookie_val) is None and not DEV_MODE:
            return RedirectResponse(url="/login", status_code=302)
        # DEV_MODE: dejar pasar para que get_portal_issuer pueda devolver el issuer demo

    response = await call_next(request)
    return response


@app.get("/")
def root():
    return RedirectResponse(url="/portal/home")


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, token: str = Query("", alias="token")):
    """
    Página de login: si se pasa ?token= válido, inicia sesión (cookie) y redirige al portal sin token.
    Si no hay token o es inválido, muestra formulario simple para ingresar token.
    """
    if token and token.strip():
        try:
            issuer = get_issuer_by_token(token.strip())
        except ValueError:
            return _render_login(request, error="Token inválido o inactivo.")
        resp = RedirectResponse(url="/portal/home", status_code=302)
        resp.set_cookie(SESSION_COOKIE_NAME, _sign_session(issuer["id"]), **_session_cookie_params(request))
        return resp
    return _render_login(request)


def _render_login(request: Request, error: Optional[str] = None):
    """Vista simple de login (ingresar token una vez)."""
    return templates.TemplateResponse(
        "login.html",
        {"request": request, "error": error},
    )


@app.get("/logout")
@app.post("/logout")
def logout():
    """Borra la cookie de sesión y redirige a la página pública."""
    resp = RedirectResponse(url="/", status_code=302)
    resp.delete_cookie(SESSION_COOKIE_NAME, path="/")
    return resp


@app.get("/portal")
def portal_root():
    """Redirige al home del portal (la auth se resuelve por cookie en cada ruta)."""
    return RedirectResponse(url="/portal/home")


# Cotizaciones: registrado aquí al inicio para evitar 404 (orden de rutas)
from fastapi.responses import HTMLResponse
from fastapi import Request

@app.get("/portal/quotations", response_class=HTMLResponse)
@app.get("/portal/cotizaciones", response_class=HTMLResponse)
def portal_quotations(request: Request, issuer: dict = Depends(get_portal_issuer)):
    return _portal_quotations_impl(request, issuer)


@app.get("/portal/cotizaciones/ping")
def portal_cotizaciones_ping():
    """Comprueba que la app tiene la ruta de cotizaciones (respuesta en texto)."""
    return Response(content="cotizaciones-ok", media_type="text/plain")

@app.get("/portal/home", response_class=HTMLResponse)
def portal_home(request: Request, issuer: dict = Depends(get_portal_issuer)):
    try:
        issuer_id = issuer["id"]
        ym = ym_now()

        count_issued = db_rows(
            """
            SELECT COUNT(*) AS n FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
              AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
            """,
            (issuer_id, ym),
        )
        count_received = db_rows(
            """
            SELECT COUNT(*) AS n FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
              AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01
              AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
            """,
            (issuer_id, ym),
        )

        # Totales del mes en curso: ingresos/gastos sin IVA, IVA recibido (neto = cobrado − retenciones), IVA pagado
        tot_issued = _get_month_totals(issuer_id, ym, "issued")
        tot_received = _get_month_totals(issuer_id, ym, "received")
        ingresos_sin_iva = tot_issued["total_base"]
        gastos_sin_iva = tot_received["total_base"]
        iva_retenciones = tot_issued["total_retenciones"]
        # Siempre usar total_iva_neto para "IVA recibido" (cobrado − retenciones)
        iva_recibido_neto = tot_issued["total_iva_neto"]
        iva_pagado = tot_received["total_iva"]

        activities = db_rows(
            """
            SELECT direction, fecha_emision, nombre FROM (
              SELECT direction, fecha_emision, nombre_receptor AS nombre FROM sat_cfdi
              WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                AND (total IS NULL OR total >= 0.01)
              UNION ALL
              SELECT direction, fecha_emision, nombre_emisor AS nombre FROM sat_cfdi
              WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                AND total IS NOT NULL AND total >= 0.01
                AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
            ) ORDER BY fecha_emision DESC LIMIT 3
            """,
            (issuer_id, issuer_id),
        )
        today = date.today()
        for a in activities:
            try:
                fd = datetime.strptime((a["fecha_emision"] or "")[:10], "%Y-%m-%d").date()
                d = (today - fd).days
                a["time_ago"] = "Hoy" if d == 0 else "Ayer" if d == 1 else f"Hace {d} días"
            except (ValueError, TypeError):
                a["time_ago"] = a.get("fecha_emision", "")[:10] or "-"

        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_home.html",
            active_page="home",
            title="Inicio",
            extra={
                "count_issued": count_issued[0]["n"] if count_issued else 0,
                "count_received": count_received[0]["n"] if count_received else 0,
                "activities": activities,
                "ym_label": ym_to_label(ym),
                "ym": ym,
                "ingresos_sin_iva": ingresos_sin_iva,
                "gastos_sin_iva": gastos_sin_iva,
                "iva_recibido_neto": iva_recibido_neto,
                "iva_retenciones": iva_retenciones,
                "iva_pagado": iva_pagado,
            },
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)
    # DEV bypass: allow opening the portal in Safari without a real token
    if DEV_MODE and (not token or token == DEV_TOKEN):
        return {
            "id": -1,
            "alias": "Contaneta",
            "rfc": "XIA190128J61",
            "regimen_fiscal": None,
            "facturapi_org_id": None,
            "active": 1,
        }

    conn = db()
    row = conn.execute(
        """
        SELECT
            i.id,
            i.rfc,
            i.razon_social,
            i.regimen_fiscal,
            i.active,
            i.facturapi_org_id,
            t.token
        FROM issuer_tokens t
        JOIN issuers i ON i.id = t.issuer_id
        WHERE t.token = ? AND t.active = 1 AND i.active = 1
        LIMIT 1
        """,
        (token,),
    ).fetchone()
    conn.close()

    if not row:
        raise ValueError("Token inválido o inactivo.")

    # Normaliza el issuer a dict y crea campos "compatibles" con el portal
    d = dict(row)
    regimen = (d.get("regimen_fiscal") or "").strip().upper()
    return {
        "id": d["id"],
        "rfc": d.get("rfc") or "",
        "alias": d.get("razon_social") or d.get("rfc") or "Emisor",
        "regimen_fiscal": regimen or None,
        "facturapi_org_id": d.get("facturapi_org_id"),  # puede no existir (None)
        "active": d.get("active", 1),
    }


@app.on_event("startup")
def _startup() -> None:
    """Ejecuta migraciones al arrancar. El schema se define únicamente por migraciones."""
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    try:
        apply_migrations(DB_PATH)
    except Exception as e:
        logging.exception("Migrations failed: %s", e)
        raise


# Etiquetas de régimen fiscal por usuario (RESICO, AE) -> código SAT para CFDI
REGIMEN_LABEL_TO_CODE = {"RESICO": "626", "AE": "612"}

# NOTA: Las funciones ensure_* fueron eliminadas porque están completamente cubiertas por
# migrations/001_baseline.sql. El schema se define únicamente por migraciones.


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    r = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ? LIMIT 1",
        (name,),
    ).fetchone()
    return bool(r)

def _has_column(conn: sqlite3.Connection, table: str, column: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r[1] == column for r in rows)


def _safe_update(conn: sqlite3.Connection, table: str, row_id: int, data: dict) -> None:
    """Update only columns that exist in current DB schema."""
    if not data:
        return
    cols = {r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    payload = {k: v for k, v in data.items() if k in cols}
    if not payload:
        return
    set_sql = ", ".join([f"{k} = ?" for k in payload.keys()])
    vals = list(payload.values()) + [row_id]
    conn.execute(f"UPDATE {table} SET {set_sql} WHERE id = ?", vals)


# ----------------------------
# Catalog helpers (catalogs.db)
# ----------------------------
def db_catalogs() -> sqlite3.Connection:
    if not os.path.exists(CATALOGS_DB):
        raise FileNotFoundError(f"No existe catalogs.db en: {CATALOGS_DB}")
    conn = sqlite3.connect(CATALOGS_DB)
    conn.row_factory = sqlite3.Row
    return conn


def _table_columns(con: sqlite3.Connection, table: str) -> set[str]:
    rows = con.execute(f"PRAGMA table_info({table})").fetchall()
    return {r[1] for r in rows}  # column name is index 1


def _pick_column(cols: set[str], candidates: list[str]) -> str:
    for c in candidates:
        if c in cols:
            return c
    raise KeyError(f"No se encontró ninguna columna de {candidates} en la tabla")


def _list_catalog(table: str, order_by: str | None = None) -> list[dict]:
    """
    Returns [{"key": "...", "label": "..."}] from a SAT catalog table.
    Tries to be resilient to column name differences.
    """
    con = db_catalogs()
    try:
        cols = _table_columns(con, table)
        key_col = _pick_column(cols, ["id", "clave", "key", "c_Clave"])
        label_col = _pick_column(cols, ["texto", "descripcion", "description", "value", "nombre"])
        ob = order_by or key_col
        rows = con.execute(
            f"SELECT {key_col} AS k, {label_col} AS v FROM {table} ORDER BY {ob}"
        ).fetchall()
        return [{"key": str(r["k"]), "label": str(r["v"])} for r in rows]
    finally:
        con.close()


def _search_catalog(table: str, q: str, limit: int = 20) -> list[dict]:
    """
    Search in big catalogs (ProdServ / Unidad).
    Returns [{"key": "...", "label": "..."}]
    """
    con = db_catalogs()
    try:
        cols = _table_columns(con, table)
        key_col = _pick_column(cols, ["id", "clave", "key"])
        label_col = _pick_column(cols, ["texto", "descripcion", "description", "value", "nombre"])

        rows = con.execute(
            f"""
            SELECT {key_col} AS k, {label_col} AS v
            FROM {table}
            WHERE ({label_col} LIKE ? OR {key_col} LIKE ?)
            LIMIT ?
            """,
            (f"%{q}%", f"%{q}%", int(limit)),
        ).fetchall()

        return [{"key": str(r["k"]), "label": str(r["v"])} for r in rows]
    finally:
        con.close()


# ----------------------------
# Form parsing
# ----------------------------
def parse_items_from_form(form) -> List[dict]:
    items: List[dict] = []

    idxs = set()
    for k in form.keys():
        m = re.match(r'^(qty|desc|key|price|iva|disc|unit)_(\d+)$', str(k))
        if m:
            idxs.add(int(m.group(2)))

    for i in sorted(idxs):
        qty = (form.get(f"qty_{i}") or "").strip()
        desc = (form.get(f"desc_{i}") or "").strip()
        key = (form.get(f"key_{i}") or "").strip()
        price = (form.get(f"price_{i}") or "").strip()
        iva = (form.get(f"iva_{i}") or "0.16").strip()
        unit = (form.get(f"unit_{i}") or "").strip()
        disc_pct = (form.get(f"disc_{i}") or "").strip()
        isr_ret = (form.get(f"isr_ret_{i}") or "0").strip()
        iva_ret = (form.get(f"iva_ret_{i}") or "0").strip()

        if not (qty or desc or key or price):
            continue

        if not (qty and desc and key and price and unit):
            raise ValueError(f"Concepto {i}: falta información.")

        disc_pct_num = 0.0
        if disc_pct.strip():
            disc_pct_num = float(disc_pct)
        disc_pct_num = max(0.0, min(100.0, disc_pct_num))

        qty_num = float(qty)
        price_base = float(price)
        iva_rate = float(iva)
        isr_ret_rate = float(isr_ret) if isr_ret else 0.0
        iva_ret_rate = float(iva_ret) if iva_ret else 0.0

        # Workaround to guarantee PDF totals match "base + IVA":
        # Send the price WITH taxes included and let Facturapi break it down.
        # If iva_rate is 0, price stays the same.
        price_to_send = price_base * (1.0 + iva_rate) if iva_rate else price_base

        line_base = qty_num * price_to_send
        line_discount = (disc_pct_num / 100.0) * line_base

        items.append(
            {
                "quantity": qty_num,
                # Facturapi supports discount at item level (amount, not %)
                "discount": float(line_discount),
                "product": {
                    "description": desc,
                    "product_key": key,
                    "price": float(price_to_send),
                    "tax_included": True,
                    "taxes": (
                        [{"type": "IVA", "rate": iva_rate}] +
                        ([{"type": "ISR", "rate": isr_ret_rate, "withholding": True}] if isr_ret_rate > 0 else []) +
                        ([{"type": "IVA", "rate": iva_ret_rate, "withholding": True}] if iva_ret_rate > 0 else [])
                    ),
                    "unit_key": unit,
                },
            }
        )

    if not items:
        raise ValueError("Debes capturar al menos un concepto completo.")

    return items


# ----------------------------
# Payments (CFDI P) form parser
# ----------------------------

def parse_payments_from_form(form) -> list[dict]:
    """Parse payment lines from the form for CFDI type P.

    Expected fields (repeated):
      pay_uuid_1, pay_amount_1, pay_uuid_2, pay_amount_2, ...
    Optional:
      pay_date (ISO datetime-local), pay_currency (defaults to MXN)

    Returns a list suitable for Facturapi `payments` payload.
    """
    idxs = set()
    for k in form.keys():
        m = re.match(r"^pay_uuid_(\d+)$", str(k))
        if m:
            idxs.add(int(m.group(1)))

    related_documents: list[dict] = []
    for i in sorted(idxs):
        uuid = (form.get(f"pay_uuid_{i}") or "").strip()
        amt = (form.get(f"pay_amount_{i}") or "").strip()
        if not (uuid or amt):
            continue
        if not (uuid and amt):
            raise ValueError(f"Pago {i}: falta UUID o monto.")
        amount = float(amt)
        if amount <= 0:
            raise ValueError(f"Pago {i}: monto inválido.")
        related_documents.append({"uuid": uuid, "amount": amount})

    if not related_documents:
        raise ValueError("En Pago (CFDI P) debes seleccionar al menos una factura (UUID) y monto.")

    pay_date = (form.get("pay_date") or "").strip() or None
    pay_currency = (form.get("pay_currency") or "").strip().upper() or None

    payment = {
        # Facturapi expects each payment to include related_documents.
        "related_documents": related_documents,
    }
    if pay_date:
        payment["date"] = pay_date
    if pay_currency:
        payment["currency"] = pay_currency

    return [payment]


# ----------------------------
# Pages
# ----------------------------
def _render_portal(
    request: Request,
    *,
    issuer: dict,
    template_name: str,
    active_page: str,
    title: str,
    extra: Optional[dict] = None,
    error: Optional[str] = None,
):
    has_nomina = False
    if issuer.get("id", 0) > 0:
        r = db_rows(
            "SELECT 1 FROM sat_cfdi WHERE issuer_id = ? AND direction = 'received' AND UPPER(TRIM(COALESCE(tipo_comprobante,''))) = 'N' LIMIT 1",
            (issuer["id"],),
        )
        has_nomina = bool(r)
    regimen_label = issuer.get("regimen_fiscal") or ""
    issuer_tax_code = REGIMEN_LABEL_TO_CODE.get(regimen_label) if regimen_label else ""
    payload = {
        "request": request,
        "token": "",  # Ya no se usa en URL; la sesión va por cookie
        "issuer_alias": issuer["alias"],
        "issuer_rfc": issuer["rfc"],
        "issuer_tax_system": issuer_tax_code,
        "issuer_regimen_label": regimen_label or "",
        "active_page": active_page,
        "title": title,
        "error": error,
        "has_nomina": has_nomina,
    }
    if extra:
        payload.update(extra)
    return templates.TemplateResponse(template_name, payload)


# Backwards compatible URL (legacy: ?token= sigue funcionando vía dependency en /portal/create)
@app.get("/facturar", response_class=HTMLResponse)
def facturar(request: Request, token: str = Query("")):
    if token:
        return RedirectResponse(url=f"/login?token={token}", status_code=302)
    return RedirectResponse(url="/portal/create", status_code=302)


@app.get("/portal/create", response_class=HTMLResponse)
def portal_create(
    request: Request,
    issuer: dict = Depends(get_portal_issuer),
    quote_id: Optional[int] = Query(None),
    customer_rfc: Optional[str] = Query(None),
    customer_legal_name: Optional[str] = Query(None),
    customer_zip: Optional[str] = Query(None),
    customer_tax_system: Optional[str] = Query(None),
    customer_email: Optional[str] = Query(None),
    concept_desc: Optional[str] = Query(None),
    concept_key: Optional[str] = Query(None),
    concept_unit: Optional[str] = Query(None),
    concept_price: Optional[str] = Query(None),
    concept_iva: Optional[str] = Query(None),
):
    """Generar factura; opcionalmente prellenar desde quote_id o desde query params."""
    customer_prefill = None
    concept_prefill = None
    quote_items = None

    if quote_id is not None:
        try:
            conn = db()
            row = conn.execute(
                "SELECT customer_rfc, customer_legal_name, customer_email FROM quotations WHERE issuer_id = ? AND id = ?",
                (issuer["id"], quote_id),
            ).fetchone()
            if row:
                r = dict(row)
                customer_prefill = {
                    "customer_rfc": (r.get("customer_rfc") or "").strip(),
                    "customer_legal_name": (r.get("customer_legal_name") or "").strip(),
                    "customer_zip": "",
                    "customer_tax_system": "",
                    "customer_email": (r.get("customer_email") or "").strip(),
                }
                items = conn.execute(
                    "SELECT description, quantity, unit_price, iva_rate FROM quotation_items WHERE quotation_id = ? ORDER BY sort_order, id",
                    (quote_id,),
                ).fetchall()
                conn.close()
                quote_items = [
                    {
                        "description": x["description"],
                        "quantity": float(x["quantity"] or 0),
                        "unit_price": float(x["unit_price"] or 0),
                        "iva_rate": float(x["iva_rate"] or 0.16),
                    }
                    for x in items
                ]
                if quote_items:
                    concept_prefill = {
                        "description": quote_items[0]["description"],
                        "product_key": "",
                        "unit_key": "E48",
                        "unit_price": str(quote_items[0]["unit_price"]),
                        "iva_rate": str(quote_items[0]["iva_rate"]),
                    }
            else:
                conn.close()
        except ValueError:
            pass

    if customer_prefill is None and (customer_rfc or customer_legal_name or customer_zip or customer_tax_system or customer_email):
        customer_prefill = {
            "customer_rfc": (customer_rfc or "").strip(),
            "customer_legal_name": (customer_legal_name or "").strip(),
            "customer_zip": (customer_zip or "").strip(),
            "customer_tax_system": (customer_tax_system or "").strip(),
            "customer_email": (customer_email or "").strip(),
        }
    if concept_prefill is None and (concept_desc or concept_key or concept_unit or concept_price or concept_iva):
        concept_prefill = {
            "description": (concept_desc or "").strip(),
            "product_key": (concept_key or "").strip(),
            "unit_key": (concept_unit or "").strip() or "E48",
            "unit_price": (concept_price or "").strip(),
            "iva_rate": (concept_iva or "0.16").strip(),
        }
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="form.html",
            active_page="create",
            title="Factura nueva",
            extra={
                "create_mode": "normal",
                "customer_prefill": customer_prefill,
                "concept_prefill": concept_prefill,
                "quote_items": quote_items,
            },
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/portal/create/quick", response_class=HTMLResponse)
def portal_create_quick(request: Request, issuer: dict = Depends(get_portal_issuer)):
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="form.html",
            active_page="create_quick",
            title="Factura rápida",
            extra={"create_mode": "quick"},
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/portal/create/multi", response_class=HTMLResponse)
def portal_create_multi(request: Request, issuer: dict = Depends(get_portal_issuer)):
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="form.html",
            active_page="create_multi",
            title="Factura múltiple",
            extra={"create_mode": "multi"},
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)

MESES_ES = (
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
)

def ym_now():
    return datetime.now().strftime("%Y-%m")

def ym_to_label(ym: str) -> str:
    """Convert 2026-01 to 'Enero 2026'."""
    try:
        y, m = ym.split("-")
        return f"{MESES_ES[int(m) - 1]} {y}"
    except (ValueError, IndexError):
        return ym

def shift_ym(ym: str, delta_months: int) -> str:
    y, m = ym.split("-")
    y = int(y); m = int(m)
    m += delta_months
    while m <= 0:
        m += 12
        y -= 1
    while m >= 13:
        m -= 12
        y += 1
    return f"{y:04d}-{m:02d}"


def _get_month_totals(issuer_id: int, ym: str, direction: str) -> dict:
    """Totales del mes para emitidas o recibidas: base (subtotal), IVA y retenciones."""
    conn = db()
    try:
        base_where = (
            "issuer_id = ? AND direction = ? AND fecha_emision IS NOT NULL AND substr(fecha_emision,1,7) = ?"
        )
        if direction == "issued":
            base_where += " AND (total IS NULL OR total >= 0.01)"
        else:
            base_where += " AND total IS NOT NULL AND total >= 0.01 AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')"
        params = (issuer_id, direction, ym)
        has_retenciones = _has_column(conn, "sat_cfdi", "retenciones")
        if has_retenciones:
            row = conn.execute(
                f"""
                SELECT
                    COALESCE(SUM(COALESCE(subtotal, total)), 0) AS total_base,
                    COALESCE(SUM(COALESCE(impuestos, 0)), 0) AS total_iva,
                    COALESCE(SUM(COALESCE(retenciones, 0)), 0) AS total_retenciones
                FROM sat_cfdi
                WHERE {base_where}
                """,
                params,
            ).fetchone()
        else:
            row = conn.execute(
                f"""
                SELECT
                    COALESCE(SUM(COALESCE(subtotal, total)), 0) AS total_base,
                    COALESCE(SUM(COALESCE(impuestos, 0)), 0) AS total_iva
                FROM sat_cfdi
                WHERE {base_where}
                """,
                params,
            ).fetchone()
        total_base = float(row[0] or 0) if row else 0.0
        total_iva = float(row[1] or 0) if row else 0.0
        total_retenciones = float(row[2] or 0) if (has_retenciones and row and len(row) >= 3) else 0.0
        return {
            "total_base": total_base,
            "total_iva": total_iva,
            "total_retenciones": total_retenciones,
            # Para emitidas: IVA cobrado neto = cobrado − retenciones (siempre usar en UI como "IVA recibido")
            "total_iva_neto": max(0.0, total_iva - total_retenciones) if direction == "issued" else total_iva,
        }
    finally:
        conn.close()


@app.get("/portal/invoices", response_class=HTMLResponse)
def portal_invoices(request: Request, issuer: dict = Depends(get_portal_issuer)):
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_invoices.html",
            active_page="issued",
            title="Mis facturas",
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/portal/invoices/issued", response_class=HTMLResponse)
def portal_invoices_issued(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
    try:
        issuer_id = issuer["id"]

        if not ym:
            ym = ym_now()

        rows = db_rows("""
            SELECT uuid, fecha_emision, rfc_receptor, nombre_receptor, total, moneda, status, xml_path,
                   serie, folio, concepto, forma_pago, metodo_pago, uso_cfdi, subtotal, descuento, impuestos,
                   COALESCE(retenciones, 0) AS retenciones,
                   tipo_comprobante, xml_status
            FROM sat_cfdi
            WHERE issuer_id = ?
              AND direction = 'issued'
              AND fecha_emision IS NOT NULL
              AND substr(fecha_emision,1,7) = ?
              AND (total IS NULL OR total >= 0.01)
            ORDER BY fecha_emision DESC
            LIMIT 300;
        """, (issuer_id, ym))

        months = db_rows("""
            SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
            FROM sat_cfdi
            WHERE issuer_id = ?
              AND direction = 'issued'
              AND fecha_emision IS NOT NULL
              AND (total IS NULL OR total >= 0.01)
            GROUP BY ym
            ORDER BY ym DESC;
        """, (issuer_id,))

        for m in months:
            m["label"] = ym_to_label(m["ym"])
        month_totals = _get_month_totals(issuer_id, ym, "issued")
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_issued.html",
            active_page="issued",
            title="Facturas emitidas",
            extra={
                "rows": rows,
                "ym": ym,
                "ym_label": ym_to_label(ym),
                "prev_ym": shift_ym(ym, -1),
                "next_ym": shift_ym(ym, +1),
                "months": months,
                "month_totals": month_totals,
            },
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/portal/invoices/received", response_class=HTMLResponse)
def portal_invoices_received(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
    try:
        issuer_id = issuer["id"]

        if not ym:
            ym = ym_now()

        rows = db_rows("""
            SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, total, moneda, status, xml_path,
                   serie, folio, concepto, forma_pago, metodo_pago, uso_cfdi, subtotal, descuento, impuestos,
                   COALESCE(retenciones, 0) AS retenciones,
                   tipo_comprobante, xml_status
            FROM sat_cfdi
            WHERE issuer_id = ?
              AND direction = 'received'
              AND fecha_emision IS NOT NULL
              AND substr(fecha_emision,1,7) = ?
              AND total IS NOT NULL AND total >= 0.01
              AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
            ORDER BY fecha_emision DESC
            LIMIT 300;
        """, (issuer_id, ym))

        months = db_rows("""
            SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
            FROM sat_cfdi
            WHERE issuer_id = ?
              AND direction = 'received'
              AND fecha_emision IS NOT NULL
              AND total IS NOT NULL AND total >= 0.01
              AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
            GROUP BY ym
            ORDER BY ym DESC;
        """, (issuer_id,))

        for m in months:
            m["label"] = ym_to_label(m["ym"])
        month_totals = _get_month_totals(issuer_id, ym, "received")
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_received.html",
            active_page="received",
            title="Facturas recibidas",
            extra={
                "rows": rows,
                "ym": ym,
                "ym_label": ym_to_label(ym),
                "prev_ym": shift_ym(ym, -1),
                "next_ym": shift_ym(ym, +1),
                "months": months,
                "month_totals": month_totals,
            },
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/portal/invoices/nomina", response_class=HTMLResponse)
def portal_invoices_nomina(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
    try:
        issuer_id = issuer["id"]

        if not ym:
            ym = ym_now()

        rows = db_rows("""
            SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, total, moneda, status, xml_path,
                   serie, folio, concepto, forma_pago, metodo_pago, uso_cfdi, subtotal, descuento, impuestos,
                   tipo_comprobante, xml_status
            FROM sat_cfdi
            WHERE issuer_id = ?
              AND direction = 'received'
              AND UPPER(TRIM(COALESCE(tipo_comprobante,''))) = 'N'
              AND fecha_emision IS NOT NULL
              AND substr(fecha_emision,1,7) = ?
            ORDER BY fecha_emision DESC
            LIMIT 300;
        """, (issuer_id, ym))

        months = db_rows("""
            SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
            FROM sat_cfdi
            WHERE issuer_id = ?
              AND direction = 'received'
              AND UPPER(TRIM(COALESCE(tipo_comprobante,''))) = 'N'
              AND fecha_emision IS NOT NULL
            GROUP BY ym
            ORDER BY ym DESC;
        """, (issuer_id,))

        for m in months:
            m["label"] = ym_to_label(m["ym"])
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_nomina.html",
            active_page="nomina",
            title="Nómina recibida",
            extra={
                "rows": rows,
                "ym": ym,
                "ym_label": ym_to_label(ym),
                "prev_ym": shift_ym(ym, -1),
                "next_ym": shift_ym(ym, +1),
                "months": months,
            },
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


# ----------------------------
# SAT XML viewer (from sat_sync storage)
# ----------------------------

def _safe_abs_path(path_like: str) -> str:
    """Resolve a stored path to an absolute path under BASE_DIR (prevent path traversal)."""
    if not path_like:
        raise ValueError("XML no disponible")

    # DB usually stores relative paths like: storage/xml/1/issued/2026/01/<uuid>.xml
    p = path_like
    if not os.path.isabs(p):
        p = os.path.join(BASE_DIR, p)

    abs_p = os.path.abspath(p)
    base = os.path.abspath(BASE_DIR)
    if not abs_p.startswith(base + os.sep):
        raise ValueError("Ruta XML inválida")
    return abs_p


@app.get("/portal/sat/xml/{uuid}")
def portal_sat_xml(uuid: str, issuer: dict = Depends(get_portal_issuer)):
    """Download/view the raw SAT XML stored by sat_sync for the current issuer. 404 si no existe."""
    conn = db()
    row = conn.execute(
        "SELECT xml_path, uuid FROM sat_cfdi WHERE issuer_id = ? AND uuid = ? LIMIT 1",
        (issuer["id"], (uuid or "").strip()),
    ).fetchone()
    conn.close()
    if not row or not row["xml_path"]:
        raise HTTPException(status_code=404, detail="XML no encontrado para este UUID")
    try:
        abs_path = _safe_abs_path(row["xml_path"])
    except ValueError:
        raise HTTPException(status_code=404, detail="Ruta XML inválida")
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="Archivo XML no existe en disco")
    with open(abs_path, "rb") as f:
        xml_bytes = f.read()
    return Response(
        content=xml_bytes,
        media_type="application/xml",
        headers={"Content-Disposition": f'inline; filename="{row["uuid"]}.xml"'},
    )


@app.get("/portal/sat/pdf/{uuid}")
def portal_sat_pdf(uuid: str, issuer: dict = Depends(get_portal_issuer), dl: int = 0):
    """Genera y devuelve un PDF detallado del CFDI a partir del XML. dl=1 fuerza descarga. 404 si no hay XML."""
    uuid_clean = (uuid or "").strip().split()[0] if uuid else ""
    if not uuid_clean:
        raise HTTPException(status_code=404, detail="UUID no válido")
    conn = db()
    row = conn.execute(
        "SELECT xml_path, uuid FROM sat_cfdi WHERE issuer_id = ? AND uuid = ? LIMIT 1",
        (issuer["id"], uuid_clean),
    ).fetchone()
    conn.close()
    if not row or not row["xml_path"]:
        raise HTTPException(status_code=404, detail="XML no encontrado para este UUID")
    try:
        abs_path = _safe_abs_path(row["xml_path"])
    except ValueError:
        raise HTTPException(status_code=404, detail="Ruta XML inválida")
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="Archivo XML no existe en disco")
    try:
        from cfdi_pdf import parse_cfdi_xml, build_pdf
        data = parse_cfdi_xml(abs_path)
        pdf_bytes = build_pdf(data)
    except Exception as e:
        err_msg = str(e)
        hint = " Instala dependencias: pip install -r requirements.txt" if "reportlab" in err_msg.lower() else ""
        return HTMLResponse(
            f'<!DOCTYPE html><html><head><meta charset="utf-8"><title>Error PDF</title></head>'
            f'<body id="pdf-error" style="margin:1rem;font-family:system-ui,sans-serif;">'
            f'<p id="pdf-error-msg">No se pudo generar el PDF: {err_msg}</p>'
            f'<p class="pdf-error-hint" style="color:#666;">{hint}</p></body></html>',
            status_code=500,
        )
    if not pdf_bytes:
        raise HTTPException(status_code=500, detail="La generación del PDF devolvió vacío")
    filename = f"cfdi-{uuid_clean[:8]}.pdf"
    disposition = "attachment" if dl else "inline"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )


def _get_cfdi_by_uuid(issuer_id: int, uuid: str, direction: str):
    """Obtiene un CFDI de sat_cfdi por (issuer_id, uuid). direction: 'issued' o 'received'."""
    conn = db()
    row = conn.execute(
        """
        SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, rfc_receptor, nombre_receptor,
               total, moneda, tipo_comprobante, status, xml_path, xml_status,
               serie, folio, forma_pago, metodo_pago, uso_cfdi, concepto,
               subtotal, descuento, impuestos, COALESCE(retenciones, 0) AS retenciones
        FROM sat_cfdi
        WHERE issuer_id = ? AND uuid = ? AND direction = ?
        LIMIT 1
        """,
        (issuer_id, (uuid or "").strip(), direction),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


@app.get("/portal/cfdi/issued/{uuid}", response_class=HTMLResponse)
def portal_cfdi_detail_issued(request: Request, uuid: str, issuer: dict = Depends(get_portal_issuer)):
    """Vista detalle de un CFDI emitido (por UUID)."""
    cfdi = _get_cfdi_by_uuid(issuer["id"], uuid, "issued")
    if not cfdi:
        raise HTTPException(status_code=404, detail="CFDI no encontrado")
    return _render_portal(
        request,
        issuer=issuer,
        template_name="portal_cfdi_detail.html",
        active_page="issued",
        title="Detalle CFDI emitido",
        extra={"cfdi": cfdi, "direction": "issued"},
    )


@app.get("/portal/cfdi/received/{uuid}", response_class=HTMLResponse)
def portal_cfdi_detail_received(request: Request, uuid: str, issuer: dict = Depends(get_portal_issuer)):
    """Vista detalle de un CFDI recibido (por UUID)."""
    cfdi = _get_cfdi_by_uuid(issuer["id"], uuid, "received")
    if not cfdi:
        raise HTTPException(status_code=404, detail="CFDI no encontrado")
    return _render_portal(
        request,
        issuer=issuer,
        template_name="portal_cfdi_detail.html",
        active_page="received",
        title="Detalle CFDI recibido",
        extra={"cfdi": cfdi, "direction": "received"},
    )


@app.get("/download/xml/{uuid}")
def download_xml(uuid: str, issuer: dict = Depends(get_portal_issuer)):
    """Descarga el XML del CFDI por UUID (issuer_id desde sesión). 404 si no existe."""
    try:
        conn = db()
        row = conn.execute(
            "SELECT xml_path, uuid FROM sat_cfdi WHERE issuer_id = ? AND uuid = ? LIMIT 1",
            (issuer["id"], (uuid or "").strip()),
        ).fetchone()
        conn.close()
        if not row or not row["xml_path"]:
            raise HTTPException(status_code=404, detail="XML no encontrado para este UUID")
        abs_path = _safe_abs_path(row["xml_path"])
        if not os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="Archivo XML no existe en disco")
        with open(abs_path, "rb") as f:
            xml_bytes = f.read()
        return Response(
            content=xml_bytes,
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="{row["uuid"]}.xml"'},
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/download/pdf/{uuid}")
def download_pdf(uuid: str, issuer: dict = Depends(get_portal_issuer), dl: int = 0):
    """Genera/sirve PDF del CFDI por UUID. 404 si no hay XML."""
    try:
        uuid_clean = (uuid or "").strip().split()[0] if uuid else ""
        if not uuid_clean:
            raise HTTPException(status_code=404, detail="UUID no válido")
        conn = db()
        row = conn.execute(
            "SELECT xml_path, uuid FROM sat_cfdi WHERE issuer_id = ? AND uuid = ? LIMIT 1",
            (issuer["id"], uuid_clean),
        ).fetchone()
        conn.close()
        if not row or not row["xml_path"]:
            raise HTTPException(status_code=404, detail="XML no encontrado; no se puede generar PDF")
        abs_path = _safe_abs_path(row["xml_path"])
        if not os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="Archivo XML no existe en disco")
        from cfdi_pdf import parse_cfdi_xml, build_pdf
        data = parse_cfdi_xml(abs_path)
        pdf_bytes = build_pdf(data)
        if not pdf_bytes:
            raise HTTPException(status_code=500, detail="Error al generar PDF")
        filename = f"cfdi-{uuid_clean[:8]}.pdf"
        disposition = "attachment" if dl else "inline"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'{disposition}; filename="{filename}"',
                "Content-Length": str(len(pdf_bytes)),
            },
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/portal/clients", response_class=HTMLResponse)
def portal_clients(request: Request, issuer: dict = Depends(get_portal_issuer)):
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_clients.html",
            active_page="clients",
            title="Clientes",
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/portal/providers", response_class=HTMLResponse)
def portal_providers(request: Request, issuer: dict = Depends(get_portal_issuer)):
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_providers.html",
            active_page="providers",
            title="Proveedores",
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/portal/products", response_class=HTMLResponse)
def portal_products(request: Request, issuer: dict = Depends(get_portal_issuer)):
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_products.html",
            active_page="products",
            title="Productos",
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


def _portal_quotations_impl(request: Request, issuer: dict):
    """Implementación de la página de cotizaciones."""
    try:
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_quotations.html",
            active_page="quotations",
            title="Cotizaciones",
        )
    except Exception as e:
        return HTMLResponse(
            f"<h3>Error</h3><p>No se pudo cargar la página de cotizaciones.</p><p><small>{str(e)}</small></p>",
            status_code=400,
        )


@app.get("/portal/summary", response_class=HTMLResponse)
def portal_summary(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
    try:
        issuer_id = issuer["id"]
        if not ym:
            ym = ym_now()
        tot_issued = _get_month_totals(issuer_id, ym, "issued")
        tot_received = _get_month_totals(issuer_id, ym, "received")
        ingresos_sin_iva = tot_issued["total_base"]
        gastos_sin_iva = tot_received["total_base"]
        iva_retenciones = tot_issued["total_retenciones"]
        # Siempre usar total_iva_neto para "IVA recibido" (cobrado − retenciones)
        iva_recibido_neto = tot_issued["total_iva_neto"]
        iva_pagado = tot_received["total_iva"]
        # Meses con datos (emitidos o recibidos) para el selector
        months_issued = db_rows(
            """
            SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
            FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
            GROUP BY ym ORDER BY ym DESC
            """,
            (issuer_id,),
        )
        months_received = db_rows(
            """
            SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
            FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
            GROUP BY ym ORDER BY ym DESC
            """,
            (issuer_id,),
        )
        ym_counts = {}
        for m in months_issued + months_received:
            ym_counts[m["ym"]] = ym_counts.get(m["ym"], 0) + m["n"]
        if ym not in ym_counts:
            ym_counts[ym] = 0
        months = [{"ym": y, "n": n, "label": ym_to_label(y)} for y, n in sorted(ym_counts.items(), reverse=True)]
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_summary.html",
            active_page="summary",
            title="Resumen",
            extra={
                "ym": ym,
                "ym_label": ym_to_label(ym),
                "prev_ym": shift_ym(ym, -1),
                "next_ym": shift_ym(ym, +1),
                "months": months,
                "ingresos_sin_iva": ingresos_sin_iva,
                "gastos_sin_iva": gastos_sin_iva,
                "iva_recibido_neto": iva_recibido_neto,
                "iva_retenciones": iva_retenciones,
                "iva_pagado": iva_pagado,
            },
        )
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.post("/submit", response_class=HTMLResponse)
async def submit(
    request: Request,
    issuer: dict = Depends(get_portal_issuer),
    customer_rfc: str = Form(...),
    customer_legal_name: str = Form(...),
    customer_zip: str = Form(...),
    customer_tax_system: str = Form(...),
    cfdi_use: str = Form(...),
    customer_email: str = Form(""),
    currency: str = Form("MXN"),
    exchange_rate: str = Form(""),
    payment_method: str = Form("PUE"),
    payment_form: str = Form(...),
    receiver_is_pm: str = Form(""),
    issuer_tax_system: str = Form(""),
):
    try:
        form = await request.form()
        export_code = (form.get("exportacion") or "01").strip()
        tipo_comprobante = (form.get("tipo_comprobante") or "I").strip().upper()

        # CFDI P (Pago) uses payments instead of items
        payments_payload: Optional[list[dict]] = None
        items: Optional[List[dict]] = None

        if tipo_comprobante == "P":
            # For CFDI P, enforce P01
            cfdi_use = "P01"
            payments_payload = parse_payments_from_form(form)
        else:
            items = parse_items_from_form(form)

        serie = (form.get("serie") or "").strip()
        folio = (form.get("folio") or "").strip()
        order_ref = (form.get("order_ref") or "").strip()
        issue_date = (form.get("issue_date") or "").strip()
        notes = (form.get("notes") or "").strip()
        save_customer = (form.get("save_customer") or "").strip() == "1"
        customer_alias = (form.get("customer_alias") or "").strip()

        folio_number: Optional[int] = None
        if folio:
            try:
                folio_number = int(folio)
            except Exception:
                raise ValueError("Folio inválido. Debe ser numérico.")

        # Basic validation for CFDI type in current MVP
        if tipo_comprobante not in ("I", "E", "P", "T", "N"):
            raise ValueError("Tipo de comprobante inválido.")

        exchange: Optional[float] = None
        if currency.upper() == "USD":
            if not exchange_rate.strip():
                raise ValueError("Captura tipo de cambio para USD.")
            exchange = float(exchange_rate.strip())

        conn = db()
        cur = conn.execute(
            """
            INSERT INTO invoices (
                issuer_id, currency, exchange_rate,
                payment_form, payment_method, cfdi_use,
                customer_rfc, customer_legal_name,
                customer_zip, customer_tax_system, customer_email
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                issuer["id"],
                currency.upper(),
                exchange,
                payment_form.strip(),
                payment_method.strip().upper(),
                cfdi_use.strip().upper(),
                customer_rfc.strip().upper(),
                customer_legal_name.strip(),
                customer_zip.strip(),
                customer_tax_system.strip(),
                customer_email.strip() or None,
            ),
        )

        invoice_local_id = cur.lastrowid

        _safe_update(
            conn,
            "invoices",
            invoice_local_id,
            {
                "export_code": export_code,
                "tipo_comprobante": tipo_comprobante,
                "series": serie or None,
                "folio_number": folio_number,
                "order_ref": order_ref or None,
                "issue_date": issue_date or None,
                "notes": notes or None,
            },
        )

        if items:
            for it in items:
                cols = {r[1] for r in conn.execute("PRAGMA table_info(invoice_items)").fetchall()}
                base_cols = ["invoice_id", "quantity", "description", "product_key", "unit_price", "iva_rate"]
                base_vals = [
                    invoice_local_id,
                    it["quantity"],
                    it["product"]["description"],
                    it["product"]["product_key"],
                    it["product"]["price"],
                    it["product"]["taxes"][0]["rate"],
                ]

                extra = {}
                if "unit_key" in cols:
                    extra["unit_key"] = it["product"].get("unit_key")
                if "discount" in cols:
                    extra["discount"] = it.get("discount", 0.0)

                insert_cols = base_cols + list(extra.keys())
                insert_vals = base_vals + list(extra.values())

                placeholders = ", ".join(["?"] * len(insert_cols))
                col_sql = ", ".join(insert_cols)
                conn.execute(
                    f"INSERT INTO invoice_items ({col_sql}) VALUES ({placeholders})",
                    tuple(insert_vals),
                )

        conn.commit()

        # Optionally save/update customer profile for future selection (solo si pasa validación)
        if save_customer:
            customer_errors = validate_customer(
                customer_rfc.strip().upper(),
                customer_legal_name.strip(),
                customer_zip.strip(),
                customer_tax_system.strip(),
                (customer_email.strip() or None),
            )
            if customer_errors:
                print(f"[submit] save_customer skipped (validación): {customer_errors}")
            else:
                try:
                    conn.execute(
                        """
                        INSERT INTO customer_profiles (
                            issuer_id, rfc, legal_name, zip, tax_system, email, alias, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                        ON CONFLICT(issuer_id, rfc) DO UPDATE SET
                            legal_name = excluded.legal_name,
                            zip = excluded.zip,
                            tax_system = excluded.tax_system,
                            email = excluded.email,
                            alias = excluded.alias,
                            updated_at = CURRENT_TIMESTAMP
                        """,
                        (
                            issuer["id"],
                            customer_rfc.strip().upper(),
                            customer_legal_name.strip(),
                            customer_zip.strip(),
                            customer_tax_system.strip(),
                            customer_email.strip() or None,
                            customer_alias or None,
                        ),
                    )
                    conn.commit()
                except Exception as e:
                    print(f"[submit] save_customer failed: {e}")

        payload = {
            "type": tipo_comprobante,
            "export": export_code or "01",
            "customer": {
                "legal_name": customer_legal_name.strip(),
                "email": customer_email.strip() or None,
                "tax_id": customer_rfc.strip().upper(),
                "tax_system": customer_tax_system.strip(),
                "address": {"zip": customer_zip.strip()},
            },
            **({"items": items} if items is not None else {}),
            **({"payments": payments_payload} if payments_payload is not None else {}),
            "use": cfdi_use.strip().upper(),
            "payment_form": payment_form.strip(),
            "payment_method": payment_method.strip().upper(),
            "currency": currency.upper(),
        }

        if serie:
            payload["series"] = serie
        if folio_number is not None:
            payload["folio_number"] = folio_number
        if issue_date:
            payload["date"] = issue_date
        if order_ref:
            payload["external_id"] = order_ref
        if notes:
            payload["conditions"] = notes

        if exchange is not None:
            payload["exchange"] = exchange

        # Prevent accidental timbrado when using DEV bypass token
        if issuer.get("facturapi_org_id") in (None, "", 0) or issuer.get("id") == -1:
            raise ValueError("DEV_MODE activo: token de prueba. Configura un token real/issuer para timbrar.")
        invoice = create_invoice(issuer["facturapi_org_id"], payload)

        fact_id = invoice.get("id")
        uuid = invoice.get("uuid")
        total = invoice.get("total")

        conn.execute(
            """
            UPDATE invoices
            SET facturapi_invoice_id = ?, uuid = ?, total = ?
            WHERE id = ?
            """,
            (fact_id, uuid, total, invoice_local_id),
        )

        # If CFDI P, store local relations (best-effort)
        if tipo_comprobante == "P" and payments_payload:
            try:
                if _table_exists(conn, "payment_relations"):
                    # Map related UUIDs to local invoice ids when possible
                    for p in payments_payload:
                        for rd in p.get("related_documents", []):
                            r_uuid = (rd.get("uuid") or "").strip()
                            r_amt = float(rd.get("amount") or 0)
                            if not r_uuid or r_amt <= 0:
                                continue
                            rel = conn.execute(
                                "SELECT id FROM invoices WHERE issuer_id = ? AND uuid = ? LIMIT 1",
                                (issuer["id"], r_uuid),
                            ).fetchone()
                            related_local_id = int(rel["id"]) if rel else None
                            if related_local_id:
                                conn.execute(
                                    """
                                    INSERT INTO payment_relations (payment_invoice_id, related_invoice_id, related_uuid, amount)
                                    VALUES (?, ?, ?, ?)
                                    """,
                                    (invoice_local_id, related_local_id, r_uuid, r_amt),
                                )
                    conn.commit()
            except Exception as e:
                print(f"[submit] storing payment_relations failed: {e}")

        conn.commit()
        conn.close()

        return templates.TemplateResponse(
            "success.html",
            {
                "request": request,
                "token": "",
                "facturapi_invoice_id": fact_id,
                "uuid": uuid,
                "total": total,
            },
        )

    except FacturapiError as fe:
        return HTMLResponse(f"<h3>Error Facturapi</h3><p>{str(fe)}</p>", status_code=400)
    except Exception as e:
        return HTMLResponse(f"<h3>Error</h3><p>{str(e)}</p>", status_code=400)


@app.get("/download/{fmt}/{invoice_id}")
def download(fmt: str, invoice_id: str, issuer: dict = Depends(get_portal_issuer)):
    fmt = fmt.lower()

    if fmt not in ("pdf", "xml", "zip"):
        return HTMLResponse("Formato inválido", status_code=400)

    try:

        if issuer.get("facturapi_org_id") in (None, "", 0) or issuer.get("id") == -1:
            raise ValueError("DEV_MODE activo: no hay issuer real para descargar. Usa token real.")
        blob = download_invoice(issuer["facturapi_org_id"], invoice_id, fmt)

        media = {"pdf": "application/pdf", "xml": "application/xml", "zip": "application/zip"}[fmt]
        filename = f"invoice_{invoice_id}.{fmt}"

        return Response(
            content=blob,
            media_type=media,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception as e:
        return HTMLResponse(f"Error descargando: {str(e)}", status_code=400)



@app.get("/api/customers")
def api_customers(issuer: dict = Depends(get_portal_issuer)):
    """Return saved customers for the issuer linked to session."""
    try:
        conn = db()
        rows = conn.execute(
            """
            SELECT id, rfc, legal_name, zip, tax_system, email, alias
            FROM customer_profiles
            WHERE issuer_id = ?
            ORDER BY COALESCE(alias, ''), rfc
            """,
            (issuer["id"],),
        ).fetchall()
        conn.close()

        return [
            {
                "id": r["id"],
                "rfc": r["rfc"],
                "legal_name": r["legal_name"],
                "zip": r["zip"],
                "tax_system": r["tax_system"],
                "email": r["email"],
                "alias": r["alias"],
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/customers/create")
def api_customers_create(payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    """Create or update a customer for the issuer linked to session."""
    try:
        rfc = (payload.get("rfc") or "").strip().upper()
        legal_name = (payload.get("legal_name") or "").strip()
        zip_val = (payload.get("zip") or "").strip() or ""
        tax_val = (payload.get("tax_system") or "").strip() or ""
        email = (payload.get("email") or "").strip() or None
        alias = (payload.get("alias") or "").strip() or None

        errors = validate_customer(rfc, legal_name, zip_val, tax_val, email)
        if errors:
            raise HTTPException(status_code=400, detail="; ".join(errors))

        conn = db()
        conn.execute(
            """
            INSERT INTO customer_profiles (issuer_id, rfc, legal_name, zip, tax_system, email, alias, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(issuer_id, rfc) DO UPDATE SET
                legal_name = excluded.legal_name,
                zip = excluded.zip,
                tax_system = excluded.tax_system,
                email = excluded.email,
                alias = excluded.alias,
                updated_at = CURRENT_TIMESTAMP
            """,
            (issuer["id"], rfc, legal_name, zip_val, tax_val, email, alias),
        )
        conn.commit()
        conn.close()
        return {"ok": True, "rfc": rfc}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/customers/delete")
def api_customers_delete(payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    """Delete a customer for the issuer linked to session (by RFC)."""
    try:
        rfc = (payload.get("rfc") or "").strip().upper()
        if not rfc:
            raise HTTPException(status_code=400, detail="RFC requerido")

        conn = db()
        cur = conn.execute(
            "DELETE FROM customer_profiles WHERE issuer_id = ? AND rfc = ?",
            (issuer["id"], rfc),
        )
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        return {"ok": True, "rfc": rfc}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/products")
def api_products(issuer: dict = Depends(get_portal_issuer)):
    """Return saved products for the issuer (conceptos reutilizables para facturación)."""
    try:
        rows = db_rows(
            """
            SELECT id, description, product_key, unit_key, unit_price, iva_rate, created_at
            FROM issuer_products
            WHERE issuer_id = ?
            ORDER BY description
            """,
            (issuer["id"],),
        )
        return [
            {
                "id": r["id"],
                "description": r["description"],
                "product_key": r["product_key"],
                "unit_key": r["unit_key"],
                "unit_price": float(r["unit_price"] or 0),
                "iva_rate": float(r["iva_rate"] or 0.16),
                "created_at": r["created_at"],
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/products/create")
def api_products_create(payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    """Create a product for the issuer (descripción, ClaveProdServ, unidad, precio, IVA)."""
    try:
        description = (payload.get("description") or "").strip()
        product_key_raw = (payload.get("product_key") or "").strip()
        # Permitir "12345678 — Descripción" desde la UI
        product_key = product_key_raw.split("—")[0].strip() if "—" in product_key_raw else product_key_raw
        unit_key = (payload.get("unit_key") or "").strip() or "E48"
        unit_price = payload.get("unit_price")
        iva_rate = payload.get("iva_rate", 0.16)
        try:
            unit_price = float(unit_price)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Precio unitario inválido")
        try:
            iva_rate = float(iva_rate)
        except (TypeError, ValueError):
            iva_rate = 0.16
        iva_rate = max(0, min(1, iva_rate))

        errors = validate_product(description, product_key_raw, unit_key, unit_price)
        if errors:
            raise HTTPException(status_code=400, detail="; ".join(errors))

        conn = db()
        conn.execute(
            """
            INSERT INTO issuer_products (issuer_id, description, product_key, unit_key, unit_price, iva_rate)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (issuer["id"], description, product_key, unit_key, unit_price, iva_rate),
        )
        conn.commit()
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.close()
        return {"ok": True, "id": rid}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ----------------------------
# Quotations (cotizaciones)
# ----------------------------

QUOTATION_STATUSES = ("draft", "sent", "accepted", "rejected", "converted", "expired")


@app.get("/api/quotations")
def api_quotations_list(issuer: dict = Depends(get_portal_issuer)):
    """List quotations for the issuer (session required)."""
    try:
        issuer_id = issuer["id"]
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))
    rows = db_rows(
        """
        SELECT q.id, q.folio, q.customer_rfc, q.customer_legal_name, q.customer_email,
               q.status, q.public_token, q.valid_until, q.notes, q.responded_at,
               q.created_at, q.updated_at,
               (SELECT COALESCE(SUM(
                   (qi.quantity * qi.unit_price) * (1 + COALESCE(qi.iva_rate, 0))
               ), 0) FROM quotation_items qi WHERE qi.quotation_id = q.id) AS total
        FROM quotations q
        WHERE q.issuer_id = ?
        ORDER BY q.created_at DESC
        """,
        (issuer_id,),
    )
    return [
        {
            "id": r["id"],
            "folio": r.get("folio"),
            "customer_rfc": r["customer_rfc"],
            "customer_legal_name": r["customer_legal_name"],
            "customer_email": r["customer_email"],
            "status": r["status"],
            "public_token": r["public_token"],
            "valid_until": r["valid_until"],
            "notes": r["notes"],
            "responded_at": r["responded_at"],
            "created_at": r["created_at"],
            "updated_at": r["updated_at"],
            "total": float(r["total"] or 0),
        }
        for r in rows
    ]


@app.post("/api/quotations/create")
def api_quotations_create(payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    """Create a quotation with items. Generates public_token for client link."""
    try:
        issuer_id = issuer["id"]

        customer_rfc = (payload.get("customer_rfc") or "").strip().upper()
        customer_legal_name = (payload.get("customer_legal_name") or "").strip()
        customer_email = (payload.get("customer_email") or "").strip() or None
        notes = (payload.get("notes") or "").strip() or None
        status = (payload.get("status") or "draft").strip().lower()
        if status not in QUOTATION_STATUSES:
            status = "draft"

        if not customer_legal_name:
            raise HTTPException(status_code=400, detail="Nombre del cliente es obligatorio")

        items = payload.get("items") or []
        if not items:
            raise HTTPException(status_code=400, detail="Agrega al menos un concepto a la cotización")

        public_token = secrets.token_urlsafe(32)
        iva_rate_quote = float(payload.get("iva_rate") or 0.16)
        currency = (payload.get("currency") or "MXN").strip() or "MXN"
        notes_default = (
            "Condiciones: Esta cotización tiene una vigencia de 30 días. "
            "Los precios están expresados en pesos mexicanos (MXN) e incluyen IVA según se indique. "
            "Para proceder, acepte esta cotización y nos pondremos en contacto."
        )
        notes = (payload.get("notes") or "").strip() or notes_default

        conn = db()
        year = datetime.now().strftime("%Y")
        prefix = f"Q-{year}-"
        next_num = conn.execute(
            """
            SELECT COALESCE(MAX(CAST(SUBSTR(folio, LENGTH(?) + 1) AS INTEGER)), 0) + 1
            FROM quotations WHERE issuer_id = ? AND (folio IS NOT NULL AND folio LIKE ?)
            """,
            (prefix, issuer_id, prefix + "%"),
        ).fetchone()[0]
        folio = f"{prefix}{next_num:04d}"
        sent_at = datetime.now().isoformat() if status == "sent" else None

        conn.execute(
            """
            INSERT INTO quotations (issuer_id, folio, customer_rfc, customer_legal_name, customer_email,
                status, public_token, notes, iva_rate, currency, sent_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (issuer_id, folio, customer_rfc or "", customer_legal_name, customer_email, status, public_token, notes, iva_rate_quote, currency, sent_at),
        )
        qid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        for idx, it in enumerate(items):
            desc = (it.get("description") or "").strip()
            if not desc:
                continue
            qty = float(it.get("quantity") or 1)
            unit_price = float(it.get("unit_price") or 0)
            iva_rate = float(it.get("iva_rate") or 0.16)
            product_id = it.get("product_id")
            if product_id is not None:
                try:
                    product_id = int(product_id)
                except (TypeError, ValueError):
                    product_id = None
            conn.execute(
                """
                INSERT INTO quotation_items (quotation_id, description, quantity, unit_price, iva_rate, product_id, sort_order)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (qid, desc, qty, unit_price, iva_rate, product_id, idx),
            )
        conn.commit()
        conn.close()
        return {"ok": True, "id": qid, "public_token": public_token}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/quotations/{qid}")
def api_quotations_get(qid: int, issuer: dict = Depends(get_portal_issuer)):
    """Get one quotation with items (for portal)."""
    issuer_id = issuer["id"]
    conn = db()
    row = conn.execute(
        "SELECT id, customer_rfc, customer_legal_name, customer_email, status, public_token, valid_until, notes, responded_at, created_at, updated_at FROM quotations WHERE issuer_id = ? AND id = ?",
        (issuer_id, qid),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    items = conn.execute(
        "SELECT id, description, quantity, unit_price, iva_rate, product_id, sort_order FROM quotation_items WHERE quotation_id = ? ORDER BY sort_order, id",
        (qid,),
    ).fetchall()
    conn.close()
    total = 0.0
    items_list = []
    for r in items:
        subtotal = float(r["quantity"] or 0) * float(r["unit_price"] or 0)
        iva = subtotal * float(r["iva_rate"] or 0)
        total += subtotal + iva
        items_list.append({
            "id": r["id"],
            "description": r["description"],
            "quantity": float(r["quantity"] or 0),
            "unit_price": float(r["unit_price"] or 0),
            "iva_rate": float(r["iva_rate"] or 0.16),
            "subtotal": subtotal,
            "total_line": subtotal + iva,
        })
    return {
        "id": dict(row)["id"],
        "customer_rfc": dict(row)["customer_rfc"],
        "customer_legal_name": dict(row)["customer_legal_name"],
        "customer_email": dict(row)["customer_email"],
        "status": dict(row)["status"],
        "public_token": dict(row)["public_token"],
        "valid_until": dict(row)["valid_until"],
        "notes": dict(row)["notes"],
        "responded_at": dict(row)["responded_at"],
        "created_at": dict(row)["created_at"],
        "updated_at": dict(row)["updated_at"],
        "items": items_list,
        "total": round(total, 2),
    }


@app.post("/api/quotations/update-status")
def api_quotations_update_status(payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    """Update quotation status (e.g. draft -> sent). Session required."""
    try:
        qid = payload.get("id")
        status = (payload.get("status") or "").strip().lower()
        if status not in QUOTATION_STATUSES:
            raise HTTPException(status_code=400, detail="Estatus inválido")
        if qid is None:
            raise HTTPException(status_code=400, detail="id requerido")
        try:
            qid = int(qid)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="id inválido")
        conn = db()
        cur = conn.execute(
            "UPDATE quotations SET status = ?, updated_at = datetime('now') WHERE issuer_id = ? AND id = ?",
            (status, issuer["id"], qid),
        )
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")
        return {"ok": True, "id": qid, "status": status}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/quotations/respond")
def api_quotations_respond(request: Request, payload: dict = Body(...)):
    """Public endpoint: client accepts or rejects a quotation (by public_token). No portal token."""
    public_token = (payload.get("public_token") or "").strip()
    action = (payload.get("action") or "").strip().lower()
    if not public_token:
        raise HTTPException(status_code=400, detail="Link inválido")
    if action not in ("accept", "reject", "aceptar", "rechazar"):
        raise HTTPException(status_code=400, detail="Acción inválida")
    status = "accepted" if action in ("accept", "aceptar") else "rejected"
    reason = (payload.get("rejection_reason") or "").strip() or None
    client_ip = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent")
    now = datetime.now().isoformat()
    conn = db()
    row = conn.execute(
        "SELECT id, status FROM quotations WHERE public_token = ?",
        (public_token,),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Cotización no encontrada o link expirado")
    if dict(row)["status"] not in ("draft", "sent"):
        conn.close()
        raise HTTPException(status_code=400, detail="Esta cotización ya fue respondida")
    qid = dict(row)["id"]
    if status == "accepted":
        conn.execute(
            """UPDATE quotations SET status = ?, responded_at = datetime('now'), updated_at = datetime('now'),
               accepted_at = ?, decision_ip = ?, decision_user_agent = ? WHERE id = ?""",
            (status, now, client_ip, user_agent, qid),
        )
    else:
        conn.execute(
            """UPDATE quotations SET status = ?, responded_at = datetime('now'), updated_at = datetime('now'),
               rejected_at = ?, decision_ip = ?, decision_user_agent = ?, rejection_reason = ? WHERE id = ?""",
            (status, now, client_ip, user_agent, reason, qid),
        )
    conn.commit()
    conn.close()
    return {"ok": True, "status": status}


def _get_quotation_by_public_token(public_token: str) -> Optional[dict]:
    """Load quotation + issuer + items by public_token. Returns None if not found."""
    conn = db()
    row = conn.execute(
        """
        SELECT q.id, q.issuer_id, q.folio, q.customer_rfc, q.customer_legal_name, q.customer_email,
               q.status, q.public_token, q.notes, q.responded_at, q.created_at, q.iva_rate AS quote_iva_rate,
               q.currency, q.rejection_reason,
               i.razon_social AS issuer_name, i.rfc AS issuer_rfc
        FROM quotations q
        JOIN issuers i ON i.id = q.issuer_id
        WHERE q.public_token = ?
        """,
        (public_token,),
    ).fetchone()
    if not row:
        conn.close()
        return None
    d = dict(row)
    items = conn.execute(
        "SELECT description, quantity, unit_price, iva_rate FROM quotation_items WHERE quotation_id = ? ORDER BY sort_order, id",
        (d["id"],),
    ).fetchall()
    subtotal_sum = 0.0
    items_list = []
    for r in items:
        line_sub = float(r["quantity"] or 0) * float(r["unit_price"] or 0)
        iva_rate = float(r["iva_rate"]) if r.get("iva_rate") is not None else float(d.get("quote_iva_rate") or 0.16)
        iva_line = line_sub * iva_rate
        subtotal_sum += line_sub
        items_list.append({
            "description": r["description"],
            "quantity": float(r["quantity"] or 0),
            "unit_price": float(r["unit_price"] or 0),
            "iva_rate": iva_rate,
            "subtotal": line_sub,
            "total_line": round(line_sub + iva_line, 2),
        })
    conn.close()
    iva_total = sum((it["subtotal"] * it["iva_rate"]) for it in items_list)
    d["items"] = items_list
    d["subtotal"] = round(subtotal_sum, 2)
    d["iva_total"] = round(iva_total, 2)
    d["total"] = round(subtotal_sum + iva_total, 2)
    d["issuer_name"] = d.get("issuer_name") or d.get("issuer_rfc") or "Emisor"
    return d


@app.get("/q/{public_token}", response_class=HTMLResponse)
@app.get("/public/cotizacion/{public_token}", response_class=HTMLResponse)
def public_quotation_view(request: Request, public_token: str):
    """Vista pública: el cliente ve la cotización y puede aceptar o rechazar. /q/{token} es alias corto."""
    quote = _get_quotation_by_public_token(public_token)
    if not quote:
        return HTMLResponse(
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>No encontrada</title></head><body><p>Cotización no encontrada o link expirado.</p></body></html>",
            status_code=404,
        )
    if quote["status"] not in ("draft", "sent"):
        return templates.TemplateResponse(
            "public_quotation_responded.html",
            {"request": request, "quotation": quote},
        )
    return templates.TemplateResponse(
        "public_quotation.html",
        {"request": request, "quotation": quote},
    )


@app.post("/public/cotizacion/respond", response_class=HTMLResponse)
def public_quotation_respond(
    request: Request,
    public_token: str = Form(alias="public_token", default=""),
    action: str = Form(default=""),
    rejection_reason: str = Form(default=""),
):
    """Form POST: cliente acepta o rechaza. Guarda IP, user-agent y motivo de rechazo."""
    token = (public_token or "").strip()
    act = (action or request.form.get("action") or "").strip().lower()
    if not token:
        return HTMLResponse("<p>Link inválido.</p>", status_code=400)
    if act not in ("accept", "reject", "aceptar", "rechazar"):
        return HTMLResponse("<p>Acción inválida.</p>", status_code=400)
    status = "accepted" if act in ("accept", "aceptar") else "rejected"
    reason = (rejection_reason or request.form.get("rejection_reason") or "").strip() or None
    client_ip = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent")
    now = datetime.now().isoformat()
    conn = db()
    row = conn.execute("SELECT id, status FROM quotations WHERE public_token = ?", (token,)).fetchone()
    if not row:
        conn.close()
        return HTMLResponse("<p>Cotización no encontrada.</p>", status_code=404)
    if dict(row)["status"] not in ("draft", "sent"):
        conn.close()
        quote = _get_quotation_by_public_token(token)
        return templates.TemplateResponse(
            "public_quotation_responded.html",
            {"request": request, "quotation": quote or {}},
        )
    qid = dict(row)["id"]
    if status == "accepted":
        conn.execute(
            """UPDATE quotations SET status = ?, responded_at = datetime('now'), updated_at = datetime('now'),
               accepted_at = ?, decision_ip = ?, decision_user_agent = ? WHERE id = ?""",
            (status, now, client_ip, user_agent, qid),
        )
    else:
        conn.execute(
            """UPDATE quotations SET status = ?, responded_at = datetime('now'), updated_at = datetime('now'),
               rejected_at = ?, decision_ip = ?, decision_user_agent = ?, rejection_reason = ? WHERE id = ?""",
            (status, now, client_ip, user_agent, reason, qid),
        )
    conn.commit()
    conn.close()
    quote = _get_quotation_by_public_token(token)
    return templates.TemplateResponse(
        "public_quotation_thanks.html",
        {"request": request, "quotation": quote, "action": status},
    )


def _build_quotation_pdf(quote: dict) -> bytes:
    """Genera PDF de cotización en español (ReportLab): encabezado, datos, tabla, totales, términos."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="QuoteTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
    body_style = ParagraphStyle(name="QuoteBody", parent=styles["Normal"], fontSize=10, spaceAfter=4)

    issuer_name = (quote.get("issuer_name") or "Cotización").replace("<", " ").replace(">", " ")
    folio = quote.get("folio") or f"Q-{quote.get('id', '')}"
    fecha = (quote.get("created_at") or "")[:10] if quote.get("created_at") else ""
    if fecha:
        try:
            y, m, d = fecha.split("-")
            fecha = f"{d}/{m}/{y}"
        except Exception:
            pass
    vigencia = "30 días"
    cliente = (quote.get("customer_legal_name") or "—").replace("<", " ").replace(">", " ")
    currency = quote.get("currency") or "MXN"

    story = [
        Paragraph(f"<b>{issuer_name}</b>", title_style),
        Paragraph("Cotización", body_style),
        Spacer(1, 12),
        Paragraph(f"<b>Folio:</b> {folio} &nbsp;&nbsp; <b>Fecha:</b> {fecha} &nbsp;&nbsp; <b>Vigencia:</b> {vigencia}", body_style),
        Paragraph(f"<b>Cliente:</b> {cliente}", body_style),
        Spacer(1, 14),
    ]

    headers = ["Descripción", "Cantidad", "Precio unitario", "Importe"]
    data = [headers]
    for it in quote.get("items") or []:
        desc = (it.get("description") or "—")[:50] + ("…" if len(str(it.get("description") or "")) > 50 else "")
        qty = float(it.get("quantity") or 0)
        pu = float(it.get("unit_price") or 0)
        importe = qty * pu
        data.append([desc, f"{qty:,.2f}", f"${pu:,.2f}", f"${importe:,.2f}"])
    subtotal = quote.get("subtotal") or sum(float(it.get("quantity") or 0) * float(it.get("unit_price") or 0) for it in quote.get("items") or [])
    if isinstance(subtotal, (int, float)):
        iva_val = quote.get("iva_total", 0)
        total_val = quote.get("total") or (subtotal + iva_val)
    else:
        iva_val = quote.get("iva_total", 0)
        total_val = quote.get("total", 0)
    data.append(["", "", "Subtotal", f"${float(subtotal):,.2f}"])
    data.append(["", "", "IVA", f"${float(iva_val):,.2f}"])
    data.append(["", "", "Total", f"${float(total_val):,.2f} {currency}"])

    t = Table(data, colWidths=[3.2 * inch, 0.8 * inch, 1.0 * inch, 1.0 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8e8e8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -4), 0.5, colors.grey),
        ("LINEABOVE", (0, -3), (-1, -1), 1, colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -4), [colors.white, colors.HexColor("#f8f8f8")]),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    terms = (quote.get("notes") or "").strip() or (
        "Condiciones: Esta cotización tiene una vigencia de 30 días. "
        "Los precios están expresados en pesos mexicanos (MXN) e incluyen IVA según se indique."
    )
    story.append(Paragraph("<b>Términos y condiciones</b>", body_style))
    story.append(Paragraph(terms.replace("\n", "<br/>")[:800], body_style))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Para proceder, acepte esta cotización a través del enlace proporcionado.", body_style))
    doc.build(story)
    return buf.getvalue()


@app.get("/portal/quotations/{qid}/pdf")
def portal_quotation_pdf(qid: int, issuer: dict = Depends(get_portal_issuer), download: str = Query("0", alias="download")):
    """PDF de la cotización (portal, requiere sesión). Por defecto inline; ?download=1 para descargar."""
    conn = db()
    row = conn.execute(
        "SELECT id, public_token FROM quotations WHERE issuer_id = ? AND id = ?",
        (issuer["id"], qid),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    conn.close()
    quote = _get_quotation_by_public_token(dict(row)["public_token"])
    if not quote:
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    try:
        pdf_bytes = _build_quotation_pdf(quote)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al generar el PDF: {str(e)}",
        )
    filename = f"cotizacion-{quote.get('folio', qid)}.pdf"
    disposition = "attachment" if (download and download.lower() in ("1", "true", "yes")) else "inline"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )


@app.get("/q/{public_token}/pdf")
def public_quotation_pdf(public_token: str):
    """PDF de cotización por link público (solo lectura)."""
    quote = _get_quotation_by_public_token(public_token)
    if not quote:
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    pdf_bytes = _build_quotation_pdf(quote)
    filename = f"cotizacion-{quote.get('folio', '')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )


@app.get("/portal/quotations/{qid}", response_class=HTMLResponse)
def portal_quotation_detail(request: Request, qid: int, issuer: dict = Depends(get_portal_issuer)):
    """Detalle de cotización: datos, link, PDF, botón Facturar si está aceptada."""
    issuer_id = issuer["id"]
    conn = db()
    row = conn.execute(
        """SELECT id, folio, customer_rfc, customer_legal_name, customer_email, status, public_token,
                  notes, responded_at, created_at, updated_at, rejection_reason
           FROM quotations WHERE issuer_id = ? AND id = ?""",
        (issuer_id, qid),
    ).fetchone()
    if not row:
        conn.close()
        return HTMLResponse("<p>Cotización no encontrada.</p>", status_code=404)
    items = conn.execute(
        "SELECT description, quantity, unit_price, iva_rate FROM quotation_items WHERE quotation_id = ? ORDER BY sort_order, id",
        (qid,),
    ).fetchall()
    conn.close()
    d = dict(row)
    subtotal = 0.0
    items_list = []
    for r in items:
        line_sub = float(r["quantity"] or 0) * float(r["unit_price"] or 0)
        iva = line_sub * float(r["iva_rate"] or 0.16)
        subtotal += line_sub
        items_list.append({
            "description": r["description"],
            "quantity": float(r["quantity"] or 0),
            "unit_price": float(r["unit_price"] or 0),
            "subtotal": line_sub,
            "total_line": round(line_sub + iva, 2),
        })
    iva_total = sum(
        it["subtotal"] * float(items[i].get("iva_rate") or 0.16)
        for i, it in enumerate(items_list)
    )
    d["items"] = items_list
    d["subtotal"] = round(subtotal, 2)
    d["iva_total"] = round(iva_total, 2)
    d["total"] = round(subtotal + iva_total, 2)
    return _render_portal(
        request,
        issuer=issuer,
        template_name="quote_detail.html",
        active_page="quotations",
        title="Cotización",
        extra={"quote": d},
    )


@app.get("/api/provider-invoices")
@app.get("/api/providers/invoices")
def api_provider_invoices(issuer: dict = Depends(get_portal_issuer), rfc: str = ""):
    """Lista de facturas recibidas de un proveedor (por RFC emisor) para el mini panel."""
    try:
        iid = issuer["id"]
        rfc_norm = (rfc or "").strip().upper()
        if not rfc_norm:
            return []
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))

    try:
        rows = db_rows(
            """
            SELECT uuid, fecha_emision, total, moneda, status, xml_path, concepto
            FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'received'
              AND UPPER(TRIM(COALESCE(rfc_emisor,''))) = ?
              AND (tipo_comprobante IS NULL OR UPPER(TRIM(COALESCE(tipo_comprobante,''))) != 'N')
              AND total IS NOT NULL AND total >= 0.01
            ORDER BY fecha_emision DESC
            LIMIT 100
            """,
            (iid, rfc_norm),
        )
        return [
            {
                "uuid": r.get("uuid"),
                "fecha_emision": r.get("fecha_emision"),
                "concepto": (r.get("concepto") or "")[:80] + ("…" if len(r.get("concepto") or "") > 80 else ""),
                "total": float(r.get("total") or 0),
                "moneda": r.get("moneda") or "MXN",
                "status": r.get("status"),
                "has_pdf": bool(r.get("xml_path")),
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def _provider_report_rows(issuer_id: int, rfc_norm: str):
    """Lista de facturas recibidas del proveedor con columnas completas para el reporte."""
    return db_rows(
        """
        SELECT uuid, fecha_emision, nombre_emisor, serie, folio, concepto,
               subtotal, descuento, impuestos, total, moneda, status,
               forma_pago, metodo_pago, uso_cfdi
        FROM sat_cfdi
        WHERE issuer_id = ? AND direction = 'received'
          AND UPPER(TRIM(COALESCE(rfc_emisor,''))) = ?
          AND (tipo_comprobante IS NULL OR UPPER(TRIM(COALESCE(tipo_comprobante,''))) != 'N')
          AND total IS NOT NULL AND total >= 0.01
        ORDER BY fecha_emision DESC
        """,
        (issuer_id, rfc_norm),
    )


def _build_provider_report_pdf(issuer: dict, provider_name: str, rows: list) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="ReportTitle", parent=styles["Heading1"], fontSize=14, spaceAfter=12)
    body_style = ParagraphStyle(name="ReportBody", parent=styles["Normal"], fontSize=10, spaceAfter=6)
    receptor_name = (issuer.get("alias") or issuer.get("rfc") or "Receptor").replace("<", " ").replace(">", " ")
    provider_safe = (provider_name or "Proveedor").replace("<", " ").replace(">", " ")

    story = [
        Paragraph(f"<b>Facturas recibidas de</b> {provider_safe}", title_style),
        Paragraph(f"<b>Receptor:</b> {receptor_name} — RFC: {issuer.get('rfc') or '—'}", body_style),
        Spacer(1, 14),
    ]
    # Tabla: Fecha, UUID, Concepto, Subtotal, Descuento, Impuestos, Total, Moneda, Estado (sin Serie ni Folio)
    headers = ["Fecha", "UUID", "Concepto", "Subtotal", "Descuento", "Impuestos", "Total", "Moneda", "Estado"]
    data = [headers]
    # Concepto truncado para que no se sobreponga a la columna Subtotal (corta en vez de desbordar)
    max_concepto_len = 18
    for r in rows:
        fecha = (r.get("fecha_emision") or "")[:10] if r.get("fecha_emision") else "—"
        uuid_short = (r.get("uuid") or "—")[:8] + "…" if r.get("uuid") and len(r.get("uuid", "")) > 8 else (r.get("uuid") or "—")
        raw_concepto = str(r.get("concepto") or "—")
        concepto = (raw_concepto[:max_concepto_len] + "…") if len(raw_concepto) > max_concepto_len else raw_concepto
        data.append([
            fecha,
            uuid_short,
            concepto,
            f"{float(r.get('subtotal') or 0):,.2f}",
            f"{float(r.get('descuento') or 0):,.2f}",
            f"{float(r.get('impuestos') or 0):,.2f}",
            f"{float(r.get('total') or 0):,.2f}",
            str(r.get("moneda") or "MXN"),
            "Vigente" if r.get("status") == "1" else ("Cancelada" if r.get("status") == "0" else "—"),
        ])
    t = Table(data, colWidths=[inch * 0.9, inch * 1.0, inch * 1.4, 0.7 * inch, 0.6 * inch, 0.6 * inch, 0.7 * inch, 0.5 * inch, 0.6 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8e8e8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (2, -1), "LEFT"),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8f8")]),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def _build_provider_report_xlsx(issuer: dict, provider_name: str, rows: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas recibidas"
    receptor_name = issuer.get("alias") or issuer.get("rfc") or "Receptor"
    provider_safe = provider_name or "Proveedor"
    ws["A1"] = f"Facturas recibidas de {provider_safe}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Receptor: {receptor_name} — RFC: {issuer.get('rfc') or '—'}"
    ws["A2"].font = Font(size=10)
    headers = ["Fecha", "UUID", "Concepto", "Subtotal", "Descuento", "Impuestos", "Total", "Moneda", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
    for i, r in enumerate(rows, 5):
        fecha = (r.get("fecha_emision") or "")[:10] if r.get("fecha_emision") else "—"
        ws.cell(row=i, column=1, value=fecha)
        ws.cell(row=i, column=2, value=r.get("uuid") or "—")
        ws.cell(row=i, column=3, value=(r.get("concepto") or "—")[:200])
        ws.cell(row=i, column=4, value=float(r.get("subtotal") or 0))
        ws.cell(row=i, column=5, value=float(r.get("descuento") or 0))
        ws.cell(row=i, column=6, value=float(r.get("impuestos") or 0))
        ws.cell(row=i, column=7, value=float(r.get("total") or 0))
        ws.cell(row=i, column=8, value=r.get("moneda") or "MXN")
        ws.cell(row=i, column=9, value="Vigente" if r.get("status") == "1" else ("Cancelada" if r.get("status") == "0" else "—"))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/api/provider-invoices/report")
def api_provider_invoices_report(
    issuer: dict = Depends(get_portal_issuer),
    rfc: str = Query(...),
    format: str = Query("pdf", alias="format"),
):
    """Reporte detallado: facturas recibidas de un proveedor en PDF o Excel (nombre emisor, receptor, tabla)."""
    if format not in ("pdf", "xlsx"):
        raise HTTPException(status_code=400, detail="format debe ser pdf o xlsx")
    try:
        rfc_norm = (rfc or "").strip().upper()
        if not rfc_norm:
            raise HTTPException(status_code=400, detail="RFC de proveedor requerido")
        rows = _provider_report_rows(issuer["id"], rfc_norm)
        provider_name = (rows[0].get("nombre_emisor") or "").strip() if rows else ""
        if not provider_name:
            provider_name = rfc_norm
        if format == "pdf":
            content = _build_provider_report_pdf(issuer, provider_name, rows)
            filename = f"facturas-recibidas-{rfc_norm[:8]}.pdf"
            media_type = "application/pdf"
        else:
            content = _build_provider_report_xlsx(issuer, provider_name, rows)
            filename = f"facturas-recibidas-{rfc_norm[:8]}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return Response(
            content=content,
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(content)),
            },
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/providers")
def api_providers(issuer: dict = Depends(get_portal_issuer)):
    """Return providers: combinación de supplier_profiles (guardados) + sat_cfdi (facturas recibidas)."""
    try:
        iid = issuer["id"]

        # 1) Proveedores guardados manualmente (supplier_profiles)
        saved = {}
        conn = db()
        try:
            if _table_exists(conn, "supplier_profiles"):
                for r in db_rows(
                    "SELECT rfc, legal_name, email, alias FROM supplier_profiles WHERE issuer_id = ?",
                    (iid,),
                ):
                    rfc_norm = (r["rfc"] or "").strip().upper()
                    if rfc_norm:
                        saved[rfc_norm] = {
                            "rfc": rfc_norm,
                            "legal_name": r["legal_name"] or "",
                            "email": r.get("email"),
                            "alias": r.get("alias"),
                            "facturas_count": 0,
                            "total_recibido": 0.0,
                            "source": "saved",
                        }
        finally:
            conn.close()

        # 2) Proveedores desde sat_cfdi (facturas recibidas)
        from_sat = db_rows(
            """
            SELECT
                UPPER(TRIM(rfc_emisor)) AS rfc,
                MAX(nombre_emisor) AS legal_name,
                COUNT(*) AS facturas_count,
                COALESCE(SUM(total), 0) AS total_recibido
            FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'received'
              AND rfc_emisor IS NOT NULL AND TRIM(rfc_emisor) != ''
              AND total IS NOT NULL AND total >= 0.01
              AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
            GROUP BY UPPER(TRIM(rfc_emisor))
            """,
            (iid,),
        )

        for r in from_sat:
            rfc = (r["rfc"] or "").strip()
            if not rfc:
                continue
            if rfc in saved:
                saved[rfc]["facturas_count"] = r["facturas_count"]
                saved[rfc]["total_recibido"] = float(r["total_recibido"] or 0)
                saved[rfc]["source"] = "both"
            else:
                saved[rfc] = {
                    "rfc": rfc,
                    "legal_name": r["legal_name"] or "",
                    "email": None,
                    "alias": None,
                    "facturas_count": r["facturas_count"],
                    "total_recibido": float(r["total_recibido"] or 0),
                    "source": "sat",
                }

        # Ordenar: facturas_count DESC, total_recibido DESC, luego por alias/rfc
        out = sorted(
            saved.values(),
            key=lambda x: (-x["facturas_count"], -x["total_recibido"], (x.get("alias") or x["rfc"]).lower()),
        )
        return out
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/providers/create")
def api_providers_create(payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    """Crear o actualizar un proveedor para el issuer (guardar manualmente)."""
    try:
        rfc = (payload.get("rfc") or "").strip().upper()
        legal_name = (payload.get("legal_name") or "").strip()
        if not rfc or not legal_name:
            raise HTTPException(status_code=400, detail="RFC y razón social son obligatorios")

        zip_val = (payload.get("zip") or "").strip() or None
        tax_val = (payload.get("tax_system") or "").strip() or None
        email = (payload.get("email") or "").strip() or None
        alias = (payload.get("alias") or "").strip() or None

        conn = db()
        conn.execute(
            """
            INSERT INTO supplier_profiles (issuer_id, rfc, legal_name, zip, tax_system, email, alias, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(issuer_id, rfc) DO UPDATE SET
                legal_name = excluded.legal_name,
                zip = excluded.zip,
                tax_system = excluded.tax_system,
                email = excluded.email,
                alias = excluded.alias,
                updated_at = CURRENT_TIMESTAMP
            """,
            (issuer["id"], rfc, legal_name, zip_val, tax_val, email, alias),
        )
        conn.commit()
        conn.close()
        return {"ok": True, "rfc": rfc}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ----------------------------
# API endpoint: pending PPD invoices for payments (CFDI P)
# ----------------------------

@app.get("/api/invoices/pending")
def api_pending_invoices(issuer: dict = Depends(get_portal_issuer)):
    """Return invoices for this issuer that are likely pending payment (PPD)."""
    try:
        conn = db()
        cols = {r[1] for r in conn.execute("PRAGMA table_info(invoices)").fetchall()}

        where = ["issuer_id = ?", "uuid IS NOT NULL", "payment_method = 'PPD'"]
        params: list = [issuer["id"]]

        # If a status column exists, exclude canceled
        if "status" in cols:
            where.append("COALESCE(status,'') != 'canceled'")
        if "cancelled" in cols:
            where.append("COALESCE(cancelled,0) = 0")

        rows = conn.execute(
            f"""
            SELECT id, uuid, total, customer_legal_name, customer_rfc, issue_date, created_at
            FROM invoices
            WHERE {' AND '.join(where)}
            ORDER BY COALESCE(issue_date, created_at) DESC
            LIMIT 200
            """,
            tuple(params),
        ).fetchall()

        conn.close()

        out = []
        for r in rows:
            out.append(
                {
                    "id": r["id"],
                    "uuid": r["uuid"],
                    "total": r["total"],
                    "customer_legal_name": r["customer_legal_name"],
                    "customer_rfc": r["customer_rfc"],
                    "date": r["issue_date"] or r["created_at"],
                }
            )
        return out
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ----------------------------
# SAT catalog endpoints (PRO) - CFDI 4.0 tables
# ----------------------------
@app.get("/api/catalogs/forma_pago")
def api_forma_pago():
    try:
        return _list_catalog("cfdi_40_formas_pago")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/catalogs/metodo_pago")
def api_metodo_pago():
    try:
        return _list_catalog("cfdi_40_metodos_pago")
    except Exception:
        return [
            {"key": "PUE", "label": "Pago en una sola exhibición"},
            {"key": "PPD", "label": "Pago en parcialidades o diferido"},
        ]


@app.get("/api/catalogs/uso_cfdi")
def api_uso_cfdi():
    try:
        return _list_catalog("cfdi_40_usos_cfdi")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/catalogs/regimen_fiscal")
def api_regimen_fiscal():
    try:
        return _list_catalog("cfdi_40_regimenes_fiscales")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/catalogs/moneda")
def api_moneda():
    try:
        return _list_catalog("cfdi_40_monedas")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/catalogs/prodserv")
def api_prodserv(q: str = Query(..., min_length=1), limit: int = Query(20, ge=1, le=50)):
    try:
        return _search_catalog("cfdi_40_productos_servicios", q=q, limit=limit)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/catalogs/unidad")
def api_unidad(q: str = Query(..., min_length=1), limit: int = Query(20, ge=1, le=50)):
    try:
        return _search_catalog("cfdi_40_claves_unidades", q=q, limit=limit)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))