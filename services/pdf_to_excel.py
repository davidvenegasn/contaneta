import logging
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Iterable, Optional

logger = logging.getLogger(__name__)


_DATE_PATTERNS = [
    re.compile(r"^(?P<d>\d{2})[\/\-](?P<m>\d{2})[\/\-](?P<y>\d{2,4})\b"),
    re.compile(r"^(?P<y>\d{4})[\/\-](?P<m>\d{2})[\/\-](?P<d>\d{2})\b"),
    re.compile(r"^(?P<d>\d{2})[\/\-\s](?P<mon>[A-Za-zÁÉÍÓÚÜÑ\.]{3,})[\/\-\s](?P<y>\d{2,4})\b"),
]

_MONTHS = {
    # ES
    "ENE": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SEP": 9,
    "SET": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
    # EN
    "JAN": 1,
    "APR": 4,
    "AUG": 8,
    "DEC": 12,
}

_BANK_KEYWORDS = [
    ("BBVA", ["BBVA", "BANCOMER"]),
    ("SANTANDER", ["SANTANDER"]),
    ("BANORTE", ["BANORTE"]),
    ("CITIBANAMEX", ["CITIBANAMEX", "BANAMEX"]),
    ("HSBC", ["HSBC"]),
    ("SCOTIABANK", ["SCOTIABANK", "SCOTIA"]),
    ("INBURSA", ["INBURSA"]),
    ("AZTECA", ["BANCO AZTECA", "AZTECA"]),
]

_STOPWORDS_CP = {
    "PAGO",
    "PAGOS",
    "COMPRA",
    "COMPRAS",
    "CARGO",
    "ABONO",
    "DEPOSITO",
    "DEPÓSITO",
    "DEP",
    "TRANSFERENCIA",
    "TRANSFER",
    "TRANSF",
    "TRASPASO",
    "TRASP",
    "SPEI",
    "SPID",
    "COMISION",
    "COMISIONES",
    "IVA",
    "REF",
    "REFERENCIA",
    "FOLIO",
    "AUT",
    "AUTORIZACION",
    "AUTORIZACIÓN",
    "ID",
    "NUM",
    "NO",
    "CTA",
    "CUENTA",
    "TARJ",
    "TARJETA",
    "DEBITO",
    "DÉBITO",
    "CREDITO",
    "CRÉDITO",
    "POS",
    "TPV",
}

_KNOWN_COUNTERPARTIES = [
    "AMAZON",
    "UBER",
    "DIDI",
    "CFE",
    "TELCEL",
    "TELMEX",
    "AT&T",
    "MOVISTAR",
    "TOTALPLAY",
    "IZZI",
    "NETFLIX",
    "SPOTIFY",
    "GOOGLE",
    "APPLE",
    "MICROSOFT",
    "AWS",
    "OPENAI",
    "ADOBE",
    "WALMART",
    "COSTCO",
    "SORIANA",
    "OXXO",
    "7-ELEVEN",
    "SEVEN",
    "PEMEX",
]


def get_storage_root(base_dir: str) -> str:
    """
    Root de storage. Respeta env APP_STORAGE_PATH si existe.
    - Si APP_STORAGE_PATH es relativo, se resuelve contra base_dir.
    - Siempre regresa ruta absoluta normalizada.
    """
    raw = (os.environ.get("APP_STORAGE_PATH") or "").strip()
    if raw:
        root = raw if os.path.isabs(raw) else os.path.join(base_dir, raw)
    else:
        root = os.path.join(base_dir, "storage")
    return os.path.normpath(os.path.abspath(root))


def safe_join(root_abs: str, *parts: str) -> str:
    """Une paths y asegura que queden bajo root_abs (previene path traversal)."""
    root_abs = os.path.normpath(os.path.abspath(root_abs))
    p = os.path.join(root_abs, *[str(x) for x in parts])
    abs_p = os.path.normpath(os.path.abspath(p))
    if abs_p == root_abs:
        return abs_p
    if not abs_p.startswith(root_abs + os.sep):
        raise ValueError("Ruta inválida (path traversal)")
    return abs_p


def ensure_parent_dir(path_abs: str) -> None:
    os.makedirs(os.path.dirname(path_abs), exist_ok=True)


def _strip_accents(s: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", s or "") if not unicodedata.combining(ch))


def _norm_text(s: str) -> str:
    t = _strip_accents(str(s or ""))
    t = re.sub(r"\s+", " ", t).strip().upper()
    return t


def _detect_bank_name(all_text_norm: str) -> str:
    t = all_text_norm or ""
    for name, keys in _BANK_KEYWORDS:
        if any(k in t for k in keys):
            return name
    return "DESCONOCIDO"


def _detect_account_last4(all_text_norm: str) -> str:
    t = all_text_norm or ""
    m = re.search(r"(?:\*{2,}|X{2,})\s*(\d{4})\b", t)
    if m:
        return m.group(1)
    m = re.search(r"(?:TERMINACION|TERMINACIÓN|ULTIMOS|ÚLTIMOS|ULT|FINAL)\s*(\d{4})\b", t)
    if m:
        return m.group(1)
    m = re.search(r"\bCUENTA\s*(\d{4})\b", t)
    if m:
        return m.group(1)
    return ""


def _detect_period(all_text_norm: str) -> tuple[str, str]:
    """
    Intenta detectar rango de periodo en el PDF (encabezado del estado de cuenta).
    Devuelve (period_start, period_end) como YYYY-MM-DD o ("","") si no detecta.
    Prioriza el texto explícito "del ... al ..." / "periodo ..." sobre fechas de movimientos.
    """
    t = all_text_norm or ""
    date_part = r"(\d{2}[\/\-]\d{2}[\/\-]\d{2,4})"
    # 1) "DEL 01/01/2026 AL 31/01/2026" o "PERIODO DEL 01/01/2026 AL 31/01/2026"
    m = re.search(
        r"\b(?:PERIODO\s+)?DEL\s+" + date_part + r"\s+(?:AL|A)\s+" + date_part + r"\b",
        t,
    )
    if m:
        d1, d2 = _detect_date(m.group(1)), _detect_date(m.group(2))
        if d1 and d2:
            return (d1, d2)
    # 2) "PERIODO 01/01/2026 AL 31/01/2026" o "01/01/2026 AL 31/01/2026" (sin DEL)
    m = re.search(
        r"\b(?:PERIODO\s*:?\s*)?(\d{2}[\/\-]\d{2}[\/\-]\d{2,4})\s+(?:AL|A)\s+(\d{2}[\/\-]\d{2}[\/\-]\d{2,4})\b",
        t,
    )
    if m:
        d1, d2 = _detect_date(m.group(1)), _detect_date(m.group(2))
        if d1 and d2:
            return (d1, d2)
    # 3) "DE 01/01/2026 A 31/01/2026"
    m = re.search(r"\bDE\s+" + date_part + r"\s+A\s+" + date_part + r"\b", t)
    if m:
        d1, d2 = _detect_date(m.group(1)), _detect_date(m.group(2))
        if d1 and d2:
            return (d1, d2)
    return ("", "")


def detect_statement_period_from_text(raw_text: str) -> tuple[str, str]:
    """
    Detecta periodo del estado de cuenta desde texto crudo (ej. páginas del PDF concatenadas).
    Útil para el pipeline de preview sin depender del Excel.
    Returns (period_start, period_end) en YYYY-MM-DD o ("", "").
    """
    if not (raw_text or "").strip():
        return ("", "")
    all_text = " ".join((raw_text or "").split())
    all_text_norm = _norm_text(all_text)
    return _detect_period(all_text_norm)


def _detect_date(line: str) -> Optional[str]:
    s = (line or "").strip()
    for pat in _DATE_PATTERNS:
        m = pat.match(s)
        if m:
            d = m.groupdict()
            try:
                if "mon" in d and d.get("mon"):
                    mon_raw = _strip_accents(d["mon"]).upper().strip().strip(".")
                    mon_raw = mon_raw[:3]
                    m_num = _MONTHS.get(mon_raw)
                    if not m_num:
                        return None
                    y = int(d["y"])
                    if y < 100:
                        y = 2000 + y
                    dt = date(int(y), int(m_num), int(d["d"]))
                    return dt.isoformat()
                y = int(d["y"])
                if y < 100:
                    y = 2000 + y
                dt = date(int(y), int(d["m"]), int(d["d"]))
                return dt.isoformat()
            except Exception:
                return None
    return None


def _parse_amount(token: str) -> Optional[float]:
    """
    Parse monto tipo $1,234.56 o (1,234.56).
    Retorna float (puede ser negativo si viene entre paréntesis o con signo).
    """
    if token is None:
        return None
    t = str(token).strip()
    if not t:
        return None
    neg = False
    if t.startswith("(") and t.endswith(")"):
        neg = True
        t = t[1:-1].strip()
    t = t.replace("$", "").replace(" ", "")
    if t.startswith("-"):
        neg = True
        t = t[1:]
    # Normalizar separadores: 1,234.56 | 1.234,56 | 1234,56 | 1234.56
    if not re.match(r"^\d[\d\.,]*$", t):
        return None
    if "," in t and "." in t:
        # decimal = último separador
        if t.rfind(",") > t.rfind("."):
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
    elif "," in t and "." not in t:
        # si termina con ,dd entonces es decimal
        if re.match(r"^\d{1,3}(?:\.\d{3})*,\d{1,2}$", t) or re.match(r"^\d+,(\d{1,2})$", t):
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
    else:
        # solo punto: dejarlo como decimal
        pass
    if not re.match(r"^\d+(\.\d{1,2})?$", t):
        return None
    try:
        n = float(t)
        return -n if neg else n
    except Exception:
        return None


def _split_columns(line: str) -> list[str]:
    # muchos estados separan columnas por 2+ espacios
    parts = re.split(r"\s{2,}", (line or "").strip())
    return [p.strip() for p in parts if p and p.strip()]


def _extract_amounts_from_end(s: str, max_amounts: int = 3) -> tuple[str, list[float]]:
    """
    Extrae hasta N montos al final del string y regresa (texto_sin_montos, montos_en_orden_original).
    """
    txt = (s or "").rstrip()
    found_rev: list[float] = []
    for _ in range(max_amounts):
        m = re.search(r"(?:\s+|^)(\(?-?\$?\d[\d\.,]*\)?)\s*$", txt)
        if not m:
            break
        tok = m.group(1)
        a = _parse_amount(tok)
        if a is None:
            break
        found_rev.append(a)
        txt = txt[: m.start(1)].rstrip()
    found = list(reversed(found_rev))
    return txt.strip(), found


def _extract_referencia(desc_norm: str) -> str:
    d = desc_norm or ""
    m = re.search(r"\b(?:REF|REFERENCIA|FOLIO|AUT|AUTORIZACION|AUTORIZACIÓN|ID)\s*[:\-]?\s*([A-Z0-9]{4,})\b", d)
    return m.group(1) if m else ""


def _metodo_pago_hint(desc_norm: str) -> str:
    d = desc_norm or ""
    if "SPEI" in d:
        return "SPEI"
    if "TPV" in d or "POS" in d:
        return "TPV"
    if "TARJ" in d or "TARJETA" in d or "DEBITO" in d or "DÉBITO" in d or "CREDITO" in d or "CRÉDITO" in d:
        return "TARJETA"
    if "COMISION" in d or "COMISIONES" in d:
        return "COMISION"
    if "EFECTIVO" in d or "RETIRO" in d or "CAJERO" in d:
        return "EFECTIVO"
    if "TRANSF" in d or "TRANSFER" in d or "TRASP" in d:
        return "TRANSFER"
    return ""


def _contraparte_hint(desc_norm: str) -> str:
    d = desc_norm or ""
    for k in _KNOWN_COUNTERPARTIES:
        if k in d:
            return k
    # fallback: primeras 1-3 "palabras útiles"
    words = [w for w in re.split(r"[^A-Z0-9&\-]+", d) if w]
    useful: list[str] = []
    for w in words:
        if len(w) < 3:
            continue
        if w in _STOPWORDS_CP:
            continue
        useful.append(w)
        if len(useful) >= 3:
            break
    return " ".join(useful[:2]) if useful else ""


def _clasificar(desc_norm: str) -> tuple[str, str, int, int, int]:
    """
    Regresa (categoria, subcategoria, es_comision_bancaria, es_impuesto_bancario, posible_facturable).
    posible_facturable se calcula con heurística simple, principalmente para gastos.
    """
    d = desc_norm or ""
    cat = "OTROS"
    sub = ""

    def has_any(keys: Iterable[str]) -> bool:
        return any(k in d for k in keys)

    if has_any(["COMISION", "MANEJO CTA", "ANUALIDAD", "MEMBERSHIP"]):
        cat = "COMISIONES BANCARIAS"
    elif has_any(["SAT", "HACIENDA", "IMPUESTO", "ISR", "IVA"]):
        cat = "IMPUESTOS"
    elif has_any(["CFE", "AGUA", "GAS", "TELMEX", "IZZI", "TOTALPLAY", "TELCEL", "AT&T", "MOVISTAR"]):
        cat = "SERVICIOS"
    elif has_any(["UBER", "DIDI", "GASOLINA", "PEMEX", "OXXO GAS", "ESTACION"]):
        cat = "TRANSPORTE"
    elif has_any(["OXXO", "7-ELEVEN", "SORIANA", "WALMART", "COSTCO", "REST", "RESTAUR", "CAFE", "CAFÉ"]):
        cat = "ALIMENTOS"
    elif has_any(["GOOGLE", "APPLE", "MICROSOFT", "AWS", "OPENAI", "ADOBE", "NETFLIX", "SPOTIFY"]):
        cat = "SOFTWARE/SUSCRIPCIONES"
    elif has_any(["HONORARIOS", "CONSULTORIA", "CONSULTOR", "FACTURA", "RFC", "SERVICIO PROFESIONAL", "SERVICIOS PROFESIONALES"]):
        cat = "HONORARIOS/PROVEEDORES"
    elif has_any(["RETIRO", "CAJERO"]):
        cat = "RETIRO/EFECTIVO"
        sub = "CAJERO"
    elif has_any(["SPEI", "TRANSF", "TRANSFER", "TRASPASO", "TRASP"]):
        cat = "TRANSFERENCIAS"
        if "SPEI" in d:
            sub = "SPEI"
        elif "TRASP" in d or "TRASPASO" in d:
            sub = "TRASPASO"
        else:
            sub = "TRANSFER"

    es_com = 1 if cat == "COMISIONES BANCARIAS" else 0
    es_imp = 1 if cat == "IMPUESTOS" else 0

    # posible_facturable: heurística rápida (sin depender de proveedores)
    fact = 0
    if cat not in ("COMISIONES BANCARIAS", "IMPUESTOS", "RETIRO/EFECTIVO"):
        if any(x in d for x in ["RFC", "FACTURA", "S A DE C V", "S.A. DE C.V", "SAPI", "SA CV"]):
            fact = 1
        elif any(x in d for x in ["WALMART", "COSTCO", "SORIANA", "CFE", "TELCEL", "TELMEX", "IZZI", "TOTALPLAY", "GOOGLE", "APPLE", "MICROSOFT", "AWS", "ADOBE"]):
            fact = 1
    return cat, sub, es_com, es_imp, fact


def _to_date(s: str | None) -> Optional[date]:
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _amount_in_tolerance(mov: float, cfdi: float) -> bool:
    if mov <= 0 or cfdi <= 0:
        return False
    diff = abs(mov - cfdi)
    tol = max(2.0, mov * 0.005)
    return diff <= tol


def _score_cfdi(mov_date: date, mov_amount: float, contraparte_hint: str, cfdi: dict) -> int:
    score = 0
    try:
        cfdi_total = float(cfdi.get("total") or 0)
    except Exception:
        cfdi_total = 0.0
    if _amount_in_tolerance(mov_amount, cfdi_total):
        score += 60
    cfdi_date = _to_date(cfdi.get("fecha_emision"))
    if cfdi_date:
        d = abs((cfdi_date - mov_date).days)
        if d <= 2:
            score += 20
    cp = _norm_text(contraparte_hint or "")
    if cp:
        hay = _norm_text(f"{cfdi.get('rfc_emisor') or ''} {cfdi.get('nombre_emisor') or ''}")
        # match por substring; si cp es multi-palabra, basta que una palabra útil aparezca
        if cp in hay:
            score += 20
        else:
            tokens = [t for t in cp.split() if len(t) >= 4 and t not in _STOPWORDS_CP]
            if any(t in hay for t in tokens):
                score += 20
    return score


def _best_cfdi_for_movement(mov_date: date, mov_amount: float, contraparte_hint: str, cfdis: list[dict]) -> tuple[Optional[dict], int]:
    best = None
    best_score = -1
    for c in cfdis:
        c_date = _to_date(c.get("fecha_emision"))
        if not c_date:
            continue
        if abs((c_date - mov_date).days) > 7:
            continue
        s = _score_cfdi(mov_date, mov_amount, contraparte_hint, c)
        if s > best_score:
            best = c
            best_score = s
    return best, (best_score if best_score >= 0 else 0)


@dataclass
class ConvertMeta:
    rows: int
    raw_lines: int
    mode: str  # 'parsed' | 'raw'


def convert_pdf_to_xlsx(
    pdf_path_abs: str,
    xlsx_path_abs: str,
    issuer_id: int | None = None,
    statement_id: int | None = None,
) -> dict[str, Any]:
    """
    Convierte un PDF de estado de cuenta a Excel.
    - Usa parser Banorte (solo montos con 2 decimales, sección DETALLE DE MOVIMIENTOS (PESOS)).
    - Si statement_id y issuer_id se pasan, persiste movimientos en bank_movements con dedupe por hash.
    - Genera XLSX: Movimientos, Gastos, Ingresos, Resumen, RAW.
    """
    try:
        import pdfplumber  # lazy import (dependencia opcional hasta instalar)
    except ModuleNotFoundError:
        pdfplumber = None
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if not os.path.isfile(pdf_path_abs):
        raise FileNotFoundError("PDF no encontrado")

    HEAD_MOV = [
        "fecha",
        "descripcion",
        "deposito",
        "retiro",
        "saldo",
        "tipo",
        "categoria",
        "contraparte_hint",
        "metodo_hint",
        "referencia",
        "cve_rastreo",
        "rfc_encontrado",
        "confidence_score",
        "source_page_first",
    ]

    def _freeze_filter(ws, ncols: int, nrows: int) -> None:
        ws.freeze_panes = "A2"
        if ncols >= 1 and nrows >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"

    def _autosize(ws, ncols: int, max_width: int = 60) -> None:
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            best = 0
            for cell in ws[letter]:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > best:
                    best = len(s)
            ws.column_dimensions[letter].width = min(max(10, best + 2), max_width)

    def _write_table(ws, headers: list[str], rows: list[list[Any]], money_cols: Iterable[int] = ()) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in rows:
            ws.append(r)
        nrows = 1 + len(rows)
        ncols = len(headers)
        _freeze_filter(ws, ncols, nrows)
        for cidx in money_cols:
            for r in range(2, nrows + 1):
                cell = ws.cell(row=r, column=cidx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
        _autosize(ws, ncols)

    if pdfplumber is None:
        ensure_parent_dir(xlsx_path_abs)
        wb = Workbook()
        wb.remove(wb.active)
        _write_table(wb.create_sheet("Movimientos"), HEAD_MOV, [], money_cols=[3, 4, 5])
        _write_table(wb.create_sheet("Gastos"), HEAD_MOV, [], money_cols=[3, 4, 5])
        _write_table(wb.create_sheet("Ingresos"), HEAD_MOV, [], money_cols=[3, 4, 5])
        ws_r = wb.create_sheet("Resumen")
        ws_r.append(["campo", "valor"])
        ws_r["A1"].font = Font(bold=True)
        ws_r["B1"].font = Font(bold=True)
        ws_r.append(["error", "Falta dependencia: pdfplumber (instala requirements.txt)"])
        ws_r.freeze_panes = "A2"
        _write_table(wb.create_sheet("RAW"), ["page", "line", "text"], [], money_cols=[])
        wb.save(xlsx_path_abs)
        return {
            "rows": 0,
            "raw_lines": 0,
            "mode": "raw",
            "error": "pdfplumber_missing",
            "period_start": "",
            "period_end": "",
            "bank_name": "",
            "account_last4": "",
            "transactions": [],
            "processed_count": 0,
            "total_ingresos": 0.0,
            "total_gastos": 0.0,
            "sin_factura_count": 0,
            "movements_count": 0,
            "ingresos_total": 0.0,
            "gastos_total": 0.0,
            "sin_parse_count": 0,
        }

    raw_rows: list[dict[str, Any]] = []
    with pdfplumber.open(pdf_path_abs) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            lines = [ln.strip() for ln in text.splitlines() if (ln or "").strip()]
            for li, ln in enumerate(lines, start=1):
                raw_rows.append({"Page": page_idx, "Line": li, "Text": ln})

    all_text = " ".join(str(r.get("Text") or "") for r in raw_rows)
    all_text_norm = _norm_text(all_text)
    period_start, period_end = _detect_period(all_text_norm)
    bank_name = _detect_bank_name(all_text_norm)
    account_last4 = _detect_account_last4(all_text_norm)

    ensure_parent_dir(xlsx_path_abs)

    # -------- pipeline robusta (Banorte o similares) --------
    from config import DEV_MODE
    from services.bank.bank_statement_parser import parse_bank_statement, write_debug_json

    debug_on = bool(DEV_MODE) and (os.environ.get("BANK_PARSER_DEBUG", "0").strip() == "1")
    parsed = parse_bank_statement(raw_rows, debug=debug_on)
    txs: list[dict[str, Any]] = parsed.transactions or []
    metrics: dict[str, Any] = parsed.metrics or {}

    # fallback (para bancos que no tengan DETALLE DE MOVIMIENTOS): parser simple por línea con fecha
    if int(metrics.get("movements_count") or 0) <= 0:
        simple_txs: list[dict[str, Any]] = []
        prev_saldo: Optional[float] = None
        rfc_re = re.compile(r"\b([A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{3})\b")
        rast_re = re.compile(r"\b(?:CVE\\s*RAST(?:REO)?)\s*[:=\-]?\s*([A-Z0-9]{8,40})\b")
        for rr in raw_rows:
            ln = str(rr.get("Text") or "").strip()
            dt = _detect_date(ln)
            if not dt:
                continue
            rest = re.sub(r"^\s*\d{2}[\/\-]\d{2}[\/\-]\d{2,4}\s+", "", ln.strip())
            rest = re.sub(r"^\s*\d{4}[\/\-]\d{2}[\/\-]\d{2}\s+", "", rest)
            rest = re.sub(r"^\s*\d{2}[\/\-\s][A-Za-zÁÉÍÓÚÜÑ\.]{3,}[\/\-\s]\d{2,4}\s+", "", rest)
            cols = _split_columns(rest)
            rest_for_amounts = " ".join(cols) if (cols and len(cols) >= 2) else rest
            desc_wo_amounts, amounts = _extract_amounts_from_end(rest_for_amounts, max_amounts=3)
            desc_raw = (desc_wo_amounts or "").strip()
            if not desc_raw:
                continue
            cargo = None
            abono = None
            saldo = None
            if len(amounts) >= 3:
                cargo, abono, saldo = amounts[-3], amounts[-2], amounts[-1]
            elif len(amounts) == 2:
                cargo, saldo = amounts[-2], amounts[-1]
            elif len(amounts) == 1:
                cargo = amounts[-1]
            if saldo is not None:
                prev_saldo = saldo
            desc_norm = _norm_text(desc_raw)
            categoria, _, _, _, _ = _clasificar(desc_norm)
            contraparte = _contraparte_hint(desc_norm)
            metodo_raw = _metodo_pago_hint(desc_norm)
            if metodo_raw in ("SPEI",):
                metodo_hint = "SPEI"
            elif metodo_raw in ("TARJETA", "TPV"):
                metodo_hint = "TARJETA"
            elif metodo_raw in ("EFECTIVO",):
                metodo_hint = "EFECTIVO"
            else:
                metodo_hint = "OTRO"
            tipo = "DESCONOCIDO"
            deposito = abs(float(abono)) if (abono is not None and float(abono or 0) > 0) else None
            retiro = abs(float(cargo)) if (cargo is not None and float(cargo or 0) > 0) else None
            if deposito:
                tipo = "INGRESO"
            elif retiro:
                tipo = "GASTO"
            ref = _extract_referencia(desc_norm)
            rast = (rast_re.search(desc_norm).group(1) if rast_re.search(desc_norm) else "")
            rfc = (rfc_re.search(desc_norm).group(1) if rfc_re.search(desc_norm) else "")
            score = 35
            if deposito or retiro:
                score += 25
            if saldo is not None:
                score += 15
            if dt:
                score += 15
            score = min(100, score)
            pg = int(rr.get("Page") or 0)
            simple_txs.append(
                {
                    "fecha": dt,
                    "descripcion_full": " ".join(desc_raw.split()),
                    "descripcion_norm": desc_norm,
                    "deposito": deposito,
                    "retiro": retiro,
                    "saldo": float(saldo) if saldo is not None else None,
                    "tipo": tipo,
                    "categoria": categoria,
                    "contraparte_hint": contraparte,
                    "metodo_hint": metodo_hint,
                    "referencia": ref,
                    "cve_rastreo": rast,
                    "rfc_encontrado": rfc,
                    "confidence_score": score,
                    "source_page_first": pg,
                    "source_page_last": pg,
                }
            )
        txs = simple_txs
        metrics = {
            "sections_detected": 0,
            "transactions_grouped": len(simple_txs),
            "movements_count": sum(1 for t in simple_txs if (t.get("deposito") or 0) or (t.get("retiro") or 0)),
            "sin_parse_count": 0,
            "saldo_count": sum(1 for t in simple_txs if isinstance(t.get("saldo"), (int, float))),
            "rfc_count": sum(1 for t in simple_txs if (t.get("rfc_encontrado") or "").strip()),
            "rastreo_count": sum(1 for t in simple_txs if (t.get("cve_rastreo") or "").strip()),
            "avg_confidence": (sum(float(t.get("confidence_score") or 0) for t in simple_txs) / len(simple_txs)) if simple_txs else 0.0,
            "low_confidence_count": sum(1 for t in simple_txs if int(t.get("confidence_score") or 0) < 60),
            "total_ingresos": sum(float(t.get("deposito") or 0) for t in simple_txs),
            "total_gastos": sum(float(t.get("retiro") or 0) for t in simple_txs),
        }

    movements_count = int(metrics.get("movements_count") or 0)
    sin_parse_count = int(metrics.get("sin_parse_count") or 0)
    total_ingresos = float(metrics.get("total_ingresos") or 0.0)
    total_gastos = float(metrics.get("total_gastos") or 0.0)
    low_confidence_count = int(metrics.get("low_confidence_count") or 0)
    mode = "parsed" if movements_count > 0 else "raw"

    # Persistir movimientos en DB con dedupe (solo INGRESO/GASTO)
    if statement_id and issuer_id and txs:
        try:
            from services.bank.bank_statement_parser import upsert_bank_movements
            upsert_bank_movements(int(issuer_id), int(statement_id), txs)
        except Exception:
            logger.exception("bank_parser: no se pudieron guardar movimientos en DB")

    logger.info(
        "bank_parser: sections=%s grouped=%s movements=%s saldo=%s rfc=%s rastreo=%s avg_conf=%.1f",
        int(metrics.get("sections_detected") or 0),
        int(metrics.get("transactions_grouped") or 0),
        movements_count,
        int(metrics.get("saldo_count") or 0),
        int(metrics.get("rfc_count") or 0),
        int(metrics.get("rastreo_count") or 0),
        float(metrics.get("avg_confidence") or 0.0),
    )

    if debug_on and parsed.debug_payload:
        try:
            write_debug_json(parsed.debug_payload, xlsx_path_abs + ".debug.json")
        except Exception:
            logger.exception("bank_parser: no se pudo escribir debug json")

    # -------- export Excel (5 hojas) --------
    wb = Workbook()
    wb.remove(wb.active)

    rows_mov: list[list[Any]] = []
    for t in txs:
        rows_mov.append(
            [
                t.get("fecha") or "",
                t.get("descripcion_full") or "",
                t.get("deposito") if isinstance(t.get("deposito"), (int, float)) else "",
                t.get("retiro") if isinstance(t.get("retiro"), (int, float)) else "",
                t.get("saldo") if isinstance(t.get("saldo"), (int, float)) else "",
                t.get("tipo") or "DESCONOCIDO",
                t.get("categoria") or "",
                t.get("contraparte_hint") or "",
                t.get("metodo_hint") or "OTRO",
                t.get("referencia") or "",
                t.get("cve_rastreo") or "",
                t.get("rfc_encontrado") or "",
                int(t.get("confidence_score") or 0),
                int(t.get("source_page_first") or 0),
            ]
        )

    ws_mov = wb.create_sheet("Movimientos")
    _write_table(ws_mov, HEAD_MOV, rows_mov, money_cols=[3, 4, 5])

    ws_g = wb.create_sheet("Gastos")
    gastos_rows = [r for r in rows_mov if r[5] == "GASTO"]
    _write_table(ws_g, HEAD_MOV, gastos_rows, money_cols=[3, 4, 5])

    ws_i = wb.create_sheet("Ingresos")
    ingresos_rows = [r for r in rows_mov if r[5] == "INGRESO"]
    _write_table(ws_i, HEAD_MOV, ingresos_rows, money_cols=[3, 4, 5])

    ws_r = wb.create_sheet("Resumen")
    ws_r.append(["campo", "valor"])
    ws_r["A1"].font = Font(bold=True)
    ws_r["B1"].font = Font(bold=True)
    ws_r.append(["movements_count", movements_count])
    ws_r.append(["sin_parse_count", sin_parse_count])
    ws_r.append(["total_ingresos", total_ingresos])
    ws_r.append(["total_gastos", total_gastos])
    ws_r.append(["neto", total_ingresos - total_gastos])
    ws_r.append(["saldo_count", int(metrics.get("saldo_count") or 0)])
    ws_r.append(["rfc_count", int(metrics.get("rfc_count") or 0)])
    ws_r.append(["rastreo_count", int(metrics.get("rastreo_count") or 0)])
    ws_r.append(["avg_confidence", float(metrics.get("avg_confidence") or 0.0)])
    for r in range(2, ws_r.max_row + 1):
        k = ws_r.cell(row=r, column=1).value
        v = ws_r.cell(row=r, column=2).value
        if isinstance(v, (int, float)) and k in ("total_ingresos", "total_gastos", "neto"):
            ws_r.cell(row=r, column=2).number_format = '"$"#,##0.00'
    ws_r.freeze_panes = "A2"
    ws_r.column_dimensions["A"].width = 28
    ws_r.column_dimensions["B"].width = 18

    def _top(rows: list[list[Any]], key_idx: int, amt_idx: int, topn: int = 10) -> list[tuple[str, float]]:
        agg: dict[str, float] = {}
        for rr in rows:
            k = (rr[key_idx] or "").strip() or "—"
            amt = rr[amt_idx] if isinstance(rr[amt_idx], (int, float)) else 0.0
            agg[k] = agg.get(k, 0.0) + float(amt)
        return sorted(agg.items(), key=lambda x: x[1], reverse=True)[:topn]

    top_cats = _top(gastos_rows, 6, 3, 10)  # categoria, retiro
    ws_r.append([])
    ws_r.append(["Top categorías (gasto)", "monto"])
    ws_r["A" + str(ws_r.max_row)].font = Font(bold=True)
    ws_r["B" + str(ws_r.max_row)].font = Font(bold=True)
    for k, v in top_cats:
        ws_r.append([k, v])
        ws_r.cell(row=ws_r.max_row, column=2).number_format = '"$"#,##0.00'

    ws_raw = wb.create_sheet("RAW")
    raw_table = [[r["Page"], r["Line"], r["Text"]] for r in raw_rows]
    _write_table(ws_raw, ["page", "line", "text"], raw_table, money_cols=[])

    wb.save(xlsx_path_abs)

    transactions_for_db = [
        {
            "fecha": t.get("fecha") or "",
            "descripcion": (t.get("descripcion_full") or "")[:2000],
            "deposito": t.get("deposito") if isinstance(t.get("deposito"), (int, float)) else None,
            "retiro": t.get("retiro") if isinstance(t.get("retiro"), (int, float)) else None,
            "saldo": t.get("saldo") if isinstance(t.get("saldo"), (int, float)) else None,
            "tipo": t.get("tipo") or "DESCONOCIDO",
            "categoria": (t.get("categoria") or "")[:200],
            "metodo_hint": (t.get("metodo_hint") or "")[:64],
            "contraparte_hint": (t.get("contraparte_hint") or "")[:200],
            "rfc_encontrado": (t.get("rfc_encontrado") or "")[:20],
            "confidence_score": int(t.get("confidence_score") or 0),
            "source_page_first": int(t.get("source_page_first") or 0),
        }
        for t in txs
    ]

    return {
        "rows": len(rows_mov),
        "raw_lines": len(raw_rows),
        "mode": mode,
        "period_start": period_start,
        "period_end": period_end,
        "bank_name": bank_name,
        "account_last4": account_last4,
        "transactions": transactions_for_db,
        # compat UI previa
        "processed_count": movements_count,
        "total_ingresos": total_ingresos,
        "total_gastos": total_gastos,
        "sin_factura_count": sin_parse_count,
        # métricas solicitadas
        "movements_count": movements_count,
        "ingresos_total": total_ingresos,
        "gastos_total": total_gastos,
        "sin_parse_count": sin_parse_count,
        "low_confidence_count": low_confidence_count,
        "quality": {
            "sections_detected": int(metrics.get("sections_detected") or 0),
            "transactions_grouped": int(metrics.get("transactions_grouped") or 0),
            "saldo_count": int(metrics.get("saldo_count") or 0),
            "rfc_count": int(metrics.get("rfc_count") or 0),
            "rastreo_count": int(metrics.get("rastreo_count") or 0),
            "avg_confidence": float(metrics.get("avg_confidence") or 0.0),
            "low_confidence_count": low_confidence_count,
        },
    }

