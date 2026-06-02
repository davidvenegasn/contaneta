"""Generate Excel workbook with fiscal papers of work for a given month."""
import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from database import db_rows
from services.fiscal.calculators import calc_iva, calc_pfae_general, calc_resico_pf
from services.fiscal.deductibility import compute_deductible_totals
from services.invoices import foreign_invoices as fi
from services.sat.sat_sync import get_month_totals
from services.ym_helpers import ym_to_label

logger = logging.getLogger(__name__)

# Brand colors
_HEADER_FILL = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BOLD = Font(name="Calibri", bold=True, size=11)
_CURRENCY = "#,##0.00"
_PCT = "0.00%"
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def _add_header_row(ws, headers: list[str], row: int = 1):
    """Write styled header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    """Set column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def build_fiscal_excel(issuer_id: int, ym: str, issuer_alias: str = "") -> bytes:
    """Generate Excel report with fiscal papers for the month.

    Sheets: Resumen, Ingresos (issued CFDIs), Gastos (received CFDIs), Deducciones.
    """
    wb = Workbook()
    label = ym_to_label(ym)

    # ── Gather data ──────────────────────────────────────────────
    issued = get_month_totals(issuer_id, ym, "issued")
    received = get_month_totals(issuer_id, ym, "received")

    deduct = compute_deductible_totals(issuer_id, ym)
    gastos_deducibles = deduct["gastos_deducibles"]
    iva_acreditable = deduct["iva_acreditable"]
    deducible_detail = deduct["detail"]

    fi.ensure_table()
    fi_totals = fi.compute_totals(issuer_id, period_month=ym)

    total_ingresos = issued["total_base"] + fi_totals["sum_ingresos"]
    total_gastos = gastos_deducibles + fi_totals["sum_gastos"]

    # Regime
    regime_rows = db_rows(
        "SELECT regimen FROM issuer_fiscal_profile WHERE issuer_id = ?",
        (issuer_id,),
    )
    regimen = (regime_rows[0]["regimen"] if regime_rows else "RESICO_PF") or "RESICO_PF"

    if regimen == "RESICO_PF":
        isr = calc_resico_pf(total_ingresos)
    else:
        isr = calc_pfae_general(
            total_ingresos,
            deducciones_mes=total_gastos,
            retenciones_isr=issued.get("total_retenciones", 0),
        )
    iva = calc_iva(
        iva_causado=issued["total_iva"],
        iva_acreditable=iva_acreditable,
        iva_retenido=issued.get("total_retenciones", 0),
    )

    # ── Sheet 1: Resumen ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.cell(row=1, column=1, value=f"Papeles de Trabajo — {label}").font = Font(
        name="Calibri", bold=True, size=14
    )
    if issuer_alias:
        ws.cell(row=2, column=1, value=issuer_alias).font = _BOLD

    r = 4
    for lbl, val in [
        ("Total Ingresos", total_ingresos),
        ("Total Gastos Deducibles", total_gastos),
        ("Utilidad", total_ingresos - total_gastos),
        ("", ""),
        ("IVA Cobrado", issued["total_iva"]),
        ("IVA Acreditable", iva_acreditable),
        ("IVA Retenido", issued.get("total_retenciones", 0)),
        ("IVA a Pagar", iva.get("iva_a_pagar", 0)),
        ("Saldo a Favor IVA", iva.get("saldo_a_favor", 0)),
        ("", ""),
        ("ISR Estimado", isr.get("isr_estimado", isr.get("isr_provisional", 0))),
        ("Régimen", regimen),
    ]:
        ws.cell(row=r, column=1, value=lbl).font = _BOLD if lbl else Font()
        cell_v = ws.cell(row=r, column=2, value=val)
        if isinstance(val, (int, float)) and lbl:
            cell_v.number_format = _CURRENCY
        r += 1
    _auto_width(ws)

    # ── Sheet 2: Ingresos ────────────────────────────────────────
    ws2 = wb.create_sheet("Ingresos")
    headers = ["UUID", "Fecha", "RFC Receptor", "Receptor", "Subtotal", "IVA", "Total", "Método Pago", "Estado"]
    _add_header_row(ws2, headers)
    rows = db_rows(
        """SELECT uuid, fecha_emision, rfc_receptor, nombre_receptor,
                  COALESCE(subtotal, total) AS subtotal, COALESCE(impuestos, 0) AS iva,
                  total, metodo_pago, status
           FROM sat_cfdi
           WHERE issuer_id = ? AND direction = 'issued'
             AND fecha_emision LIKE ? || '%'
             AND COALESCE(UPPER(TRIM(status)), '') NOT IN ('C','CANCELADO','CANCELADA','0')
             AND UPPER(TRIM(COALESCE(status,''))) NOT LIKE 'CANCEL%%'
           ORDER BY fecha_emision""",
        (issuer_id, ym),
    )
    for i, row in enumerate(rows, 2):
        ws2.cell(row=i, column=1, value=row.get("uuid", ""))
        ws2.cell(row=i, column=2, value=row.get("fecha_emision", ""))
        ws2.cell(row=i, column=3, value=row.get("rfc_receptor", ""))
        ws2.cell(row=i, column=4, value=row.get("nombre_receptor", ""))
        ws2.cell(row=i, column=5, value=float(row.get("subtotal") or 0)).number_format = _CURRENCY
        ws2.cell(row=i, column=6, value=float(row.get("iva") or 0)).number_format = _CURRENCY
        ws2.cell(row=i, column=7, value=float(row.get("total") or 0)).number_format = _CURRENCY
        ws2.cell(row=i, column=8, value=row.get("metodo_pago", ""))
        ws2.cell(row=i, column=9, value=row.get("status", ""))
    _auto_width(ws2)

    # ── Sheet 3: Gastos ──────────────────────────────────────────
    ws3 = wb.create_sheet("Gastos")
    headers = ["UUID", "Fecha", "RFC Emisor", "Emisor", "Subtotal", "IVA", "Total", "Tipo", "Estado"]
    _add_header_row(ws3, headers)
    rows = db_rows(
        """SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor,
                  COALESCE(subtotal, total) AS subtotal, COALESCE(impuestos, 0) AS iva,
                  total, tipo_comprobante, status
           FROM sat_cfdi
           WHERE issuer_id = ? AND direction = 'received'
             AND fecha_emision LIKE ? || '%'
             AND total IS NOT NULL AND total >= 0.01
             AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
             AND COALESCE(UPPER(TRIM(status)), '') NOT IN ('C','CANCELADO','CANCELADA','0')
             AND UPPER(TRIM(COALESCE(status,''))) NOT LIKE 'CANCEL%%'
           ORDER BY fecha_emision""",
        (issuer_id, ym),
    )
    for i, row in enumerate(rows, 2):
        ws3.cell(row=i, column=1, value=row.get("uuid", ""))
        ws3.cell(row=i, column=2, value=row.get("fecha_emision", ""))
        ws3.cell(row=i, column=3, value=row.get("rfc_emisor", ""))
        ws3.cell(row=i, column=4, value=row.get("nombre_emisor", ""))
        ws3.cell(row=i, column=5, value=float(row.get("subtotal") or 0)).number_format = _CURRENCY
        ws3.cell(row=i, column=6, value=float(row.get("iva") or 0)).number_format = _CURRENCY
        ws3.cell(row=i, column=7, value=float(row.get("total") or 0)).number_format = _CURRENCY
        ws3.cell(row=i, column=8, value=row.get("tipo_comprobante", ""))
        ws3.cell(row=i, column=9, value=row.get("status", ""))
    _auto_width(ws3)

    # ── Sheet 4: Deducciones ─────────────────────────────────────
    ws4 = wb.create_sheet("Deducciones")
    headers = ["Fecha", "RFC Emisor", "Emisor", "Concepto", "Total", "% Deducible", "Deducible", "IVA Acreditable"]
    _add_header_row(ws4, headers)
    for i, d in enumerate(deducible_detail, 2):
        ws4.cell(row=i, column=1, value=d.get("fecha", ""))
        ws4.cell(row=i, column=2, value=d.get("rfc_emisor", ""))
        ws4.cell(row=i, column=3, value=d.get("nombre_emisor", ""))
        ws4.cell(row=i, column=4, value=d.get("concepto", ""))
        ws4.cell(row=i, column=5, value=float(d.get("total") or 0)).number_format = _CURRENCY
        pct = float(d.get("deductibility_pct", 100)) / 100
        ws4.cell(row=i, column=6, value=pct).number_format = _PCT
        ws4.cell(row=i, column=7, value=float(d.get("deducible", 0))).number_format = _CURRENCY
        ws4.cell(row=i, column=8, value=float(d.get("iva_acreditable", 0))).number_format = _CURRENCY
    _auto_width(ws4)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
