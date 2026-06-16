"""Excel export for reports using openpyxl."""
import io
import logging

logger = logging.getLogger(__name__)


def monthly_to_excel(report: dict) -> bytes:
    """Convert a monthly report to Excel bytes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {report['periodo']}"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8E8E8")

    # Summary section
    ws.append(["Reporte Fiscal Mensual", report["periodo"]])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Concepto", "Subtotal", "IVA", "Retenciones", "Total"])
    for cell in ws[3]:
        cell.font = bold
        cell.fill = header_fill

    ing = report["ingresos"]
    ws.append(["Ingresos", ing["subtotal"], ing["iva"], ing["retenciones"], ing["total"]])
    gas = report["gastos_neto"]
    ws.append(["Gastos deducibles", gas["subtotal"], gas["iva_acreditable"], gas.get("retenciones", 0), gas["total"]])
    nc = report["notas_credito"]
    ws.append(["Notas de crédito recibidas", nc["subtotal"], nc["iva"], 0, nc["total"]])
    ws.append([])
    ws.append(["Utilidad fiscal", report["utilidad_fiscal"]])
    ws.append(["IVA neto a pagar", report["iva_neto"]])
    ws.append(["ISR estimado", report["isr_estimado"]])

    # Column widths
    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 22

    # Emitidos sheet
    if report.get("cfdi_emitidos"):
        ws2 = wb.create_sheet("Emitidas")
        ws2.append(["Fecha", "UUID", "Receptor", "RFC", "Concepto", "Subtotal", "IVA", "Retenciones", "Total", "Moneda"])
        for cell in ws2[1]:
            cell.font = bold
            cell.fill = header_fill
        for r in report["cfdi_emitidos"]:
            ws2.append([
                (r.get("fecha_emision") or "")[:10], r.get("uuid", ""),
                r.get("nombre_receptor", ""), r.get("rfc_receptor", ""),
                r.get("concepto", ""), r.get("subtotal", 0), r.get("impuestos", 0),
                r.get("retenciones", 0), r.get("total", 0), r.get("moneda", "MXN"),
            ])

    # Recibidos sheet
    if report.get("cfdi_recibidos"):
        ws3 = wb.create_sheet("Recibidas")
        ws3.append(["Fecha", "UUID", "Emisor", "RFC", "Concepto", "Subtotal", "IVA", "Retenciones", "Total", "Moneda"])
        for cell in ws3[1]:
            cell.font = bold
            cell.fill = header_fill
        for r in report["cfdi_recibidos"]:
            ws3.append([
                (r.get("fecha_emision") or "")[:10], r.get("uuid", ""),
                r.get("nombre_emisor", ""), r.get("rfc_emisor", ""),
                r.get("concepto", ""), r.get("subtotal", 0), r.get("impuestos", 0),
                r.get("retenciones", 0), r.get("total", 0), r.get("moneda", "MXN"),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def annual_to_excel(report: dict) -> bytes:
    """Convert an annual report to Excel bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Anual {report['year']}"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8E8E8")

    ws.append([f"Reporte Fiscal Anual {report['year']}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["Mes", "Ingresos", "IVA cobrado", "Retenciones", "Gastos", "IVA pagado",
               "Utilidad", "IVA neto", "ISR estimado"]
    ws.append(headers)
    for cell in ws[3]:
        cell.font = bold
        cell.fill = header_fill

    month_names = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                   "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    for i, m in enumerate(report["months"]):
        ws.append([
            month_names[i], m["ingresos_subtotal"], m["ingresos_iva"],
            m["ingresos_retenciones"], m["gastos_subtotal"], m["gastos_iva"],
            m["utilidad"], m["iva_neto"], m["isr_estimado"],
        ])

    ws.append([])
    t = report["totals"]
    ws.append(["TOTAL", t["ingresos_subtotal"], t["ingresos_iva"],
               t["ingresos_retenciones"], t["gastos_subtotal"], t["gastos_iva"],
               t["utilidad"], t["iva_neto"], t["isr_provisionales"]])
    for cell in ws[ws.max_row]:
        cell.font = bold

    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
