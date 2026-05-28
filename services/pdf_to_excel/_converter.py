"""Main PDF-to-Excel conversion function for bank statements."""
import logging
import os
from dataclasses import dataclass
from typing import Any

from services.pdf_to_excel._fallback import fallback_simple_parse
from services.pdf_to_excel._helpers import (
    _detect_account_last4,
    _detect_bank_name,
    _detect_period,
    _norm_text,
)
from services.pdf_to_excel._storage import ensure_parent_dir

logger = logging.getLogger(__name__)

HEAD_MOV = [
    "fecha",
    "descripcion",
    "deposito",
    "retiro",
    "saldo",
    "tipo",
    "categoria",
    "contraparte_hint",
    "metodo_hint",
    "referencia",
    "cve_rastreo",
    "rfc_encontrado",
    "confidence_score",
    "source_page_first",
]


@dataclass
class ConvertMeta:
    rows: int
    raw_lines: int
    mode: str  # 'parsed' | 'raw'


def convert_pdf_to_xlsx(
    pdf_path_abs: str,
    xlsx_path_abs: str,
    issuer_id: int | None = None,
    statement_id: int | None = None,
) -> dict[str, Any]:
    """
    Convierte un PDF de estado de cuenta a Excel.
    - Usa parser Banorte (solo montos con 2 decimales, sección DETALLE DE MOVIMIENTOS (PESOS)).
    - Si statement_id y issuer_id se pasan, persiste movimientos en bank_movements con dedupe por hash.
    - Genera XLSX: Movimientos, Gastos, Ingresos, Resumen, RAW.
    """
    try:
        import pdfplumber  # lazy import (dependencia opcional hasta instalar)
    except ModuleNotFoundError:
        pdfplumber = None
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if not os.path.isfile(pdf_path_abs):
        raise FileNotFoundError("PDF no encontrado")

    def _freeze_filter(ws, ncols: int, nrows: int) -> None:
        ws.freeze_panes = "A2"
        if ncols >= 1 and nrows >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"

    def _autosize(ws, ncols: int, max_width: int = 60) -> None:
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            best = 0
            for cell in ws[letter]:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > best:
                    best = len(s)
            ws.column_dimensions[letter].width = min(max(10, best + 2), max_width)

    def _write_table(ws, headers: list[str], rows: list[list[Any]], money_cols=()) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in rows:
            ws.append(r)
        nrows = 1 + len(rows)
        ncols = len(headers)
        _freeze_filter(ws, ncols, nrows)
        for cidx in money_cols:
            for r in range(2, nrows + 1):
                cell = ws.cell(row=r, column=cidx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
        _autosize(ws, ncols)

    if pdfplumber is None:
        ensure_parent_dir(xlsx_path_abs)
        wb = Workbook()
        wb.remove(wb.active)
        _write_table(wb.create_sheet("Movimientos"), HEAD_MOV, [], money_cols=[3, 4, 5])
        _write_table(wb.create_sheet("Gastos"), HEAD_MOV, [], money_cols=[3, 4, 5])
        _write_table(wb.create_sheet("Ingresos"), HEAD_MOV, [], money_cols=[3, 4, 5])
        ws_r = wb.create_sheet("Resumen")
        ws_r.append(["campo", "valor"])
        ws_r["A1"].font = Font(bold=True)
        ws_r["B1"].font = Font(bold=True)
        ws_r.append(["error", "Falta dependencia: pdfplumber (instala requirements.txt)"])
        ws_r.freeze_panes = "A2"
        _write_table(wb.create_sheet("RAW"), ["page", "line", "text"], [], money_cols=[])
        wb.save(xlsx_path_abs)
        return {
            "rows": 0, "raw_lines": 0, "mode": "raw", "error": "pdfplumber_missing",
            "period_start": "", "period_end": "", "bank_name": "", "account_last4": "",
            "transactions": [], "processed_count": 0,
            "total_ingresos": 0.0, "total_gastos": 0.0,
            "sin_factura_count": 0, "movements_count": 0,
            "ingresos_total": 0.0, "gastos_total": 0.0, "sin_parse_count": 0,
        }

    raw_rows: list[dict[str, Any]] = []
    with pdfplumber.open(pdf_path_abs) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            lines = [ln.strip() for ln in text.splitlines() if (ln or "").strip()]
            for li, ln in enumerate(lines, start=1):
                raw_rows.append({"Page": page_idx, "Line": li, "Text": ln})

    all_text = " ".join(str(r.get("Text") or "") for r in raw_rows)
    all_text_norm = _norm_text(all_text)
    period_start, period_end = _detect_period(all_text_norm)
    bank_name = _detect_bank_name(all_text_norm)
    account_last4 = _detect_account_last4(all_text_norm)

    ensure_parent_dir(xlsx_path_abs)

    # -------- pipeline robusta (Banorte o similares) --------
    from config import DEV_MODE
    from services.bank.bank_statement_parser import parse_bank_statement, write_debug_json

    debug_on = bool(DEV_MODE) and (os.environ.get("BANK_PARSER_DEBUG", "0").strip() == "1")
    parsed = parse_bank_statement(raw_rows, debug=debug_on)
    txs: list[dict[str, Any]] = parsed.transactions or []
    metrics: dict[str, Any] = parsed.metrics or {}

    # fallback (para bancos que no tengan DETALLE DE MOVIMIENTOS): parser simple por línea con fecha
    if int(metrics.get("movements_count") or 0) <= 0:
        txs, metrics = fallback_simple_parse(raw_rows)

    movements_count = int(metrics.get("movements_count") or 0)
    sin_parse_count = int(metrics.get("sin_parse_count") or 0)
    total_ingresos = float(metrics.get("total_ingresos") or 0.0)
    total_gastos = float(metrics.get("total_gastos") or 0.0)
    low_confidence_count = int(metrics.get("low_confidence_count") or 0)
    mode = "parsed" if movements_count > 0 else "raw"

    # Persistir movimientos en DB con dedupe (solo INGRESO/GASTO)
    if statement_id and issuer_id and txs:
        try:
            from services.bank.bank_statement_parser import upsert_bank_movements
            upsert_bank_movements(int(issuer_id), int(statement_id), txs)
        except Exception:
            logger.exception("bank_parser: no se pudieron guardar movimientos en DB")

    logger.info(
        "bank_parser: sections=%s grouped=%s movements=%s saldo=%s rfc=%s rastreo=%s avg_conf=%.1f",
        int(metrics.get("sections_detected") or 0),
        int(metrics.get("transactions_grouped") or 0),
        movements_count,
        int(metrics.get("saldo_count") or 0),
        int(metrics.get("rfc_count") or 0),
        int(metrics.get("rastreo_count") or 0),
        float(metrics.get("avg_confidence") or 0.0),
    )

    if debug_on and parsed.debug_payload:
        try:
            write_debug_json(parsed.debug_payload, xlsx_path_abs + ".debug.json")
        except Exception:
            logger.exception("bank_parser: no se pudo escribir debug json")

    # -------- export Excel (5 hojas) --------
    wb = Workbook()
    wb.remove(wb.active)

    rows_mov: list[list[Any]] = []
    for t in txs:
        rows_mov.append(
            [
                t.get("fecha") or "",
                t.get("descripcion_full") or "",
                t.get("deposito") if isinstance(t.get("deposito"), (int, float)) else "",
                t.get("retiro") if isinstance(t.get("retiro"), (int, float)) else "",
                t.get("saldo") if isinstance(t.get("saldo"), (int, float)) else "",
                t.get("tipo") or "DESCONOCIDO",
                t.get("categoria") or "",
                t.get("contraparte_hint") or "",
                t.get("metodo_hint") or "OTRO",
                t.get("referencia") or "",
                t.get("cve_rastreo") or "",
                t.get("rfc_encontrado") or "",
                int(t.get("confidence_score") or 0),
                int(t.get("source_page_first") or 0),
            ]
        )

    ws_mov = wb.create_sheet("Movimientos")
    _write_table(ws_mov, HEAD_MOV, rows_mov, money_cols=[3, 4, 5])

    ws_g = wb.create_sheet("Gastos")
    gastos_rows = [r for r in rows_mov if r[5] == "GASTO"]
    _write_table(ws_g, HEAD_MOV, gastos_rows, money_cols=[3, 4, 5])

    ws_i = wb.create_sheet("Ingresos")
    ingresos_rows = [r for r in rows_mov if r[5] == "INGRESO"]
    _write_table(ws_i, HEAD_MOV, ingresos_rows, money_cols=[3, 4, 5])

    ws_r = wb.create_sheet("Resumen")
    ws_r.append(["campo", "valor"])
    ws_r["A1"].font = Font(bold=True)
    ws_r["B1"].font = Font(bold=True)
    ws_r.append(["movements_count", movements_count])
    ws_r.append(["sin_parse_count", sin_parse_count])
    ws_r.append(["total_ingresos", total_ingresos])
    ws_r.append(["total_gastos", total_gastos])
    ws_r.append(["neto", total_ingresos - total_gastos])
    ws_r.append(["saldo_count", int(metrics.get("saldo_count") or 0)])
    ws_r.append(["rfc_count", int(metrics.get("rfc_count") or 0)])
    ws_r.append(["rastreo_count", int(metrics.get("rastreo_count") or 0)])
    ws_r.append(["avg_confidence", float(metrics.get("avg_confidence") or 0.0)])
    for r in range(2, ws_r.max_row + 1):
        k = ws_r.cell(row=r, column=1).value
        v = ws_r.cell(row=r, column=2).value
        if isinstance(v, (int, float)) and k in ("total_ingresos", "total_gastos", "neto"):
            ws_r.cell(row=r, column=2).number_format = '"$"#,##0.00'
    ws_r.freeze_panes = "A2"
    ws_r.column_dimensions["A"].width = 28
    ws_r.column_dimensions["B"].width = 18

    def _top(rows: list[list[Any]], key_idx: int, amt_idx: int, topn: int = 10) -> list[tuple[str, float]]:
        agg: dict[str, float] = {}
        for rr in rows:
            k = (rr[key_idx] or "").strip() or "—"
            amt = rr[amt_idx] if isinstance(rr[amt_idx], (int, float)) else 0.0
            agg[k] = agg.get(k, 0.0) + float(amt)
        return sorted(agg.items(), key=lambda x: x[1], reverse=True)[:topn]

    top_cats = _top(gastos_rows, 6, 3, 10)  # categoria, retiro
    ws_r.append([])
    ws_r.append(["Top categorías (gasto)", "monto"])
    ws_r["A" + str(ws_r.max_row)].font = Font(bold=True)
    ws_r["B" + str(ws_r.max_row)].font = Font(bold=True)
    for k, v in top_cats:
        ws_r.append([k, v])
        ws_r.cell(row=ws_r.max_row, column=2).number_format = '"$"#,##0.00'

    ws_raw = wb.create_sheet("RAW")
    raw_table = [[r["Page"], r["Line"], r["Text"]] for r in raw_rows]
    _write_table(ws_raw, ["page", "line", "text"], raw_table, money_cols=[])

    wb.save(xlsx_path_abs)

    transactions_for_db = [
        {
            "fecha": t.get("fecha") or "",
            "descripcion": (t.get("descripcion_full") or "")[:2000],
            "deposito": t.get("deposito") if isinstance(t.get("deposito"), (int, float)) else None,
            "retiro": t.get("retiro") if isinstance(t.get("retiro"), (int, float)) else None,
            "saldo": t.get("saldo") if isinstance(t.get("saldo"), (int, float)) else None,
            "tipo": t.get("tipo") or "DESCONOCIDO",
            "categoria": (t.get("categoria") or "")[:200],
            "metodo_hint": (t.get("metodo_hint") or "")[:64],
            "contraparte_hint": (t.get("contraparte_hint") or "")[:200],
            "rfc_encontrado": (t.get("rfc_encontrado") or "")[:20],
            "confidence_score": int(t.get("confidence_score") or 0),
            "source_page_first": int(t.get("source_page_first") or 0),
        }
        for t in txs
    ]

    return {
        "rows": len(rows_mov),
        "raw_lines": len(raw_rows),
        "mode": mode,
        "period_start": period_start,
        "period_end": period_end,
        "bank_name": bank_name,
        "account_last4": account_last4,
        "transactions": transactions_for_db,
        # compat UI previa
        "processed_count": movements_count,
        "total_ingresos": total_ingresos,
        "total_gastos": total_gastos,
        "sin_factura_count": sin_parse_count,
        # métricas solicitadas
        "movements_count": movements_count,
        "ingresos_total": total_ingresos,
        "gastos_total": total_gastos,
        "sin_parse_count": sin_parse_count,
        "low_confidence_count": low_confidence_count,
        "quality": {
            "sections_detected": int(metrics.get("sections_detected") or 0),
            "transactions_grouped": int(metrics.get("transactions_grouped") or 0),
            "saldo_count": int(metrics.get("saldo_count") or 0),
            "rfc_count": int(metrics.get("rfc_count") or 0),
            "rastreo_count": int(metrics.get("rastreo_count") or 0),
            "avg_confidence": float(metrics.get("avg_confidence") or 0.0),
            "low_confidence_count": low_confidence_count,
        },
    }
