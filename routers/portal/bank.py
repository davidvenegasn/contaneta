"""Portal bank routes."""
import hashlib
import io
import json
import logging
import os
import secrets
from datetime import datetime, timezone
from typing import Any, Optional

from fastapi import Body, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response

from config import (
    BASE_DIR,
)
from database import db, has_column, table_exists
from routers.deps import get_portal_issuer
from routers.portal._helpers import (
    MAX_LIST_OFFSET,
    _db_row_to_dict,
    _strip_date_from_description,
    render_portal,
    ym_now,
)
from services import audit, file_access_log
from services.action_log import log_action
from services.auth import csrf as csrf_service
from services.auth import rate_limit as rate_limit_service
from services.bank.bank_accounts import create_account as bank_create_account
from services.bank.bank_accounts import delete_account as bank_delete_account
from services.bank.bank_accounts import list_active_accounts as bank_list_accounts
from services.bank.bank_accounts import list_active_accounts_raw as bank_list_accounts_raw
from services.bank.bank_accounts import list_all_accounts as bank_list_all_accounts
from services.bank.bank_accounts import update_account as bank_update_account
from services.bank.bank_cfdi_matching import confirm_match as match_confirm
from services.bank.bank_cfdi_matching import find_cfdi_candidates, save_suggested_matches
from services.bank.bank_cfdi_matching import reject_match as match_reject
from services.bank.bank_own_accounts import detect_own_account_transfer, reclassify_own_transfers_by_rfc
from services.bank.bank_parse_preview import parse_bank_pdf_to_movements_preview, reclassify_movements
from services.bank.bank_preview_models import compute_dedupe_fingerprint
from services.bank.bank_preview_pipeline import parse_bank_statement_preview
from services.bank.bank_statement_ingest import (
    commit_preview_to_db,
    extract_statement_metadata,
    ingest_bank_statement,
)
from services.pdf_to_excel import convert_pdf_to_xlsx, ensure_parent_dir, get_storage_root, safe_join
from services.portal_errors import portal_error_type
from services.sat.sat_sync import get_month_totals, get_sat_sync_status
from services.ym_helpers import sanitize_ym, shift_ym, ym_to_label

logger = logging.getLogger(__name__)

_get_month_totals = get_month_totals
_get_sat_sync_status = get_sat_sync_status


def register_bank_routes(router, templates):
    """Register Bank routes on the portal router."""

    def _render_portal(request, **kwargs):
        return render_portal(templates, request, **kwargs)

    @router.get("/bancos", response_class=RedirectResponse)
    def portal_bancos_redirect():
        """Redirigir a la página de convertir estado de cuenta."""
        return RedirectResponse(url="/portal/convertir-edo-cuenta", status_code=302)

    @router.get("/convertir-edo-cuenta", response_class=HTMLResponse)
    def portal_convertir_edo_cuenta(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Página única: arrastrar PDF, convertir a Excel y ver movimientos (sin pestañas ni hub)."""
        try:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_bank_pdf_to_excel.html",
                active_page="convertir_edo_cuenta",
                title="Convertir Edo. de Cuenta",
            )
        except Exception as e:
            logger.exception("convertir-edo-cuenta: error en render completo (%s), usando página mínima", e)
            try:
                return templates.TemplateResponse(
                    request,
                    "portal_convertir_edo_cuenta_minimal.html",
                    {
                        "csrf_token": csrf_service.generate_csrf_token(),
                        "preview_movements": [],
                        "preview_summary": {},
                    },
                    status_code=200,
                )
            except Exception as e2:
                logger.exception("convertir-edo-cuenta: fallback mínima también falló: %s", e2)
                raise

    # ---------- Bank: PDF → Excel (estado de cuenta) ----------
    MAX_BANK_PDF_SIZE = 15 * 1024 * 1024  # 15MB
    MAX_BANK_PDF_FILES = 10
    MAX_BANK_PDF_TOTAL_SIZE = 50 * 1024 * 1024  # 50MB total multi-upload

    def _ensure_bank_exports_table(conn) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bank_pdf_exports (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              issuer_id INTEGER NOT NULL,
              file_id TEXT NOT NULL,
              pdf_path TEXT NOT NULL,
              xlsx_path TEXT NOT NULL,
              meta_json TEXT,
              created_at TEXT NOT NULL DEFAULT (datetime('now')),
              UNIQUE(issuer_id, file_id)
            );
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_pdf_exports_issuer ON bank_pdf_exports(issuer_id, created_at);")

    def _ensure_bank_statements_table(conn) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bank_statements (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              issuer_id INTEGER NOT NULL,
              bank_name TEXT,
              account_last4 TEXT,
              period_start TEXT,
              period_end TEXT,
              source_pdf_path TEXT NOT NULL,
              source_pdf_sha256 TEXT NOT NULL,
              created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            """
        )
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_bank_statements_issuer_sha ON bank_statements(issuer_id, source_pdf_sha256);")

    def _movement_dedup_hash(issuer_id: int, fecha: str, descripcion: str, deposito: Optional[float], retiro: Optional[float]) -> str:
        """Hash para deduplicar movimientos: mismo issuer + fecha + concepto + montos = mismo movimiento."""
        dep = f"{float(deposito or 0):.2f}"
        ret = f"{float(retiro or 0):.2f}"
        desc = (descripcion or "").strip()[:500].replace("\r", " ").replace("\n", " ")
        payload = f"{issuer_id}|{fecha or ''}|{desc}|{dep}|{ret}"
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def _ensure_bank_movements_table(conn) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bank_movements (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              issuer_id INTEGER NOT NULL,
              statement_file_id TEXT NOT NULL DEFAULT '0',
              movement_hash TEXT,
              fecha TEXT,
              descripcion TEXT,
              raw_description TEXT,
              normalized_description TEXT,
              deposito REAL,
              retiro REAL,
              saldo REAL,
              tipo TEXT,
              categoria TEXT,
              metodo_hint TEXT,
              contraparte_hint TEXT,
              reference_text TEXT,
              rfc_encontrado TEXT,
              confidence_score INTEGER,
              source_page_first INTEGER,
              period_month TEXT,
              created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            """
        )
        # Add columns that may be missing in older databases
        for col, coltype in [
            ("movement_hash", "TEXT"),
            ("raw_description", "TEXT"),
            ("normalized_description", "TEXT"),
            ("reference_text", "TEXT"),
            ("period_month", "TEXT"),
        ]:
            if not has_column(conn, "bank_movements", col):
                try:
                    conn.execute(f"ALTER TABLE bank_movements ADD COLUMN {col} {coltype};")
                except Exception:
                    pass
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_statement ON bank_movements(issuer_id, statement_file_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_tipo ON bank_movements(issuer_id, tipo);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_categoria ON bank_movements(issuer_id, categoria);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_fecha ON bank_movements(issuer_id, fecha);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_confidence ON bank_movements(issuer_id, confidence_score);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_period ON bank_movements(issuer_id, period_month);")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_bank_movements_issuer_hash ON bank_movements(issuer_id, movement_hash) WHERE movement_hash IS NOT NULL;")

    @router.get("/bank/pdf-to-excel", response_class=RedirectResponse)
    def portal_bank_pdf_to_excel_redirect():
        """Redirigir al nombre canónico de la página."""
        return RedirectResponse(url="/portal/convertir-edo-cuenta", status_code=302)

    @router.post("/bank/pdf-to-excel/preview-json", response_class=JSONResponse)
    async def portal_bank_pdf_preview_json(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        file: UploadFile = File(...),
    ):
        """Parsea PDF Banorte y devuelve movimientos en JSON. Sin DB, sin guardar. Para mostrar listado en la misma página."""
        if rate_limit_service.is_rate_limited(request, "bank_pdf_preview", max_attempts=10, window_seconds=60):
            raise HTTPException(status_code=429, detail="Demasiadas solicitudes. Intenta en un momento.")
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        filename = (file.filename or "").strip().lower()
        if not filename.endswith(".pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser .pdf")
        content_type = (file.content_type or "").lower().strip()
        if content_type and content_type not in ("application/pdf", "application/x-pdf"):
            raise HTTPException(status_code=400, detail="Tipo de archivo inválido (solo PDF).")
        size = 0
        chunks: list[bytes] = []
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BANK_PDF_SIZE:
                raise HTTPException(status_code=400, detail="El PDF excede el máximo de 15MB.")
            chunks.append(chunk)
        if size <= 0:
            raise HTTPException(status_code=400, detail="El PDF está vacío.")
        # Magic bytes mínimo: evita 500 en el parser legacy por archivo no-PDF
        if chunks and not chunks[0].startswith(b"%PDF-"):
            raise HTTPException(status_code=400, detail="El archivo no parece ser un PDF válido.")
        import tempfile
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                for ch in chunks:
                    tmp.write(ch)
                tmp_path = tmp.name
            try:
                result = parse_bank_pdf_to_movements_preview(tmp_path, preset="conservative")
            except Exception:
                raise HTTPException(status_code=400, detail="No pudimos leer el PDF. Verifica que sea un estado de cuenta válido.")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
        movements = result.get("movements") or []
        summary = result.get("summary") or {}
        if summary.get("error"):
            return JSONResponse(
                {"ok": False, "detail": summary.get("error"), "movements": [], "summary": summary},
                status_code=400,
            )
        if not movements:
            logger.warning(
                "preview-json: 0 movements from PDF '%s' (rows=%s, sections=%s, txs=%s, bank=%s)",
                filename,
                summary.get("raw_rows_count", "?"),
                summary.get("sections_detected", "?"),
                summary.get("txs_grouped_count", "?"),
                summary.get("bank_name", "?"),
            )
        return JSONResponse({"ok": True, "movements": movements, "summary": summary})

    @router.post("/bank/pdf-to-excel/preview-multi", response_class=JSONResponse)
    async def portal_bank_pdf_preview_multi(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        files: list[UploadFile] = File(..., description="PDFs de estados de cuenta"),
    ):
        """Multi-PDF preview: procesa varios PDFs, consolidado en memoria. Sin DB."""
        if rate_limit_service.is_rate_limited(request, "bank_pdf_preview", max_attempts=10, window_seconds=60):
            raise HTTPException(status_code=429, detail="Demasiadas solicitudes. Intenta en un momento.")
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        if not files or len(files) > MAX_BANK_PDF_FILES:
            raise HTTPException(
                status_code=400,
                detail=f"Envía entre 1 y {MAX_BANK_PDF_FILES} archivos PDF.",
            )
        all_movements: list[dict] = []
        files_summary: list[dict] = []
        file_errors: list[dict] = []
        file_warnings: list[dict] = []
        total_size = 0
        for idx, uf in enumerate(files):
            fn = (uf.filename or "").strip()
            if not fn.lower().endswith(".pdf"):
                file_errors.append({"file_name": fn or f"archivo_{idx + 1}", "error": "El archivo debe ser .pdf"})
                continue
            chunks = []
            size = 0
            while True:
                chunk = await uf.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_BANK_PDF_SIZE:
                    file_errors.append({"file_name": fn, "error": "El PDF excede el máximo de 15MB."})
                    break
                chunks.append(chunk)
            if size > MAX_BANK_PDF_SIZE:
                continue
            total_size += size
            if total_size > MAX_BANK_PDF_TOTAL_SIZE:
                file_errors.append({"file_name": fn, "error": "Se superó el tamaño total permitido (50MB)."})
                continue
            if size <= 0:
                file_errors.append({"file_name": fn, "error": "El PDF está vacío."})
                continue
            pdf_bytes = b"".join(chunks)
            if not pdf_bytes.startswith(b"%PDF"):
                file_errors.append({"file_name": fn, "error": "El archivo no es un PDF válido."})
                continue
            result = parse_bank_statement_preview(pdf_bytes, file_name=fn, file_index=idx)
            movs = result.get("movements") or []
            fs = result.get("file_summary") or {}
            err = result.get("file_error")
            warns = result.get("file_warnings") or []
            if err:
                file_errors.append({"file_name": fn, "error": err})
            if warns:
                file_warnings.append({"file_name": fn, "warnings": warns})
            files_summary.append(fs)
            base_idx = len(all_movements)
            for i, m in enumerate(movs):
                m["_global_idx"] = base_idx + i + 1
            all_movements.extend(movs)
        # Marcar duplicados por fingerprint (misma fecha+monto+concepto+archivo)
        fp_counts: dict[str, int] = {}
        for m in all_movements:
            fp = m.get("dedupe_fingerprint") or compute_dedupe_fingerprint(m)
            m["dedupe_fingerprint"] = fp
            fp_counts[fp] = fp_counts.get(fp, 0) + 1
        for m in all_movements:
            if fp_counts.get(m["dedupe_fingerprint"], 0) > 1:
                m["posible_duplicado"] = True
                w = m.get("warnings") or []
                if "Posible duplicado en esta carga" not in w:
                    w.append("Posible duplicado en esta carga")
                m["warnings"] = w
        issuer_id = int(issuer.get("id") or 0)
        user_accounts = bank_list_accounts_raw(issuer_id) if issuer_id > 0 else []
        statement_owner_name = None
        statement_owner_rfc = None
        if files_summary:
            fs = next((f for f in files_summary if f.get("account_holder_name")), None)
            if fs:
                statement_owner_name = fs.get("account_holder_name")
                statement_owner_rfc = fs.get("account_holder_rfc")
        issuer_rfc = (issuer.get("rfc") or "").strip().upper()
        for m in all_movements:
            detect_own_account_transfer(m, user_accounts, statement_owner_name, statement_owner_rfc, issuer_rfc=issuer_rfc)
        total_ing = sum(m.get("monto_deposito") or 0 for m in all_movements)
        total_gas = sum(m.get("monto_retiro") or 0 for m in all_movements)
        total_ing_impactan = sum(
            (m.get("monto_deposito") or 0) for m in all_movements
            if m.get("impacta_contabilidad", True) and (m.get("tipo_movimiento") or "").upper() == "INGRESO"
        )
        total_gas_impactan = sum(
            (m.get("monto_retiro") or 0) for m in all_movements
            if m.get("impacta_contabilidad", True) and (m.get("tipo_movimiento") or "").upper() == "GASTO"
        )
        count_in = sum(1 for m in all_movements if (m.get("tipo_movimiento") or "").upper() == "INGRESO")
        count_out = sum(1 for m in all_movements if (m.get("tipo_movimiento") or "").upper() == "GASTO")
        count_info = sum(1 for m in all_movements if (m.get("tipo_movimiento") or "").upper() == "INFO")
        count_fin = sum(1 for m in all_movements if m.get("es_movimiento_financiero"))
        low_conf = sum(1 for m in all_movements if int(m.get("confianza_clasificacion") or 0) < 60)
        count_revisar = sum(1 for m in all_movements if m.get("requiere_revision"))
        count_duplicados = sum(1 for m in all_movements if m.get("posible_duplicado"))
        global_summary = {
            "files_processed": len(files_summary),
            "files_with_errors": len(file_errors),
            "total_movements": len(all_movements),
            "total_ingresos": round(total_ing, 2),
            "total_gastos": round(total_gas, 2),
            "total_ingresos_que_impactan": round(total_ing_impactan, 2),
            "total_gastos_que_impactan": round(total_gas_impactan, 2),
            "count_ingreso": count_in,
            "count_gasto": count_out,
            "count_info": count_info,
            "count_financiero": count_fin,
            "count_low_confidence": low_conf,
            "count_requiere_revision": count_revisar,
            "count_duplicados": count_duplicados,
        }
        # Auto-save movements to DB (no manual step, no bank account required)
        auto_save_result = {"saved": False}
        if all_movements and issuer_id > 0:
            try:
                from database import db as _db
                from services.bank.bank_statement_parser import upsert_bank_movements
                _conn = _db()
                try:
                    _ensure_bank_movements_table(_conn)
                    _conn.commit()
                finally:
                    _conn.close()
                # Map preview fields → upsert_bank_movements format
                mapped = []
                for m in all_movements:
                    tipo_raw = (m.get("tipo_movimiento") or "INFO").upper()
                    if tipo_raw not in ("INGRESO", "GASTO"):
                        continue
                    mapped.append({
                        "tipo": tipo_raw,
                        "fecha": m.get("fecha") or "",
                        "descripcion": m.get("concepto_resumen") or m.get("raw_text_original") or "",
                        "descripcion_full": m.get("raw_text_original") or m.get("concepto_resumen") or "",
                        "descripcion_norm": m.get("raw_text_normalized") or m.get("concepto_resumen") or "",
                        "deposito": m.get("monto_deposito") or 0,
                        "retiro": m.get("monto_retiro") or 0,
                        "saldo": m.get("saldo"),
                        "categoria": m.get("categoria_sugerida") or "",
                        "metodo_hint": m.get("canal") or "",
                        "contraparte_hint": m.get("contraparte_nombre") or "",
                        "referencia": m.get("referencia") or m.get("cve_rastreo") or "",
                        "rfc_encontrado": m.get("rfc_detectado") or "",
                        "confidence_score": m.get("confianza_clasificacion") or 0,
                        "source_page_first": m.get("page_number"),
                    })
                if mapped:
                    count = upsert_bank_movements(issuer_id, statement_id=0, transactions=mapped)
                    # Derive period from first movement
                    first_fecha = (mapped[0].get("fecha") or "")[:7]
                    auto_save_result = {
                        "saved": True,
                        "inserted_count": count,
                        "duplicate_movements_count": 0,
                        "period_month": first_fecha or None,
                    }
                    logger.info("Auto-saved %d movements for issuer %s", count, issuer_id)
            except Exception:
                logger.exception("Auto-save bank movements failed for issuer %s", issuer_id)
        return JSONResponse({
            "ok": True,
            "movements": all_movements,
            "global_summary": global_summary,
            "files_summary": files_summary,
            "file_errors": file_errors,
            "file_warnings": file_warnings,
            "auto_save": auto_save_result,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        })

    @router.get("/bank/accounts", response_class=JSONResponse)
    def portal_bank_accounts_list(issuer: dict = Depends(get_portal_issuer)):
        """Lista cuentas bancarias del usuario (para detectar cuentas propias)."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        accounts = bank_list_accounts(int(issuer["id"]))
        return JSONResponse({"ok": True, "accounts": accounts})

    @router.get("/bank/accounts/manage", response_class=HTMLResponse)
    def portal_bank_accounts_manage(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Pantalla simple: Mis cuentas bancarias (config para detectar traspasos propios)."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        accounts = bank_list_all_accounts(issuer_id)
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_bank_accounts.html",
            active_page="bank_accounts",
            title="Mis cuentas bancarias",
            extra={"accounts": accounts or []},
        )

    @router.post("/bank/accounts", response_class=JSONResponse)
    def portal_bank_accounts_create(
        issuer: dict = Depends(get_portal_issuer),
        alias: str = Body(..., embed=True),
        bank_name: str = Body(..., embed=True),
        clabe: Optional[str] = Body(None, embed=True),
        account_last4: Optional[str] = Body(None, embed=True),
        holder_name: Optional[str] = Body(None, embed=True),
        rfc_titular: Optional[str] = Body(None, embed=True),
        is_active: bool = Body(True, embed=True),
    ):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        created = bank_create_account(
            int(issuer["id"]), alias=alias, bank_name=bank_name,
            clabe=clabe, account_last4=account_last4, holder_name=holder_name,
            rfc_titular=rfc_titular, is_active=is_active,
        )
        if created.get("error"):
            raise HTTPException(status_code=500, detail=created["error"])
        return JSONResponse({"ok": True, "account": created})

    @router.put("/bank/accounts/{account_id}", response_class=JSONResponse)
    def portal_bank_accounts_update(
        account_id: int,
        issuer: dict = Depends(get_portal_issuer),
        payload: dict = Body(...),
    ):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        allowed = {"alias", "bank_name", "clabe", "account_last4", "holder_name", "rfc_titular", "is_active"}
        kwargs = {k: v for k, v in payload.items() if k in allowed}
        if "account_last4" in kwargs and kwargs["account_last4"]:
            kwargs["account_last4"] = str(kwargs["account_last4"]).strip()[:4]
        updated = bank_update_account(account_id, int(issuer["id"]), **kwargs)
        if not updated:
            raise HTTPException(status_code=404, detail="Cuenta no encontrada")
        return JSONResponse({"ok": True, "account": updated})

    @router.delete("/bank/accounts/{account_id}", response_class=JSONResponse)
    def portal_bank_accounts_delete(account_id: int, issuer: dict = Depends(get_portal_issuer)):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        deleted = bank_delete_account(account_id, int(issuer["id"]))
        if not deleted:
            raise HTTPException(status_code=404, detail="Cuenta no encontrada")
        return JSONResponse({"ok": True})

    @router.post("/bank/statements/ingest", response_class=JSONResponse)
    async def portal_bank_statements_ingest(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        file: UploadFile = File(...),
        bank_account_id: int = Form(..., description="ID de cuenta bancaria del issuer"),
    ):
        """Ingesta estado de cuenta con validación RFC/cuenta. Fases 2+3."""
        if rate_limit_service.is_rate_limited(request, "bank_ingest", max_attempts=10, window_seconds=60):
            raise HTTPException(status_code=429, detail="Demasiadas solicitudes. Intenta en un momento.")
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        fn = (file.filename or "").strip()
        if not fn.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser .pdf")
        size = 0
        chunks = []
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BANK_PDF_SIZE:
                raise HTTPException(status_code=400, detail="El PDF excede el máximo de 15MB.")
            chunks.append(chunk)
        if size <= 0:
            raise HTTPException(status_code=400, detail="El PDF está vacío.")
        pdf_bytes = b"".join(chunks)
        if not pdf_bytes.startswith(b"%PDF"):
            raise HTTPException(status_code=400, detail="El archivo no es un PDF válido.")
        sha = hashlib.sha256(pdf_bytes).hexdigest()
        result = parse_bank_statement_preview(pdf_bytes, file_name=fn, file_index=0)
        if result.get("file_error"):
            return JSONResponse(
                {"ok": False, "rejection_reason": result["file_error"], "status": "parse_error"},
                status_code=400,
            )
        storage_root = get_storage_root(BASE_DIR)
        uploads_rel = os.path.join("uploads", str(issuer_id), "bank")
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        pdf_name = f"{stamp}_{sha[:12]}.pdf"
        pdf_rel_path = os.path.join(uploads_rel, pdf_name)
        pdf_abs_path = safe_join(storage_root, pdf_rel_path)
        ensure_parent_dir(pdf_abs_path)
        with open(pdf_abs_path, "wb") as f:
            f.write(pdf_bytes)
        expected_rfc = (issuer.get("rfc") or "").strip()
        ingest_result = ingest_bank_statement(
            issuer_id=issuer_id,
            bank_account_id=bank_account_id,
            pdf_path=pdf_rel_path,
            pdf_sha256=sha,
            source_file_name=fn,
            preview_result=result,
            expected_issuer_rfc=expected_rfc,
        )
        if not ingest_result.get("ok"):
            return JSONResponse(
                {
                    "ok": False,
                    "rejection_reason": ingest_result.get("rejection_reason", "Error desconocido"),
                    "status": ingest_result.get("status", "error"),
                },
                status_code=400,
            )
        statement_id = ingest_result.get("statement_id")
        movements_count = ingest_result.get("movements_count", 0)
        if statement_id and movements_count > 0 and table_exists(db(), "bank_invoice_matches"):
            try:
                conn = db()
                rows = conn.execute(
                    "SELECT id, deposito, retiro, amount, fecha, descripcion, rfc_encontrado, counterparty_rfc_detected, requires_cfdi FROM bank_movements WHERE issuer_id = ? AND bank_statement_id = ?",
                    (issuer_id, statement_id),
                ).fetchall()
                conn.close()
                for r in rows:
                    r = _db_row_to_dict(r)
                    if int(r.get("requires_cfdi") or 0):
                        mov = r
                        candidates = find_cfdi_candidates(issuer_id, mov, direction="received", limit=5)
                        if candidates:
                            save_suggested_matches(issuer_id, int(r["id"]), candidates, "payment")
            except Exception as e:
                logger.exception("bank ingest: matching post-insert failed: %s", e)
        log_action(request, "bank_statement_ingest", issuer_id=issuer_id, entity_id=statement_id)
        return JSONResponse({
            "ok": True,
            "statement_id": statement_id,
            "movements_count": movements_count,
            "inserted_count": ingest_result.get("inserted_count", movements_count),
            "duplicate_movements_count": ingest_result.get("duplicate_movements_count", 0),
            "duplicate": ingest_result.get("duplicate", False),
        })

    @router.post("/bank/matches/{match_id}/confirm", response_class=JSONResponse)
    def portal_bank_match_confirm(match_id: int, issuer: dict = Depends(get_portal_issuer)):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        ok = match_confirm(match_id, int(issuer["id"]))
        if not ok:
            raise HTTPException(status_code=404, detail="Match no encontrado")
        return JSONResponse({"ok": True, "status": "confirmed"})

    @router.post("/bank/matches/{match_id}/reject", response_class=JSONResponse)
    def portal_bank_match_reject(match_id: int, issuer: dict = Depends(get_portal_issuer)):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        ok = match_reject(match_id, int(issuer["id"]))
        if not ok:
            raise HTTPException(status_code=404, detail="Match no encontrado")
        return JSONResponse({"ok": True, "status": "rejected"})

    @router.patch("/bank/movements/{movement_id}", response_class=JSONResponse)
    async def portal_bank_movement_update(
        movement_id: int,
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
    ):
        """Actualiza descripción y/o categoría de un movimiento (mismo comportamiento que en convertir edo. de cuenta)."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        try:
            body = await request.json() if request.headers.get("content-type", "").strip().startswith("application/json") else {}
        except Exception:
            body = {}
        descripcion = body.get("descripcion")
        categoria = body.get("categoria")
        if descripcion is None and categoria is None:
            return JSONResponse({"ok": True, "updated": False})
        conn = db()
        try:
            row = conn.execute(
                "SELECT id FROM bank_movements WHERE id = ? AND issuer_id = ?",
                (movement_id, issuer_id),
            ).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Movimiento no encontrado")
            updates = []
            params: list = []
            if descripcion is not None:
                updates.append("descripcion = ?")
                params.append(str(descripcion).strip() if descripcion else "")
            if categoria is not None:
                updates.append("categoria = ?")
                params.append(str(categoria).strip() if categoria else "")
            if not updates:
                return JSONResponse({"ok": True, "updated": False})
            params.extend([movement_id, issuer_id])
            conn.execute(
                "UPDATE bank_movements SET " + ", ".join(updates) + " WHERE id = ? AND issuer_id = ?",
                params,
            )
            conn.commit()
            return JSONResponse({"ok": True, "updated": True})
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    @router.post("/bank/movements/delete-all", response_class=JSONResponse)
    def portal_bank_movements_delete_all(issuer: dict = Depends(get_portal_issuer)):
        """Borra todos los movimientos bancarios del emisor actual. Requiere confirmación en el cliente."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        conn = db()
        try:
            cur = conn.execute("DELETE FROM bank_movements WHERE issuer_id = ?", (issuer_id,))
            deleted = cur.rowcount
            conn.commit()
            return JSONResponse({"ok": True, "deleted": deleted})
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    @router.post("/bank/pdf-to-excel/upload")
    async def portal_bank_pdf_to_excel_upload(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        file: UploadFile = File(...),
        preview: Optional[str] = Form(None),
    ):
        if rate_limit_service.is_rate_limited(request, "bank_pdf_upload", max_attempts=10, window_seconds=60):
            raise HTTPException(status_code=429, detail="Demasiadas solicitudes. Intenta en un momento.")
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")

        # Validaciones básicas
        filename = (file.filename or "").strip()
        name_l = filename.lower()
        if not name_l.endswith(".pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser .pdf")
        content_type = (file.content_type or "").lower().strip()
        if content_type and content_type not in ("application/pdf", "application/x-pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser un PDF válido (MIME application/pdf)")

        # Leer PDF en memoria
        sha = hashlib.sha256()
        size = 0
        chunks: list[bytes] = []
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BANK_PDF_SIZE:
                raise HTTPException(status_code=400, detail="El PDF excede el máximo de 15MB.")
            sha.update(chunk)
            chunks.append(chunk)
        if size <= 0:
            raise HTTPException(status_code=400, detail="El PDF está vacío.")
        pdf_bytes_head = chunks[0] if chunks else b""
        if not pdf_bytes_head.startswith(b"%PDF"):
            raise HTTPException(status_code=400, detail="El archivo no es un PDF válido.")

        # Vista previa Banorte: solo parsear y devolver HTML (no DB, no XLSX)
        if preview and str(preview).strip().lower() in ("1", "true", "yes"):
            import tempfile
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    for ch in chunks:
                        tmp.write(ch)
                    tmp_path = tmp.name
                result = parse_bank_pdf_to_movements_preview(tmp_path)
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_bank_pdf_to_excel.html",
                active_page="bank_pdf_to_excel",
                title="Convertir Edo. de Cuenta",
                extra={
                    "preview_movements": result.get("movements") or [],
                    "preview_summary": result.get("summary") or {},
                },
            )

        storage_root = get_storage_root(BASE_DIR)
        uploads_rel_dir = os.path.join("uploads", str(issuer_id), "bank")
        exports_rel_dir = os.path.join("exports", str(issuer_id), "bank_statements")

        pdf_sha256 = sha.hexdigest()
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        pdf_name = f"{stamp}_{pdf_sha256[:12]}.pdf"
        pdf_rel_path = os.path.join(uploads_rel_dir, pdf_name)
        pdf_abs_path = safe_join(storage_root, pdf_rel_path)
        ensure_parent_dir(pdf_abs_path)
        with open(pdf_abs_path, "wb") as f:
            for ch in chunks:
                f.write(ch)

        pdf_bytes = b"".join(chunks)
        # Get or create bank_statement (dedupe por mismo PDF). Validar RFC antes de guardar.
        conn = db()
        upload_metadata = {}
        result_preview = parse_bank_statement_preview(pdf_bytes, file_name=filename or "documento.pdf", file_index=0)
        if not result_preview.get("file_error"):
            upload_metadata = extract_statement_metadata(result_preview)
            detected_rfc = (upload_metadata.get("detected_holder_rfc") or "").strip().upper().replace(" ", "")
            expected_rfc = (issuer.get("rfc") or "").strip().upper().replace(" ", "")
            if expected_rfc and detected_rfc and expected_rfc != detected_rfc:
                raise HTTPException(
                    status_code=400,
                    detail="El RFC del estado de cuenta no coincide con el RFC de tu cuenta. No se puede procesar este PDF.",
                )
        try:
            _ensure_bank_statements_table(conn)
            row = conn.execute(
                "SELECT id FROM bank_statements WHERE issuer_id = ? AND source_pdf_sha256 = ? LIMIT 1",
                (issuer_id, pdf_sha256),
            ).fetchone()
            if row:
                statement_id = int(row["id"])
            else:
                conn.execute(
                    """
                    INSERT INTO bank_statements (issuer_id, source_pdf_path, source_pdf_sha256, created_at)
                    VALUES (?, ?, ?, datetime('now'))
                    """,
                    (issuer_id, pdf_rel_path, pdf_sha256),
                )
                statement_id = conn.execute("SELECT last_insert_rowid() AS rid").fetchone()["rid"]
                conn.commit()
            period_month = (upload_metadata.get("period_month") or "")[:7]
            if period_month and has_column(conn, "bank_statements", "period_month"):
                conn.execute(
                    "UPDATE bank_statements SET period_month = ?, bank_name = ?, account_last4 = ? WHERE id = ? AND issuer_id = ?",
                    (period_month, upload_metadata.get("bank_name"), upload_metadata.get("account_last4"), statement_id, issuer_id),
                )
                conn.commit()
        finally:
            conn.close()

        file_id = secrets.token_urlsafe(16)
        xlsx_name = f"{stamp}_{file_id[:10]}.xlsx"
        xlsx_rel_path = os.path.join(exports_rel_dir, xlsx_name)
        xlsx_abs_path = safe_join(storage_root, xlsx_rel_path)

        try:
            meta = convert_pdf_to_xlsx(
                pdf_abs_path,
                xlsx_abs_path,
                issuer_id=issuer_id,
                statement_id=statement_id,
            )
        except Exception as e:
            logger.exception("bank pdf-to-excel: error convirtiendo issuer=%s pdf=%s", issuer_id, pdf_rel_path)
            portal_error_type("parse_fail", log_context={"issuer_id": issuer_id, "pdf": pdf_rel_path})

        meta_for_storage = {k: v for k, v in (meta or {}).items() if k != "transactions"}
        meta_json_str = json.dumps(meta_for_storage, ensure_ascii=False)[:4000]

        conn = db()
        try:
            _ensure_bank_exports_table(conn)
            _ensure_bank_movements_table(conn)
            conn.execute(
                """
                INSERT INTO bank_pdf_exports (issuer_id, file_id, pdf_path, xlsx_path, meta_json)
                VALUES (?, ?, ?, ?, ?)
                """,
                (issuer_id, file_id, pdf_rel_path, xlsx_rel_path, meta_json_str),
            )
            default_period = (upload_metadata.get("period_month") or (meta or {}).get("period_start") or "")[:7]
            mov_has_period = has_column(conn, "bank_movements", "period_month")
            mov_has_hash = has_column(conn, "bank_movements", "movement_hash")
            user_accounts = bank_list_accounts_raw(int(issuer_id)) if int(issuer_id) > 0 else []
            statement_owner_name = (upload_metadata.get("detected_holder_name") or "").strip() or None
            statement_owner_rfc = (upload_metadata.get("detected_holder_rfc") or "").strip() or None
            # Post-process: apply category rules (own-account transfers, card payments)
            # then update movements already inserted by upsert_bank_movements()
            for t in (meta or {}).get("transactions") or []:
                fecha = (t.get("fecha") or "")[:32]
                descripcion = (t.get("descripcion") or "")[:2000]
                deposito = t.get("deposito")
                retiro = t.get("retiro")
                tipo = (t.get("tipo") or "").strip().upper()
                if tipo not in ("INGRESO", "GASTO"):
                    continue
                try:
                    desc_norm = (descripcion or "").strip().upper()
                    mov_hint = {
                        "raw_text_original": descripcion,
                        "raw_text_normalized": desc_norm,
                        "referencia": "",
                        "contraparte_nombre": (t.get("contraparte_hint") or "").strip(),
                        "rfc_detectado": (t.get("rfc_encontrado") or "").strip(),
                        "tipo_movimiento": tipo,
                        "monto_deposito": float(deposito or 0),
                        "monto_retiro": float(retiro or 0),
                        "categoria_sugerida": (t.get("categoria") or "").strip(),
                        "warnings": [],
                    }
                    issuer_rfc_auto = (issuer.get("rfc") or "").strip().upper()
                    detect_own_account_transfer(mov_hint, user_accounts, statement_owner_name, statement_owner_rfc, issuer_rfc=issuer_rfc_auto)
                    new_cat = None
                    if mov_hint.get("es_transferencia_propia_probable"):
                        new_cat = "CUENTA_PROPIA"
                    else:
                        metodo = (t.get("metodo_hint") or "").upper()
                        if (
                            ("PAGO CONCENTRACION" in desc_norm or "PAGO TARJETA" in desc_norm or "TARJETA DE CRED" in desc_norm)
                            or ("TARJETA" in metodo and "PAGO" in desc_norm)
                        ):
                            new_cat = "FINANCIERO_PAGO_TARJETA"
                    if new_cat and mov_has_hash:
                        from services.bank.bank_statement_parser import _movement_hash
                        h = _movement_hash(issuer_id, t)
                        conn.execute(
                            "UPDATE bank_movements SET categoria = ? WHERE issuer_id = ? AND movement_hash = ?",
                            (new_cat, issuer_id, h),
                        )
                except Exception:
                    pass
            conn.commit()
            # Retroactive: reclassify any movements matching issuer RFC
            _auto_rfc = (issuer.get("rfc") or "").strip().upper()
            if _auto_rfc:
                reclassify_own_transfers_by_rfc(conn, int(issuer_id), _auto_rfc)
        finally:
            conn.close()

        log_action(request, "bank_pdf_to_excel", issuer_id=issuer_id, entity_id=file_id[:32])
        # Campos resumidos para UI (además de meta completo)
        try:
            processed_count = int((meta or {}).get("processed_count") or 0)
        except Exception:
            processed_count = 0
        try:
            total_ingresos = float((meta or {}).get("total_ingresos") or 0)
        except Exception:
            total_ingresos = 0.0
        try:
            total_gastos = float((meta or {}).get("total_gastos") or 0)
        except Exception:
            total_gastos = 0.0
        try:
            sin_factura_count = int((meta or {}).get("sin_factura_count") or 0)
        except Exception:
            sin_factura_count = 0
        try:
            movements_count = int((meta or {}).get("movements_count") or 0)
        except Exception:
            movements_count = 0
        try:
            ingresos_total = float((meta or {}).get("ingresos_total") or 0)
        except Exception:
            ingresos_total = 0.0
        try:
            gastos_total = float((meta or {}).get("gastos_total") or 0)
        except Exception:
            gastos_total = 0.0
        try:
            sin_parse_count = int((meta or {}).get("sin_parse_count") or 0)
        except Exception:
            sin_parse_count = 0
        quality = (meta or {}).get("quality") if isinstance((meta or {}).get("quality"), dict) else None
        try:
            low_confidence_count = int((meta or {}).get("low_confidence_count") or 0)
        except Exception:
            low_confidence_count = 0
        return JSONResponse(
            {
                "ok": True,
                "file_id": file_id,
                "statement_id": statement_id,
                "meta": meta,
                "processed_count": processed_count,
                "total_ingresos": total_ingresos,
                "total_gastos": total_gastos,
                "sin_factura_count": sin_factura_count,
                "movements_count": movements_count,
                "ingresos_total": ingresos_total,
                "gastos_total": gastos_total,
                "sin_parse_count": sin_parse_count,
                "low_confidence_count": low_confidence_count,
                "quality": quality,
                "download_url": f"/portal/bank/pdf-to-excel/download/{file_id}",
            }
        )

    def _build_preview_export_xlsx(movements: list[dict[str, Any]]) -> bytes:
        """Genera XLSX en memoria desde lista de movimientos (preview editados). Sin DB."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        wb.remove(wb.active)
        headers = ["idx", "fecha", "concepto", "descripcion_raw", "deposito", "retiro", "saldo", "direction", "method", "category", "bucket", "deductible_hint", "needs_review", "confidence", "notes"]
        ws_mov = wb.create_sheet("Movimientos", 0)
        for col, h in enumerate(headers, 1):
            cell = ws_mov.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        for i, m in enumerate(movements, start=2):
            concept = (m.get("concept") or m.get("description_short") or "")[:500]
            raw = (m.get("description_raw") or "")[:2000]
            notes = (m.get("notes") or "")[:1000]
            ws_mov.append([
                m.get("idx"),
                m.get("date"),
                concept,
                raw,
                m.get("deposit"),
                m.get("withdraw"),
                m.get("balance"),
                m.get("direction"),
                m.get("method"),
                m.get("category"),
                m.get("bucket"),
                m.get("deductible_hint"),
                m.get("needs_review"),
                m.get("confidence"),
                notes,
            ])
        gastos = [m for m in movements if (m.get("withdraw") or 0) > 0]
        ingresos = [m for m in movements if (m.get("deposit") or 0) > 0]
        ws_g = wb.create_sheet("Gastos")
        ws_g.append(headers)
        for m in gastos:
            concept = (m.get("concept") or m.get("description_short") or "")[:500]
            ws_g.append([m.get("idx"), m.get("date"), concept, m.get("description_raw"), m.get("deposit"), m.get("withdraw"), m.get("balance"), m.get("direction"), m.get("method"), m.get("category"), m.get("bucket"), m.get("deductible_hint"), m.get("needs_review"), m.get("confidence"), m.get("notes")])
        ws_i = wb.create_sheet("Ingresos")
        ws_i.append(headers)
        for m in ingresos:
            concept = (m.get("concept") or m.get("description_short") or "")[:500]
            ws_i.append([m.get("idx"), m.get("date"), concept, m.get("description_raw"), m.get("deposit"), m.get("withdraw"), m.get("balance"), m.get("direction"), m.get("method"), m.get("category"), m.get("bucket"), m.get("deductible_hint"), m.get("needs_review"), m.get("confidence"), m.get("notes")])
        total_dep = sum(float(m.get("deposit") or 0) for m in movements)
        total_wd = sum(float(m.get("withdraw") or 0) for m in movements)
        ws_r = wb.create_sheet("Resumen")
        ws_r.append(["campo", "valor"])
        ws_r.append(["total_depositos", total_dep])
        ws_r.append(["total_retiros", total_wd])
        ws_r.append(["neto", total_dep - total_wd])
        ws_r.append(["count_movimientos", len(movements)])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @router.post("/bank/preview/export")
    def portal_bank_preview_export(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Recibe JSON con movimientos editados, devuelve XLSX. Sin DB."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        movements = body.get("movements") if isinstance(body.get("movements"), list) else []
        if not movements:
            raise HTTPException(status_code=400, detail="No hay movimientos para exportar.")
        try:
            xlsx_bytes = _build_preview_export_xlsx(movements)
        except Exception as e:
            logger.exception("bank preview export: %s", e)
            raise HTTPException(status_code=500, detail="Error al generar el Excel.")
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="movimientos_preview.xlsx"'},
        )

    @router.post("/bank/preview/reclassify")
    def portal_bank_preview_reclassify(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Re-clasifica movimientos con preset (conservative | aggressive). Sin DB."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        movements = body.get("movements") if isinstance(body.get("movements"), list) else []
        preset = (body.get("preset") or "conservative").strip().lower()
        if preset not in ("conservative", "aggressive"):
            preset = "conservative"
        if not movements:
            return JSONResponse({"movements": [], "preset": preset})
        try:
            out = reclassify_movements(movements, preset=preset)
        except Exception as e:
            logger.exception("bank preview reclassify: %s", e)
            raise HTTPException(status_code=500, detail="Error al re-clasificar.")
        return JSONResponse({"movements": out, "preset": preset})

    @router.post("/bank/preview/commit", response_class=JSONResponse)
    def portal_bank_preview_commit(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Persiste movimientos del preview (editados en frontend) sin re-subir PDF."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        bank_account_id = body.get("bank_account_id")
        if bank_account_id is None:
            raise HTTPException(status_code=400, detail="Falta bank_account_id")
        try:
            bank_account_id = int(bank_account_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="bank_account_id debe ser un número")
        files = body.get("files") if isinstance(body.get("files"), list) else []
        if not files:
            raise HTTPException(status_code=400, detail="Falta lista de archivos (files) con file_summary y movements.")
        expected_rfc = (issuer.get("rfc") or "").strip()
        results = []
        total_inserted = 0
        total_duplicate_movements = 0
        last_statement_id = None
        period_month = None
        for item in files:
            file_summary = item.get("file_summary") if isinstance(item.get("file_summary"), dict) else {}
            movements = item.get("movements") if isinstance(item.get("movements"), list) else []
            if not movements:
                continue
            out = commit_preview_to_db(
                issuer_id=issuer_id,
                bank_account_id=bank_account_id,
                file_summary=file_summary,
                movements=movements,
                expected_issuer_rfc=expected_rfc,
            )
            if not out.get("ok"):
                return JSONResponse(
                    {"ok": False, "rejection_reason": out.get("rejection_reason", "Error al guardar"), "status": out.get("status", "error")},
                    status_code=400,
                )
            results.append({
                "statement_id": out.get("statement_id"),
                "inserted_count": out.get("inserted_count", 0),
                "duplicate_statement": out.get("duplicate_statement", False),
                "duplicate_movements_count": out.get("duplicate_movements_count", 0),
            })
            total_inserted += out.get("inserted_count", 0)
            total_duplicate_movements += out.get("duplicate_movements_count", 0)
            if out.get("statement_id"):
                last_statement_id = out["statement_id"]
            if out.get("period_month"):
                period_month = out["period_month"]
        # Retroactive: reclassify movements matching issuer RFC
        _commit_rfc = (issuer.get("rfc") or "").strip().upper()
        if _commit_rfc and total_inserted > 0:
            try:
                _cconn = db()
                reclassify_own_transfers_by_rfc(_cconn, issuer_id, _commit_rfc)
                _cconn.close()
            except Exception:
                pass
        log_action(request, "bank_preview_commit", issuer_id=issuer_id, entity_id=last_statement_id)
        return JSONResponse({
            "ok": True,
            "statement_id": last_statement_id,
            "inserted_count": total_inserted,
            "duplicate_movements_count": total_duplicate_movements,
            "results": results,
            "period_month": period_month,
        })

    @router.get("/bank/pdf-to-excel/download/{file_id}")
    def portal_bank_pdf_to_excel_download(
        request: Request,
        file_id: str,
        issuer: dict = Depends(get_portal_issuer),
    ):
        issuer_id = int(issuer.get("id") or 0)
        fid = (file_id or "").strip()
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        if not fid:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")

        conn = db()
        try:
            _ensure_bank_exports_table(conn)
            row = conn.execute(
                "SELECT xlsx_path FROM bank_pdf_exports WHERE issuer_id = ? AND file_id = ? LIMIT 1",
                (issuer_id, fid),
            ).fetchone()
        finally:
            conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")

        storage_root = get_storage_root(BASE_DIR)
        xlsx_rel_path = (row["xlsx_path"] or "").strip()
        if not xlsx_rel_path:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        try:
            xlsx_abs_path = safe_join(storage_root, xlsx_rel_path)
        except ValueError:
            raise HTTPException(status_code=404, detail="Ruta inválida")
        if not os.path.exists(xlsx_abs_path):
            raise HTTPException(status_code=404, detail="El archivo ya no existe en disco")

        filename = f"estado_cuenta_{fid[:8]}.xlsx"
        file_access_log.log_file_access(
            request=request,
            action="download_bank_xlsx",
            issuer_id=issuer_id,
            user_id=getattr(request.state, "user_id", None),
            file_path=xlsx_rel_path,
            entity="bank_pdf_exports",
            entity_id=fid[:64],
        )
        return FileResponse(
            path=xlsx_abs_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    @router.get("/bank/statements", response_class=HTMLResponse)
    def portal_bank_statements(request: Request, issuer: dict = Depends(get_portal_issuer)):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        statements: list = []
        conn = None
        try:
            conn = db()
            conn.row_factory = lambda cursor, row: dict(zip([c[0] for c in cursor.description], row))
            _ensure_bank_exports_table(conn)
            _ensure_bank_movements_table(conn)
            rows = conn.execute(
                "SELECT file_id, pdf_path, xlsx_path, meta_json, created_at FROM bank_pdf_exports WHERE issuer_id = ? ORDER BY created_at DESC",
                (issuer_id,),
            ).fetchall()
            statements = []
            for r in rows:
                r = _db_row_to_dict(r)
                meta = {}
                if r.get("meta_json"):
                    try:
                        meta = json.loads(r["meta_json"] or "{}")
                    except Exception:
                        pass
                period_start = meta.get("period_start") or ""
                period_end = meta.get("period_end") or ""
                bank_name = meta.get("bank_name") or "—"
                account_last4 = meta.get("account_last4") or "—"
                movements_count = int(meta.get("movements_count") or 0)
                total_gastos = float(meta.get("gastos_total") or meta.get("total_gastos") or 0)
                total_ingresos = float(meta.get("ingresos_total") or meta.get("total_ingresos") or 0)
                period_label = f"{period_start} – {period_end}" if (period_start or period_end) else "—"
                statements.append({
                    "file_id": r["file_id"],
                    "statement_key": r["file_id"],
                    "created_at": r["created_at"] or "",
                    "period_label": period_label,
                    "bank_name": bank_name,
                    "account_last4": account_last4,
                    "movements_count": movements_count,
                    "total_gastos": total_gastos,
                    "total_ingresos": total_ingresos,
                    "source": "export",
                })
            if table_exists(conn, "bank_statements"):
                has_pm = has_column(conn, "bank_statements", "period_month")
                has_tm = has_column(conn, "bank_statements", "total_movements")
                if has_pm and has_tm:
                    st_rows = conn.execute(
                        "SELECT id, period_month, bank_name, account_last4, total_movements, status, created_at FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                else:
                    st_rows = conn.execute(
                        "SELECT id, bank_name, account_last4, period_start, period_end, created_at FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                for r in st_rows:
                    r = _db_row_to_dict(r)
                    if has_pm:
                        pm = r.get("period_month") or ""
                    else:
                        pm = (r.get("period_start") or "")[:7]
                    period_label = pm if pm else ((r.get("created_at") or "")[:7] or "—")
                    statements.append({
                        "file_id": None,
                        "statement_key": f"stmt_{r['id']}",
                        "statement_id": r["id"],
                        "created_at": r.get("created_at") or "",
                        "period_label": period_label,
                        "bank_name": r.get("bank_name") or "—",
                        "account_last4": r.get("account_last4") or "—",
                        "movements_count": int(r.get("total_movements") or 0) if has_tm else 0,
                        "total_gastos": 0,
                        "total_ingresos": 0,
                        "source": "ingest",
                    })
            statements.sort(key=lambda x: (x.get("created_at") or ""), reverse=True)
        except Exception as e:
            logger.warning("portal bank/statements: error cargando lista (%s), mostrando vacío", e)
            statements = []
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_bank_statements.html",
            active_page="bank_statements",
            title="Estados de cuenta",
            statements=statements,
        )

    @router.get("/movimientos", response_class=HTMLResponse)
    @router.get("/bank/movements", response_class=HTMLResponse)
    def portal_bank_movements(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: Optional[str] = Query(None, description="Mes YYYY-MM (selector como emitidas/recibidas)"),
        statement_id: Optional[str] = Query(None, description="Filtrar por estado de cuenta (file_id o stmt_N)"),
        period_month: Optional[str] = Query(None, description="YYYY-MM (legacy, usa ym si no viene)"),
        tipo: Optional[str] = Query(None, description="INGRESO, GASTO, INFO"),
        categoria: Optional[str] = Query(None),
        cfdi_match_status: Optional[str] = Query(None, description="pending, suggested, confirmed, rejected"),
        match_filter: Optional[str] = Query(None, description="none|probable (conciliación)"),
        min_confidence: Optional[int] = Query(None, ge=0, le=100),
        search: Optional[str] = Query(None),
        hide_own_transfers: Optional[int] = Query(None, description="1 para ocultar traspasos propios"),
        hide_financial: Optional[int] = Query(None, description="1 para ocultar pagos/cargos financieros"),
        only_real_expenses: Optional[int] = Query(None, description="1 para solo gastos reales"),
        limit: int = Query(200, ge=1, le=500),
        offset: int = Query(0, ge=0, le=MAX_LIST_OFFSET),
    ):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        # Mes: prioridad ym (selector) > period_month (legacy) > mes actual
        explicit_ym = sanitize_ym(ym or period_month or "", "")
        period_month = explicit_ym or ym_now()
        movements: list = []
        total_count = 0
        sum_ingresos = 0.0
        sum_gastos = 0.0
        cuenta_propia_entradas = 0.0
        cuenta_propia_salidas = 0.0
        statements_opt: list = []
        months_with_movements: list[dict] = []
        conn = None
        try:
            conn = db()
            conn.row_factory = lambda cursor, row: dict(zip([c[0] for c in cursor.description], row))
            _ensure_bank_movements_table(conn)
            _ensure_bank_exports_table(conn)
            # Heal pass: reclassify old movements matching issuer RFC
            _heal_rfc = (issuer.get("rfc") or "").strip().upper()
            if _heal_rfc:
                reclassify_own_transfers_by_rfc(conn, issuer_id, _heal_rfc)
            has_matches = table_exists(conn, "bank_invoice_matches") and table_exists(conn, "sat_cfdi")

            # Default to current month so users always see today's context.
            # (Previously auto-selected most recent month with data, which was confusing.)

            params: list = [issuer_id]
            where_clauses = ["issuer_id = ?"]

            if statement_id:
                sid = statement_id.strip()
                if sid.startswith("stmt_"):
                    try:
                        bid = int(sid.replace("stmt_", ""))
                        if has_column(conn, "bank_movements", "bank_statement_id"):
                            where_clauses.append("bank_statement_id = ?")
                            params.append(bid)
                        else:
                            where_clauses.append("statement_file_id = ?")
                            params.append(sid)
                    except ValueError:
                        where_clauses.append("statement_file_id = ?")
                        params.append(sid)
                else:
                    if has_column(conn, "bank_movements", "statement_file_id"):
                        where_clauses.append("statement_file_id = ?")
                        params.append(sid)
            if has_column(conn, "bank_movements", "period_month"):
                where_clauses.append("period_month = ?")
                params.append(period_month)
            if tipo:
                where_clauses.append("tipo = ?")
                params.append(tipo.strip().upper())
            if categoria:
                where_clauses.append("categoria = ?")
                params.append(categoria.strip())
            if hide_own_transfers:
                where_clauses.append("COALESCE(categoria,'') != 'CUENTA_PROPIA'")
            if hide_financial:
                where_clauses.append("COALESCE(categoria,'') NOT IN ('FINANCIERO_PAGO_TARJETA','MOVIMIENTO_FINANCIERO','COMISIONES BANCARIAS','COMISIONES_BANCARIAS','COMISION_BANCARIA')")
            if only_real_expenses:
                where_clauses.append(
                    "COALESCE(categoria,'') NOT IN ('CUENTA_PROPIA','FINANCIERO_PAGO_TARJETA','MOVIMIENTO_FINANCIERO','COMISIONES BANCARIAS','COMISIONES_BANCARIAS','COMISION_BANCARIA','TRASPASO_PROPIO')"
                )
            if cfdi_match_status and has_column(conn, "bank_movements", "cfdi_match_status"):
                where_clauses.append("cfdi_match_status = ?")
                params.append(cfdi_match_status.strip().lower())
            if match_filter and has_matches:
                mf = (match_filter or "").strip().lower()
                if mf == "probable":
                    where_clauses.append(
                        """EXISTS (
                             SELECT 1 FROM bank_invoice_matches bim
                             WHERE bim.issuer_id = bank_movements.issuer_id
                               AND bim.bank_movement_id = bank_movements.id
                               AND bim.status IN ('suggested','confirmed')
                               AND COALESCE(bim.score,0) >= 80
                           )"""
                    )
                elif mf == "revisar":
                    where_clauses.append(
                        """EXISTS (
                             SELECT 1 FROM bank_invoice_matches bim
                             WHERE bim.issuer_id = bank_movements.issuer_id
                               AND bim.bank_movement_id = bank_movements.id
                               AND bim.status IN ('suggested','confirmed')
                               AND COALESCE(bim.score,0) BETWEEN 50 AND 79
                           )"""
                    )
                elif mf == "none":
                    where_clauses.append(
                        """NOT EXISTS (
                             SELECT 1 FROM bank_invoice_matches bim
                             WHERE bim.issuer_id = bank_movements.issuer_id
                               AND bim.bank_movement_id = bank_movements.id
                               AND bim.status IN ('suggested','confirmed')
                               AND COALESCE(bim.score,0) >= 50
                           )"""
                    )
            if min_confidence is not None:
                where_clauses.append("confidence_score >= ?")
                params.append(min_confidence)
            if search and search.strip():
                q = f"%{search.strip()}%"
                if has_column(conn, "bank_movements", "raw_description"):
                    where_clauses.append("(descripcion LIKE ? OR contraparte_hint LIKE ? OR raw_description LIKE ?)")
                    params.extend([q, q, q])
                else:
                    where_clauses.append("(descripcion LIKE ? OR contraparte_hint LIKE ?)")
                    params.extend([q, q])

            where_sql = " AND ".join(where_clauses)

            total_count_row = conn.execute(
                f"SELECT COUNT(*) AS c FROM bank_movements WHERE {where_sql}",
                params,
            ).fetchone()
            total_count = int(_db_row_to_dict(total_count_row).get("c", 0) or 0)

            _impacta_filter = " AND COALESCE(impacta_contabilidad, 1) = 1" if has_column(conn, "bank_movements", "impacta_contabilidad") else " AND COALESCE(categoria,'') != 'CUENTA_PROPIA'"
            sum_row = conn.execute(
                f"SELECT COALESCE(SUM(deposito), 0) AS ing, COALESCE(SUM(retiro), 0) AS gas FROM bank_movements WHERE {where_sql}{_impacta_filter}",
                params,
            ).fetchone()
            sum_row_d = _db_row_to_dict(sum_row)
            sum_ingresos = float(sum_row_d.get("ing", 0) or 0)
            sum_gastos = float(sum_row_d.get("gas", 0) or 0)

            # Conciliation stats (always for full period, ignoring user filters)
            concil_stats = {"matched": 0, "unmatched": 0, "total_real": 0}
            try:
                _concil_base = "issuer_id = ? AND COALESCE(impacta_contabilidad, 1) = 1" if has_column(conn, "bank_movements", "impacta_contabilidad") else "issuer_id = ? AND COALESCE(categoria,'') != 'CUENTA_PROPIA'"
                _concil_params: list = [issuer_id]
                if has_column(conn, "bank_movements", "period_month"):
                    _concil_base += " AND period_month = ?"
                    _concil_params.append(period_month)
                if table_exists(conn, "bank_invoice_matches"):
                    matched_row = conn.execute(
                        f"""SELECT COUNT(*) AS n FROM bank_movements
                            WHERE {_concil_base}
                            AND EXISTS (
                              SELECT 1 FROM bank_invoice_matches bim
                              WHERE bim.issuer_id = bank_movements.issuer_id
                                AND bim.bank_movement_id = bank_movements.id
                                AND bim.status IN ('suggested','confirmed')
                                AND COALESCE(bim.score,0) >= 80
                            )""",
                        _concil_params,
                    ).fetchone()
                    concil_stats["matched"] = int(_db_row_to_dict(matched_row).get("n", 0) or 0)
                total_real_row = conn.execute(
                    f"SELECT COUNT(*) AS n FROM bank_movements WHERE {_concil_base}",
                    _concil_params,
                ).fetchone()
                concil_stats["total_real"] = int(_db_row_to_dict(total_real_row).get("n", 0) or 0)
                concil_stats["unmatched"] = concil_stats["total_real"] - concil_stats["matched"]
            except Exception:
                pass

            # Cuenta propia sums (always for issuer + period, ignoring other filters)
            _cp_where = "issuer_id = ? AND COALESCE(categoria,'') = 'CUENTA_PROPIA'"
            _cp_params: list = [issuer_id]
            if has_column(conn, "bank_movements", "period_month"):
                _cp_where += " AND period_month = ?"
                _cp_params.append(period_month)
            cp_row = conn.execute(
                f"SELECT COALESCE(SUM(deposito), 0) AS cp_ing, COALESCE(SUM(retiro), 0) AS cp_gas FROM bank_movements WHERE {_cp_where}",
                _cp_params,
            ).fetchone()
            cp_row_d = _db_row_to_dict(cp_row)
            cuenta_propia_entradas = float(cp_row_d.get("cp_ing", 0) or 0)
            cuenta_propia_salidas = float(cp_row_d.get("cp_gas", 0) or 0)

            # Construir SELECT solo con columnas que existan (compatibilidad con distintos esquemas)
            sel = ["id"]
            if has_column(conn, "bank_movements", "statement_file_id"):
                sel.append("statement_file_id")
            elif has_column(conn, "bank_movements", "statement_id"):
                sel.append("statement_id AS statement_file_id")
            sel.append("fecha")
            if has_column(conn, "bank_movements", "descripcion"):
                sel.append("descripcion")
            elif has_column(conn, "bank_movements", "descripcion_norm"):
                sel.append("descripcion_norm AS descripcion")
            else:
                sel.append("descripcion_raw AS descripcion")
            sel.extend(["deposito", "retiro", "saldo", "tipo", "categoria", "metodo_hint", "contraparte_hint"])
            if has_column(conn, "bank_movements", "rfc_encontrado"):
                sel.append("rfc_encontrado")
            elif has_column(conn, "bank_movements", "rfc_detectado"):
                sel.append("rfc_detectado AS rfc_encontrado")
            sel.append("confidence_score")
            if has_column(conn, "bank_movements", "bank_statement_id"):
                sel.append("bank_statement_id")
                sel.append("(SELECT bs.bank_name FROM bank_statements bs WHERE bs.id = bank_movements.bank_statement_id LIMIT 1) AS statement_bank_name")
            if has_column(conn, "bank_movements", "cfdi_match_status"):
                sel.append("cfdi_match_status")
            if has_matches:
                sel.append(
                    """(
                        SELECT sc.uuid
                        FROM bank_invoice_matches bim
                        JOIN sat_cfdi sc ON sc.id = bim.cfdi_id
                        WHERE bim.issuer_id = bank_movements.issuer_id
                          AND bim.bank_movement_id = bank_movements.id
                          AND bim.status IN ('suggested','confirmed')
                        ORDER BY bim.score DESC, bim.id DESC
                        LIMIT 1
                    ) AS probable_cfdi_uuid"""
                )
                sel.append(
                    """(
                        SELECT bim.score
                        FROM bank_invoice_matches bim
                        WHERE bim.issuer_id = bank_movements.issuer_id
                          AND bim.bank_movement_id = bank_movements.id
                          AND bim.status IN ('suggested','confirmed')
                        ORDER BY bim.score DESC, bim.id DESC
                        LIMIT 1
                    ) AS probable_cfdi_score"""
                )
                sel.append(
                    """(
                        SELECT bim.status
                        FROM bank_invoice_matches bim
                        WHERE bim.issuer_id = bank_movements.issuer_id
                          AND bim.bank_movement_id = bank_movements.id
                          AND bim.status IN ('suggested','confirmed')
                        ORDER BY bim.score DESC, bim.id DESC
                        LIMIT 1
                    ) AS probable_cfdi_status"""
                )
            select_cols = ", ".join(sel)
            params_ext = params + [limit, offset]
            movements = conn.execute(
                f"SELECT {select_cols} FROM bank_movements WHERE {where_sql} ORDER BY fecha DESC, id DESC LIMIT ? OFFSET ?",
                params_ext,
            ).fetchall()
            movements = [_db_row_to_dict(r) for r in movements]
            for row in movements:
                row.setdefault("fecha", None)
                row.setdefault("descripcion", None)
                row.setdefault("deposito", None)
                row.setdefault("retiro", None)
                row.setdefault("saldo", None)
                row.setdefault("tipo", None)
                row.setdefault("categoria", None)
                row.setdefault("metodo_hint", None)
                row.setdefault("contraparte_hint", None)
                row.setdefault("rfc_encontrado", None)
                row.setdefault("confidence_score", None)
                row.setdefault("cfdi_match_status", None)
                row.setdefault("bank_statement_id", None)
                row.setdefault("statement_bank_name", None)
                row.setdefault("probable_cfdi_uuid", None)
                row.setdefault("probable_cfdi_score", None)
                row.setdefault("probable_cfdi_status", None)
                # Asegurar que montos sean numéricos para el formato en plantilla
                for key in ("deposito", "retiro", "saldo", "confidence_score", "probable_cfdi_score"):
                    if row.get(key) is not None and row[key] != "":
                        try:
                            if key == "confidence_score":
                                row[key] = int(float(row[key]))
                            elif key == "probable_cfdi_score":
                                row[key] = int(float(row[key]))
                            else:
                                row[key] = float(row[key])
                        except (TypeError, ValueError):
                            row[key] = None if key != "confidence_score" else 0
                # Concepto = descripción sin prefijo de fecha (igual que en convertir edo. de cuenta)
                row["concepto"] = _strip_date_from_description(row.get("descripcion")) or (row.get("descripcion") or "").strip()

            statements_opt = []
            for r in conn.execute(
                "SELECT file_id, meta_json, created_at FROM bank_pdf_exports WHERE issuer_id = ? ORDER BY created_at DESC",
                (issuer_id,),
            ).fetchall():
                r = _db_row_to_dict(r)
                meta = {}
                if r.get("meta_json"):
                    try:
                        meta = json.loads(r["meta_json"] or "{}")
                    except Exception:
                        pass
                p_start = meta.get("period_start") or ""
                p_end = meta.get("period_end") or ""
                if p_start or p_end:
                    label = f"{p_start} – {p_end}"
                else:
                    label = (r.get("created_at") or "")[:16] or (r["file_id"][:12] + "…")
                statements_opt.append({"statement_id": r["file_id"], "label": label})
            if table_exists(conn, "bank_statements"):
                has_pm = has_column(conn, "bank_statements", "period_month")
                if has_pm:
                    st_opt_rows = conn.execute(
                        "SELECT id, period_month, total_movements FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                else:
                    st_opt_rows = conn.execute(
                        "SELECT id, period_start FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                for r in st_opt_rows:
                    r = _db_row_to_dict(r)
                    if has_pm:
                        pm = r.get("period_month") or ""
                    else:
                        pm = (r.get("period_start") or "")[:7]
                    label = pm if pm else f"Estado #{r['id']}"
                    statements_opt.append({"statement_id": f"stmt_{r['id']}", "label": label})
            # Meses con movimientos (para el selector como emitidas/recibidas)
            if has_column(conn, "bank_movements", "period_month"):
                months_rows = conn.execute(
                    """SELECT period_month AS ym, COUNT(*) AS n FROM bank_movements
                       WHERE issuer_id = ? AND period_month IS NOT NULL AND TRIM(period_month) != ''
                       GROUP BY period_month ORDER BY period_month DESC""",
                    (issuer_id,),
                ).fetchall()
                for r in months_rows:
                    r = _db_row_to_dict(r)
                    ym_val = r.get("ym") or ""
                    if ym_val:
                        months_with_movements.append({"ym": ym_val, "n": int(r.get("n") or 0), "label": ym_to_label(ym_val)})
            else:
                # Sin columna period_month: usar meses de bank_statements si existen
                if table_exists(conn, "bank_statements") and has_column(conn, "bank_statements", "period_month"):
                    months_rows = conn.execute(
                        """SELECT period_month AS ym, COALESCE(SUM(total_movements), 0) AS n FROM bank_statements
                           WHERE issuer_id = ? AND period_month IS NOT NULL AND TRIM(period_month) != ''
                           GROUP BY period_month ORDER BY period_month DESC""",
                        (issuer_id,),
                    ).fetchall()
                    for r in months_rows:
                        r = _db_row_to_dict(r)
                        ym_val = r.get("ym") or ""
                        if ym_val:
                            months_with_movements.append({"ym": ym_val, "n": int(r.get("n") or 0), "label": ym_to_label(ym_val)})
        except Exception as e:
            logger.warning("portal movimientos: error cargando datos (%s), mostrando lista vacía", e)
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

        ym_safe = sanitize_ym(period_month or "", ym_now())

        # Auto-run matching on page load (lightweight — skips if already done)
        if total_count > 0:
            try:
                from services.bank import bank_cfdi_matching as _bcm
                _bcm.refresh_suggestions_for_month(issuer_id, ym_safe)
            except Exception as _me:
                logger.debug("auto-matching on movimientos page: %s", _me)

        try:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_bank_movements.html",
                active_page="bank_movements",
                title="Movimientos",
                movements=movements,
                total_count=total_count,
                sum_ingresos=sum_ingresos,
                sum_gastos=sum_gastos,
                cuenta_propia_entradas=cuenta_propia_entradas,
                cuenta_propia_salidas=cuenta_propia_salidas,
                limit=limit,
                offset=offset,
                statement_id=statement_id or "",
                period_month=ym_safe,
                ym=ym_safe,
                ym_label=ym_to_label(ym_safe),
                prev_ym=shift_ym(ym_safe, -1),
                next_ym=shift_ym(ym_safe, +1),
                months=months_with_movements,
                tipo=tipo or "",
                categoria=categoria or "",
                cfdi_match_status=cfdi_match_status or "",
                match_filter=(match_filter or ""),
                min_confidence=min_confidence,
                search=search or "",
                hide_own_transfers=1 if hide_own_transfers else 0,
                hide_financial=1 if hide_financial else 0,
                only_real_expenses=1 if only_real_expenses else 0,
                statements_opt=statements_opt,
                concil_stats=concil_stats,
                csrf_token=csrf_service.generate_csrf_token(),
            )
        except Exception as e:
            logger.exception("portal movimientos (render): %s", e)
            raise HTTPException(status_code=500, detail=f"Error al mostrar la página: {e!s}")

    @router.get("/bank/movements/export")
    def portal_bank_movements_export(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: Optional[str] = Query(None),
        statement_id: Optional[str] = Query(None),
        tipo: Optional[str] = Query(None),
        categoria: Optional[str] = Query(None),
        cfdi_match_status: Optional[str] = Query(None),
        match_filter: Optional[str] = Query(None),
        search: Optional[str] = Query(None),
        hide_own_transfers: Optional[int] = Query(None),
        hide_financial: Optional[int] = Query(None),
        only_real_expenses: Optional[int] = Query(None),
    ):
        """Export filtered movements to XLSX."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        period_month = sanitize_ym(ym or "", ym_now())
        conn = db()
        conn.row_factory = lambda cursor, row: dict(zip([c[0] for c in cursor.description], row))
        try:
            _ensure_bank_movements_table(conn)
            params: list = [issuer_id]
            where_clauses = ["issuer_id = ?"]
            if has_column(conn, "bank_movements", "period_month"):
                where_clauses.append("period_month = ?")
                params.append(period_month)
            if statement_id:
                sid = statement_id.strip()
                if sid.startswith("stmt_"):
                    try:
                        bid = int(sid.replace("stmt_", ""))
                        if has_column(conn, "bank_movements", "bank_statement_id"):
                            where_clauses.append("bank_statement_id = ?")
                            params.append(bid)
                        else:
                            where_clauses.append("statement_file_id = ?")
                            params.append(sid)
                    except ValueError:
                        where_clauses.append("statement_file_id = ?")
                        params.append(sid)
                else:
                    if has_column(conn, "bank_movements", "statement_file_id"):
                        where_clauses.append("statement_file_id = ?")
                        params.append(sid)
            if tipo:
                where_clauses.append("tipo = ?")
                params.append(tipo.strip().upper())
            if hide_own_transfers:
                where_clauses.append("COALESCE(categoria,'') != 'CUENTA_PROPIA'")
            if hide_financial:
                where_clauses.append("COALESCE(categoria,'') NOT IN ('FINANCIERO_PAGO_TARJETA','MOVIMIENTO_FINANCIERO','COMISIONES BANCARIAS','COMISIONES_BANCARIAS','COMISION_BANCARIA')")
            if only_real_expenses:
                where_clauses.append(
                    "COALESCE(categoria,'') NOT IN ('CUENTA_PROPIA','FINANCIERO_PAGO_TARJETA','MOVIMIENTO_FINANCIERO','COMISIONES BANCARIAS','COMISIONES_BANCARIAS','COMISION_BANCARIA','TRASPASO_PROPIO')"
                )
            if cfdi_match_status and has_column(conn, "bank_movements", "cfdi_match_status"):
                where_clauses.append("cfdi_match_status = ?")
                params.append(cfdi_match_status.strip().lower())
            if search and search.strip():
                q = f"%{search.strip()}%"
                if has_column(conn, "bank_movements", "raw_description"):
                    where_clauses.append("(descripcion LIKE ? OR contraparte_hint LIKE ? OR raw_description LIKE ?)")
                    params.extend([q, q, q])
                else:
                    where_clauses.append("(descripcion LIKE ? OR contraparte_hint LIKE ?)")
                    params.extend([q, q])
            where_sql = " AND ".join(where_clauses)
            rows = conn.execute(
                f"SELECT fecha, descripcion, tipo, deposito, retiro, saldo, categoria, contraparte_hint, rfc_encontrado FROM bank_movements WHERE {where_sql} ORDER BY fecha DESC, id DESC",
                params,
            ).fetchall()
        finally:
            conn.close()

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        headers = ["Fecha", "Descripción", "Tipo", "Depósito", "Retiro", "Saldo", "Categoría", "Contraparte", "RFC"]
        ws.append(headers)
        for r in rows:
            ws.append([
                r.get("fecha"),
                _strip_date_from_description(r.get("descripcion")) or r.get("descripcion", ""),
                r.get("tipo"),
                r.get("deposito"),
                r.get("retiro"),
                r.get("saldo"),
                r.get("categoria"),
                r.get("contraparte_hint"),
                r.get("rfc_encontrado"),
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"movimientos_{period_month}.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @router.post("/bank/movements/reconcile", response_class=RedirectResponse)
    def portal_bank_movements_reconcile(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: str = Form(...),
        csrf_token: str | None = Form(None),
    ):
        token_val = (csrf_token or request.headers.get("X-CSRF-Token") or "").strip()
        if not csrf_service.verify_csrf_token(token_val):
            raise HTTPException(status_code=403, detail="Token CSRF inválido o expirado")
        if rate_limit_service.is_rate_limited(request, "bank_reconcile"):
            raise HTTPException(status_code=429, detail="Demasiados intentos. Espera un minuto.")
        issuer_id = int(issuer.get("id") or 0)
        from services.bank import bank_cfdi_matching as bank_cfdi_matching_service

        bank_cfdi_matching_service.refresh_suggestions_for_month(issuer_id, ym)
        audit.log(
            action="bank_reconcile_run",
            user_id=getattr(request.state, "user_id", 0) or 0,
            issuer_id=issuer_id,
            request=request,
            entity="bank_movements",
            entity_id=ym,
        )
        log_action(request, "bank_reconcile_run", issuer_id=issuer_id, ym=ym)
        return RedirectResponse(url=f"/portal/bank/movements?ym={ym}", status_code=302)

