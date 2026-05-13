"""Bank preview operations — export, reclassify, commit, and PDF download."""
import io
import logging
import os
from typing import Any

from fastapi import Body, Depends, Request
from fastapi.responses import FileResponse, JSONResponse, Response
from starlette.exceptions import HTTPException

from config import BASE_DIR
from database import db
from routers.deps import get_portal_issuer
from routers.portal.bank._bank_helpers import ensure_bank_exports_table
from services import file_access_log
from services.action_log import log_action
from services.bank.bank_own_accounts import reclassify_own_transfers_by_rfc
from services.bank.bank_parse_preview import reclassify_movements
from services.bank.bank_statement_ingest import commit_preview_to_db
from services.pdf_to_excel import get_storage_root, safe_join

logger = logging.getLogger(__name__)


def _build_preview_export_xlsx(movements: list[dict[str, Any]]) -> bytes:
    """Genera XLSX en memoria desde lista de movimientos (preview editados). Sin DB."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    headers = ["idx", "fecha", "concepto", "descripcion_raw", "deposito", "retiro", "saldo", "direction", "method", "category", "bucket", "deductible_hint", "needs_review", "confidence", "notes"]
    ws_mov = wb.create_sheet("Movimientos", 0)
    for col, h in enumerate(headers, 1):
        cell = ws_mov.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for i, m in enumerate(movements, start=2):
        concept = (m.get("concept") or m.get("description_short") or "")[:500]
        raw = (m.get("description_raw") or "")[:2000]
        notes = (m.get("notes") or "")[:1000]
        ws_mov.append([
            m.get("idx"),
            m.get("date"),
            concept,
            raw,
            m.get("deposit"),
            m.get("withdraw"),
            m.get("balance"),
            m.get("direction"),
            m.get("method"),
            m.get("category"),
            m.get("bucket"),
            m.get("deductible_hint"),
            m.get("needs_review"),
            m.get("confidence"),
            notes,
        ])
    gastos = [m for m in movements if (m.get("withdraw") or 0) > 0]
    ingresos = [m for m in movements if (m.get("deposit") or 0) > 0]
    ws_g = wb.create_sheet("Gastos")
    ws_g.append(headers)
    for m in gastos:
        concept = (m.get("concept") or m.get("description_short") or "")[:500]
        ws_g.append([m.get("idx"), m.get("date"), concept, m.get("description_raw"), m.get("deposit"), m.get("withdraw"), m.get("balance"), m.get("direction"), m.get("method"), m.get("category"), m.get("bucket"), m.get("deductible_hint"), m.get("needs_review"), m.get("confidence"), m.get("notes")])
    ws_i = wb.create_sheet("Ingresos")
    ws_i.append(headers)
    for m in ingresos:
        concept = (m.get("concept") or m.get("description_short") or "")[:500]
        ws_i.append([m.get("idx"), m.get("date"), concept, m.get("description_raw"), m.get("deposit"), m.get("withdraw"), m.get("balance"), m.get("direction"), m.get("method"), m.get("category"), m.get("bucket"), m.get("deductible_hint"), m.get("needs_review"), m.get("confidence"), m.get("notes")])
    total_dep = sum(float(m.get("deposit") or 0) for m in movements)
    total_wd = sum(float(m.get("withdraw") or 0) for m in movements)
    ws_r = wb.create_sheet("Resumen")
    ws_r.append(["campo", "valor"])
    ws_r.append(["total_depositos", total_dep])
    ws_r.append(["total_retiros", total_wd])
    ws_r.append(["neto", total_dep - total_wd])
    ws_r.append(["count_movimientos", len(movements)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def register_bank_preview_ops_routes(router, templates):
    """Register preview export/reclassify/commit and download routes."""

    @router.post("/bank/preview/export")
    def portal_bank_preview_export(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Recibe JSON con movimientos editados, devuelve XLSX. Sin DB."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesion invalida")
        movements = body.get("movements") if isinstance(body.get("movements"), list) else []
        if not movements:
            raise HTTPException(status_code=400, detail="No hay movimientos para exportar.")
        try:
            xlsx_bytes = _build_preview_export_xlsx(movements)
        except Exception as e:
            logger.exception("bank preview export: %s", e)
            raise HTTPException(status_code=500, detail="Error al generar el Excel.")
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="movimientos_preview.xlsx"'},
        )

    @router.post("/bank/preview/reclassify")
    def portal_bank_preview_reclassify(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Re-clasifica movimientos con preset (conservative | aggressive). Sin DB."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesion invalida")
        movements = body.get("movements") if isinstance(body.get("movements"), list) else []
        preset = (body.get("preset") or "conservative").strip().lower()
        if preset not in ("conservative", "aggressive"):
            preset = "conservative"
        if not movements:
            return JSONResponse({"movements": [], "preset": preset})
        try:
            out = reclassify_movements(movements, preset=preset)
        except Exception as e:
            logger.exception("bank preview reclassify: %s", e)
            raise HTTPException(status_code=500, detail="Error al re-clasificar.")
        return JSONResponse({"movements": out, "preset": preset})

    @router.post("/bank/preview/commit", response_class=JSONResponse)
    def portal_bank_preview_commit(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Persiste movimientos del preview (editados en frontend) sin re-subir PDF."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesion invalida")
        bank_account_id = body.get("bank_account_id")
        if bank_account_id is None:
            raise HTTPException(status_code=400, detail="Falta bank_account_id")
        try:
            bank_account_id = int(bank_account_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="bank_account_id debe ser un numero")
        files = body.get("files") if isinstance(body.get("files"), list) else []
        if not files:
            raise HTTPException(status_code=400, detail="Falta lista de archivos (files) con file_summary y movements.")
        expected_rfc = (issuer.get("rfc") or "").strip()
        results = []
        total_inserted = 0
        total_duplicate_movements = 0
        last_statement_id = None
        period_month = None
        for item in files:
            file_summary = item.get("file_summary") if isinstance(item.get("file_summary"), dict) else {}
            movements = item.get("movements") if isinstance(item.get("movements"), list) else []
            if not movements:
                continue
            out = commit_preview_to_db(
                issuer_id=issuer_id,
                bank_account_id=bank_account_id,
                file_summary=file_summary,
                movements=movements,
                expected_issuer_rfc=expected_rfc,
            )
            if not out.get("ok"):
                return JSONResponse(
                    {"ok": False, "rejection_reason": out.get("rejection_reason", "Error al guardar"), "status": out.get("status", "error")},
                    status_code=400,
                )
            results.append({
                "statement_id": out.get("statement_id"),
                "inserted_count": out.get("inserted_count", 0),
                "duplicate_statement": out.get("duplicate_statement", False),
                "duplicate_movements_count": out.get("duplicate_movements_count", 0),
            })
            total_inserted += out.get("inserted_count", 0)
            total_duplicate_movements += out.get("duplicate_movements_count", 0)
            if out.get("statement_id"):
                last_statement_id = out["statement_id"]
            if out.get("period_month"):
                period_month = out["period_month"]
        # Retroactive: reclassify movements matching issuer RFC
        _commit_rfc = (issuer.get("rfc") or "").strip().upper()
        if _commit_rfc and total_inserted > 0:
            try:
                _cconn = db()
                reclassify_own_transfers_by_rfc(_cconn, issuer_id, _commit_rfc)
                _cconn.close()
            except Exception:
                pass
        log_action(request, "bank_preview_commit", issuer_id=issuer_id, entity_id=last_statement_id)
        return JSONResponse({
            "ok": True,
            "statement_id": last_statement_id,
            "inserted_count": total_inserted,
            "duplicate_movements_count": total_duplicate_movements,
            "results": results,
            "period_month": period_month,
        })

    @router.get("/bank/pdf-to-excel/download/{file_id}")
    def portal_bank_pdf_to_excel_download(
        request: Request,
        file_id: str,
        issuer: dict = Depends(get_portal_issuer),
    ):
        issuer_id = int(issuer.get("id") or 0)
        fid = (file_id or "").strip()
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesion invalida")
        if not fid:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")

        conn = db()
        try:
            ensure_bank_exports_table(conn)
            row = conn.execute(
                "SELECT xlsx_path FROM bank_pdf_exports WHERE issuer_id = ? AND file_id = ? LIMIT 1",
                (issuer_id, fid),
            ).fetchone()
        finally:
            conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")

        storage_root = get_storage_root(BASE_DIR)
        xlsx_rel_path = (row["xlsx_path"] or "").strip()
        if not xlsx_rel_path:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        try:
            xlsx_abs_path = safe_join(storage_root, xlsx_rel_path)
        except ValueError:
            raise HTTPException(status_code=404, detail="Ruta invalida")
        if not os.path.exists(xlsx_abs_path):
            raise HTTPException(status_code=404, detail="El archivo ya no existe en disco")

        filename = f"estado_cuenta_{fid[:8]}.xlsx"
        file_access_log.log_file_access(
            request=request,
            action="download_bank_xlsx",
            issuer_id=issuer_id,
            user_id=getattr(request.state, "user_id", None),
            file_path=xlsx_rel_path,
            entity="bank_pdf_exports",
            entity_id=fid[:64],
        )
        return FileResponse(
            path=xlsx_abs_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )
