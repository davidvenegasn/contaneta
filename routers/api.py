"""Rutas API JSON: clientes, productos, cotizaciones, proveedores, catálogos SAT."""
import os
import json
import logging
import secrets
import hashlib
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Request, Body, Depends, Query, HTTPException

logger = logging.getLogger(__name__)

from database import db, db_rows, table_exists, has_column, list_catalog, search_catalog, safe_update
from validators import validate_customer, validate_product
from routers.deps import get_portal_issuer
from config import BASE_DIR, DEV_FIXTURES
from facturapi_client import create_invoice, download_invoice, FacturapiError
try:
    from cfdi_pdf import (
        USO_CFDI,
        REGIMEN_FISCAL,
        FORMA_PAGO,
        MONEDA,
        CLAVE_UNIDAD,
    )
except Exception:
    # Fallback si cfdi_pdf no carga (p. ej. reportlab no instalado)
    USO_CFDI = {"G03": "Gastos en general", "G01": "Adquisición de mercancías", "CN01": "Nómina"}
    REGIMEN_FISCAL = {"601": "General de Ley Personas Morales", "612": "Personas Físicas con Actividades Empresariales", "616": "Sin obligaciones fiscales", "626": "Régimen Simplificado de Confianza"}
    FORMA_PAGO = {"03": "Transferencia electrónica", "01": "Efectivo", "99": "Por definir"}
    MONEDA = {"MXN": "Peso Mexicano", "USD": "Dólar Americano"}
    CLAVE_UNIDAD = {"E48": "Unidad de servicio", "EA": "Cada uno", "H87": "Pieza"}
from services import subscription as subscription_service, csrf as csrf_service
from services.action_log import log_action
from services.http import ok, ok_list
from services.schemas import ClientCreate, ProductCreate
from services import clients_service, products_service
from services import jobs as jobs_service
from services import invoices_service

router = APIRouter(prefix="/api")

QUOTATION_STATUSES = ("draft", "sent", "accepted", "rejected", "converted", "expired")

# Paginación: nunca devolver miles de filas; siempre limit/offset con tope
DEFAULT_LIST_LIMIT = 200
MAX_LIST_LIMIT = 500  # tope duro: ningún listado devuelve más de 500 registros por petición
MAX_LIST_OFFSET = 50_000  # tope duro para evitar scans enormes en SQLite

# Fixtures para DEV_FIXTURES=1 (tests/manual_fixtures/*.json)
def _load_fixture(name: str):
    """Si DEV_FIXTURES está activo, carga JSON desde tests/manual_fixtures/{name}.json. Si no existe o falla, devuelve None."""
    if not DEV_FIXTURES:
        return None
    path = os.path.join(BASE_DIR, "tests", "manual_fixtures", f"{name}.json")
    try:
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        logger.debug("Fixture %s: %s", name, e)
    return None


# ----- Account status (checklist activación + P36 chips topbar) -----
@router.get("/account/status")
def api_account_status(request: Request, issuer: dict = Depends(get_portal_issuer)):
    """
    Estado de activación del emisor para el checklist del dropdown "Mi cuenta" y chips del topbar.
    Requiere sesión o token (get_portal_issuer).
    Retorna: issuer_ok, sat_ok, has_customer, has_product, completed, total,
             sat_status, last_sync_at, sync_status, plan_label (P36).
    """
    from services.tenant import require_issuer_id

    issuer_id = require_issuer_id(issuer)
    user_id = getattr(request.state, "user_id", 0) or 0
    issuer_ok = False
    sat_ok = False
    has_customer = False
    has_product = False
    sat_status = "none"
    last_sync_at = None
    sync_status = "ok"
    plan_label = None

    if issuer_id > 0:
        # 1) Datos fiscales: RFC, razón social, régimen no vacíos (CP opcional si existe en DB)
        ir = db_rows(
            "SELECT rfc, razon_social, regimen_fiscal FROM issuers WHERE id = ? LIMIT 1",
            (issuer_id,),
        )
        if ir:
            r = ir[0]
            rfc = (r.get("rfc") or "").strip()
            razon = (r.get("razon_social") or "").strip()
            regimen = (r.get("regimen_fiscal") or "").strip()
            issuer_ok = bool(rfc and razon and regimen)

        # 2) SAT FIEL: credenciales válidas (validation_ok = 1); P36 sat_status: ok / none / error
        sc_valid = db_rows(
            "SELECT 1 FROM sat_credentials WHERE issuer_id = ? AND validation_ok = 1 LIMIT 1",
            (issuer_id,),
        )
        sc_any = db_rows(
            "SELECT 1 FROM sat_credentials WHERE issuer_id = ? LIMIT 1",
            (issuer_id,),
        )
        sat_ok = bool(sc_valid)
        if sc_valid:
            sat_status = "ok"
        elif sc_any:
            sat_status = "error"
        else:
            sat_status = "none"

        # 3) Al menos un cliente
        cust = db_rows("SELECT COUNT(*) AS n FROM customer_profiles WHERE issuer_id = ?", (issuer_id,))
        has_customer = (cust[0]["n"] if cust else 0) >= 1

        # 4) Al menos un producto
        prod = db_rows("SELECT COUNT(*) AS n FROM issuer_products WHERE issuer_id = ?", (issuer_id,))
        has_product = (prod[0]["n"] if prod else 0) >= 1

        # P36: sync status (mismo criterio que /portal/sat/status)
        running = db_rows(
            "SELECT 1 FROM sat_jobs WHERE issuer_id = ? AND status IN ('queued','running') LIMIT 1",
            (issuer_id,),
        )
        last_ok = db_rows(
            "SELECT MAX(finished_at) AS t FROM sat_jobs WHERE issuer_id = ? AND status = 'ok'",
            (issuer_id,),
        )
        last_error = db_rows(
            "SELECT finished_at, last_error FROM sat_jobs WHERE issuer_id = ? AND status = 'error' ORDER BY finished_at DESC LIMIT 1",
            (issuer_id,),
        )
        sync_state = db_rows(
            "SELECT MAX(last_run_at) AS t FROM sat_sync_state WHERE issuer_id = ?",
            (issuer_id,),
        )
        last_sync_at = (sync_state and sync_state[0].get("t")) or (last_ok and last_ok[0].get("t")) or None
        if running:
            sync_status = "running"
        elif last_error and last_ok and last_error[0].get("t") and last_ok[0].get("t") and last_error[0]["t"] > last_ok[0]["t"]:
            sync_status = "error"
        elif last_error and not last_ok:
            sync_status = "error"
        else:
            sync_status = "ok"

        # P36: plan_label Trial / Pro
        if subscription_service.is_subscription_active(user_id):
            plan_label = "Pro"
        elif subscription_service.is_issuer_trial_active(issuer_id):
            plan_label = "Trial"
        else:
            plan_label = None

    completed = sum([issuer_ok, sat_ok, has_customer, has_product])
    return {
        "issuer_ok": issuer_ok,
        "sat_ok": sat_ok,
        "has_customer": has_customer,
        "has_product": has_product,
        "completed": completed,
        "total": 4,
        "sat_status": sat_status,
        "last_sync_at": last_sync_at,
        "sync_status": sync_status,
        "plan_label": plan_label,
    }


def _provider_report_rows(issuer_id: int, rfc_norm: str):
    return db_rows(
        """
        SELECT uuid, fecha_emision, nombre_emisor, serie, folio, concepto,
               subtotal, descuento, impuestos, total, moneda, status,
               forma_pago, metodo_pago, uso_cfdi
        FROM sat_cfdi
        WHERE issuer_id = ? AND direction = 'received'
          AND UPPER(TRIM(COALESCE(rfc_emisor,''))) = ?
          AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
          AND total IS NOT NULL AND total >= 0.01
        ORDER BY fecha_emision DESC
        """,
        (issuer_id, rfc_norm),
    )


def _build_provider_report_pdf(issuer: dict, provider_name: str, rows: list) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="ReportTitle", parent=styles["Heading1"], fontSize=14, spaceAfter=12)
    body_style = ParagraphStyle(name="ReportBody", parent=styles["Normal"], fontSize=10, spaceAfter=6)
    receptor_name = (issuer.get("alias") or issuer.get("rfc") or "Receptor").replace("<", " ").replace(">", " ")
    provider_safe = (provider_name or "Proveedor").replace("<", " ").replace(">", " ")

    story = [
        Paragraph(f"<b>Facturas recibidas de</b> {provider_safe}", title_style),
        Paragraph(f"<b>Receptor:</b> {receptor_name} — RFC: {issuer.get('rfc') or '—'}", body_style),
        Spacer(1, 14),
    ]
    headers = ["Fecha", "UUID", "Concepto", "Subtotal", "Descuento", "Impuestos", "Total", "Moneda", "Estado"]
    data = [headers]
    max_concepto_len = 18
    for r in rows:
        fecha = (r.get("fecha_emision") or "")[:10] if r.get("fecha_emision") else "—"
        uuid_short = (r.get("uuid") or "—")[:8] + "…" if r.get("uuid") and len(r.get("uuid", "")) > 8 else (r.get("uuid") or "—")
        raw_concepto = str(r.get("concepto") or "—")
        concepto = (raw_concepto[:max_concepto_len] + "…") if len(raw_concepto) > max_concepto_len else raw_concepto
        data.append([
            fecha, uuid_short, concepto,
            f"{float(r.get('subtotal') or 0):,.2f}", f"{float(r.get('descuento') or 0):,.2f}",
            f"{float(r.get('impuestos') or 0):,.2f}", f"{float(r.get('total') or 0):,.2f}",
            str(r.get("moneda") or "MXN"),
            "Vigente" if r.get("status") == "1" else ("Cancelada" if r.get("status") == "0" else "—"),
        ])
    t = Table(data, colWidths=[inch * 0.9, inch * 1.0, inch * 1.4, 0.7 * inch, 0.6 * inch, 0.6 * inch, 0.7 * inch, 0.5 * inch, 0.6 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8e8e8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (2, -1), "LEFT"),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8f8")]),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def _build_provider_report_xlsx(issuer: dict, provider_name: str, rows: list) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas recibidas"
    receptor_name = issuer.get("alias") or issuer.get("rfc") or "Receptor"
    provider_safe = provider_name or "Proveedor"
    ws["A1"] = f"Facturas recibidas de {provider_safe}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Receptor: {receptor_name} — RFC: {issuer.get('rfc') or '—'}"
    ws["A2"].font = Font(size=10)
    headers = ["Fecha", "UUID", "Concepto", "Subtotal", "Descuento", "Impuestos", "Total", "Moneda", "Estado"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h).font = Font(bold=True)
    for i, r in enumerate(rows, 5):
        fecha = (r.get("fecha_emision") or "")[:10] if r.get("fecha_emision") else "—"
        ws.cell(row=i, column=1, value=fecha)
        ws.cell(row=i, column=2, value=r.get("uuid") or "—")
        ws.cell(row=i, column=3, value=(r.get("concepto") or "—")[:200])
        ws.cell(row=i, column=4, value=float(r.get("subtotal") or 0))
        ws.cell(row=i, column=5, value=float(r.get("descuento") or 0))
        ws.cell(row=i, column=6, value=float(r.get("impuestos") or 0))
        ws.cell(row=i, column=7, value=float(r.get("total") or 0))
        ws.cell(row=i, column=8, value=r.get("moneda") or "MXN")
        ws.cell(row=i, column=9, value="Vigente" if r.get("status") == "1" else ("Cancelada" if r.get("status") == "0" else "—"))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----- Customers -----
# List endpoints: return 200 + [] when no data; only 4xx/5xx on real errors.


# ----- Jobs (genérico) -----
@router.get("/jobs")
def api_jobs(
    issuer: dict = Depends(get_portal_issuer),
    limit: int = Query(20, ge=1, le=200, description="Máximo de registros"),
):
    items = jobs_service.list_jobs(issuer["id"], limit=limit)
    total = jobs_service.count_jobs(issuer["id"])
    return ok_list(items, total=total)


@router.get("/jobs/{job_id}")
def api_job_get(job_id: int, issuer: dict = Depends(get_portal_issuer)):
    job = jobs_service.get_job_for_issuer(job_id, issuer["id"])
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado.")
    payload = {
        "id": job.get("id"),
        "issuer_id": job.get("issuer_id"),
        "name": job.get("name"),
        "status": job.get("status"),
        "progress": job.get("progress"),
        "message": job.get("message"),
        "payload": job.get("payload"),
        "result": job.get("result"),
        "created_at": job.get("created_at"),
        "updated_at": job.get("updated_at"),
    }
    return ok(payload)
@router.get("/customers")
def api_customers(
    issuer: dict = Depends(get_portal_issuer),
    limit: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Máximo de registros"),
    offset: int = Query(0, ge=0, le=MAX_LIST_OFFSET, description="Registros a saltar"),
):
    fixture = _load_fixture("clients")
    if fixture is not None:
        return fixture
    try:
        conn = db()
        issuer_id = issuer["id"]
        # Incluir clientes que están en la tabla "clients" (p. ej. backfill desde facturas emitidas)
        # para que el dropdown de factura rápida muestre los mismos que la página Contactos > Clientes.
        if table_exists(conn, "clients"):
            conn.execute(
                """
                INSERT OR IGNORE INTO customer_profiles (issuer_id, rfc, legal_name, zip, tax_system, email, alias, updated_at)
                SELECT issuer_id, rfc, COALESCE(name, ''), COALESCE(cp, ''), COALESCE(regimen_fiscal, ''), email, NULL, datetime('now')
                FROM clients WHERE issuer_id = ?
                """,
                (issuer_id,),
            )
            conn.commit()
        conn.close()
        items, total = clients_service.list_clients(issuer_id, limit=limit, offset=offset)
        return ok_list(items, total=total)
    except Exception as e:
        logger.warning("api_customers: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error al cargar la lista de clientes.")


@router.post("/customers/create")
def api_customers_create(request: Request, payload: ClientCreate = Body(...), issuer: dict = Depends(get_portal_issuer)):
    csrf_service.verify_api_csrf(request)
    try:
        rfc = payload.rfc
        legal_name = payload.legal_name
        zip_val = payload.zip or ""
        tax_val = payload.tax_system or ""
        email = payload.email or None
        alias = payload.alias or None
        errors = validate_customer(rfc, legal_name, zip_val, tax_val, email)
        if errors:
            raise HTTPException(status_code=400, detail="; ".join(errors))
        conn = db()
        conn.execute(
            """
            INSERT INTO customer_profiles (issuer_id, rfc, legal_name, zip, tax_system, email, alias, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(issuer_id, rfc) DO UPDATE SET
                legal_name = excluded.legal_name, zip = excluded.zip, tax_system = excluded.tax_system,
                email = excluded.email, alias = excluded.alias, updated_at = CURRENT_TIMESTAMP
            """,
            (issuer["id"], rfc, legal_name, zip_val, tax_val, email, alias),
        )
        conn.commit()
        conn.close()
        return {"ok": True, "data": {"rfc": rfc}}
    except HTTPException:
        raise
    except Exception:
        logger.exception("api customers create: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos guardar el cliente. Intenta de nuevo.",
        )


@router.post("/customers/delete")
def api_customers_delete(request: Request, payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    csrf_service.verify_api_csrf(request)
    try:
        rfc = (payload.get("rfc") or "").strip().upper()
        if not rfc:
            raise HTTPException(status_code=400, detail="RFC requerido")
        conn = db()
        cur = conn.execute("DELETE FROM customer_profiles WHERE issuer_id = ? AND rfc = ?", (issuer["id"], rfc))
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        return {"ok": True, "data": {"rfc": rfc}}
    except HTTPException:
        raise
    except Exception:
        logger.exception("api customers delete: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos completar la acción. Intenta de nuevo.",
        )


# ----- Products -----
@router.get("/products")
def api_products(
    issuer: dict = Depends(get_portal_issuer),
    limit: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Máximo de registros"),
    offset: int = Query(0, ge=0, le=MAX_LIST_OFFSET, description="Registros a saltar"),
):
    fixture = _load_fixture("products")
    if fixture is not None:
        return fixture
    try:
        issuer_id = issuer["id"]
        items, total = products_service.list_products(issuer_id, limit=limit, offset=offset)
        return ok_list(items, total=total)
    except Exception as e:
        logger.warning("api_products: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error al cargar la lista de productos.")


@router.post("/products/create")
def api_products_create(request: Request, payload: ProductCreate = Body(...), issuer: dict = Depends(get_portal_issuer)):
    csrf_service.verify_api_csrf(request)
    try:
        description = payload.description
        product_key_raw = payload.product_key
        # product_key ya viene normalizado en el schema (split '—'), pero conservamos el raw para validar/error.
        product_key = payload.product_key
        unit_key = payload.unit_key or "E48"
        unit_price = float(payload.unit_price)
        iva_rate = float(payload.iva_rate)
        errors = validate_product(description, product_key_raw, unit_key, unit_price)
        if errors:
            raise HTTPException(status_code=400, detail="; ".join(errors))
        conn = db()
        conn.execute(
            """INSERT INTO issuer_products (issuer_id, description, product_key, unit_key, unit_price, iva_rate)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (issuer["id"], description, product_key, unit_key, unit_price, iva_rate),
        )
        conn.commit()
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.close()
        return {"ok": True, "data": {"id": rid}}
    except HTTPException:
        raise
    except Exception:
        logger.exception("api products create: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos guardar el producto. Intenta de nuevo.",
        )


@router.post("/products/delete")
def api_products_delete(request: Request, payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    """Elimina un producto del emisor. P37: uso con modal de confirmación en el portal."""
    csrf_service.verify_api_csrf(request)
    product_id = payload.get("id") or payload.get("product_id")
    if product_id is None:
        raise HTTPException(status_code=400, detail="id o product_id es requerido.")
    try:
        product_id = int(product_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="id debe ser numérico.")
    issuer_id = issuer["id"]
    conn = db()
    cur = conn.execute(
        "DELETE FROM issuer_products WHERE issuer_id = ? AND id = ?",
        (issuer_id, product_id),
    )
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="Producto no encontrado o ya fue eliminado.")
    return {"ok": True}


# ----- Quick invoice (Home: cliente + producto → timbrar sin salir) -----
@router.get("/quick-invoice/bootstrap")
def api_quick_invoice_bootstrap(issuer: dict = Depends(get_portal_issuer)):
    """Devuelve clientes, productos, defaults y catálogos para el widget Factura rápida en Inicio."""
    try:
        conn = db()
        issuer_id = issuer["id"]
        # Clientes (misma fuente que /api/customers y Contactos)
        if table_exists(conn, "clients"):
            conn.execute(
                """
                INSERT OR IGNORE INTO customer_profiles (issuer_id, rfc, legal_name, zip, tax_system, email, alias, updated_at)
                SELECT issuer_id, rfc, COALESCE(name, ''), COALESCE(cp, ''), COALESCE(regimen_fiscal, ''), email, NULL, datetime('now')
                FROM clients WHERE issuer_id = ?
                """,
                (issuer_id,),
            )
            conn.commit()
        rows_c = conn.execute(
            """
            SELECT id, rfc, legal_name, zip, tax_system, email, alias
            FROM customer_profiles WHERE issuer_id = ? ORDER BY COALESCE(alias, ''), rfc
            LIMIT 500
            """,
            (issuer_id,),
        ).fetchall()
        clients = [
            {
                "id": r["id"],
                "rfc": r["rfc"],
                "name": r["legal_name"],
                "legal_name": r["legal_name"],
                "zip": r["zip"],
                "regimen": r["tax_system"],
                "tax_system": r["tax_system"],
                "email": r["email"],
            }
            for r in rows_c
        ]
        # Productos (misma fuente que /api/products y Productos)
        if table_exists(conn, "products"):
            rows_p = conn.execute(
                """
                SELECT id, name, clave_prod_serv, clave_unidad, unidad, default_unit_price, default_currency
                FROM products WHERE issuer_id = ? AND COALESCE(active, 1) = 1 ORDER BY name LIMIT 500
                """,
                (issuer_id,),
            ).fetchall()
            products = [
                {
                    "id": r["id"],
                    "name": r["name"] or "",
                    "description": r["name"] or "",
                    "price": float(r["default_unit_price"] or 0),
                    "unit_price": float(r["default_unit_price"] or 0),
                    "currency": (r["default_currency"] or "MXN").strip() or "MXN",
                    "prodserv": r["clave_prod_serv"] or "",
                    "product_key": r["clave_prod_serv"] or "",
                    "unit_key": r["clave_unidad"] or "E48",
                    "unit_name": r["unidad"] or "",
                    "iva_default": 0.16,
                }
                for r in rows_p
            ]
        else:
            rows_p = conn.execute(
                """
                SELECT id, description, product_key, unit_key, unit_price, iva_rate
                FROM issuer_products WHERE issuer_id = ? ORDER BY description LIMIT 500
                """,
                (issuer_id,),
            ).fetchall()
            products = [
                {
                    "id": r["id"],
                    "name": r["description"] or "",
                    "description": r["description"] or "",
                    "price": float(r["unit_price"] or 0),
                    "unit_price": float(r["unit_price"] or 0),
                    "currency": "MXN",
                    "prodserv": r["product_key"] or "",
                    "product_key": r["product_key"] or "",
                    "unit_key": r["unit_key"] or "E48",
                    "unit_name": "",
                    "iva_default": float(r["iva_rate"] or 0.16),
                }
                for r in rows_p
            ]
        conn.close()
        payload = {
            "clients": clients,
            "products": products,
            "defaults": {
                "currency": "MXN",
                "payment_form": "03",
                "payment_method": "PUE",
                "uso_cfdi": "G03",
                "series": None,
                "folio": None,
            },
            "tax_presets": {
                "ivas": [
                    {"rate": 0.16, "label": "IVA 16%"},
                    {"rate": 0.0, "label": "IVA 0%"},
                ],
                "retenciones": [
                    {"type": "ISR", "rate": 0.10, "label": "Ret ISR 10%"},
                    {"type": "IVA", "rate": 0.1067, "label": "Ret IVA 10.67%"},
                ],
            },
        }
        return ok(payload)
    except Exception as e:
        logger.warning("quick-invoice bootstrap: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error al cargar datos para factura rápida.")


@router.post("/invoices/quick")
def api_invoices_quick(
    request: Request,
    payload: dict = Body(...),
    issuer: dict = Depends(get_portal_issuer),
):
    """Crea y timbra una factura con un solo concepto desde cliente + producto.

    Permite overrides mínimos (receptor, concepto, IVA/retenciones) para el precálculo editable en Home.
    """
    csrf_service.verify_api_csrf(request)
    user_id = getattr(request.state, "user_id", 0) or 0
    if not subscription_service.can_issuer_use_sync_and_timbrado(issuer.get("id"), user_id):
        raise HTTPException(status_code=402, detail="Actualiza tu plan para emitir facturas.")
    customer_id = payload.get("customer_id")
    items_in = payload.get("items")
    product_id = payload.get("product_id")
    has_items = isinstance(items_in, list) and len(items_in) > 0
    if not customer_id:
        raise HTTPException(status_code=400, detail="customer_id es requerido.")
    if not has_items and not product_id:
        raise HTTPException(status_code=400, detail="product_id es requerido (o envía items).")
    try:
        customer_id = int(customer_id)
        if not has_items:
            product_id = int(product_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="customer_id/product_id inválidos.")

    # Single-item mode legacy fields
    quantity = None
    unit_price_override = None
    if not has_items:
        quantity = float(payload.get("quantity", 1))
        if quantity <= 0 or quantity > 999999:
            raise HTTPException(status_code=400, detail="Cantidad inválida.")
        unit_price_override = payload.get("unit_price")
        if unit_price_override is not None:
            try:
                unit_price_override = float(unit_price_override)
                if unit_price_override < 0:
                    raise HTTPException(status_code=400, detail="Precio unitario no puede ser negativo.")
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Precio unitario inválido.")

    issuer_id = issuer["id"]
    cust = db_rows(
        "SELECT id, rfc, legal_name, zip, tax_system, email FROM customer_profiles WHERE issuer_id = ? AND id = ? LIMIT 1",
        (issuer_id, customer_id),
    )
    if not cust:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
    c = cust[0]

    def _resolve_product(pid: int) -> dict:
        """Resolver producto desde issuer_products o products."""
        rows = db_rows(
            "SELECT id, description, product_key, unit_key, unit_price, iva_rate FROM issuer_products WHERE issuer_id = ? AND id = ? LIMIT 1",
            (issuer_id, pid),
        )
        if rows:
            return rows[0]
        _conn = db()
        try:
            if table_exists(_conn, "products"):
                row = _conn.execute(
                    "SELECT id, name, clave_prod_serv, clave_unidad, default_unit_price FROM products WHERE issuer_id = ? AND id = ? LIMIT 1",
                    (issuer_id, pid),
                ).fetchone()
                if row:
                    row = dict(row)
                    return {
                        "id": row["id"],
                        "description": row.get("name") or "",
                        "product_key": row.get("clave_prod_serv") or "",
                        "unit_key": row.get("clave_unidad") or "E48",
                        "unit_price": float(row.get("default_unit_price") or 0),
                        "iva_rate": 0.16,
                    }
        finally:
            _conn.close()
        raise HTTPException(status_code=404, detail=f"Producto no encontrado: {pid}")

    p = _resolve_product(int(product_id)) if not has_items else None
    # ----- Overrides (desde Home precálculo editable) -----
    customer_rfc = (c.get("rfc") or "").strip().upper()
    customer_legal_name = (
        (payload.get("customer_legal_name") or payload.get("customer_name") or c.get("legal_name") or "").strip()
        or (c.get("legal_name") or "").strip()
    )
    customer_zip = (payload.get("customer_zip") if payload.get("customer_zip") is not None else (c.get("zip") or "")).strip() or "00000"
    customer_tax_system = (payload.get("customer_tax_system") if payload.get("customer_tax_system") is not None else (c.get("tax_system") or "")).strip() or "616"
    customer_email_raw = (payload.get("customer_email") if payload.get("customer_email") is not None else (c.get("email") or "")).strip()
    customer_email = customer_email_raw or None

    def _parse_iva_rate(val, default_val: float) -> tuple[float, bool]:
        """Devuelve (iva_rate, iva_exempt). Acepta 'EXENTO'."""
        if val is None or val == "":
            return (max(0.0, min(1.0, float(default_val))), False)
        if isinstance(val, str) and val.strip().upper() == "EXENTO":
            return (0.0, True)
        try:
            n = float(val)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="IVA rate inválido.")
        n = max(0.0, min(1.0, n))
        return (n, False)

    def _parse_rate(name: str, default: float = 0.0) -> float:
        v = payload.get(name)
        if v is None or v == "":
            return default
        try:
            n = float(v)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f"{name} inválido.")
        return max(0.0, min(1.0, n))

    isr_ret_rate = _parse_rate("isr_ret_rate", 0.0)
    iva_ret_rate = _parse_rate("iva_ret_rate", 0.0)

    cfdi_use = (payload.get("cfdi_use") or payload.get("uso_cfdi") or "G03").strip().upper() or "G03"
    payment_form = (payload.get("payment_form") or "03").strip() or "03"
    payment_method = (payload.get("payment_method") or "PUE").strip().upper() or "PUE"
    currency = (payload.get("currency") or "MXN").strip().upper() or "MXN"

    # Validaciones ligeras (mismo validador que create_customer/product)
    cust_errors = validate_customer(customer_rfc, customer_legal_name, customer_zip, customer_tax_system, customer_email)
    if cust_errors:
        raise HTTPException(status_code=400, detail="; ".join(cust_errors))
    items_fact = []
    items_meta = []  # para DB invoice_items + sat_cfdi (multi-item)
    if has_items:
        for it in items_in:
            if not isinstance(it, dict):
                raise HTTPException(status_code=400, detail="items inválidos.")
            pid = it.get("product_id")
            if not pid:
                raise HTTPException(status_code=400, detail="Cada item requiere product_id.")
            try:
                pid = int(pid)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="product_id inválido en items.")
            qty = float(it.get("quantity", 1))
            if qty <= 0 or qty > 999999:
                raise HTTPException(status_code=400, detail="Cantidad inválida en items.")
            base_p = _resolve_product(pid)
            description = (it.get("description") or base_p.get("description") or "").strip() or (base_p.get("description") or "").strip()
            product_key = (it.get("product_key") or base_p.get("product_key") or "").strip() or "84111500"
            unit_key = (it.get("unit_key") or base_p.get("unit_key") or "").strip() or "E48"
            up_override = it.get("unit_price")
            if up_override is not None and up_override != "":
                try:
                    up_override = float(up_override)
                    if up_override < 0:
                        raise HTTPException(status_code=400, detail="Precio unitario no puede ser negativo.")
                except (TypeError, ValueError):
                    raise HTTPException(status_code=400, detail="Precio unitario inválido.")
            unit_price = float(up_override if up_override is not None and up_override != "" else (base_p.get("unit_price") or 0))
            iva_rate, iva_exempt = _parse_iva_rate(it.get("iva_rate"), float(base_p.get("iva_rate") or 0.16))

            prod_errors = validate_product(description, product_key, unit_key, unit_price)
            if prod_errors:
                raise HTTPException(status_code=400, detail="; ".join(prod_errors))

            price_to_send = unit_price * (1.0 + iva_rate) if iva_rate else unit_price
            taxes = []
            if not iva_exempt:
                taxes.append({"type": "IVA", "rate": iva_rate})
            if isr_ret_rate > 0:
                taxes.append({"type": "ISR", "rate": isr_ret_rate, "withholding": True})
            if iva_ret_rate > 0:
                taxes.append({"type": "IVA", "rate": iva_ret_rate, "withholding": True})
            items_fact.append(
                {
                    "quantity": qty,
                    "discount": 0.0,
                    "product": {
                        "description": description,
                        "product_key": product_key,
                        "price": round(price_to_send, 2),
                        "tax_included": True,
                        "taxes": taxes,
                        "unit_key": unit_key,
                    },
                }
            )
            items_meta.append(
                {
                    "quantity": qty,
                    "description": description,
                    "product_key": product_key,
                    "unit_key": unit_key,
                    "unit_price": unit_price,  # sin IVA
                    "iva_rate": iva_rate,
                    "price_to_send": round(price_to_send, 2),  # con IVA si aplica
                }
            )
    else:
        description = (payload.get("description") or p.get("description") or "").strip() or (p.get("description") or "").strip()
        product_key = (payload.get("product_key") or p.get("product_key") or "").strip() or "84111500"
        unit_key = (payload.get("unit_key") or p.get("unit_key") or "").strip() or "E48"
        unit_price = float(unit_price_override if unit_price_override is not None else (p.get("unit_price") or 0))
        iva_rate, iva_exempt = _parse_iva_rate(payload.get("iva_rate"), float(p.get("iva_rate") or 0.16))

        prod_errors = validate_product(description, product_key, unit_key, unit_price)
        if prod_errors:
            raise HTTPException(status_code=400, detail="; ".join(prod_errors))

        price_to_send = unit_price * (1.0 + iva_rate) if iva_rate else unit_price
        taxes = []
        if not iva_exempt:
            taxes.append({"type": "IVA", "rate": iva_rate})
        if isr_ret_rate > 0:
            taxes.append({"type": "ISR", "rate": isr_ret_rate, "withholding": True})
        if iva_ret_rate > 0:
            taxes.append({"type": "IVA", "rate": iva_ret_rate, "withholding": True})
        items_fact.append(
            {
                "quantity": quantity,
                "discount": 0.0,
                "product": {
                    "description": description,
                    "product_key": product_key,
                    "price": round(price_to_send, 2),
                    "tax_included": True,
                    "taxes": taxes,
                    "unit_key": unit_key,
                },
            }
        )
        items_meta.append(
            {
                "quantity": quantity,
                "description": description,
                "product_key": product_key,
                "unit_key": unit_key,
                "unit_price": unit_price,
                "iva_rate": iva_rate,
                "price_to_send": round(price_to_send, 2),
            }
        )
    payload_fact = invoices_service.build_invoice_payload(
        invoice_type="I",
        export_code="01",
        customer=invoices_service.build_customer(
            rfc=customer_rfc,
            legal_name=customer_legal_name,
            zip_code=customer_zip,
            tax_system=customer_tax_system,
            email=customer_email,
        ),
        items=items_fact,
        payments=None,
        cfdi_use=cfdi_use,
        payment_form=payment_form,
        payment_method=payment_method,
        currency=currency,
    )
    if issuer.get("facturapi_org_id") in (None, "", 0) or issuer.get("id") == -1:
        raise HTTPException(status_code=400, detail="Configuración de facturación no disponible.")
    conn = db()
    try:
        cur = conn.execute(
            """
            INSERT INTO invoices (
                issuer_id, currency, exchange_rate,
                payment_form, payment_method, cfdi_use,
                customer_rfc, customer_legal_name,
                customer_zip, customer_tax_system, customer_email
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                issuer_id,
                currency,
                None,
                payment_form,
                payment_method,
                cfdi_use,
                customer_rfc,
                customer_legal_name,
                customer_zip,
                customer_tax_system,
                customer_email,
            ),
        )
        invoice_local_id = cur.lastrowid
        safe_update(
            conn,
            "invoices",
            invoice_local_id,
            {"export_code": "01", "tipo_comprobante": "I"},
        )
        cols = {r[1] for r in conn.execute("PRAGMA table_info(invoice_items)").fetchall()}
        base_cols = ["invoice_id", "quantity", "description", "product_key", "unit_price", "iva_rate"]
        has_unit_key = "unit_key" in cols
        has_discount = "discount" in cols
        insert_cols = base_cols + (["unit_key"] if has_unit_key else []) + (["discount"] if has_discount else [])
        placeholders = ", ".join(["?"] * len(insert_cols))
        for it in items_meta:
            base_vals = [
                invoice_local_id,
                float(it["quantity"]),
                it["description"],
                it["product_key"],
                float(it["price_to_send"]),
                float(it["iva_rate"]),
            ]
            extra_vals = []
            if has_unit_key:
                extra_vals.append(it.get("unit_key") or "E48")
            if has_discount:
                extra_vals.append(0.0)
            conn.execute(
                f"INSERT INTO invoice_items ({', '.join(insert_cols)}) VALUES ({placeholders})",
                tuple(base_vals + extra_vals),
            )
        conn.commit()
    except Exception as e:
        conn.close()
        logger.warning("api_invoices_quick insert: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error al registrar la factura.")
    try:
        invoice = create_invoice(issuer["facturapi_org_id"], payload_fact)
    except FacturapiError as fe:
        conn.close()
        logger.warning("api invoices quick FacturapiError: issuer_id=%s %s", issuer.get("id"), fe, exc_info=True)
        raise HTTPException(
            status_code=400,
            detail="No se pudo timbrar la factura. Revisa los datos e intenta de nuevo.",
        )
    fact_id = invoice.get("id")
    uuid = invoice.get("uuid")
    total = invoice.get("total")
    conn.execute(
        "UPDATE invoices SET facturapi_invoice_id = ?, uuid = ?, total = ? WHERE id = ? AND issuer_id = ?",
        (fact_id, uuid, total, invoice_local_id, issuer_id),
    )
    conn.commit()
    conn.close()

    # ----- Guardar XML en storage + registrar en sat_cfdi para descargas /portal/sat/xml|pdf/{uuid} -----
    try:
        if uuid and fact_id:
            xml_bytes = download_invoice(issuer["facturapi_org_id"], fact_id, "xml")
            if isinstance(xml_bytes, str):
                xml_bytes = xml_bytes.encode("utf-8")
            if xml_bytes:
                now = datetime.utcnow()
                year = now.strftime("%Y")
                month = now.strftime("%m")
                rel_path = os.path.join("storage", "xml", str(issuer_id), "issued", year, month, f"{uuid}.xml")
                abs_path = os.path.normpath(os.path.abspath(os.path.join(BASE_DIR, rel_path)))
                os.makedirs(os.path.dirname(abs_path), exist_ok=True)
                with open(abs_path, "wb") as f:
                    f.write(xml_bytes)
                xml_sha256 = hashlib.sha256(xml_bytes).hexdigest()

                subtotal = sum(float(it["quantity"]) * float(it["unit_price"]) for it in items_meta)
                iva_amt = sum(float(it["quantity"]) * float(it["unit_price"]) * float(it["iva_rate"]) for it in items_meta)
                ret_isr_amt = subtotal * float(isr_ret_rate)
                ret_iva_amt = iva_amt * float(iva_ret_rate)
                ret_total = ret_isr_amt + ret_iva_amt
                concepto_txt = (
                    (items_meta[0]["description"] or "")[:220]
                    if len(items_meta) == 1
                    else f"{len(items_meta)} conceptos"
                )

                conn2 = db()
                conn2.execute(
                    """
                    INSERT INTO sat_cfdi (
                      issuer_id, direction, uuid, status, fecha_emision,
                      rfc_emisor, nombre_emisor, rfc_receptor, nombre_receptor,
                      total, moneda, tipo_comprobante, xml_path, uso_cfdi,
                      subtotal, impuestos, retenciones, concepto, metodo_pago, forma_pago,
                      xml_status, xml_sha256, xml_downloaded_at, updated_at
                    ) VALUES (
                      ?, 'issued', ?, 'V', ?,
                      ?, ?, ?, ?,
                      ?, ?, 'I', ?, ?,
                      ?, ?, ?, ?, ?, ?,
                      'ok', ?, datetime('now'), datetime('now')
                    )
                    ON CONFLICT(issuer_id, direction, uuid) DO UPDATE SET
                      xml_path = excluded.xml_path,
                      total = excluded.total,
                      moneda = excluded.moneda,
                      uso_cfdi = excluded.uso_cfdi,
                      subtotal = excluded.subtotal,
                      impuestos = excluded.impuestos,
                      retenciones = excluded.retenciones,
                      concepto = excluded.concepto,
                      metodo_pago = excluded.metodo_pago,
                      forma_pago = excluded.forma_pago,
                      xml_status = excluded.xml_status,
                      xml_sha256 = excluded.xml_sha256,
                      xml_downloaded_at = excluded.xml_downloaded_at,
                      updated_at = datetime('now')
                    """,
                    (
                        issuer_id,
                        uuid,
                        now.isoformat(timespec="seconds"),
                        (issuer.get("rfc") or "").strip().upper() or None,
                        (issuer.get("razon_social") or "").strip() or None,
                        customer_rfc,
                        customer_legal_name,
                        float(total or (subtotal + iva_amt - ret_total)),
                        currency,
                        rel_path,
                        cfdi_use,
                        float(subtotal),
                        float(iva_amt),
                        float(ret_total),
                        concepto_txt,
                        payment_method,
                        payment_form,
                        xml_sha256,
                    ),
                )
                conn2.commit()
                conn2.close()
    except Exception as e:
        logger.warning("api_invoices_quick xml/sat_cfdi: %s", e, exc_info=True)

    log_action(request, "invoice_created", issuer_id=issuer["id"], invoice_id=fact_id, uuid=(uuid or "")[:36])
    return {"ok": True, "uuid": uuid, "total": total}


@router.post("/invoices/bulk_issue")
def api_invoices_bulk_issue(
    request: Request,
    payload: dict = Body(...),
    issuer: dict = Depends(get_portal_issuer),
):
    """Emite N facturas (una por cliente) con 1 producto.

    Fase 1: simplificado (sin retenciones por cliente).
    """
    csrf_service.verify_api_csrf(request)
    user_id = getattr(request.state, "user_id", 0) or 0
    if not subscription_service.can_issuer_use_sync_and_timbrado(issuer.get("id"), user_id):
        raise HTTPException(status_code=402, detail="Actualiza tu plan para emitir facturas.")

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Payload inválido.")

    customer_ids = payload.get("customer_ids") or payload.get("client_ids") or []
    product_id = payload.get("product_id")
    qty = payload.get("qty", payload.get("quantity", 1))
    unit_price_override = payload.get("unit_price")

    if not isinstance(customer_ids, list) or not customer_ids:
        raise HTTPException(status_code=400, detail="customer_ids es requerido.")
    if not product_id:
        raise HTTPException(status_code=400, detail="product_id es requerido.")
    try:
        product_id = int(product_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="product_id inválido.")
    try:
        qty = float(qty)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="qty inválido.")
    if qty <= 0 or qty > 999999:
        raise HTTPException(status_code=400, detail="qty inválido.")
    if unit_price_override is not None and unit_price_override != "":
        try:
            unit_price_override = float(unit_price_override)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="unit_price inválido.")

    issuer_id = int(issuer.get("id") or 0)
    if issuer_id <= 0:
        raise HTTPException(status_code=401, detail="Sesión inválida.")

    results = []
    conn = db()
    try:
        for cid in customer_ids:
            try:
                client_id = int(cid)
            except (TypeError, ValueError):
                results.append({"customer_id": cid, "ok": False, "error": "ID inválido"})
                continue

            row = conn.execute(
                "SELECT id, rfc, name, cp, regimen_fiscal, email FROM clients WHERE issuer_id = ? AND id = ? LIMIT 1",
                (issuer_id, client_id),
            ).fetchone()
            if not row:
                results.append({"customer_id": client_id, "ok": False, "error": "Cliente no encontrado"})
                continue
            row = dict(row)
            rfc = (row.get("rfc") or "").strip().upper()
            legal_name = (row.get("name") or "").strip() or rfc
            zip_code = (row.get("cp") or "").strip() or "00000"
            tax_system = (row.get("regimen_fiscal") or "").strip() or "616"
            email = (row.get("email") or "").strip() or None

            if not rfc:
                results.append({"customer_id": client_id, "ok": False, "error": "RFC vacío"})
                continue

            # Asegurar customer_profile por RFC (reusa validación existente)
            cust_row = conn.execute(
                "SELECT id FROM customer_profiles WHERE issuer_id = ? AND rfc = ? LIMIT 1",
                (issuer_id, rfc),
            ).fetchone()
            if cust_row:
                customer_profile_id = int(cust_row["id"])
            else:
                cust_errors = validate_customer(rfc, legal_name, zip_code, tax_system, email)
                if cust_errors:
                    results.append({"customer_id": client_id, "ok": False, "error": "; ".join(cust_errors)})
                    continue
                cur = conn.execute(
                    """
                    INSERT INTO customer_profiles (issuer_id, rfc, legal_name, zip, tax_system, email, alias, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, NULL, datetime('now'), datetime('now'))
                    """,
                    (issuer_id, rfc, legal_name, zip_code, tax_system, email),
                )
                customer_profile_id = int(cur.lastrowid)
                conn.commit()

            try:
                quick_payload = {
                    "customer_id": customer_profile_id,
                    "product_id": product_id,
                    "quantity": qty,
                }
                if unit_price_override is not None and unit_price_override != "":
                    quick_payload["unit_price"] = unit_price_override
                r = api_invoices_quick(request, payload=quick_payload, issuer=issuer)
                results.append({"customer_id": client_id, "ok": True, "uuid": r.get("uuid"), "total": r.get("total")})
            except HTTPException as he:
                results.append({"customer_id": client_id, "ok": False, "error": he.detail})
            except Exception as e:
                logger.warning("bulk_issue: error emitiendo a client_id=%s: %s", client_id, e, exc_info=True)
                results.append({"customer_id": client_id, "ok": False, "error": "No se pudo emitir."})
    finally:
        conn.close()

    return ok({"results": results})


# ----- Quotations -----
@router.get("/quotations")
def api_quotations_list(
    issuer: dict = Depends(get_portal_issuer),
    limit: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Máximo de registros"),
    offset: int = Query(0, ge=0, le=MAX_LIST_OFFSET, description="Registros a saltar"),
):
    try:
        issuer_id = issuer["id"]
    except (ValueError, KeyError, TypeError):
        logger.exception("api quotations list: issuer inválido")
        raise HTTPException(status_code=401, detail="Sesión inválida")
    try:
        conn = db()
        total_row = conn.execute(
            "SELECT COUNT(*) AS c FROM quotations WHERE issuer_id = ?",
            (issuer_id,),
        ).fetchone()
        total = total_row[0] if total_row else 0
        rows = conn.execute(
            """
            SELECT q.id, q.folio, q.customer_rfc, q.customer_legal_name, q.customer_email,
                   q.status, q.public_token, q.valid_until, q.notes, q.responded_at, q.created_at, q.updated_at,
                   (SELECT COALESCE(SUM((qi.quantity * qi.unit_price) * (1 + COALESCE(qi.iva_rate, 0))), 0)
                    FROM quotation_items qi WHERE qi.quotation_id = q.id) AS total
            FROM quotations q WHERE q.issuer_id = ? ORDER BY q.created_at DESC
            LIMIT ? OFFSET ?
            """,
            (issuer_id, limit, offset),
        ).fetchall()
        conn.close()
        items = [{"id": r["id"], "folio": r.get("folio"), "customer_rfc": r["customer_rfc"],
                  "customer_legal_name": r["customer_legal_name"], "customer_email": r["customer_email"],
                  "status": r["status"], "public_token": r["public_token"], "valid_until": r["valid_until"],
                  "notes": r["notes"], "responded_at": r["responded_at"], "created_at": r["created_at"],
                  "updated_at": r["updated_at"], "total": float(r["total"] or 0)} for r in rows]
        return {"items": items, "total": total}
    except Exception as e:
        logger.warning("api_quotations_list: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="Error al cargar la lista de cotizaciones.")


@router.post("/quotations/create")
def api_quotations_create(request: Request, payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    csrf_service.verify_api_csrf(request)
    try:
        issuer_id = issuer["id"]
        customer_rfc = (payload.get("customer_rfc") or "").strip().upper()
        customer_legal_name = (payload.get("customer_legal_name") or "").strip()
        customer_email = (payload.get("customer_email") or "").strip() or None
        notes = (payload.get("notes") or "").strip() or None
        status = (payload.get("status") or "draft").strip().lower()
        if status not in QUOTATION_STATUSES:
            status = "draft"
        if not customer_legal_name:
            raise HTTPException(status_code=400, detail="Nombre del cliente es obligatorio")
        items = payload.get("items") or []
        if not items:
            raise HTTPException(status_code=400, detail="Agrega al menos un concepto a la cotización")
        public_token = secrets.token_urlsafe(32)
        iva_rate_quote = float(payload.get("iva_rate") or 0.16)
        currency = (payload.get("currency") or "MXN").strip() or "MXN"
        notes_default = (
            "Condiciones: Esta cotización tiene una vigencia de 30 días. "
            "Los precios están expresados en pesos mexicanos (MXN) e incluyen IVA según se indique. "
            "Para proceder, acepte esta cotización y nos pondremos en contacto."
        )
        notes = (payload.get("notes") or "").strip() or notes_default
        conn = db()
        year = datetime.now().strftime("%Y")
        prefix = f"Q-{year}-"
        next_num = conn.execute(
            """SELECT COALESCE(MAX(CAST(SUBSTR(folio, LENGTH(?) + 1) AS INTEGER)), 0) + 1
               FROM quotations WHERE issuer_id = ? AND (folio IS NOT NULL AND folio LIKE ?)""",
            (prefix, issuer_id, prefix + "%"),
        ).fetchone()[0]
        folio = f"{prefix}{next_num:04d}"
        sent_at = datetime.now().isoformat() if status == "sent" else None
        conn.execute(
            """INSERT INTO quotations (issuer_id, folio, customer_rfc, customer_legal_name, customer_email,
                status, public_token, notes, iva_rate, currency, sent_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))""",
            (issuer_id, folio, customer_rfc or "", customer_legal_name, customer_email, status, public_token, notes, iva_rate_quote, currency, sent_at),
        )
        qid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        for idx, it in enumerate(items):
            desc = (it.get("description") or "").strip()
            if not desc:
                continue
            qty = float(it.get("quantity") or 1)
            unit_price = float(it.get("unit_price") or 0)
            iva_rate = float(it.get("iva_rate") or 0.16)
            product_id = it.get("product_id")
            if product_id is not None:
                try:
                    product_id = int(product_id)
                except (TypeError, ValueError):
                    product_id = None
            conn.execute(
                """INSERT INTO quotation_items (quotation_id, description, quantity, unit_price, iva_rate, product_id, sort_order)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (qid, desc, qty, unit_price, iva_rate, product_id, idx),
            )
        # Snapshot para PDF consistente aunque se editen productos después
        items_list = []
        subtotal_sum = 0.0
        for it in items:
            desc = (it.get("description") or "").strip()
            if not desc:
                continue
            qty = float(it.get("quantity") or 1)
            unit_price = float(it.get("unit_price") or 0)
            iva_rate = float(it.get("iva_rate") or 0.16)
            line_sub = qty * unit_price
            iva_line = line_sub * iva_rate
            subtotal_sum += line_sub
            items_list.append({
                "description": desc,
                "quantity": qty,
                "unit_price": unit_price,
                "iva_rate": iva_rate,
                "subtotal": round(line_sub, 2),
                "total_line": round(line_sub + iva_line, 2),
            })
        iva_total = round(sum((x["subtotal"] * x["iva_rate"]) for x in items_list), 2)
        total = round(subtotal_sum + iva_total, 2)
        created_at = datetime.now().isoformat()
        snapshot = {
            "issuer_name": (issuer.get("razon_social") or issuer.get("rfc") or "").strip(),
            "issuer_rfc": (issuer.get("rfc") or "").strip(),
            "issuer_regimen": (issuer.get("regimen_fiscal") or "").strip(),
            "customer_rfc": customer_rfc or "",
            "customer_legal_name": customer_legal_name or "",
            "customer_email": (customer_email or "").strip() or None,
            "items": items_list,
            "subtotal": round(subtotal_sum, 2),
            "iva_total": iva_total,
            "total": total,
            "valid_until": None,
            "notes": notes or "",
            "folio": folio,
            "created_at": created_at,
        }
        if has_column(conn, "quotations", "metadata_json"):
            conn.execute(
                "UPDATE quotations SET metadata_json = ? WHERE id = ? AND issuer_id = ?",
                (json.dumps(snapshot), qid, issuer_id),
            )
        conn.commit()
        conn.close()
        return {"ok": True, "data": {"id": qid, "public_token": public_token}}
    except HTTPException:
        raise
    except Exception:
        logger.exception("api quotations create: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos crear la cotización. Intenta de nuevo.",
        )


@router.get("/quotations/{qid}")
def api_quotations_get(qid: int, issuer: dict = Depends(get_portal_issuer)):
    issuer_id = issuer["id"]
    conn = db()
    row = conn.execute(
        "SELECT id, customer_rfc, customer_legal_name, customer_email, status, public_token, valid_until, notes, responded_at, created_at, updated_at FROM quotations WHERE issuer_id = ? AND id = ?",
        (issuer_id, qid),
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    items = conn.execute(
        "SELECT id, description, quantity, unit_price, iva_rate, product_id, sort_order FROM quotation_items WHERE quotation_id = ? ORDER BY sort_order, id",
        (qid,),
    ).fetchall()
    conn.close()
    total = 0.0
    items_list = []
    for r in items:
        subtotal = float(r["quantity"] or 0) * float(r["unit_price"] or 0)
        iva = subtotal * float(r["iva_rate"] or 0)
        total += subtotal + iva
        items_list.append({"id": r["id"], "description": r["description"], "quantity": float(r["quantity"] or 0),
                           "unit_price": float(r["unit_price"] or 0), "iva_rate": float(r["iva_rate"] or 0.16),
                           "subtotal": subtotal, "total_line": subtotal + iva})
    d = dict(row)
    return {"id": d["id"], "customer_rfc": d["customer_rfc"], "customer_legal_name": d["customer_legal_name"],
            "customer_email": d["customer_email"], "status": d["status"], "public_token": d["public_token"],
            "valid_until": d["valid_until"], "notes": d["notes"], "responded_at": d["responded_at"],
            "created_at": d["created_at"], "updated_at": d["updated_at"], "items": items_list, "total": round(total, 2)}


@router.post("/quotations/update-status")
def api_quotations_update_status(request: Request, payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    csrf_service.verify_api_csrf(request)
    try:
        qid = payload.get("id")
        status = (payload.get("status") or "").strip().lower()
        if status not in QUOTATION_STATUSES:
            raise HTTPException(status_code=400, detail="Estatus inválido")
        if qid is None:
            raise HTTPException(status_code=400, detail="id requerido")
        try:
            qid = int(qid)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="id inválido")
        conn = db()
        cur = conn.execute(
            "UPDATE quotations SET status = ?, updated_at = datetime('now') WHERE issuer_id = ? AND id = ?",
            (status, issuer["id"], qid),
        )
        conn.commit()
        conn.close()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")
        return {"ok": True, "data": {"id": qid, "status": status}}
    except HTTPException:
        raise
    except Exception:
        logger.exception("api quotations status: qid=%s issuer_id=%s", qid, issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos obtener el estado. Intenta de nuevo.",
        )


@router.post("/quotations/respond")
def api_quotations_respond(request: Request, payload: dict = Body(...)):
    public_token = (payload.get("public_token") or "").strip()
    action = (payload.get("action") or "").strip().lower()
    if not public_token:
        raise HTTPException(status_code=400, detail="Link inválido")
    if action not in ("accept", "reject", "aceptar", "rechazar"):
        raise HTTPException(status_code=400, detail="Acción inválida")
    status = "accepted" if action in ("accept", "aceptar") else "rejected"
    reason = (payload.get("rejection_reason") or "").strip() or None
    client_ip = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent")
    now = datetime.now().isoformat()
    conn = db()
    row = conn.execute("SELECT id, status FROM quotations WHERE public_token = ?", (public_token,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Cotización no encontrada o link expirado")
    if dict(row)["status"] not in ("draft", "sent"):
        conn.close()
        raise HTTPException(status_code=400, detail="Esta cotización ya fue respondida")
    qid = dict(row)["id"]
    if status == "accepted":
        conn.execute(
            """UPDATE quotations SET status = ?, responded_at = datetime('now'), updated_at = datetime('now'),
               accepted_at = ?, decision_ip = ?, decision_user_agent = ? WHERE id = ?""",
            (status, now, client_ip, user_agent, qid),
        )
    else:
        conn.execute(
            """UPDATE quotations SET status = ?, responded_at = datetime('now'), updated_at = datetime('now'),
               rejected_at = ?, decision_ip = ?, decision_user_agent = ?, rejection_reason = ? WHERE id = ?""",
            (status, now, client_ip, user_agent, reason, qid),
        )
    conn.commit()
    conn.close()
    return {"ok": True, "data": {"status": status}}


# ----- Provider invoices -----
@router.get("/provider-invoices")
@router.get("/providers/invoices")
def api_provider_invoices(
    issuer: dict = Depends(get_portal_issuer),
    rfc: str = "",
    limit: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Máximo de registros"),
    offset: int = Query(0, ge=0, description="Registros a saltar"),
):
    try:
        iid = issuer["id"]
        rfc_norm = (rfc or "").strip().upper()
        if not rfc_norm:
            return {"items": [], "total": 0}
    except (ValueError, KeyError, TypeError):
        logger.exception("api cfdi issued list: issuer inválido")
        raise HTTPException(status_code=401, detail="Sesión inválida")
    try:
        base_where = """
            issuer_id = ? AND direction = 'received'
            AND UPPER(TRIM(COALESCE(rfc_emisor,''))) = ?
            AND (tipo_comprobante IS NULL OR UPPER(TRIM(COALESCE(tipo_comprobante,''))) != 'N')
            AND total IS NOT NULL AND total >= 0.01
        """
        total_row = db_rows(
            f"SELECT COUNT(*) AS c FROM sat_cfdi WHERE {base_where}",
            (iid, rfc_norm),
        )
        total = total_row[0]["c"] if total_row else 0
        rows = db_rows(
            f"""
            SELECT uuid, fecha_emision, total, moneda, status, xml_path, concepto
            FROM sat_cfdi
            WHERE {base_where}
            ORDER BY fecha_emision DESC LIMIT ? OFFSET ?
            """,
            (iid, rfc_norm, limit, offset),
        )
        items = [{"uuid": r.get("uuid"), "fecha_emision": r.get("fecha_emision"),
                  "concepto": (r.get("concepto") or "")[:80] + ("…" if len(r.get("concepto") or "") > 80 else ""),
                  "total": float(r.get("total") or 0), "moneda": r.get("moneda") or "MXN",
                  "status": r.get("status"), "has_pdf": bool(r.get("xml_path"))} for r in rows]
        return {"items": items, "total": total}
    except Exception:
        logger.exception("api cfdi issued list: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos cargar la lista. Intenta de nuevo.",
        )


@router.get("/provider-invoices/report")
def api_provider_invoices_report(
    issuer: dict = Depends(get_portal_issuer),
    rfc: str = Query(...),
    format: str = Query("pdf", alias="format"),
):
    if format not in ("pdf", "xlsx"):
        raise HTTPException(status_code=400, detail="format debe ser pdf o xlsx")
    try:
        rfc_norm = (rfc or "").strip().upper()
        if not rfc_norm:
            raise HTTPException(status_code=400, detail="RFC de proveedor requerido")
        rows = _provider_report_rows(issuer["id"], rfc_norm)
        provider_name = (rows[0].get("nombre_emisor") or "").strip() if rows else ""
        if not provider_name:
            provider_name = rfc_norm
        if format == "pdf":
            content = _build_provider_report_pdf(issuer, provider_name, rows)
            filename = f"facturas-recibidas-{rfc_norm[:8]}.pdf"
            media_type = "application/pdf"
        else:
            content = _build_provider_report_xlsx(issuer, provider_name, rows)
            filename = f"facturas-recibidas-{rfc_norm[:8]}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        from fastapi.responses import Response
        return Response(
            content=content,
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"', "Content-Length": str(len(content))},
        )
    except HTTPException:
        raise
    except (ValueError, KeyError, TypeError):
        logger.exception("api provider report: issuer inválido")
        raise HTTPException(status_code=401, detail="Sesión inválida")
    except Exception:
        logger.exception("api provider report: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos generar el reporte. Intenta de nuevo.",
        )


@router.get("/providers")
def api_providers(
    issuer: dict = Depends(get_portal_issuer),
    limit: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Máximo de registros"),
    offset: int = Query(0, ge=0, description="Registros a saltar"),
):
    try:
        iid = issuer["id"]
        saved = {}
        conn = db()
        try:
            if table_exists(conn, "supplier_profiles"):
                for r in db_rows("SELECT rfc, legal_name, email, alias FROM supplier_profiles WHERE issuer_id = ?", (iid,)):
                    rfc_norm = (r["rfc"] or "").strip().upper()
                    if rfc_norm:
                        saved[rfc_norm] = {"rfc": rfc_norm, "legal_name": r["legal_name"] or "", "email": r.get("email"),
                                           "alias": r.get("alias"), "facturas_count": 0, "total_recibido": 0.0, "source": "saved"}
        finally:
            conn.close()
        from_sat = db_rows(
            """
            SELECT UPPER(TRIM(rfc_emisor)) AS rfc, MAX(nombre_emisor) AS legal_name,
                   COUNT(*) AS facturas_count, COALESCE(SUM(total), 0) AS total_recibido
            FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'received' AND rfc_emisor IS NOT NULL AND TRIM(rfc_emisor) != ''
              AND total IS NOT NULL AND total >= 0.01
              AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
            GROUP BY UPPER(TRIM(rfc_emisor))
            """,
            (iid,),
        )
        for r in from_sat:
            rfc = (r["rfc"] or "").strip()
            if not rfc:
                continue
            if rfc in saved:
                saved[rfc]["facturas_count"] = r["facturas_count"]
                saved[rfc]["total_recibido"] = float(r["total_recibido"] or 0)
                saved[rfc]["source"] = "both"
            else:
                saved[rfc] = {"rfc": rfc, "legal_name": r["legal_name"] or "", "email": None, "alias": None,
                              "facturas_count": r["facturas_count"], "total_recibido": float(r["total_recibido"] or 0), "source": "sat"}
        out = sorted(saved.values(), key=lambda x: (-x["facturas_count"], -x["total_recibido"], (x.get("alias") or x["rfc"]).lower()))
        total = len(out)
        items = out[offset : offset + limit]
        return {"items": items, "total": total}
    except Exception as e:
        logger.warning("api_providers: %s", e, exc_info=True)
        return {"items": [], "total": 0}


@router.post("/providers/create")
def api_providers_create(request: Request, payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
    csrf_service.verify_api_csrf(request)
    try:
        rfc = (payload.get("rfc") or "").strip().upper()
        legal_name = (payload.get("legal_name") or "").strip()
        if not rfc or not legal_name:
            raise HTTPException(status_code=400, detail="RFC y razón social son obligatorios")
        zip_val = (payload.get("zip") or "").strip() or None
        tax_val = (payload.get("tax_system") or "").strip() or None
        email = (payload.get("email") or "").strip() or None
        alias = (payload.get("alias") or "").strip() or None
        conn = db()
        conn.execute(
            """
            INSERT INTO supplier_profiles (issuer_id, rfc, legal_name, zip, tax_system, email, alias, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(issuer_id, rfc) DO UPDATE SET legal_name = excluded.legal_name, zip = excluded.zip,
                tax_system = excluded.tax_system, email = excluded.email, alias = excluded.alias, updated_at = CURRENT_TIMESTAMP
            """,
            (issuer["id"], rfc, legal_name, zip_val, tax_val, email, alias),
        )
        conn.commit()
        conn.close()
        return {"ok": True, "data": {"rfc": rfc}}
    except HTTPException:
        raise
    except Exception:
        logger.exception("api providers create: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos guardar el proveedor. Intenta de nuevo.",
        )


@router.get("/invoices/issued")
def api_invoices_issued(
    issuer: dict = Depends(get_portal_issuer),
    ym: str = Query(None, description="Year-month (YYYY-MM)"),
    search: str = Query("", description="Search UUID/RFC/nombre/concepto"),
    status: str = Query("", description="Status filter: vigente, cancelada, all"),
    min_amount: float = Query(None, description="Minimum amount"),
    max_amount: float = Query(None, description="Maximum amount"),
    metodo_pago: str = Query("", description="PUE or PPD"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Items per page"),
):
    """API endpoint para facturas emitidas con filtros y paginación."""
    fixture = _load_fixture("issued")
    if fixture is not None:
        return fixture
    issuer_id = issuer["id"]
    if not ym:
        from datetime import datetime
        ym = datetime.now().strftime("%Y-%m")
    
    # Build WHERE clause
    where_parts = [
        "issuer_id = ?",
        "direction = 'issued'",
        "fecha_emision IS NOT NULL",
        "substr(fecha_emision,1,7) = ?",
        "(total IS NULL OR total >= 0.01)",
        "xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''",
    ]
    params = [issuer_id, ym]
    
    # Deduplicate subquery (same as portal route)
    dedup_subquery = """
        id IN (
            SELECT id FROM (
                SELECT id, ROW_NUMBER() OVER (
                    PARTITION BY issuer_id, direction, LOWER(TRIM(uuid))
                    ORDER BY (CASE WHEN COALESCE(total,0) >= 0.01 THEN 0 ELSE 1 END), id
                ) AS rn
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL 
                  AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
            ) WHERE rn = 1
        )
    """
    where_parts.append(dedup_subquery)
    params.extend([issuer_id, ym])
    
    # Search filter
    if search:
        search_term = f"%{search.upper()}%"
        where_parts.append(
            "(UPPER(COALESCE(uuid,'')) LIKE ? OR UPPER(COALESCE(rfc_receptor,'')) LIKE ? "
            "OR UPPER(COALESCE(nombre_receptor,'')) LIKE ? OR UPPER(COALESCE(concepto,'')) LIKE ?)"
        )
        params.extend([search_term, search_term, search_term, search_term])
    
    # Status filter
    if status and status.lower() in ("vigente", "cancelada"):
        if status.lower() == "vigente":
            where_parts.append("(status = '1' OR UPPER(TRIM(COALESCE(status,''))) = 'V' OR UPPER(TRIM(COALESCE(status,''))) = 'VIGENTE')")
        elif status.lower() == "cancelada":
            where_parts.append("(status = '0' OR UPPER(TRIM(COALESCE(status,''))) = 'C' OR UPPER(TRIM(COALESCE(status,''))) LIKE 'CANCEL%')")
    
    # Amount filters
    if min_amount is not None:
        where_parts.append("COALESCE(total, 0) >= ?")
        params.append(min_amount)
    if max_amount is not None:
        where_parts.append("COALESCE(total, 0) <= ?")
        params.append(max_amount)
    
    # Metodo pago filter
    if metodo_pago and metodo_pago.upper() in ("PUE", "PPD"):
        where_parts.append("UPPER(TRIM(COALESCE(metodo_pago,''))) = ?")
        params.append(metodo_pago.upper())
    
    where_clause = " AND ".join(where_parts)
    
    # Count total (row_factory devuelve dict; la clave es el nombre de columna, no el índice)
    try:
        conn = db()
        count_row = conn.execute(
            f"SELECT COUNT(*) AS c FROM sat_cfdi WHERE {where_clause}",
            tuple(params)
        ).fetchone()
        total_count = int(count_row.get("c", 0)) if count_row else 0
        conn.close()
    except Exception as e:
        logger.exception("Error counting invoices")
        raise HTTPException(status_code=500, detail="Error al contar facturas")
    
    # Fetch paginated results (solo columnas usadas en el listado para payload pequeño)
    try:
        offset = (page - 1) * per_page
        rows = db_rows(
            f"""
            SELECT uuid, fecha_emision, rfc_receptor, nombre_receptor, concepto, total, moneda,
                   COALESCE(impuestos, 0) AS impuestos, COALESCE(retenciones, 0) AS retenciones,
                   metodo_pago, status, xml_path
            FROM sat_cfdi
            WHERE {where_clause}
            ORDER BY fecha_emision DESC
            LIMIT ? OFFSET ?
            """,
            tuple(params) + (per_page, offset)
        )
    except Exception as e:
        logger.exception("Error fetching invoices")
        raise HTTPException(status_code=500, detail="Error al obtener facturas")

    return {
        "data": rows,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total_count,
            "pages": (total_count + per_page - 1) // per_page if total_count > 0 else 0,
        },
        "filters": {
            "ym": ym,
            "search": search,
            "status": status,
            "min_amount": min_amount,
            "max_amount": max_amount,
            "metodo_pago": metodo_pago,
        }
    }


@router.get("/invoices/received")
def api_invoices_received(
    issuer: dict = Depends(get_portal_issuer),
    ym: str = Query(None, description="Year-month (YYYY-MM)"),
    search: str = Query("", description="Search UUID/RFC/nombre/concepto"),
    status: str = Query("", description="Status filter: vigente, cancelada, all"),
    min_amount: float = Query(None, description="Minimum amount"),
    max_amount: float = Query(None, description="Maximum amount"),
    metodo_pago: str = Query("", description="PUE or PPD"),
    match_filter: str = Query("", description="Conciliación: none|probable"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Items per page"),
):
    """API endpoint para facturas recibidas con filtros y paginación."""
    fixture = _load_fixture("received")
    if fixture is not None:
        return fixture
    issuer_id = issuer["id"]
    if not ym:
        from datetime import datetime
        ym = datetime.now().strftime("%Y-%m")
    
    # Build WHERE clause
    where_parts = [
        "issuer_id = ?",
        "direction = 'received'",
        "fecha_emision IS NOT NULL",
        "substr(fecha_emision,1,7) = ?",
        "total IS NOT NULL AND total >= 0.01",
        "(tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')",
        "xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''",
    ]
    params = [issuer_id, ym]
    
    # Deduplicate subquery
    dedup_subquery = """
        id IN (
            SELECT id FROM (
                SELECT id, ROW_NUMBER() OVER (
                    PARTITION BY issuer_id, direction, LOWER(TRIM(uuid))
                    ORDER BY id
                ) AS rn
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL 
                  AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01 
                  AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
            ) WHERE rn = 1
        )
    """
    where_parts.append(dedup_subquery)
    params.extend([issuer_id, ym])
    
    # Search filter
    if search:
        search_term = f"%{search.upper()}%"
        where_parts.append(
            "(UPPER(COALESCE(uuid,'')) LIKE ? OR UPPER(COALESCE(rfc_emisor,'')) LIKE ? "
            "OR UPPER(COALESCE(nombre_emisor,'')) LIKE ? OR UPPER(COALESCE(concepto,'')) LIKE ?)"
        )
        params.extend([search_term, search_term, search_term, search_term])
    
    # Status filter
    if status and status.lower() in ("vigente", "cancelada"):
        if status.lower() == "vigente":
            where_parts.append("(status = '1' OR UPPER(TRIM(COALESCE(status,''))) = 'V' OR UPPER(TRIM(COALESCE(status,''))) = 'VIGENTE')")
        elif status.lower() == "cancelada":
            where_parts.append("(status = '0' OR UPPER(TRIM(COALESCE(status,''))) = 'C' OR UPPER(TRIM(COALESCE(status,''))) LIKE 'CANCEL%')")
    
    # Amount filters
    if min_amount is not None:
        where_parts.append("COALESCE(total, 0) >= ?")
        params.append(min_amount)
    if max_amount is not None:
        where_parts.append("COALESCE(total, 0) <= ?")
        params.append(max_amount)
    
    # Metodo pago filter
    if metodo_pago and metodo_pago.upper() in ("PUE", "PPD"):
        where_parts.append("UPPER(TRIM(COALESCE(metodo_pago,''))) = ?")
        params.append(metodo_pago.upper())

    # Conciliación (mismo modelo que bank/movements)
    mf = (match_filter or "").strip().lower()
    if mf in ("none", "probable"):
        # Solo si existe tabla; si no existe, no filtrar (degrada a 'todos')
        try:
            conn0 = db()
            has_matches = ("bank_invoice_matches" in {r[0] for r in conn0.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()})
            conn0.close()
        except Exception:
            has_matches = False
        if has_matches:
            if mf == "probable":
                where_parts.append(
                    """EXISTS (
                         SELECT 1 FROM bank_invoice_matches bim
                         WHERE bim.issuer_id = sat_cfdi.issuer_id
                           AND bim.cfdi_id = sat_cfdi.id
                           AND bim.status IN ('suggested','confirmed')
                           AND COALESCE(bim.score,0) >= 80
                       )"""
                )
            elif mf == "none":
                where_parts.append(
                    """NOT EXISTS (
                         SELECT 1 FROM bank_invoice_matches bim
                         WHERE bim.issuer_id = sat_cfdi.issuer_id
                           AND bim.cfdi_id = sat_cfdi.id
                           AND bim.status IN ('suggested','confirmed')
                           AND COALESCE(bim.score,0) >= 50
                       )"""
                )
    
    where_clause = " AND ".join(where_parts)
    
    # Count total (row_factory devuelve dict)
    try:
        conn = db()
        count_row = conn.execute(
            f"SELECT COUNT(*) AS c FROM sat_cfdi WHERE {where_clause}",
            tuple(params)
        ).fetchone()
        total_count = int(count_row.get("c", 0)) if count_row else 0
        conn.close()
    except Exception as e:
        logger.exception("Error counting invoices")
        raise HTTPException(status_code=500, detail="Error al contar facturas")
    
    # Fetch paginated results
    try:
        offset = (page - 1) * per_page
        rows = db_rows(
            f"""
            SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, concepto, total, moneda,
                   COALESCE(impuestos, 0) AS impuestos, COALESCE(retenciones, 0) AS retenciones,
                   metodo_pago, status, xml_path,
                   (
                     SELECT bm.id
                     FROM bank_invoice_matches bim
                     JOIN bank_movements bm ON bm.id = bim.bank_movement_id AND bm.issuer_id = bim.issuer_id
                     WHERE bim.issuer_id = sat_cfdi.issuer_id
                       AND bim.cfdi_id = sat_cfdi.id
                       AND bim.status IN ('suggested','confirmed')
                     ORDER BY bim.score DESC, bim.id DESC
                     LIMIT 1
                   ) AS probable_movement_id,
                   (
                     SELECT bm.fecha
                     FROM bank_invoice_matches bim
                     JOIN bank_movements bm ON bm.id = bim.bank_movement_id AND bm.issuer_id = bim.issuer_id
                     WHERE bim.issuer_id = sat_cfdi.issuer_id
                       AND bim.cfdi_id = sat_cfdi.id
                       AND bim.status IN ('suggested','confirmed')
                     ORDER BY bim.score DESC, bim.id DESC
                     LIMIT 1
                   ) AS probable_movement_fecha,
                   (
                     SELECT COALESCE(bm.deposito, 0) - COALESCE(bm.retiro, 0)
                     FROM bank_invoice_matches bim
                     JOIN bank_movements bm ON bm.id = bim.bank_movement_id AND bm.issuer_id = bim.issuer_id
                     WHERE bim.issuer_id = sat_cfdi.issuer_id
                       AND bim.cfdi_id = sat_cfdi.id
                       AND bim.status IN ('suggested','confirmed')
                     ORDER BY bim.score DESC, bim.id DESC
                     LIMIT 1
                   ) AS probable_movement_amount,
                   (
                     SELECT bm.descripcion
                     FROM bank_invoice_matches bim
                     JOIN bank_movements bm ON bm.id = bim.bank_movement_id AND bm.issuer_id = bim.issuer_id
                     WHERE bim.issuer_id = sat_cfdi.issuer_id
                       AND bim.cfdi_id = sat_cfdi.id
                       AND bim.status IN ('suggested','confirmed')
                     ORDER BY bim.score DESC, bim.id DESC
                     LIMIT 1
                   ) AS probable_movement_desc,
                   (
                     SELECT bim.score
                     FROM bank_invoice_matches bim
                     WHERE bim.issuer_id = sat_cfdi.issuer_id
                       AND bim.cfdi_id = sat_cfdi.id
                       AND bim.status IN ('suggested','confirmed')
                     ORDER BY bim.score DESC, bim.id DESC
                     LIMIT 1
                   ) AS probable_movement_score,
                   (
                     SELECT bim.status
                     FROM bank_invoice_matches bim
                     WHERE bim.issuer_id = sat_cfdi.issuer_id
                       AND bim.cfdi_id = sat_cfdi.id
                       AND bim.status IN ('suggested','confirmed')
                     ORDER BY bim.score DESC, bim.id DESC
                     LIMIT 1
                   ) AS probable_movement_status
            FROM sat_cfdi
            WHERE {where_clause}
            ORDER BY fecha_emision DESC
            LIMIT ? OFFSET ?
            """,
            tuple(params) + (per_page, offset)
        )
    except Exception as e:
        logger.exception("Error fetching invoices")
        raise HTTPException(status_code=500, detail="Error al obtener facturas")
    
    return {
        "data": rows,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total_count,
            "pages": (total_count + per_page - 1) // per_page if total_count > 0 else 0,
        },
        "filters": {
            "ym": ym,
            "search": search,
            "status": status,
            "min_amount": min_amount,
            "max_amount": max_amount,
            "metodo_pago": metodo_pago,
            "match_filter": match_filter,
        }
    }


@router.get("/invoices/pending")
def api_pending_invoices(
    issuer: dict = Depends(get_portal_issuer),
    limit: int = Query(DEFAULT_LIST_LIMIT, ge=1, le=MAX_LIST_LIMIT, description="Máximo de registros"),
    offset: int = Query(0, ge=0, description="Registros a saltar"),
):
    try:
        conn = db()
        cols = {r[1] for r in conn.execute("PRAGMA table_info(invoices)").fetchall()}
        where = ["issuer_id = ?", "uuid IS NOT NULL", "payment_method = 'PPD'"]
        params = [issuer["id"]]
        if "status" in cols:
            where.append("COALESCE(status,'') != 'canceled'")
        if "cancelled" in cols:
            where.append("COALESCE(cancelled,0) = 0")
        where_sql = " AND ".join(where)
        count_row = conn.execute(
            f"SELECT COUNT(*) AS c FROM invoices WHERE {where_sql}",
            tuple(params),
        ).fetchone()
        total = int(count_row.get("c", 0)) if count_row else 0
        rows = conn.execute(
            f"""SELECT id, uuid, total, customer_legal_name, customer_rfc, issue_date, created_at
                FROM invoices WHERE {where_sql}
                ORDER BY COALESCE(issue_date, created_at) DESC LIMIT ? OFFSET ?""",
            tuple(params) + (limit, offset),
        ).fetchall()
        conn.close()
        items = [{"id": r["id"], "uuid": r["uuid"], "total": r["total"], "customer_legal_name": r["customer_legal_name"],
                  "customer_rfc": r["customer_rfc"], "date": r["issue_date"] or r["created_at"]} for r in rows]
        return {"items": items, "total": total}
    except Exception:
        logger.exception("api invoices list: issuer_id=%s", issuer.get("id"))
        raise HTTPException(
            status_code=500,
            detail="No pudimos cargar la lista. Intenta de nuevo.",
        )


# ----- SAT catalogs -----
# Los catálogos se leen de catalogs/catalogs.db (SAT). Si no existe el archivo (p. ej. no se
# añadió el DB de un repo comunitario), se usan listas estáticas para que el formulario funcione.

def _catalog_list(d):
    """Convierte dict {clave: etiqueta} a lista [{key, label}] para los selects."""
    return [{"key": str(k), "label": str(v)} for k, v in sorted(d.items())]


# Fallbacks ampliados cuando no hay catalogs.db (lista completa para moneda/unidad y búsqueda ProdServ)
MONEDA_FALLBACK = {
    "MXN": "Peso Mexicano",
    "USD": "Dólar Americano",
    "EUR": "Euro",
    "MXV": "México Unidad de Inversión (UDI)",
    "GBP": "Libra Esterlina",
    "CAD": "Dólar Canadiense",
    "CHF": "Franco Suizo",
    "JPY": "Yen Japonés",
    "CNY": "Yuan Chino",
    "AUD": "Dólar Australiano",
    "BRL": "Real Brasileño",
    "COP": "Peso Colombiano",
    "ARS": "Peso Argentino",
    "CLP": "Peso Chileno",
    "PEN": "Sol Peruano",
    "XXX": "Los códigos asignados para transacciones en que intervenga ninguna moneda",
}
UNIDAD_FALLBACK = {
    "E48": "Unidad de servicio",
    "EA": "Cada uno",
    "H87": "Pieza",
    "ACT": "Actividad",
    "LTR": "Litro",
    "MTR": "Metro",
    "KGM": "Kilogramo",
    "GRM": "Gramo",
    "MTK": "Metro cuadrado",
    "MTQ": "Metro cúbico",
    "DAY": "Día",
    "HUR": "Hora",
    "MIN": "Minuto",
    "C62": "Unidad",
    "XBX": "Caja",
    "PA": "Paquete",
    "PK": "Paquete",
    "SET": "Conjunto",
    "PR": "Par",
    "NIU": "Número de artículos",
    "DZN": "Docena",
    "XPK": "Paquete",
    "XRO": "Rollo",
    "XCT": "Ciento",
    "XPL": "Pliego",
    "XNA": "Artículo",
    "XNE": "Kilo neto",
    "XBR": "Barra",
    "XBO": "Botella",
    "XBE": "Lata",
    "XBG": "Bolsa",
}
# Lista (clave, descripción) para búsqueda ProdServ por palabra; descripción en minúsculas para matchear.
PRODSERV_FALLBACK = [
    ("81112100", "Servicios de asesoría en negocios y comercio"),
    ("81112101", "Asesoría en negocios"),
    ("84111500", "Servicios contables (honorarios contables)"),
    ("84111501", "Servicios de contabilidad"),
    ("84111502", "Servicios de auditoría"),
    ("84111503", "Servicios de teneduría de libros"),
    ("84111600", "Servicios de impuestos"),
    ("84111800", "Servicios de consultoría en gestión"),
    ("53111500", "Servicios de alquiler o arrendamiento de equipo"),
    ("53111501", "Renta de equipo"),
    ("53111502", "Arrendamiento de maquinaria"),
    ("53131600", "Servicios de mantenimiento de equipo"),
    ("80101600", "Servicios de consultoría en negocios"),
    ("80101601", "Consultoría administrativa"),
    ("80101602", "Consultoría en gestión"),
    ("80101800", "Servicios de consultoría en tecnología"),
    ("80101801", "Consultoría en sistemas"),
    ("81101500", "Servicios de diseño"),
    ("81101501", "Diseño gráfico"),
    ("81101502", "Diseño de software"),
    ("81102200", "Servicios de desarrollo de software"),
    ("81102201", "Desarrollo de aplicaciones"),
    ("81111800", "Servicios de soporte técnico"),
    ("81112200", "Servicios de consultoría en ingeniería"),
    ("90101500", "Servicios de limpieza"),
    ("90101600", "Servicios de limpieza de edificios"),
    ("92111500", "Servicios de capacitación"),
    ("92111501", "Capacitación empresarial"),
    ("92111502", "Cursos de capacitación"),
    ("93101600", "Servicios de publicidad"),
    ("93101601", "Publicidad y promoción"),
    ("84111801", "Servicios de consultoría en recursos humanos"),
    ("84111802", "Outsourcing o subcontratación de servicios"),
    ("81112102", "Asesoría en comercio"),
    ("43211500", "Equipo de cómputo"),
    ("43211501", "Computadoras personales"),
    ("43222600", "Software"),
    ("43222601", "Software de aplicación"),
    ("44111500", "Mobiliario de oficina"),
    ("44111501", "Escritorios y mesas"),
    ("50192100", "Servicios de mensajería"),
    ("50192101", "Mensajería y paquetería"),
]


@router.get("/catalogs/forma_pago")
def api_forma_pago():
    try:
        return ok(list_catalog("cfdi_40_formas_pago"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid catalog table")
    except Exception:
        logger.warning("api catalogs forma_pago: usando fallback (catalogs.db no disponible)")
        return ok(_catalog_list(FORMA_PAGO))


@router.get("/catalogs/metodo_pago")
def api_metodo_pago():
    try:
        return ok(list_catalog("cfdi_40_metodos_pago"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid catalog table")
    except Exception:
        return ok([{"key": "PUE", "label": "Pago en una sola exhibición"}, {"key": "PPD", "label": "Pago en parcialidades o diferido"}])


@router.get("/catalogs/uso_cfdi")
def api_uso_cfdi():
    try:
        return ok(list_catalog("cfdi_40_usos_cfdi"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid catalog table")
    except Exception:
        logger.warning("api catalogs uso_cfdi: usando fallback (catalogs.db no disponible)")
        return ok(_catalog_list(USO_CFDI))


@router.get("/catalogs/regimen_fiscal")
def api_regimen_fiscal():
    try:
        return ok(list_catalog("cfdi_40_regimenes_fiscales"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid catalog table")
    except Exception:
        logger.warning("api catalogs regimen_fiscal: usando fallback (catalogs.db no disponible)")
        reg = dict(REGIMEN_FISCAL)
        reg["616"] = "Sin obligaciones fiscales"
        return ok(_catalog_list(reg))


@router.get("/catalogs/moneda")
def api_moneda():
    try:
        return ok(list_catalog("cfdi_40_monedas"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid catalog table")
    except Exception:
        logger.warning("api catalogs moneda: usando fallback (catalogs.db no disponible)")
        return ok(_catalog_list(MONEDA_FALLBACK))


@router.get("/catalogs/prodserv")
def api_prodserv(q: str = Query(..., min_length=1), limit: int = Query(20, ge=1, le=50)):
    try:
        return ok(search_catalog("cfdi_40_productos_servicios", q=q, limit=limit))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid catalog table")
    except Exception:
        logger.warning("api catalogs prodserv: usando fallback estático (catalogs.db no disponible)")
        q_lower = q.strip().lower()
        out = []
        for clave, desc in PRODSERV_FALLBACK:
            if q_lower in clave or q_lower in desc.lower():
                out.append({"key": clave, "label": desc})
                if len(out) >= limit:
                    break
        return ok(out)


@router.get("/catalogs/unidad")
def api_unidad(q: str = Query(..., min_length=1), limit: int = Query(20, ge=1, le=50)):
    try:
        return ok(search_catalog("cfdi_40_claves_unidades", q=q, limit=limit))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid catalog table")
    except Exception:
        logger.warning("api catalogs unidad: usando fallback (catalogs.db no disponible)")
        q_lower = q.strip().lower()
        items = [
            {"key": k, "label": v}
            for k, v in UNIDAD_FALLBACK.items()
            if q_lower in v.lower() or q_lower in k.lower()
        ]
        return ok(items[: int(limit)])


# ---------- Month Close API ----------

@router.get("/month-close")
def api_month_close_get(
    request: Request,
    issuer: dict = Depends(get_portal_issuer),
    ym: str = Query(..., min_length=7, max_length=7),
):
    from services import month_close as mc
    issuer_id = int(issuer.get("id") or 0)
    if issuer_id <= 0:
        raise HTTPException(status_code=401, detail="Sesión inválida")
    try:
        data = mc.get_full_month_close(issuer_id, ym)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return ok(data)


@router.post("/month-close")
def api_month_close_post(
    request: Request,
    issuer: dict = Depends(get_portal_issuer),
    body: dict = Body(...),
):
    from services import month_close as mc
    issuer_id = int(issuer.get("id") or 0)
    if issuer_id <= 0:
        raise HTTPException(status_code=401, detail="Sesión inválida")
    ym = (body.get("ym") or "").strip()
    status = body.get("status")
    checklist = body.get("checklist")
    try:
        data = mc.save_month_close(issuer_id, ym, status=status, checklist=checklist)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    log_action(request, "month_close_save", issuer_id=issuer_id, ym=ym)
    return ok(data)


# ---------- Matching Preview API ----------

@router.get("/matching/preview")
def api_matching_preview(
    request: Request,
    issuer: dict = Depends(get_portal_issuer),
    ym: str = Query(..., min_length=7, max_length=7),
):
    from services.matching import preview_month
    issuer_id = int(issuer.get("id") or 0)
    if issuer_id <= 0:
        raise HTTPException(status_code=401, detail="Sesión inválida")
    try:
        result = preview_month(issuer_id, ym)
    except Exception as e:
        logger.warning("matching preview error: %s", e)
        return ok({"ok": False, "message": str(e)})
    return ok(result)


# ---------- Notifications API ----------

@router.get("/notifications")
def api_notifications_list(
    request: Request,
    issuer: dict = Depends(get_portal_issuer),
    unread_only: bool = Query(True),
    limit: int = Query(10, ge=1, le=50),
):
    from services import notifications as notif_service
    issuer_id = int(issuer.get("id") or 0)
    if issuer_id <= 0:
        raise HTTPException(status_code=401, detail="Sesión inválida")
    items = notif_service.list_notifications(issuer_id, unread_only=unread_only, limit=limit)
    return ok_list(items, len(items))


@router.post("/notifications/{notification_id}/read")
def api_notification_mark_read(
    request: Request,
    notification_id: int,
    issuer: dict = Depends(get_portal_issuer),
):
    from services import notifications as notif_service
    issuer_id = int(issuer.get("id") or 0)
    if issuer_id <= 0:
        raise HTTPException(status_code=401, detail="Sesión inválida")
    success = notif_service.mark_read(issuer_id, notification_id)
    return ok({"marked": success})
