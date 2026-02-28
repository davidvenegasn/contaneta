# Portal HTML routes and helpers
import hashlib
import io
import json
import logging
import os
import re
import secrets
import stat
from services.errors import ExternalServiceError
from services.subprocess_utils import run_php
from datetime import datetime, date, timezone
from typing import Optional, Any

from fastapi import APIRouter, Request, Depends, Query, HTTPException, File, UploadFile, Form, Body
from fastapi.responses import HTMLResponse, Response, RedirectResponse, JSONResponse, FileResponse

from config import BASE_DIR, REGIMEN_LABEL_TO_CODE, COOKIE_DEMO_VIEW, DB_PATH, DEV_MODE, PORTAL_SHELL_V2
from database import db, db_rows, has_column, table_exists
from routers.deps import get_portal_issuer
from services import quotations as quotations_service, rate_limit as rate_limit_service, session as session_service, audit, subscription as subscription_service, csrf as csrf_service
from services import file_access_log
from services.action_log import log_action
from services.redirects import safe_next_url
from services.portal_errors import portal_error_type
from services.pdf_to_excel import convert_pdf_to_xlsx, get_storage_root, safe_join, ensure_parent_dir
from services.bank_parse_preview import parse_bank_pdf_to_movements_preview, reclassify_movements
from services.bank_preview_pipeline import parse_bank_statement_preview
from services.bank_preview_models import compute_dedupe_fingerprint
from services.catalog_from_cfdi import backfill_catalog_from_existing_cfdi
from services.bank_accounts import list_active_accounts as bank_list_accounts, list_active_accounts_raw as bank_list_accounts_raw, list_all_accounts as bank_list_all_accounts, get_account as bank_get_account, create_account as bank_create_account, update_account as bank_update_account, delete_account as bank_delete_account
from services.bank_own_accounts import detect_own_account_transfer
from services.bank_statement_ingest import ingest_bank_statement, extract_statement_metadata, validate_statement_ownership, commit_preview_to_db
from services.bank_cfdi_matching import find_cfdi_candidates, save_suggested_matches, confirm_match as match_confirm, reject_match as match_reject

logger = logging.getLogger(__name__)

# Paginación: evitar OFFSET enorme (degrada SQLite)
MAX_LIST_OFFSET = 50_000


def _db_row_to_dict(row: Any) -> dict:
    """Convierte sqlite3.Row (o cualquier fila) a dict para que .get() funcione."""
    if row is None:
        return {}
    if isinstance(row, dict):
        return row
    # sqlite3.Row no es un Mapping estándar; dict(row) puede dar claves numéricas
    if hasattr(row, "keys"):
        try:
            return dict(zip(row.keys(), row))
        except Exception:
            pass
    try:
        return dict(row)
    except Exception:
        return {}


def _strip_date_from_description(desc: Optional[str]) -> str:
    """Quita prefijo de fecha de la descripción para mostrar como concepto (ej. 31-ENE-26 o 2026-01-31)."""
    if not desc or not isinstance(desc, str):
        return desc or ""
    s = desc.strip()
    # DD-MMM-YY (ej. 31-ENE-26)
    m = re.match(r"^\d{1,2}-[A-Za-z]{3}-\d{2,4}\s*", s)
    if m:
        return s[m.end() :].strip() or s
    # YYYY-MM-DD
    m = re.match(r"^\d{4}-\d{2}-\d{2}\s*", s)
    if m:
        return s[m.end() :].strip() or s
    # DD/MM/YYYY o DD/MM/YY
    m = re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}\s*", s)
    if m:
        return s[m.end() :].strip() or s
    return s


# ----------------------------
# Helpers
# ----------------------------

MESES_ES = (
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
)


def ym_now():
    return datetime.now().strftime("%Y-%m")


def _ensure_sat_credentials_validation_columns(conn) -> None:
    """Añade columnas validation_* a sat_credentials si no existen."""
    for col, col_type in [("validation_at", "TEXT"), ("validation_ok", "INTEGER"), ("validation_message", "TEXT")]:
        if not has_column(conn, "sat_credentials", col):
            conn.execute(f"ALTER TABLE sat_credentials ADD COLUMN {col} {col_type};")


def _credentials_dir(issuer_id: int) -> str:
    """Ruta al directorio storage/credentials/{issuer_id}/ (creado si no existe)."""
    path = os.path.join(BASE_DIR, "storage", "credentials", str(issuer_id))
    os.makedirs(path, mode=0o700, exist_ok=True)
    return path


def _run_fiel_validation(issuer_id: int) -> tuple[bool, str]:
    """Ejecuta check_fiel.php para issuer_id, actualiza sat_credentials y devuelve (ok, message)."""
    php_script = os.path.join(BASE_DIR, "sat_sync", "check_fiel.php")
    if not os.path.isfile(php_script):
        return False, "No se encontró el script de validación."
    env = os.environ.copy()
    env["APP_DB_PATH"] = str(DB_PATH)
    try:
        # Cifrado at-rest: pasar credenciales desencriptadas vía env (solo al proceso PHP)
        from services.sat_credentials_secure import decrypted_fiel_env

        with decrypted_fiel_env(int(issuer_id)) as fiel_env:
            env.update(fiel_env)
        stdout, stderr = run_php(
            [php_script, str(issuer_id)],
            timeout=30,
            cwd=BASE_DIR,
            env=env,
        )
        ok = True
        message = (stdout or "").strip() or "FIEL validada correctamente."
    except ExternalServiceError as e:
        ok = False
        message = (e.internal_message or e.public_message or "Error al validar la FIEL.").strip()
    conn = db()
    try:
        _ensure_sat_credentials_validation_columns(conn)
        if has_column(conn, "sat_credentials", "validation_at"):
            conn.execute(
                """
                UPDATE sat_credentials SET validation_at = datetime('now'), validation_ok = ?, validation_message = ?
                WHERE issuer_id = ?
                """,
                (1 if ok else 0, message[:500], issuer_id),
            )
            conn.commit()
    finally:
        conn.close()
    return ok, message


def ym_to_label(ym: str) -> str:
    """Convert 2026-01 to 'Enero 2026'."""
    try:
        y, m = ym.split("-")
        return f"{MESES_ES[int(m) - 1]} {y}"
    except (ValueError, IndexError):
        return ym


def shift_ym(ym: str, delta_months: int) -> str:
    y, m = ym.split("-")
    y, m = int(y), int(m)
    m += delta_months
    while m <= 0:
        m += 12
        y -= 1
    while m >= 13:
        m -= 12
        y += 1
    return f"{y:04d}-{m:02d}"


def _get_month_totals(issuer_id: int, ym: str, direction: str) -> dict:
    """Totales del mes para emitidas o recibidas: base (subtotal), IVA y retenciones."""
    conn = db()
    try:
        base_where = (
            "issuer_id = ? AND direction = ? AND fecha_emision IS NOT NULL AND substr(fecha_emision,1,7) = ?"
        )
        if direction == "issued":
            base_where += " AND (total IS NULL OR total >= 0.01)"
        else:
            base_where += " AND total IS NOT NULL AND total >= 0.01 AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')"
        params = (issuer_id, direction, ym)
        has_retenciones = has_column(conn, "sat_cfdi", "retenciones")
        if has_retenciones:
            row = conn.execute(
                f"""
                SELECT
                    COALESCE(SUM(COALESCE(subtotal, total)), 0) AS total_base,
                    COALESCE(SUM(COALESCE(impuestos, 0)), 0) AS total_iva,
                    COALESCE(SUM(COALESCE(retenciones, 0)), 0) AS total_retenciones
                FROM sat_cfdi
                WHERE {base_where}
                """,
                params,
            ).fetchone()
        else:
            row = conn.execute(
                f"""
                SELECT
                    COALESCE(SUM(COALESCE(subtotal, total)), 0) AS total_base,
                    COALESCE(SUM(COALESCE(impuestos, 0)), 0) AS total_iva
                FROM sat_cfdi
                WHERE {base_where}
                """,
                params,
            ).fetchone()
        if not row:
            total_base = total_iva = total_retenciones = 0.0
        elif isinstance(row, dict):
            total_base = float(row.get("total_base") or 0)
            total_iva = float(row.get("total_iva") or 0)
            total_retenciones = float(row.get("total_retenciones") or 0) if has_retenciones else 0.0
        else:
            total_base = float(row[0] or 0)
            total_iva = float(row[1] or 0)
            total_retenciones = float(row[2] or 0) if (has_retenciones and len(row) >= 3) else 0.0
        return {
            "total_base": total_base,
            "total_iva": total_iva,
            "total_retenciones": total_retenciones,
            "total_iva_neto": max(0.0, total_iva - total_retenciones) if direction == "issued" else total_iva,
        }
    finally:
        conn.close()


def _safe_abs_path(path_like: str) -> str:
    """Resolve a stored path to an absolute path under BASE_DIR (prevent path traversal)."""
    if not path_like or not (path_like or "").strip():
        raise ValueError("XML no disponible")
    p = path_like.strip()
    if not os.path.isabs(p):
        p = os.path.join(BASE_DIR, p)
    abs_p = os.path.normpath(os.path.abspath(p))
    base = os.path.abspath(BASE_DIR)
    if not abs_p.startswith(base + os.sep):
        raise ValueError("Ruta XML inválida")
    # Guardrail: nunca servir llaves/certs ni credenciales aunque alguien logre inyectar un path.
    blocked = [
        os.path.join(base, "keys"),
        os.path.join(base, "storage", "credentials"),
    ]
    for b in blocked:
        b_abs = os.path.normpath(os.path.abspath(b))
        if abs_p == b_abs or abs_p.startswith(b_abs + os.sep):
            raise ValueError("Ruta XML inválida")
    return abs_p


def _get_sat_sync_status(issuer_id: int) -> dict:
    """Estado del sync SAT para un issuer: last_sync_at, status (running|ok|error), message."""
    conn = db()
    try:
        running = conn.execute(
            "SELECT 1 FROM sat_jobs WHERE issuer_id = ? AND status IN ('queued','running') LIMIT 1",
            (issuer_id,),
        ).fetchone()
        last_ok = conn.execute(
            "SELECT MAX(finished_at) AS t FROM sat_jobs WHERE issuer_id = ? AND status = 'ok'",
            (issuer_id,),
        ).fetchone()
        last_error = conn.execute(
            "SELECT finished_at, last_error FROM sat_jobs WHERE issuer_id = ? AND status = 'error' ORDER BY finished_at DESC LIMIT 1",
            (issuer_id,),
        ).fetchone()
        sync_state = conn.execute(
            "SELECT MAX(last_run_at) AS t FROM sat_sync_state WHERE issuer_id = ?",
            (issuer_id,),
        ).fetchone()
    finally:
        conn.close()
    last_sync_at = (sync_state and sync_state["t"]) or (last_ok and last_ok["t"]) or None
    if running:
        status = "running"
        message = "Sincronización en proceso"
    elif last_error and last_ok and last_error["t"] and last_ok["t"] and last_error["t"] > last_ok["t"]:
        status = "error"
        message = (last_error["last_error"] or "Error en la última sincronización")[:200]
    elif last_error and not last_ok:
        status = "error"
        message = (last_error["last_error"] or "Error en la última sincronización")[:200]
    else:
        status = "ok"
        message = None
    return {"last_sync_at": last_sync_at, "status": status, "message": message}


def _get_cfdi_by_uuid(issuer_id: int, uuid: str, direction: str):
    """Obtiene un CFDI de sat_cfdi por (issuer_id, uuid). direction: 'issued' o 'received'. Búsqueda por UUID case-insensitive."""
    u = (uuid or "").strip()
    if not u:
        return None
    conn = db()
    row = conn.execute(
        """
        SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, rfc_receptor, nombre_receptor,
               total, moneda, tipo_comprobante, status, xml_path, xml_status,
               serie, folio, forma_pago, metodo_pago, uso_cfdi, concepto,
               subtotal, descuento, impuestos, COALESCE(retenciones, 0) AS retenciones
        FROM sat_cfdi
        WHERE issuer_id = ? AND LOWER(TRIM(uuid)) = LOWER(?) AND direction = ?
        LIMIT 1
        """,
        (issuer_id, u, direction),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_portal_router(templates):
    """Build portal router with all /portal/* HTML routes. Requires Jinja2 templates instance."""

    def _render_portal(
        request: Request,
        *,
        issuer: dict,
        template_name: str,
        active_page: str,
        title: str,
        extra: Optional[dict] = None,
        extra_context: Optional[dict] = None,
        error: Optional[str] = None,
        status_code: int = 200,
        **template_vars: Any,
    ):
        def _db_row_to_dict(row: Any) -> dict:
            if row is None:
                return {}
            if isinstance(row, dict):
                return row
            if hasattr(row, "keys"):
                try:
                    return dict(zip(row.keys(), row))
                except Exception:
                    pass
            try:
                return dict(row)
            except Exception:
                return {}

        has_nomina = False
        if issuer.get("id", 0) > 0:
            try:
                r = db_rows(
                    "SELECT 1 FROM sat_cfdi WHERE issuer_id = ? AND direction = 'received' AND UPPER(TRIM(COALESCE(tipo_comprobante,''))) = 'N' LIMIT 1",
                    (issuer["id"],),
                )
                has_nomina = bool(r)
            except Exception:
                pass
        regimen_label = issuer.get("regimen_fiscal") or ""
        issuer_tax_code = REGIMEN_LABEL_TO_CODE.get(regimen_label) if regimen_label else ""
        show_welcome_popup = getattr(request.state, "issuer_is_placeholder", False) and not getattr(
            request.state, "is_demo_view", False
        )
        is_demo_view = getattr(request.state, "is_demo_view", False)
        is_impersonating = getattr(request.state, "is_impersonating", False)
        issuer_id = issuer.get("id", 0)
        menu_sat_configured = False
        menu_catalog_ok = False
        if issuer_id > 0:
            try:
                menu_sat_configured = bool(
                    db_rows("SELECT 1 FROM sat_credentials WHERE issuer_id = ? AND validation_ok = 1 LIMIT 1", (issuer_id,))
                )
                cust = db_rows("SELECT COUNT(*) AS n FROM customer_profiles WHERE issuer_id = ?", (issuer_id,))
                prod = db_rows("SELECT COUNT(*) AS n FROM issuer_products WHERE issuer_id = ?", (issuer_id,))
                menu_catalog_ok = (cust[0]["n"] if cust else 0) > 0 and (prod[0]["n"] if prod else 0) > 0
            except Exception:
                pass
        path = request.url.path

        def _nav_is_active(prefix_or_list):  # Helper para estado activo en rail/drawer por ruta
            if isinstance(prefix_or_list, str):
                return path.startswith(prefix_or_list)
            return any(path.startswith(p) for p in prefix_or_list)

        payload: dict[str, Any] = {
            "request": request,
            "token": "",
            "issuer_alias": issuer.get("alias", ""),
            "issuer_rfc": issuer.get("rfc", ""),
            "issuer_tax_system": issuer_tax_code,
            "issuer_regimen_label": regimen_label or "",
            "active_page": active_page,
            "nav_is_active": _nav_is_active,
            "title": title,
            "error": error,
            "has_nomina": has_nomina,
            "show_welcome_popup": show_welcome_popup,
            "is_demo_view": is_demo_view,
            "is_impersonating": is_impersonating,
            "menu_sat_configured": menu_sat_configured,
            "menu_catalog_ok": menu_catalog_ok,
            "dev_debug_panel": DEV_MODE,
            "portal_shell_v2": PORTAL_SHELL_V2,
        }
        if extra:
            payload.update(extra)
        if extra_context:
            payload.update(extra_context)
        if template_vars:
            payload.update(template_vars)
        payload.setdefault("csrf_token", csrf_service.generate_csrf_token())
        return templates.TemplateResponse(template_name, payload, status_code=status_code)

    def _portal_quotations_impl(request: Request, issuer: dict):
        try:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_quotations.html",
                active_page="quotations",
                title="Cotizaciones",
            )
        except Exception:
            logger.exception("portal: error renderizando cotizaciones")
            raise

    router = APIRouter(prefix="/portal", tags=["portal"])

    @router.get("")
    def portal_root():
        return RedirectResponse(url="/portal/home", status_code=302)

    @router.get("/login", response_class=RedirectResponse)
    def portal_login_redirect():
        """La ruta de login es /login, no /portal/login. Redirigir para evitar 404."""
        return RedirectResponse(url="/login", status_code=302)

    @router.get("/set-demo-view", response_class=RedirectResponse)
    def portal_set_demo_view(request: Request, _: dict = Depends(get_portal_issuer)):
        resp = RedirectResponse(url="/portal/home", status_code=302)
        resp.set_cookie(COOKIE_DEMO_VIEW, "1", max_age=86400 * 7, path="/", samesite="lax")
        return resp

    @router.get("/exit-demo-view", response_class=RedirectResponse)
    def portal_exit_demo_view(request: Request):
        resp = RedirectResponse(url="/portal/home", status_code=302)
        resp.delete_cookie(COOKIE_DEMO_VIEW, path="/")
        return resp

    @router.get("/quotations", response_class=HTMLResponse)
    @router.get("/cotizaciones", response_class=HTMLResponse)
    def portal_quotations(request: Request, issuer: dict = Depends(get_portal_issuer)):
        return _portal_quotations_impl(request, issuer)

    @router.get("/cotizaciones/ping")
    def portal_cotizaciones_ping():
        return Response(content="cotizaciones-ok", media_type="text/plain")

    @router.get("/quotations/{qid}/pdf")
    def portal_quotation_pdf(
        request: Request,
        qid: int,
        issuer: dict = Depends(get_portal_issuer),
        download: str = Query("0", alias="download"),
    ):
        conn = db()
        row = conn.execute(
            "SELECT id, public_token FROM quotations WHERE issuer_id = ? AND id = ?",
            (issuer["id"], qid),
        ).fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="Cotización no encontrada")
        conn.close()
        quote = quotations_service.get_quotation_by_public_token(dict(row)["public_token"])
        if not quote:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")
        cookie_val = request.cookies.get(session_service.get_session_cookie_name())
        data = session_service.verify_session(cookie_val)
        uid = data[0] if data and len(data) >= 1 else None
        audit.log(action="quotation_pdf", user_id=uid, issuer_id=issuer["id"], details=f"qid={qid}")
        log_action(request, "quotation_pdf", issuer_id=issuer["id"], quotation_id=qid)
        try:
            pdf_bytes = quotations_service.build_quotation_pdf(quote)
        except Exception:
            logger.exception("portal: error generando PDF de cotización qid=%s", qid)
            raise
        disposition = "attachment" if download == "1" else "inline"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'{disposition}; filename="cotizacion-{qid}.pdf"'},
        )

    @router.get("/quotations/{qid}", response_class=HTMLResponse)
    def portal_quotation_detail(request: Request, qid: int, issuer: dict = Depends(get_portal_issuer)):
        conn = db()
        row = conn.execute(
            "SELECT id, public_token, folio, customer_rfc, customer_legal_name, customer_email, status, notes, responded_at, created_at FROM quotations WHERE issuer_id = ? AND id = ?",
            (issuer["id"], qid),
        ).fetchone()
        conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")
        d = dict(row)
        quote = quotations_service.get_quotation_by_public_token(d["public_token"])
        if not quote:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")
        return _render_portal(
            request,
            issuer=issuer,
            template_name="quote_detail.html",
            active_page="quotations",
            title="Cotización",
            extra={"quote": quote},
        )

    @router.get("/home", response_class=HTMLResponse)
    def portal_home(request: Request, issuer: dict = Depends(get_portal_issuer)):
        try:
            issuer_id = issuer["id"]
            ym = ym_now()
            count_issued = db_rows(
                """
                SELECT COUNT(*) AS n FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                  AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
                """,
                (issuer_id, ym),
            )
            count_received = db_rows(
                """
                SELECT COUNT(*) AS n FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                  AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01
                  AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                """,
                (issuer_id, ym),
            )
            tot_issued = _get_month_totals(issuer_id, ym, "issued")
            tot_received = _get_month_totals(issuer_id, ym, "received")
            ingresos_sin_iva = tot_issued["total_base"]
            gastos_sin_iva = tot_received["total_base"]
            iva_retenciones = tot_issued["total_retenciones"]
            iva_recibido_neto = tot_issued["total_iva_neto"]
            iva_pagado = tot_received["total_iva"]
            activities = db_rows(
                """
                SELECT direction, fecha_emision, nombre, total, uuid FROM (
                  SELECT direction, fecha_emision, nombre_receptor AS nombre, total, uuid FROM sat_cfdi
                  WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                    AND (total IS NULL OR total >= 0.01)
                  UNION ALL
                  SELECT direction, fecha_emision, nombre_emisor AS nombre, total, uuid FROM sat_cfdi
                  WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                    AND total IS NOT NULL AND total >= 0.01
                    AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                ) ORDER BY fecha_emision DESC LIMIT 50
                """,
                (issuer_id, issuer_id),
            )
            today = date.today()
            for a in activities:
                try:
                    fd = datetime.strptime((a["fecha_emision"] or "")[:10], "%Y-%m-%d").date()
                    d = (today - fd).days
                    a["time_ago"] = "Hoy" if d == 0 else "Ayer" if d == 1 else f"Hace {d} días"
                except (ValueError, TypeError):
                    a["time_ago"] = a.get("fecha_emision", "")[:10] or "-"
            # Onboarding: RFC completo, FIEL/CSD, clientes, productos (para ocultar banner cuando todo está listo)
            rfc_val = (issuer.get("rfc") or "").strip().upper()
            razon = (issuer.get("razon_social") or "").strip()
            rfc_configured = bool(rfc_val and rfc_val != "PENDIENTE" and razon)
            has_fiel = bool(
                db_rows("SELECT 1 FROM sat_credentials WHERE issuer_id = ? LIMIT 1", (issuer_id,))
            )
            cust_count = db_rows("SELECT COUNT(*) AS n FROM customer_profiles WHERE issuer_id = ?", (issuer_id,))
            prod_count = db_rows("SELECT COUNT(*) AS n FROM issuer_products WHERE issuer_id = ?", (issuer_id,))
            any_issued = db_rows(
                "SELECT 1 FROM sat_cfdi WHERE issuer_id = ? AND direction = 'issued' AND (total IS NULL OR total >= 0.01) LIMIT 1",
                (issuer_id,),
            )
            onboarding = {
                "rfc_configured": rfc_configured,
                "has_fiel": has_fiel,
                "count_customers": cust_count[0]["n"] if cust_count else 0,
                "count_products": prod_count[0]["n"] if prod_count else 0,
                "has_any_issued": bool(any_issued),
            }
            # Listas para Factura rápida (dropdowns y datos precargados)
            quick_customers = db_rows(
                """
                SELECT id, rfc, legal_name, zip, tax_system, email, alias
                FROM customer_profiles WHERE issuer_id = ? ORDER BY COALESCE(alias, ''), rfc
                """,
                (issuer_id,),
            ) or []
            quick_products = db_rows(
                """
                SELECT id, description, product_key, unit_key, unit_price, iva_rate, created_at
                FROM issuer_products WHERE issuer_id = ? ORDER BY description
                """,
                (issuer_id,),
            ) or []

            def _serialize_row(r):
                d = dict(r) if hasattr(r, "keys") else r
                out = {}
                for k, v in d.items():
                    if hasattr(v, "isoformat"):
                        out[k] = v.isoformat()
                    elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                        try:
                            out[k] = float(v)
                        except (TypeError, ValueError):
                            out[k] = v
                    else:
                        out[k] = v
                return out

            # Escapar para incrustar en <script>: evitar que </script> en datos cierre el tag
            def _script_safe(s: str) -> str:
                return (s or "").replace("</", "<\\/")
            quick_customers_json = _script_safe(json.dumps([_serialize_row(r) for r in quick_customers]))
            quick_products_json = _script_safe(json.dumps([_serialize_row(r) for r in quick_products]))

            # Notificaciones (motor simple)
            notifications = []
            try:
                from services import notifications as notifications_service

                notifications_service.refresh_for_issuer(int(issuer_id))
                notifications = notifications_service.list_notifications(int(issuer_id), unread_only=True, limit=10) or []
            except Exception:
                notifications = []

            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_home.html",
                active_page="home",
                title="Inicio",
                extra={
                    "issuer_id": issuer_id,
                    "count_issued": count_issued[0]["n"] if count_issued else 0,
                    "count_received": count_received[0]["n"] if count_received else 0,
                    "activities": activities,
                    "ym_label": ym_to_label(ym),
                    "ym": ym,
                    "ingresos_sin_iva": ingresos_sin_iva,
                    "gastos_sin_iva": gastos_sin_iva,
                    "iva_recibido_neto": iva_recibido_neto,
                    "iva_retenciones": iva_retenciones,
                    "iva_pagado": iva_pagado,
                    "onboarding": onboarding,
                    "quick_customers": quick_customers,
                    "quick_products": quick_products,
                    "quick_customers_json": quick_customers_json,
                    "quick_products_json": quick_products_json,
                    "sat_sync_status": _get_sat_sync_status(issuer_id),
                    "has_fiel_validated": has_fiel and bool(
                        db_rows(
                            "SELECT 1 FROM sat_credentials WHERE issuer_id = ? AND validation_ok = 1 LIMIT 1",
                            (issuer_id,),
                        )
                    ),
                    "notifications": notifications,
                    "csrf_token": csrf_service.generate_csrf_token(),
                },
            )
        except Exception:
            logger.exception("portal: error renderizando home")
            raise

    @router.get("/qa", response_class=HTMLResponse)
    def portal_qa(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Página interna de QA: checks básicos y enlaces rápidos. Solo visible con DEV_MODE=1 o ENV=dev."""
        if not DEV_MODE:
            raise HTTPException(status_code=404, detail="Not Found")
        try:
            issuer_id = issuer["id"]
            ym = ym_now()
            count_issued = db_rows(
                """
                SELECT COUNT(*) AS n FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                  AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
                """,
                (issuer_id, ym),
            )
            count_received = db_rows(
                """
                SELECT COUNT(*) AS n FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                  AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01
                  AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                """,
                (issuer_id, ym),
            )
            sat_sync_status = _get_sat_sync_status(issuer_id)
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_qa.html",
                active_page="home",
                title="QA (solo dev)",
                extra={
                    "ym": ym,
                    "ym_label": ym_to_label(ym),
                    "count_issued": count_issued[0]["n"] if count_issued else 0,
                    "count_received": count_received[0]["n"] if count_received else 0,
                    "sat_last_sync": sat_sync_status.get("last_sync_at"),
                    "sat_status": sat_sync_status.get("status"),
                },
            )
        except HTTPException:
            raise
        except Exception:
            logger.exception("portal: error renderizando /portal/qa")
            raise HTTPException(status_code=500, detail="Error al cargar QA")

    @router.get("/create", response_class=HTMLResponse)
    def portal_create(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        quote_id: Optional[int] = Query(None),
        customer_rfc: Optional[str] = Query(None),
        customer_legal_name: Optional[str] = Query(None),
        customer_zip: Optional[str] = Query(None),
        customer_tax_system: Optional[str] = Query(None),
        customer_email: Optional[str] = Query(None),
        concept_desc: Optional[str] = Query(None),
        concept_key: Optional[str] = Query(None),
        concept_unit: Optional[str] = Query(None),
        concept_price: Optional[str] = Query(None),
        concept_iva: Optional[str] = Query(None),
    ):
        customer_prefill = None
        concept_prefill = None
        quote_items = None
        if quote_id is not None:
            try:
                conn = db()
                row = conn.execute(
                    "SELECT customer_rfc, customer_legal_name, customer_email FROM quotations WHERE issuer_id = ? AND id = ?",
                    (issuer["id"], quote_id),
                ).fetchone()
                if row:
                    r = dict(row)
                    customer_prefill = {
                        "customer_rfc": (r.get("customer_rfc") or "").strip(),
                        "customer_legal_name": (r.get("customer_legal_name") or "").strip(),
                        "customer_zip": "",
                        "customer_tax_system": "",
                        "customer_email": (r.get("customer_email") or "").strip(),
                    }
                    items = conn.execute(
                        "SELECT description, quantity, unit_price, iva_rate FROM quotation_items WHERE quotation_id = ? ORDER BY sort_order, id",
                        (quote_id,),
                    ).fetchall()
                    conn.close()
                    quote_items = [
                        {
                            "description": x["description"],
                            "quantity": float(x["quantity"] or 0),
                            "unit_price": float(x["unit_price"] or 0),
                            "iva_rate": float(x["iva_rate"] or 0.16),
                        }
                        for x in items
                    ]
                    if quote_items:
                        concept_prefill = {
                            "description": quote_items[0]["description"],
                            "product_key": "",
                            "unit_key": "E48",
                            "unit_price": str(quote_items[0]["unit_price"]),
                            "iva_rate": str(quote_items[0]["iva_rate"]),
                        }
                else:
                    conn.close()
            except ValueError:
                pass
        if customer_prefill is None and (customer_rfc or customer_legal_name or customer_zip or customer_tax_system or customer_email):
            customer_prefill = {
                "customer_rfc": (customer_rfc or "").strip(),
                "customer_legal_name": (customer_legal_name or "").strip(),
                "customer_zip": (customer_zip or "").strip(),
                "customer_tax_system": (customer_tax_system or "").strip(),
                "customer_email": (customer_email or "").strip(),
            }
        if concept_prefill is None and (concept_desc or concept_key or concept_unit or concept_price or concept_iva):
            concept_prefill = {
                "description": (concept_desc or "").strip(),
                "product_key": (concept_key or "").strip(),
                "unit_key": (concept_unit or "").strip() or "E48",
                "unit_price": (concept_price or "").strip(),
                "iva_rate": (concept_iva or "0.16").strip(),
            }
        try:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="form.html",
                active_page="create",
                title="Factura nueva",
                extra={
                    "create_mode": "normal",
                    "customer_prefill": customer_prefill,
                    "concept_prefill": concept_prefill,
                    "quote_items": quote_items,
                    "csrf_token": csrf_service.generate_csrf_token(),
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/create")
            raise

    @router.get("/create/quick", response_class=HTMLResponse)
    def portal_create_quick(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        customer_id: Optional[int] = Query(None),
        product_id: Optional[int] = Query(None),
    ):
        # Sin cliente y producto: página para elegir (misma fuente que Clientes y Productos: /api/customers, /api/products)
        if customer_id is None or product_id is None:
            try:
                return _render_portal(
                    request,
                    issuer=issuer,
                    template_name="portal_create_quick_choose.html",
                    active_page="create",
                    title="Factura rápida",
                )
            except Exception:
                logger.exception("portal: error renderizando selector factura rápida")
                raise
        customer_prefill = None
        concept_prefill = None
        issuer_id = issuer["id"]
        cust = db_rows(
            "SELECT id, rfc, legal_name, zip, tax_system, email FROM customer_profiles WHERE issuer_id = ? AND id = ? LIMIT 1",
            (issuer_id, customer_id),
        )
        prod = db_rows(
            "SELECT id, description, product_key, unit_key, unit_price, iva_rate FROM issuer_products WHERE issuer_id = ? AND id = ? LIMIT 1",
            (issuer_id, product_id),
        )
        if cust and prod:
            c = cust[0]
            p = prod[0]
            customer_prefill = {
                "customer_rfc": (c.get("rfc") or "").strip(),
                "customer_legal_name": (c.get("legal_name") or "").strip(),
                "customer_zip": (c.get("zip") or "").strip(),
                "customer_tax_system": (c.get("tax_system") or "").strip(),
                "customer_email": (c.get("email") or "").strip(),
            }
            concept_prefill = {
                "description": (p.get("description") or "").strip(),
                "product_key": (p.get("product_key") or "").strip(),
                "unit_key": (p.get("unit_key") or "").strip() or "E48",
                "unit_price": str(p.get("unit_price") or ""),
                "iva_rate": str(p.get("iva_rate") or "0.16"),
            }
        try:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="form.html",
                active_page="create_quick",
                title="Factura rápida",
                extra={
                    "create_mode": "quick",
                    "csrf_token": csrf_service.generate_csrf_token(),
                    "customer_prefill": customer_prefill,
                    "concept_prefill": concept_prefill,
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/create/quick")
            raise

    @router.get("/create/multi", response_class=HTMLResponse)
    def portal_create_multi(request: Request, issuer: dict = Depends(get_portal_issuer)):
        try:
            return _render_portal(
                request, issuer=issuer, template_name="form.html", active_page="create_multi", title="Factura múltiple", extra={"create_mode": "multi", "csrf_token": csrf_service.generate_csrf_token()}
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/create/multi")
            raise

    @router.get("/invoices", response_class=HTMLResponse)
    def portal_invoices(request: Request, issuer: dict = Depends(get_portal_issuer)):
        try:
            return _render_portal(
                request, issuer=issuer, template_name="portal_invoices.html", active_page="issued", title="Mis facturas"
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/invoices")
            raise

    @router.get("/invoices/issued", response_class=HTMLResponse)
    def portal_invoices_issued(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
        try:
            issuer_id = issuer["id"]
            if not ym:
                ym = ym_now()
            rows = db_rows("""
                SELECT uuid, fecha_emision, rfc_receptor, nombre_receptor, concepto, total, moneda,
                       COALESCE(impuestos, 0) AS impuestos, COALESCE(retenciones, 0) AS retenciones,
                       metodo_pago, status, xml_path
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                  AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                  AND id IN (
                    SELECT id FROM (
                      SELECT id, ROW_NUMBER() OVER (
                        PARTITION BY issuer_id, direction, LOWER(TRIM(uuid))
                        ORDER BY (CASE WHEN COALESCE(total,0) >= 0.01 THEN 0 ELSE 1 END), id
                      ) AS rn
                      FROM sat_cfdi
                      WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
                        AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                    ) WHERE rn = 1
                  )
                ORDER BY fecha_emision DESC LIMIT 300;
            """, (issuer_id, ym, issuer_id, ym))
            months = db_rows("""
                SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                  AND (total IS NULL OR total >= 0.01)
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                GROUP BY ym ORDER BY ym DESC;
            """, (issuer_id,))
            for m in months:
                m["label"] = ym_to_label(m["ym"])
            month_totals = _get_month_totals(issuer_id, ym, "issued")
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_issued.html",
                active_page="issued",
                title="Facturas emitidas",
                extra={
                    "rows": rows,
                    "ym": ym,
                    "ym_label": ym_to_label(ym),
                    "prev_ym": shift_ym(ym, -1),
                    "next_ym": shift_ym(ym, +1),
                    "months": months,
                    "month_totals": month_totals,
                    "sat_sync_status": _get_sat_sync_status(issuer_id),
                    "has_fiel_validated": bool(db_rows("SELECT 1 FROM sat_credentials WHERE issuer_id = ? AND validation_ok = 1 LIMIT 1", (issuer_id,))),
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/invoices/issued ym=%s", ym)
            raise

    @router.get("/invoices/received", response_class=HTMLResponse)
    def portal_invoices_received(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
        try:
            issuer_id = issuer["id"]
            if not ym:
                ym = ym_now()
            rows = db_rows("""
                SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, concepto, total, moneda,
                       COALESCE(impuestos, 0) AS impuestos, COALESCE(retenciones, 0) AS retenciones,
                       metodo_pago, status, xml_path
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                  AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01
                  AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                  AND id IN (
                    SELECT id FROM (
                      SELECT id, ROW_NUMBER() OVER (
                        PARTITION BY issuer_id, direction, LOWER(TRIM(uuid))
                        ORDER BY id
                      ) AS rn
                      FROM sat_cfdi
                      WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01 AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                        AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                    ) WHERE rn = 1
                  )
                ORDER BY fecha_emision DESC LIMIT 300;
            """, (issuer_id, ym, issuer_id, ym))
            months = db_rows("""
                SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                  AND total IS NOT NULL AND total >= 0.01
                  AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                GROUP BY ym ORDER BY ym DESC;
            """, (issuer_id,))
            for m in months:
                m["label"] = ym_to_label(m["ym"])
            month_totals = _get_month_totals(issuer_id, ym, "received")
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_received.html",
                active_page="received",
                title="Facturas recibidas",
                extra={
                    "rows": rows,
                    "ym": ym,
                    "ym_label": ym_to_label(ym),
                    "prev_ym": shift_ym(ym, -1),
                    "next_ym": shift_ym(ym, +1),
                    "months": months,
                    "month_totals": month_totals,
                    "sat_sync_status": _get_sat_sync_status(issuer_id),
                    "has_fiel_validated": bool(db_rows("SELECT 1 FROM sat_credentials WHERE issuer_id = ? AND validation_ok = 1 LIMIT 1", (issuer_id,))),
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/invoices/received ym=%s", ym)
            raise

    # ---------- Hubs (navegación agrupada con tabs) ----------
    @router.get("/facturas", response_class=HTMLResponse)
    def portal_facturas_hub(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        tab: str = Query("issued", description="issued|received|ppd"),
        ym: str = None,
    ):
        """Hub Facturas: tabs Emitidas / Recibidas / PPD. Reutiliza misma lógica y datos que rutas legacy."""
        try:
            if tab not in ("issued", "received", "ppd"):
                tab = "issued"
            issuer_id = issuer["id"]
            if not ym:
                ym = ym_now()
            month_picker_base_url = f"/portal/facturas?tab={tab}"
            sat_sync_status = _get_sat_sync_status(issuer_id)
            has_fiel_validated = bool(db_rows("SELECT 1 FROM sat_credentials WHERE issuer_id = ? AND validation_ok = 1 LIMIT 1", (issuer_id,)))
            base_extra = {
                "ym": ym,
                "ym_label": ym_to_label(ym),
                "prev_ym": shift_ym(ym, -1),
                "next_ym": shift_ym(ym, +1),
                "sat_sync_status": sat_sync_status,
                "has_fiel_validated": has_fiel_validated,
                "month_picker_base_url": month_picker_base_url,
            }
            if tab == "issued":
                rows = db_rows("""
                    SELECT uuid, fecha_emision, rfc_receptor, nombre_receptor, concepto, total, moneda,
                           COALESCE(impuestos, 0) AS impuestos, COALESCE(retenciones, 0) AS retenciones,
                           metodo_pago, status, xml_path
                    FROM sat_cfdi
                    WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                      AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
                      AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                      AND id IN (
                        SELECT id FROM (
                          SELECT id, ROW_NUMBER() OVER (
                            PARTITION BY issuer_id, direction, LOWER(TRIM(uuid))
                            ORDER BY (CASE WHEN COALESCE(total,0) >= 0.01 THEN 0 ELSE 1 END), id
                          ) AS rn
                          FROM sat_cfdi
                          WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
                            AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                        ) WHERE rn = 1
                      )
                    ORDER BY fecha_emision DESC LIMIT 300;
                """, (issuer_id, ym, issuer_id, ym))
                months = db_rows("""
                    SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
                    FROM sat_cfdi
                    WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                      AND (total IS NULL OR total >= 0.01)
                      AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                    GROUP BY ym ORDER BY ym DESC;
                """, (issuer_id,))
                for m in months:
                    m["label"] = ym_to_label(m["ym"])
                month_totals = _get_month_totals(issuer_id, ym, "issued")
                base_extra.update({
                    "rows": rows,
                    "months": months,
                    "month_totals": month_totals,
                })
            else:
                # received y ppd usan los mismos datos (recibidas); PPD se filtra en front con metodo_pago
                rows = db_rows("""
                    SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, concepto, total, moneda,
                           COALESCE(impuestos, 0) AS impuestos, COALESCE(retenciones, 0) AS retenciones,
                           metodo_pago, status, xml_path
                    FROM sat_cfdi
                    WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                      AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01
                      AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                      AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                      AND id IN (
                        SELECT id FROM (
                          SELECT id, ROW_NUMBER() OVER (
                            PARTITION BY issuer_id, direction, LOWER(TRIM(uuid))
                            ORDER BY id
                          ) AS rn
                          FROM sat_cfdi
                          WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01 AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                            AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                        ) WHERE rn = 1
                      )
                    ORDER BY fecha_emision DESC LIMIT 300;
                """, (issuer_id, ym, issuer_id, ym))
                months = db_rows("""
                    SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
                    FROM sat_cfdi
                    WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                      AND total IS NOT NULL AND total >= 0.01
                      AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
                      AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                    GROUP BY ym ORDER BY ym DESC;
                """, (issuer_id,))
                for m in months:
                    m["label"] = ym_to_label(m["ym"])
                month_totals = _get_month_totals(issuer_id, ym, "received")
                base_extra.update({
                    "rows": rows,
                    "months": months,
                    "month_totals": month_totals,
                    "default_metodo_pago": "PPD" if tab == "ppd" else None,
                    "list_title": "Facturas recibidas (PPD)" if tab == "ppd" else "Facturas recibidas",
                })
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_facturas.html",
                active_page="facturas_hub",
                title="Facturas",
                extra={
                    **base_extra,
                    "active_tab": tab,
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/facturas tab=%s", tab)
            raise

    @router.get("/contactos", response_class=HTMLResponse)
    def portal_contactos_hub(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        tab: str = Query("clientes", description="clientes|proveedores"),
        q: str = Query(""),
        page: int = Query(1, ge=1),
        per_page: int = Query(200, ge=1, le=500),
    ):
        """Hub Contactos: tabs Clientes / Proveedores."""
        try:
            if tab not in ("clientes", "proveedores"):
                tab = "clientes"
            issuer_id = int(issuer.get("id") or 0)
            rows = []
            total = 0
            pages = 0
            query = (q or "").strip()
            # Guardrail: no permitir offsets enormes por page muy alta
            max_page = (MAX_LIST_OFFSET // max(1, int(per_page))) + 1
            if page > max_page:
                page = max_page
            if tab == "clientes" and issuer_id > 0:
                conn = db()
                try:
                    if table_exists(conn, "clients"):
                        if query:
                            like = f"%{query}%"
                            total_row = conn.execute(
                                "SELECT COUNT(*) AS c FROM clients WHERE issuer_id = ? AND (rfc LIKE ? OR COALESCE(name,'') LIKE ?)",
                                (issuer_id, like, like),
                            ).fetchone()
                            total = int(total_row.get("c") or total_row.get("n") or 0) if isinstance(total_row, dict) else (int(total_row[0]) if total_row else 0)
                            offset = (page - 1) * per_page
                            rows = conn.execute(
                                """
                                SELECT id, rfc, name, cp, regimen_fiscal, uso_cfdi_default, email, phone, last_seen_at
                                FROM clients
                                WHERE issuer_id = ? AND (rfc LIKE ? OR COALESCE(name,'') LIKE ?)
                                ORDER BY COALESCE(last_seen_at, created_at) DESC
                                LIMIT ? OFFSET ?
                                """,
                                (issuer_id, like, like, per_page, offset),
                            ).fetchall()
                        else:
                            total_row = conn.execute(
                                "SELECT COUNT(*) AS c FROM clients WHERE issuer_id = ?", (issuer_id,)
                            ).fetchone()
                            total = int(total_row.get("c") or total_row.get("n") or 0) if isinstance(total_row, dict) else (int(total_row[0]) if total_row else 0)
                            offset = (page - 1) * per_page
                            rows = conn.execute(
                                """
                                SELECT id, rfc, name, cp, regimen_fiscal, uso_cfdi_default, email, phone, last_seen_at
                                FROM clients
                                WHERE issuer_id = ?
                                ORDER BY COALESCE(last_seen_at, created_at) DESC
                                LIMIT ? OFFSET ?
                                """,
                                (issuer_id, per_page, offset),
                            ).fetchall()
                        rows = [_db_row_to_dict(r) for r in rows]
                        pages = (total + per_page - 1) // per_page if total > 0 else 0
                finally:
                    conn.close()
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_contactos.html",
                active_page="contactos_hub",
                title="Contactos",
                extra={
                    "active_tab": tab,
                    "rows": rows,
                    "q": query,
                    "total": total,
                    "page": page,
                    "per_page": per_page,
                    "pages": pages,
                    "hub_base": "/portal/contactos",
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/contactos")
            raise

    @router.get("/bancos", response_class=RedirectResponse)
    def portal_bancos_redirect():
        """Redirigir a la página de convertir estado de cuenta."""
        return RedirectResponse(url="/portal/convertir-edo-cuenta", status_code=302)

    @router.get("/convertir-edo-cuenta", response_class=HTMLResponse)
    def portal_convertir_edo_cuenta(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Página única: arrastrar PDF, convertir a Excel y ver movimientos (sin pestañas ni hub)."""
        try:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_bank_pdf_to_excel.html",
                active_page="convertir_edo_cuenta",
                title="Convertir Edo. de Cuenta",
            )
        except Exception as e:
            logger.exception("convertir-edo-cuenta: error en render completo (%s), usando página mínima", e)
            try:
                return templates.TemplateResponse(
                    "portal_convertir_edo_cuenta_minimal.html",
                    {
                        "request": request,
                        "csrf_token": csrf_service.generate_csrf_token(),
                        "preview_movements": [],
                        "preview_summary": {},
                    },
                    status_code=200,
                )
            except Exception as e2:
                logger.exception("convertir-edo-cuenta: fallback mínima también falló: %s", e2)
                raise

    @router.get("/invoices/nomina", response_class=HTMLResponse)
    def portal_invoices_nomina(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
        try:
            issuer_id = issuer["id"]
            if not ym:
                ym = ym_now()
            rows = db_rows("""
                SELECT uuid, fecha_emision, rfc_emisor, nombre_emisor, total, moneda, status, xml_path,
                       serie, folio, concepto, forma_pago, metodo_pago, uso_cfdi, subtotal, descuento, impuestos,
                       tipo_comprobante, xml_status
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received'
                  AND UPPER(TRIM(COALESCE(tipo_comprobante,''))) = 'N'
                  AND fecha_emision IS NOT NULL AND substr(fecha_emision,1,7) = ?
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                ORDER BY fecha_emision DESC LIMIT 300;
            """, (issuer_id, ym))
            months = db_rows("""
                SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received'
                  AND UPPER(TRIM(COALESCE(tipo_comprobante,''))) = 'N'
                  AND fecha_emision IS NOT NULL
                  AND xml_path IS NOT NULL AND TRIM(COALESCE(xml_path,'')) != ''
                GROUP BY ym ORDER BY ym DESC;
            """, (issuer_id,))
            for m in months:
                m["label"] = ym_to_label(m["ym"])
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_nomina.html",
                active_page="nomina",
                title="Nómina recibida",
                extra={
                    "rows": rows,
                    "ym": ym,
                    "ym_label": ym_to_label(ym),
                    "prev_ym": shift_ym(ym, -1),
                    "next_ym": shift_ym(ym, +1),
                    "months": months,
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/invoices/nomina ym=%s", ym)
            raise

    def _audit_user_issuer(request: Request):
        cookie_val = request.cookies.get(session_service.get_session_cookie_name())
        data = session_service.verify_session(cookie_val)
        user_id = data[0] if data and len(data) >= 1 else None
        issuer_id = data[1] if data and len(data) >= 2 else None
        return user_id, issuer_id

    @router.get("/sat/xml/{uuid}")
    def portal_sat_xml(request: Request, uuid: str, issuer: dict = Depends(get_portal_issuer)):
        u = (uuid or "").strip()
        if not u:
            raise HTTPException(status_code=404, detail="UUID no válido")
        conn = db()
        row = conn.execute(
            "SELECT xml_path, uuid FROM sat_cfdi WHERE issuer_id = ? AND LOWER(TRIM(uuid)) = LOWER(?) LIMIT 1",
            (issuer["id"], u),
        ).fetchone()
        conn.close()
        if not row or not row["xml_path"]:
            raise HTTPException(status_code=404, detail="XML no encontrado para este UUID")
        try:
            abs_path = _safe_abs_path(row["xml_path"])
        except ValueError:
            raise HTTPException(status_code=404, detail="Ruta XML inválida")
        if not os.path.exists(abs_path):
            portal_error_type("file_missing", log_context={"issuer_id": issuer["id"], "uuid": u[:36]})
        uid, iid = _audit_user_issuer(request)
        if not subscription_service.can_issuer_use_sync_and_timbrado(issuer["id"], uid or 0):
            raise HTTPException(status_code=402, detail="Actualiza tu plan para descargar XML. Ve a /pricing.")
        audit.log(
            action="download_xml",
            user_id=uid,
            issuer_id=issuer["id"],
            details=u[:36],
            request=request,
            entity="cfdi",
            entity_id=u,
        )
        log_action(request, "download_xml", issuer_id=issuer["id"], entity_id=u[:36])
        file_access_log.log_file_access(
            request=request,
            action="download_xml",
            issuer_id=issuer["id"],
            user_id=uid,
            file_path=row.get("xml_path") if isinstance(row, dict) else None,
            entity="cfdi",
            entity_id=u[:36],
        )
        with open(abs_path, "rb") as f:
            xml_bytes = f.read()
        return Response(
            content=xml_bytes,
            media_type="application/xml",
            headers={"Content-Disposition": f'inline; filename="{row["uuid"]}.xml"'},
        )

    @router.get("/sat/pdf/{uuid}")
    def portal_sat_pdf(request: Request, uuid: str, issuer: dict = Depends(get_portal_issuer), dl: int = 0):
        uuid_clean = (uuid or "").strip().split()[0] if uuid else ""
        if not uuid_clean:
            raise HTTPException(status_code=404, detail="UUID no válido")
        conn = db()
        row = conn.execute(
            "SELECT xml_path, uuid FROM sat_cfdi WHERE issuer_id = ? AND LOWER(TRIM(uuid)) = LOWER(?) LIMIT 1",
            (issuer["id"], uuid_clean),
        ).fetchone()
        conn.close()
        if not row or not row["xml_path"]:
            raise HTTPException(status_code=404, detail="XML no encontrado para este UUID")
        try:
            abs_path = _safe_abs_path(row["xml_path"])
        except ValueError:
            raise HTTPException(status_code=404, detail="Ruta XML inválida")
        if not os.path.exists(abs_path):
            portal_error_type("file_missing", log_context={"issuer_id": issuer["id"], "uuid": uuid_clean[:36]})
        uid, _ = _audit_user_issuer(request)
        if not subscription_service.can_issuer_use_sync_and_timbrado(issuer["id"], uid or 0):
            raise HTTPException(status_code=402, detail="Actualiza tu plan para descargar PDF. Ve a /pricing.")
        audit.log(
            action="download_pdf",
            user_id=uid,
            issuer_id=issuer["id"],
            details=uuid_clean[:36],
            request=request,
            entity="cfdi",
            entity_id=uuid_clean,
        )
        log_action(request, "download_pdf", issuer_id=issuer["id"], entity_id=uuid_clean[:36])
        file_access_log.log_file_access(
            request=request,
            action="download_pdf",
            issuer_id=issuer["id"],
            user_id=uid,
            file_path=row.get("xml_path") if isinstance(row, dict) else None,
            entity="cfdi",
            entity_id=uuid_clean[:36],
        )
        try:
            from cfdi_pdf import parse_cfdi_xml, build_pdf
            data = parse_cfdi_xml(abs_path)
            pdf_bytes = build_pdf(data)
        except ImportError:
            portal_error_type("reportlab_missing", log_context={"issuer_id": issuer["id"], "uuid": uuid_clean[:36]})
        except Exception:
            logger.exception(
                "portal: error generando PDF issuer_id=%s uuid=%s",
                issuer["id"],
                uuid_clean[:36],
            )
            portal_error_type("server_error", log_context={"issuer_id": issuer["id"], "uuid": uuid_clean[:36]}, override_message="No se pudo generar el PDF. Intenta de nuevo.")
        if not pdf_bytes:
            portal_error_type("server_error", log_context={"issuer_id": issuer["id"]}, override_message="La generación del PDF devolvió vacío.")
        filename = f"cfdi-{uuid_clean[:8]}.pdf"
        disposition = "attachment" if dl else "inline"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'{disposition}; filename="{filename}"',
                "Content-Length": str(len(pdf_bytes)),
            },
        )

    @router.get("/cfdi/issued/{uuid}", response_class=HTMLResponse)
    def portal_cfdi_detail_issued(request: Request, uuid: str, issuer: dict = Depends(get_portal_issuer)):
        cfdi = _get_cfdi_by_uuid(issuer["id"], uuid, "issued")
        if not cfdi:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_cfdi_detail.html",
                active_page="issued",
                title="CFDI no encontrado",
                extra={"cfdi": None, "direction": "issued", "error": "not_found", "requested_uuid": uuid},
                status_code=404,
            )
        uid = getattr(request.state, "user_id", None) or 0
        audit.log(
            action="cfdi_view",
            user_id=uid if uid else None,
            issuer_id=issuer["id"],
            details=f"direction=issued uuid={(uuid or '')[:36]}",
            request=request,
            entity="cfdi",
            entity_id=(uuid or "").strip()[:36],
        )
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_cfdi_detail.html",
            active_page="issued",
            title="Detalle CFDI emitido",
            extra={"cfdi": cfdi, "direction": "issued"},
        )

    @router.get("/cfdi/received/{uuid}", response_class=HTMLResponse)
    def portal_cfdi_detail_received(request: Request, uuid: str, issuer: dict = Depends(get_portal_issuer)):
        cfdi = _get_cfdi_by_uuid(issuer["id"], uuid, "received")
        if not cfdi:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_cfdi_detail.html",
                active_page="received",
                title="CFDI no encontrado",
                extra={"cfdi": None, "direction": "received", "error": "not_found", "requested_uuid": uuid},
                status_code=404,
            )
        uid = getattr(request.state, "user_id", None) or 0
        audit.log(
            action="cfdi_view",
            user_id=uid if uid else None,
            issuer_id=issuer["id"],
            details=f"direction=received uuid={(uuid or '')[:36]}",
            request=request,
            entity="cfdi",
            entity_id=(uuid or "").strip()[:36],
        )
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_cfdi_detail.html",
            active_page="received",
            title="Detalle CFDI recibido",
            extra={"cfdi": cfdi, "direction": "received"},
        )

    @router.get("/clients", response_class=HTMLResponse)
    def portal_clients(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        q: str = Query(""),
        page: int = Query(1, ge=1),
        per_page: int = Query(200, ge=1, le=500),
    ):
        try:
            issuer_id = int(issuer.get("id") or 0)
            query = (q or "").strip()
            max_page = (MAX_LIST_OFFSET // max(1, int(per_page))) + 1
            if page > max_page:
                page = max_page
            rows = []
            total = 0
            if issuer_id > 0:
                conn = db()
                try:
                    if not table_exists(conn, "clients"):
                        pages = 0
                    else:
                        if query:
                            like = f"%{query}%"
                            total_row = conn.execute(
                                "SELECT COUNT(*) AS c FROM clients WHERE issuer_id = ? AND (rfc LIKE ? OR COALESCE(name,'') LIKE ?)",
                                (issuer_id, like, like),
                            ).fetchone()
                            total = int(total_row.get("c") or total_row.get("n") or 0) if isinstance(total_row, dict) else (int(total_row[0]) if total_row else 0)
                            offset = (page - 1) * per_page
                            rows = conn.execute(
                                """
                                SELECT id, rfc, name, cp, regimen_fiscal, uso_cfdi_default, email, phone, last_seen_at
                                FROM clients
                                WHERE issuer_id = ? AND (rfc LIKE ? OR COALESCE(name,'') LIKE ?)
                                ORDER BY COALESCE(last_seen_at, created_at) DESC
                                LIMIT ? OFFSET ?
                                """,
                                (issuer_id, like, like, per_page, offset),
                            ).fetchall()
                        else:
                            total_row = conn.execute(
                                "SELECT COUNT(*) AS c FROM clients WHERE issuer_id = ?", (issuer_id,)
                            ).fetchone()
                            total = int(total_row.get("c") or total_row.get("n") or 0) if isinstance(total_row, dict) else (int(total_row[0]) if total_row else 0)
                            offset = (page - 1) * per_page
                            rows = conn.execute(
                                """
                                SELECT id, rfc, name, cp, regimen_fiscal, uso_cfdi_default, email, phone, last_seen_at
                                FROM clients
                                WHERE issuer_id = ?
                                ORDER BY COALESCE(last_seen_at, created_at) DESC
                                LIMIT ? OFFSET ?
                                """,
                                (issuer_id, per_page, offset),
                            ).fetchall()
                        pages = (total + per_page - 1) // per_page if total > 0 else 0
                finally:
                    conn.close()
            else:
                pages = 0
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_clients.html",
                active_page="clients",
                title="Clientes",
                extra={
                    "rows": [dict(r) for r in rows] if rows else [],
                    "q": query,
                    "total": total,
                    "page": page,
                    "per_page": per_page,
                    "pages": pages,
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/clients")
            raise

    @router.get("/providers", response_class=HTMLResponse)
    def portal_providers(request: Request, issuer: dict = Depends(get_portal_issuer)):
        try:
            return _render_portal(
                request, issuer=issuer, template_name="portal_providers.html", active_page="providers", title="Proveedores"
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/providers")
            raise

    @router.get("/products", response_class=HTMLResponse)
    def portal_products(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        q: str = Query(""),
        page: int = Query(1, ge=1),
        per_page: int = Query(200, ge=1, le=500),
    ):
        try:
            issuer_id = int(issuer.get("id") or 0)
            query = (q or "").strip()
            max_page = (MAX_LIST_OFFSET // max(1, int(per_page))) + 1
            if page > max_page:
                page = max_page
            rows = []
            total = 0
            if issuer_id > 0:
                conn = db()
                try:
                    if query:
                        like = f"%{query}%"
                        total_row = conn.execute(
                            "SELECT COUNT(*) AS c FROM products WHERE issuer_id = ? AND (COALESCE(name,'') LIKE ? OR COALESCE(clave_prod_serv,'') LIKE ?)",
                            (issuer_id, like, like),
                        ).fetchone()
                        total = int(total_row.get("c") or total_row.get("n") or 0) if isinstance(total_row, dict) else (int(total_row[0]) if total_row else 0)
                        offset = (page - 1) * per_page
                        rows = conn.execute(
                            """
                            SELECT id, name, clave_prod_serv, clave_unidad, unidad,
                                   default_unit_price, default_currency, active, updated_at
                            FROM products
                            WHERE issuer_id = ?
                              AND (COALESCE(name,'') LIKE ? OR COALESCE(clave_prod_serv,'') LIKE ?)
                            ORDER BY active DESC, updated_at DESC
                            LIMIT ? OFFSET ?
                            """,
                            (issuer_id, like, like, per_page, offset),
                        ).fetchall()
                    else:
                        total_row = conn.execute(
                            "SELECT COUNT(*) AS c FROM products WHERE issuer_id = ?", (issuer_id,)
                        ).fetchone()
                        total = int(total_row.get("c") or total_row.get("n") or 0) if isinstance(total_row, dict) else (int(total_row[0]) if total_row else 0)
                        offset = (page - 1) * per_page
                        rows = conn.execute(
                            """
                            SELECT id, name, clave_prod_serv, clave_unidad, unidad,
                                   default_unit_price, default_currency, active, updated_at
                            FROM products
                            WHERE issuer_id = ?
                            ORDER BY active DESC, updated_at DESC
                            LIMIT ? OFFSET ?
                            """,
                            (issuer_id, per_page, offset),
                        ).fetchall()
                finally:
                    conn.close()
            pages = (total + per_page - 1) // per_page if total > 0 else 0
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_products.html",
                active_page="products",
                title="Productos",
                extra={
                    "rows": [dict(r) for r in rows] if rows else [],
                    "q": query,
                    "total": total,
                    "page": page,
                    "per_page": per_page,
                    "pages": pages,
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/products")
            raise

    @router.get("/products/suggestions", response_class=HTMLResponse)
    def portal_products_suggestions(request: Request, issuer: dict = Depends(get_portal_issuer), q: str = Query("")):
        try:
            issuer_id = int(issuer.get("id") or 0)
            query = (q or "").strip()
            rows = []
            if issuer_id > 0:
                conn = db()
                try:
                    if query:
                        like = f"%{query}%"
                        rows = conn.execute(
                            """
                            SELECT id, clave_prod_serv, raw_description, clave_unidad, unidad,
                                   unit_price_hint, currency, times_seen, last_seen_at
                            FROM product_observations
                            WHERE issuer_id = ?
                              AND (COALESCE(raw_description,'') LIKE ? OR COALESCE(clave_prod_serv,'') LIKE ?)
                            ORDER BY times_seen DESC, last_seen_at DESC
                            LIMIT 500
                            """,
                            (issuer_id, like, like),
                        ).fetchall()
                    else:
                        rows = conn.execute(
                            """
                            SELECT id, clave_prod_serv, raw_description, clave_unidad, unidad,
                                   unit_price_hint, currency, times_seen, last_seen_at
                            FROM product_observations
                            WHERE issuer_id = ?
                            ORDER BY times_seen DESC, last_seen_at DESC
                            LIMIT 500
                            """,
                            (issuer_id,),
                        ).fetchall()
                finally:
                    conn.close()
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_product_suggestions.html",
                active_page="product_suggestions",
                title="Productos sugeridos",
                extra={"rows": [dict(r) for r in rows] if rows else [], "q": query},
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/products/suggestions")
            raise

    @router.post("/products/suggestions/{observation_id}/convert", response_class=JSONResponse)
    def portal_products_suggestions_convert(
        request: Request,
        observation_id: int,
        payload: dict = Body(default_factory=dict),
        issuer: dict = Depends(get_portal_issuer),
    ):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        try:
            obs_id = int(observation_id)
        except Exception:
            raise HTTPException(status_code=400, detail="ID inválido")

        name_override = (payload.get("name") if isinstance(payload, dict) else "") or ""
        name_override = str(name_override).strip()

        conn = db()
        try:
            obs = conn.execute(
                """
                SELECT id, clave_prod_serv, clave_unidad, unidad, raw_description, unit_price_hint, currency
                FROM product_observations
                WHERE issuer_id = ? AND id = ?
                LIMIT 1
                """,
                (issuer_id, obs_id),
            ).fetchone()
            if not obs:
                raise HTTPException(status_code=404, detail="Sugerencia no encontrada")

            name = name_override or (obs["raw_description"] or "").strip()
            if not name:
                raise HTTPException(status_code=400, detail="Nombre del producto es obligatorio")

            cps = (obs["clave_prod_serv"] or "").strip() or None
            cu = (obs["clave_unidad"] or "").strip() or None
            unidad = (obs["unidad"] or "").strip() or None
            price = obs["unit_price_hint"]
            currency = (obs["currency"] or "").strip().upper() or None

            conn.execute(
                """
                INSERT INTO products (
                  issuer_id, name, clave_prod_serv, clave_unidad, unidad,
                  default_unit_price, default_currency, active, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, datetime('now'), datetime('now'))
                ON CONFLICT(issuer_id, name, clave_prod_serv, clave_unidad) DO UPDATE SET
                  unidad = COALESCE(excluded.unidad, products.unidad),
                  default_unit_price = COALESCE(excluded.default_unit_price, products.default_unit_price),
                  default_currency = COALESCE(excluded.default_currency, products.default_currency),
                  active = 1,
                  updated_at = datetime('now')
                """,
                (issuer_id, name, cps, cu, unidad, price, currency),
            )
            conn.commit()
            row = conn.execute(
                """
                SELECT id FROM products
                WHERE issuer_id = ? AND name = ? AND COALESCE(clave_prod_serv,'') = COALESCE(?, '')
                  AND COALESCE(clave_unidad,'') = COALESCE(?, '')
                ORDER BY id DESC LIMIT 1
                """,
                (issuer_id, name, cps, cu),
            ).fetchone()
            prod_id = int(row["id"]) if row else None
        finally:
            conn.close()

        log_action(request, "product_converted_from_observation", issuer_id=issuer_id, entity_id=str(obs_id))
        return JSONResponse({"ok": True, "product_id": prod_id})

    @router.post("/products/save", response_class=JSONResponse)
    def portal_products_save(
        request: Request,
        payload: dict = Body(...),
        issuer: dict = Depends(get_portal_issuer),
    ):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Payload inválido")

        prod_id = payload.get("id")
        name = str(payload.get("name") or "").strip()
        clave_prod_serv = str(payload.get("clave_prod_serv") or "").strip() or None
        clave_unidad = str(payload.get("clave_unidad") or "").strip() or None
        unidad = str(payload.get("unidad") or "").strip() or None
        default_currency = str(payload.get("default_currency") or "").strip().upper() or None
        active = 1 if str(payload.get("active") or "1").strip() not in ("0", "false", "False") else 0
        default_unit_price = payload.get("default_unit_price")
        try:
            default_unit_price = float(default_unit_price) if default_unit_price not in (None, "", "null") else None
        except Exception:
            raise HTTPException(status_code=400, detail="Precio inválido")

        if not name:
            raise HTTPException(status_code=400, detail="Nombre es obligatorio")

        conn = db()
        try:
            if prod_id is not None and str(prod_id).strip():
                pid = int(prod_id)
                cur = conn.execute(
                    """
                    UPDATE products SET
                      name = ?,
                      clave_prod_serv = ?,
                      clave_unidad = ?,
                      unidad = ?,
                      default_unit_price = ?,
                      default_currency = ?,
                      active = ?,
                      updated_at = datetime('now')
                    WHERE issuer_id = ? AND id = ?
                    """,
                    (name, clave_prod_serv, clave_unidad, unidad, default_unit_price, default_currency, active, issuer_id, pid),
                )
                if cur.rowcount == 0:
                    raise HTTPException(status_code=404, detail="Producto no encontrado")
                conn.commit()
                return JSONResponse({"ok": True, "id": pid})

            conn.execute(
                """
                INSERT INTO products (
                  issuer_id, name, clave_prod_serv, clave_unidad, unidad,
                  default_unit_price, default_currency, active, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
                ON CONFLICT(issuer_id, name, clave_prod_serv, clave_unidad) DO UPDATE SET
                  unidad = COALESCE(excluded.unidad, products.unidad),
                  default_unit_price = COALESCE(excluded.default_unit_price, products.default_unit_price),
                  default_currency = COALESCE(excluded.default_currency, products.default_currency),
                  active = excluded.active,
                  updated_at = datetime('now')
                """,
                (issuer_id, name, clave_prod_serv, clave_unidad, unidad, default_unit_price, default_currency, active),
            )
            conn.commit()
            row = conn.execute(
                """
                SELECT id FROM products
                WHERE issuer_id = ? AND name = ? AND COALESCE(clave_prod_serv,'') = COALESCE(?, '')
                  AND COALESCE(clave_unidad,'') = COALESCE(?, '')
                ORDER BY id DESC LIMIT 1
                """,
                (issuer_id, name, clave_prod_serv, clave_unidad),
            ).fetchone()
            return JSONResponse({"ok": True, "id": int(row["id"]) if row else None})
        finally:
            conn.close()

    @router.post("/products/{product_id}/toggle", response_class=JSONResponse)
    def portal_products_toggle(request: Request, product_id: int, issuer: dict = Depends(get_portal_issuer)):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        pid = int(product_id)
        conn = db()
        try:
            row = conn.execute(
                "SELECT active FROM products WHERE issuer_id = ? AND id = ? LIMIT 1",
                (issuer_id, pid),
            ).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Producto no encontrado")
            new_val = 0 if int(row["active"] or 0) == 1 else 1
            conn.execute(
                "UPDATE products SET active = ?, updated_at = datetime('now') WHERE issuer_id = ? AND id = ?",
                (new_val, issuer_id, pid),
            )
            conn.commit()
        finally:
            conn.close()
        return JSONResponse({"ok": True, "active": new_val})

    @router.post("/products/{product_id}/delete", response_class=JSONResponse)
    def portal_products_delete(request: Request, product_id: int, issuer: dict = Depends(get_portal_issuer)):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        pid = int(product_id)
        conn = db()
        try:
            cur = conn.execute(
                "DELETE FROM products WHERE issuer_id = ? AND id = ?",
                (issuer_id, pid),
            )
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Producto no encontrado")
        finally:
            conn.close()
        return JSONResponse({"ok": True})

    @router.post("/clients/save", response_class=JSONResponse)
    def portal_clients_save(request: Request, payload: dict = Body(...), issuer: dict = Depends(get_portal_issuer)):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Payload inválido")

        cid = payload.get("id")
        name = str(payload.get("name") or "").strip() or None
        cp = str(payload.get("cp") or "").strip() or None
        regimen = str(payload.get("regimen_fiscal") or "").strip() or None
        uso = str(payload.get("uso_cfdi_default") or "").strip().upper() or None
        email = str(payload.get("email") or "").strip() or None
        phone = str(payload.get("phone") or "").strip() or None

        if cid is None or not str(cid).strip():
            raise HTTPException(status_code=400, detail="ID requerido")
        client_id = int(cid)

        conn = db()
        try:
            cur = conn.execute(
                """
                UPDATE clients SET
                  name = COALESCE(?, name),
                  cp = COALESCE(?, cp),
                  regimen_fiscal = COALESCE(?, regimen_fiscal),
                  uso_cfdi_default = COALESCE(?, uso_cfdi_default),
                  email = COALESCE(?, email),
                  phone = COALESCE(?, phone),
                  updated_at = datetime('now')
                WHERE issuer_id = ? AND id = ?
                """,
                (name, cp, regimen, uso, email, phone, issuer_id, client_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Cliente no encontrado")
            conn.commit()
        finally:
            conn.close()

    @router.post("/clients/{client_id}/delete", response_class=JSONResponse)
    def portal_clients_delete(request: Request, client_id: int, issuer: dict = Depends(get_portal_issuer)):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        cid = int(client_id)
        conn = db()
        try:
            cur = conn.execute(
                "DELETE FROM clients WHERE issuer_id = ? AND id = ?",
                (issuer_id, cid),
            )
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Cliente no encontrado")
        finally:
            conn.close()
        return JSONResponse({"ok": True})
        return JSONResponse({"ok": True, "id": client_id})

    @router.get("/datos-fiscales", response_class=HTMLResponse)
    def portal_datos_fiscales(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Vista read-only de datos fiscales del emisor (RFC, razón social, régimen)."""
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_datos_fiscales.html",
            active_page="datos_fiscales",
            title="Datos fiscales",
            extra={
                "issuer_razon_social": issuer.get("alias") or issuer.get("rfc") or "—",
            },
        )

    @router.get("/summary", response_class=HTMLResponse)
    def portal_summary(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str = None):
        try:
            issuer_id = issuer["id"]
            if not ym:
                ym = ym_now()
            tot_issued = _get_month_totals(issuer_id, ym, "issued")
            tot_received = _get_month_totals(issuer_id, ym, "received")
            ingresos_sin_iva = tot_issued["total_base"]
            gastos_sin_iva = tot_received["total_base"]
            iva_retenciones = tot_issued["total_retenciones"]
            iva_recibido_neto = tot_issued["total_iva_neto"]
            iva_pagado = tot_received["total_iva"]
            months_issued = db_rows(
                """
                SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
                GROUP BY ym ORDER BY ym DESC
                """,
                (issuer_id,),
            )
            months_received = db_rows(
                """
                SELECT substr(fecha_emision,1,7) AS ym, count(*) AS n
                FROM sat_cfdi
                WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
                GROUP BY ym ORDER BY ym DESC
                """,
                (issuer_id,),
            )
            ym_counts = {}
            for m in months_issued + months_received:
                ym_counts[m["ym"]] = ym_counts.get(m["ym"], 0) + m["n"]
            if ym not in ym_counts:
                ym_counts[ym] = 0
            months = [{"ym": y, "n": n, "label": ym_to_label(y)} for y, n in sorted(ym_counts.items(), reverse=True)]
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_summary.html",
                active_page="summary",
                title="Resumen",
                extra={
                    "ym": ym,
                    "ym_label": ym_to_label(ym),
                    "prev_ym": shift_ym(ym, -1),
                    "next_ym": shift_ym(ym, +1),
                    "months": months,
                    "ingresos_sin_iva": ingresos_sin_iva,
                    "gastos_sin_iva": gastos_sin_iva,
                    "iva_recibido_neto": iva_recibido_neto,
                    "iva_retenciones": iva_retenciones,
                    "iva_pagado": iva_pagado,
                },
            )
        except Exception:
            logger.exception("portal: error renderizando /portal/summary ym=%s", ym)
            raise

    # ---------- Month Close (cierre mensual PF) ----------
    @router.get("/month-close", response_class=HTMLResponse)
    def portal_month_close(request: Request, issuer: dict = Depends(get_portal_issuer), ym: str | None = Query(None)):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        ym_val = (ym or ym_now()).strip()[:7] or ym_now()
        from services import month_close as month_close_service

        status = month_close_service.get_status(issuer_id, ym_val)
        ov = status.get("overrides") if isinstance(status.get("overrides"), dict) else {}

        issued_count = db_rows(
            """
            SELECT COUNT(*) AS n FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'issued' AND fecha_emision IS NOT NULL
              AND substr(fecha_emision,1,7) = ? AND (total IS NULL OR total >= 0.01)
            """,
            (issuer_id, ym_val),
        )
        received_count = db_rows(
            """
            SELECT COUNT(*) AS n FROM sat_cfdi
            WHERE issuer_id = ? AND direction = 'received' AND fecha_emision IS NOT NULL
              AND substr(fecha_emision,1,7) = ? AND total IS NOT NULL AND total >= 0.01
              AND (tipo_comprobante IS NULL OR UPPER(TRIM(tipo_comprobante)) != 'N')
            """,
            (issuer_id, ym_val),
        )
        n_issued = int(issued_count[0]["n"] if issued_count else 0)
        n_received = int(received_count[0]["n"] if received_count else 0)

        movements_count = 0
        try:
            r = db_rows("SELECT COUNT(*) AS n FROM bank_movements WHERE issuer_id = ? AND period_month = ?", (issuer_id, ym_val))
            movements_count = int(r[0]["n"] if r else 0)
        except Exception:
            movements_count = 0

        tot_issued = _get_month_totals(issuer_id, ym_val, "issued")
        tot_received = _get_month_totals(issuer_id, ym_val, "received")
        iva_est = {
            "iva_recibido_neto": float(tot_issued.get("total_iva_neto") or 0),
            "iva_pagado": float(tot_received.get("total_iva") or 0),
            "iva_estimado_a_pagar": round(float(tot_issued.get("total_iva_neto") or 0) - float(tot_received.get("total_iva") or 0), 2),
        }

        has_acuse = month_close_service.pdf_exists(issuer_id=issuer_id, ym=ym_val, kind="acuse")
        has_opinion = month_close_service.pdf_exists(issuer_id=issuer_id, ym=ym_val, kind="opinion")

        items = [
            {"key": "sync_issued", "label": "Facturas emitidas sincronizadas", "ok": bool(n_issued > 0), "meta": f"{n_issued} este mes"},
            {"key": "sync_received", "label": "Facturas recibidas sincronizadas", "ok": bool(n_received > 0), "meta": f"{n_received} este mes"},
            {"key": "bank_movements", "label": "Movimientos bancarios cargados", "ok": bool(movements_count > 0), "meta": f"{movements_count} este mes"},
            {"key": "reconciliation", "label": "Conciliación: gastos sin factura / facturas sin movimiento", "ok": False, "meta": "MVP: en progreso"},
            {"key": "tax_estimate", "label": "Estimación de impuestos (IVA)", "ok": bool(n_issued > 0 or n_received > 0), "meta": f"IVA est.: {iva_est['iva_estimado_a_pagar']:.2f}"},
            {"key": "acuse", "label": "Subir acuse de declaración (PDF)", "ok": bool(has_acuse), "meta": "PDF"},
            {"key": "opinion", "label": "Subir opinión de cumplimiento (PDF)", "ok": bool(has_opinion), "meta": "PDF"},
        ]
        for it in items:
            if it["key"] in ov:
                it["ok"] = bool(ov[it["key"]])

        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_month_close.html",
            active_page="month_close",
            title="Cierre del mes",
            extra={
                "ym": ym_val,
                "ym_label": ym_to_label(ym_val),
                "items": items,
                "status": status,
                "iva_est": iva_est,
                "csrf_token": csrf_service.generate_csrf_token(),
                "has_acuse": has_acuse,
                "has_opinion": has_opinion,
                "month_status": month_close_service.get_month_status_enum(issuer_id, ym_val),
            },
        )

    @router.post("/month-close/status", response_class=RedirectResponse)
    def portal_month_close_status(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: str = Form(...),
        status: str = Form(...),
        csrf_token: str | None = Form(None),
    ):
        token_val = (csrf_token or request.headers.get("X-CSRF-Token") or "").strip()
        if not csrf_service.verify_csrf_token(token_val):
            raise HTTPException(status_code=403, detail="Token CSRF inválido o expirado")
        issuer_id = int(issuer.get("id") or 0)
        from services import month_close as month_close_service

        try:
            month_close_service.save_month_close(issuer_id, ym, status=status)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        log_action(request, "month_close_status_change", issuer_id=issuer_id, ym=ym, status=status)
        return RedirectResponse(url=f"/portal/month-close?ym={ym}", status_code=302)

    @router.post("/month-close/override", response_class=RedirectResponse)
    def portal_month_close_override(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: str = Form(...),
        key: str = Form(...),
        value: str = Form("0"),
        csrf_token: str | None = Form(None),
    ):
        token_val = (csrf_token or request.headers.get("X-CSRF-Token") or "").strip()
        if not csrf_service.verify_csrf_token(token_val):
            raise HTTPException(status_code=403, detail="Token CSRF inválido o expirado")
        issuer_id = int(issuer.get("id") or 0)
        from services import month_close as month_close_service

        month_close_service.set_override(issuer_id, ym, key, value in ("1", "true", "on", "yes"))
        return RedirectResponse(url=f"/portal/month-close?ym={ym}", status_code=302)

    @router.post("/month-close/upload", response_class=RedirectResponse)
    async def portal_month_close_upload(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: str = Form(...),
        kind: str = Form(...),
        pdf: UploadFile = File(...),
        csrf_token: str | None = Form(None),
    ):
        token_val = (csrf_token or request.headers.get("X-CSRF-Token") or "").strip()
        if not csrf_service.verify_csrf_token(token_val):
            raise HTTPException(status_code=403, detail="Token CSRF inválido o expirado")
        if rate_limit_service.is_rate_limited(request, "month_close_upload"):
            raise HTTPException(status_code=429, detail="Demasiados intentos. Espera un minuto.")
        issuer_id = int(issuer.get("id") or 0)
        kind_norm = (kind or "").strip().lower()
        if kind_norm not in ("acuse", "opinion"):
            raise HTTPException(status_code=400, detail="Tipo inválido")
        body = await pdf.read()
        if not body or len(body) < 10:
            raise HTTPException(status_code=400, detail="Archivo vacío")
        if len(body) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="PDF demasiado grande (máx 10 MB)")
        if not body.startswith(b"%PDF"):
            raise HTTPException(status_code=400, detail="El archivo debe ser PDF")
        from services import month_close as month_close_service

        rel = month_close_service.write_pdf_to_storage(issuer_id=issuer_id, ym=ym, kind=kind_norm, pdf_bytes=body)
        audit.log(action="month_close_upload", user_id=getattr(request.state, "user_id", 0) or 0, issuer_id=issuer_id, request=request, entity="month_close", entity_id=f"{ym}:{kind_norm}")
        log_action(request, "month_close_upload", issuer_id=issuer_id, ym=ym, kind=kind_norm)
        file_access_log.log_file_access(
            request=request,
            action="upload_month_close_pdf",
            issuer_id=issuer_id,
            user_id=getattr(request.state, "user_id", None),
            file_path=rel,
            entity="month_close",
            entity_id=f"{ym}:{kind_norm}",
        )
        return RedirectResponse(url=f"/portal/month-close?ym={ym}", status_code=302)

    @router.get("/month-close/download/{ym}/{kind}", response_class=Response)
    def portal_month_close_download(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: str = "",
        kind: str = "",
        dl: int = 0,
    ):
        issuer_id = int(issuer.get("id") or 0)
        from services import month_close as month_close_service

        try:
            abs_path, rel = month_close_service.get_pdf_abs_path(issuer_id=issuer_id, ym=ym, kind=kind)
        except ValueError:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        if not os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        disposition = "attachment" if int(dl or 0) == 1 else "inline"
        file_access_log.log_file_access(
            request=request,
            action="download_month_close_pdf",
            issuer_id=issuer_id,
            user_id=getattr(request.state, "user_id", None),
            file_path=rel,
            entity="month_close",
            entity_id=f"{ym}:{kind}",
        )
        filename = f"{kind}_{ym}.pdf"
        return FileResponse(path=abs_path, media_type="application/pdf", filename=filename, headers={"Content-Disposition": f"{disposition}; filename=\"{filename}\""})

    @router.get("/plan", response_class=HTMLResponse)
    def portal_plan(request: Request, issuer: dict = Depends(get_portal_issuer), success: str = Query(""), canceled: str = Query("")):
        user_id = getattr(request.state, "user_id", None) or 0
        issuer_id = int(issuer.get("id") or 0)
        subscription = subscription_service.get_subscription_by_user_id(user_id) if user_id else None
        is_active = subscription_service.is_subscription_active(user_id)
        from services import plans as plans_service
        plan_summary = plans_service.get_plan_summary(issuer_id) if issuer_id > 0 else {
            "plan": "free", "plan_label": "Gratis", "price_mxn": 0,
            "limits": {"invoices": {"used": 0, "limit": 5}, "sat_syncs": {"used": 0, "limit": 0}, "bank_imports": {"used": 0, "limit": 2}, "bank_accounts": {"limit": 1}, "month_close": False, "matching": False},
            "all_plans": plans_service.get_all_plans() if issuer_id > 0 else {},
        }
        # Ensure all_plans always has data
        if not plan_summary.get("all_plans"):
            plan_summary["all_plans"] = {k: {"label": v["label"], "price_mxn": v["price_mxn"], "invoices": v["invoices_per_month"], "sat_syncs": v["sat_syncs_per_month"]} for k, v in plans_service.PLANS.items()}
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_plan.html",
            active_page="plan",
            title="Mi plan",
            extra={
                "subscription": subscription,
                "is_active": is_active,
                "success": success == "1",
                "canceled": canceled == "1",
                "plan_summary": plan_summary,
            },
        )

    @router.get("/info", response_class=RedirectResponse)
    def portal_info():
        """Redirige a la página pública de seguridad (misma para usuarios y visitantes)."""
        return RedirectResponse(url="/seguridad", status_code=302)

    # ---------- SAT sync (encolar desde UI; el worker procesa) ----------
    @router.post("/sat/sync", response_class=JSONResponse)
    def portal_sat_sync(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Encola sincronización SAT (issued + received). Requiere FIEL configurada y validada."""
        if rate_limit_service.is_rate_limited(request, "sat_sync"):
            return JSONResponse({"ok": False, "message": "Demasiados intentos. Espera un minuto."}, status_code=429)
        issuer_id = issuer["id"]
        user_id = getattr(request.state, "user_id", 0) or 0
        if not subscription_service.can_issuer_use_sync_and_timbrado(issuer_id, user_id):
            return JSONResponse(
                {"ok": False, "message": "Tu periodo de prueba ha terminado. Actualiza tu plan para seguir sincronizando."},
                status_code=402,
            )
        conn = db()
        try:
            _ensure_sat_credentials_validation_columns(conn)
            cred = conn.execute(
                "SELECT validation_ok FROM sat_credentials WHERE issuer_id = ?",
                (issuer_id,),
            ).fetchone()
            if not cred:
                return JSONResponse({"ok": False, "message": "Configura y valida tu FIEL en Ajustes primero."}, status_code=400)
            if cred["validation_ok"] != 1:
                return JSONResponse({"ok": False, "message": "Valida tu FIEL en Ajustes antes de sincronizar."}, status_code=400)
            # No encolar si ya hay jobs en cola o en ejecución para este issuer
            pending = conn.execute(
                "SELECT 1 FROM sat_jobs WHERE issuer_id = ? AND status IN ('queued','running') LIMIT 1",
                (issuer_id,),
            ).fetchone()
            if pending:
                return JSONResponse({"ok": False, "message": "Ya hay una sincronización en curso. Espera a que termine."}, status_code=409)
            conn.execute(
                """
                INSERT INTO sat_jobs (issuer_id, job_type, direction, status, created_at, updated_at)
                VALUES (?, 'xml', 'issued', 'queued', datetime('now'), datetime('now')),
                       (?, 'xml', 'received', 'queued', datetime('now'), datetime('now'))
                """,
                (issuer_id, issuer_id),
            )
            conn.commit()
        finally:
            conn.close()
        # Auditoría: inicio de sync SAT
        audit.log(
            action="sat_sync_started",
            user_id=user_id,
            issuer_id=issuer_id,
            request=request,
            entity="sat_jobs",
            entity_id=str(issuer_id),
        )
        log_action(request, "sat_sync_started", user_id=user_id, issuer_id=issuer_id)
        return JSONResponse({"ok": True, "message": "Sincronización iniciada."})

    @router.get("/sat/status", response_class=JSONResponse)
    def portal_sat_status(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Estado del sync SAT para este issuer: último sync, en proceso / ok / error."""
        issuer_id = issuer["id"]
        conn = db()
        try:
            running = conn.execute(
                "SELECT 1 FROM sat_jobs WHERE issuer_id = ? AND status IN ('queued','running') LIMIT 1",
                (issuer_id,),
            ).fetchone()
            last_ok = conn.execute(
                "SELECT MAX(finished_at) AS t FROM sat_jobs WHERE issuer_id = ? AND status = 'ok'",
                (issuer_id,),
            ).fetchone()
            last_error = conn.execute(
                "SELECT finished_at, last_error FROM sat_jobs WHERE issuer_id = ? AND status = 'error' ORDER BY finished_at DESC LIMIT 1",
                (issuer_id,),
            ).fetchone()
            sync_state = conn.execute(
                "SELECT MAX(last_run_at) AS t FROM sat_sync_state WHERE issuer_id = ?",
                (issuer_id,),
            ).fetchone()
        finally:
            conn.close()
        last_sync_at = (sync_state and sync_state["t"]) or (last_ok and last_ok["t"]) or None
        if running:
            status = "running"
            message = "Sincronización en proceso"
        elif last_error and last_ok and last_error["t"] and last_ok["t"] and last_error["t"] > last_ok["t"]:
            status = "error"
            message = (last_error["last_error"] or "Error en la última sincronización")[:200]
        elif last_error and not last_ok:
            status = "error"
            message = (last_error["last_error"] or "Error en la última sincronización")[:200]
        else:
            status = "ok"
            message = None
        return JSONResponse({
            "ok": True,
            "last_sync_at": last_sync_at,
            "status": status,
            "message": message,
        })

    # ---------- Catálogo sugerido desde CFDI emitidos (auto-captura) ----------
    @router.post("/catalog/backfill", response_class=JSONResponse)
    def portal_catalog_backfill(
        request: Request,
        payload: dict = Body(default_factory=dict),
        issuer: dict = Depends(get_portal_issuer),
    ):
        """
        Dispara backfill de catálogo sugerido (clients + product_observations) desde CFDI emitidos ya guardados.
        Responde JSON con métricas. Multi-issuer: siempre filtra por issuer_id actual.
        """
        if rate_limit_service.is_rate_limited(request, "catalog_backfill"):
            return JSONResponse({"ok": False, "detail": "Demasiados intentos. Espera un minuto."}, status_code=429)
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")

        limit = payload.get("limit") if isinstance(payload, dict) else None
        since = payload.get("since") if isinstance(payload, dict) else None
        try:
            result = backfill_catalog_from_existing_cfdi(
                issuer_id,
                limit=int(limit) if limit is not None and str(limit).strip() else None,
                since=str(since).strip() if since is not None and str(since).strip() else None,
            )
        except Exception as e:
            logger.exception("portal catalog backfill: issuer=%s", issuer_id)
            raise HTTPException(status_code=500, detail="No se pudo ejecutar el backfill. Revisa que existan XML emitidos y que las migraciones estén aplicadas.")

        return JSONResponse(
            {
                "ok": True,
                "processed": result.processed,
                "clients_upserted": result.clients_upserted,
                "observations_upserted": result.observations_upserted,
                "errors_count": result.errors_count,
                "errors_sample": (result.errors or [])[:10],
            }
        )

    # ---------- Bank: PDF → Excel (estado de cuenta) ----------
    MAX_BANK_PDF_SIZE = 15 * 1024 * 1024  # 15MB
    MAX_BANK_PDF_FILES = 10
    MAX_BANK_PDF_TOTAL_SIZE = 50 * 1024 * 1024  # 50MB total multi-upload

    def _ensure_bank_exports_table(conn) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bank_pdf_exports (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              issuer_id INTEGER NOT NULL,
              file_id TEXT NOT NULL,
              pdf_path TEXT NOT NULL,
              xlsx_path TEXT NOT NULL,
              meta_json TEXT,
              created_at TEXT NOT NULL DEFAULT (datetime('now')),
              UNIQUE(issuer_id, file_id)
            );
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_pdf_exports_issuer ON bank_pdf_exports(issuer_id, created_at);")

    def _ensure_bank_statements_table(conn) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bank_statements (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              issuer_id INTEGER NOT NULL,
              bank_name TEXT,
              account_last4 TEXT,
              period_start TEXT,
              period_end TEXT,
              source_pdf_path TEXT NOT NULL,
              source_pdf_sha256 TEXT NOT NULL,
              created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            """
        )
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_bank_statements_issuer_sha ON bank_statements(issuer_id, source_pdf_sha256);")

    def _movement_dedup_hash(issuer_id: int, fecha: str, descripcion: str, deposito: Optional[float], retiro: Optional[float]) -> str:
        """Hash para deduplicar movimientos: mismo issuer + fecha + concepto + montos = mismo movimiento."""
        dep = "" if deposito is None else f"{float(deposito):.2f}"
        ret = "" if retiro is None else f"{float(retiro):.2f}"
        desc = (descripcion or "").strip()[:500].replace("\r", " ").replace("\n", " ")
        payload = f"{issuer_id}|{fecha or ''}|{desc}|{dep}|{ret}"
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def _ensure_bank_movements_table(conn) -> None:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bank_movements (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              issuer_id INTEGER NOT NULL,
              statement_file_id TEXT NOT NULL,
              fecha TEXT,
              descripcion TEXT,
              deposito REAL,
              retiro REAL,
              saldo REAL,
              tipo TEXT,
              categoria TEXT,
              metodo_hint TEXT,
              contraparte_hint TEXT,
              rfc_encontrado TEXT,
              confidence_score INTEGER,
              source_page_first INTEGER,
              created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            """
        )
        if not has_column(conn, "bank_movements", "movement_hash"):
            try:
                conn.execute("ALTER TABLE bank_movements ADD COLUMN movement_hash TEXT;")
            except Exception:
                pass
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_statement ON bank_movements(issuer_id, statement_file_id);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_tipo ON bank_movements(issuer_id, tipo);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_categoria ON bank_movements(issuer_id, categoria);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_fecha ON bank_movements(issuer_id, fecha);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bank_movements_issuer_confidence ON bank_movements(issuer_id, confidence_score);")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_bank_movements_issuer_hash ON bank_movements(issuer_id, movement_hash) WHERE movement_hash IS NOT NULL;")

    @router.get("/bank/pdf-to-excel", response_class=RedirectResponse)
    def portal_bank_pdf_to_excel_redirect():
        """Redirigir al nombre canónico de la página."""
        return RedirectResponse(url="/portal/convertir-edo-cuenta", status_code=302)

    @router.post("/bank/pdf-to-excel/preview-json", response_class=JSONResponse)
    async def portal_bank_pdf_preview_json(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        file: UploadFile = File(...),
    ):
        """Parsea PDF Banorte y devuelve movimientos en JSON. Sin DB, sin guardar. Para mostrar listado en la misma página."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        filename = (file.filename or "").strip().lower()
        if not filename.endswith(".pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser .pdf")
        content_type = (file.content_type or "").lower().strip()
        if content_type and content_type not in ("application/pdf", "application/x-pdf"):
            raise HTTPException(status_code=400, detail="Tipo de archivo inválido (solo PDF).")
        size = 0
        chunks: list[bytes] = []
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BANK_PDF_SIZE:
                raise HTTPException(status_code=400, detail="El PDF excede el máximo de 15MB.")
            chunks.append(chunk)
        if size <= 0:
            raise HTTPException(status_code=400, detail="El PDF está vacío.")
        # Magic bytes mínimo: evita 500 en el parser legacy por archivo no-PDF
        if chunks and not chunks[0].startswith(b"%PDF-"):
            raise HTTPException(status_code=400, detail="El archivo no parece ser un PDF válido.")
        import tempfile
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                for ch in chunks:
                    tmp.write(ch)
                tmp_path = tmp.name
            try:
                result = parse_bank_pdf_to_movements_preview(tmp_path, preset="conservative")
            except Exception:
                raise HTTPException(status_code=400, detail="No pudimos leer el PDF. Verifica que sea un estado de cuenta válido.")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
        movements = result.get("movements") or []
        summary = result.get("summary") or {}
        if summary.get("error"):
            return JSONResponse(
                {"ok": False, "detail": summary.get("error"), "movements": [], "summary": summary},
                status_code=400,
            )
        return JSONResponse({"ok": True, "movements": movements, "summary": summary})

    @router.post("/bank/pdf-to-excel/preview-multi", response_class=JSONResponse)
    async def portal_bank_pdf_preview_multi(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        files: list[UploadFile] = File(..., description="PDFs de estados de cuenta"),
    ):
        """Multi-PDF preview: procesa varios PDFs, consolidado en memoria. Sin DB."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        if not files or len(files) > MAX_BANK_PDF_FILES:
            raise HTTPException(
                status_code=400,
                detail=f"Envía entre 1 y {MAX_BANK_PDF_FILES} archivos PDF.",
            )
        all_movements: list[dict] = []
        files_summary: list[dict] = []
        file_errors: list[dict] = []
        file_warnings: list[dict] = []
        total_size = 0
        for idx, uf in enumerate(files):
            fn = (uf.filename or "").strip()
            if not fn.lower().endswith(".pdf"):
                file_errors.append({"file_name": fn or f"archivo_{idx + 1}", "error": "El archivo debe ser .pdf"})
                continue
            chunks = []
            size = 0
            while True:
                chunk = await uf.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_BANK_PDF_SIZE:
                    file_errors.append({"file_name": fn, "error": "El PDF excede el máximo de 15MB."})
                    break
                chunks.append(chunk)
            if size > MAX_BANK_PDF_SIZE:
                continue
            total_size += size
            if total_size > MAX_BANK_PDF_TOTAL_SIZE:
                file_errors.append({"file_name": fn, "error": "Se superó el tamaño total permitido (50MB)."})
                continue
            if size <= 0:
                file_errors.append({"file_name": fn, "error": "El PDF está vacío."})
                continue
            pdf_bytes = b"".join(chunks)
            result = parse_bank_statement_preview(pdf_bytes, file_name=fn, file_index=idx)
            movs = result.get("movements") or []
            fs = result.get("file_summary") or {}
            err = result.get("file_error")
            warns = result.get("file_warnings") or []
            if err:
                file_errors.append({"file_name": fn, "error": err})
            if warns:
                file_warnings.append({"file_name": fn, "warnings": warns})
            files_summary.append(fs)
            base_idx = len(all_movements)
            for i, m in enumerate(movs):
                m["_global_idx"] = base_idx + i + 1
            all_movements.extend(movs)
        # Marcar duplicados por fingerprint (misma fecha+monto+concepto+archivo)
        fp_counts: dict[str, int] = {}
        for m in all_movements:
            fp = m.get("dedupe_fingerprint") or compute_dedupe_fingerprint(m)
            m["dedupe_fingerprint"] = fp
            fp_counts[fp] = fp_counts.get(fp, 0) + 1
        for m in all_movements:
            if fp_counts.get(m["dedupe_fingerprint"], 0) > 1:
                m["posible_duplicado"] = True
                w = m.get("warnings") or []
                if "Posible duplicado en esta carga" not in w:
                    w.append("Posible duplicado en esta carga")
                m["warnings"] = w
        issuer_id = int(issuer.get("id") or 0)
        user_accounts = bank_list_accounts_raw(issuer_id) if issuer_id > 0 else []
        statement_owner_name = None
        statement_owner_rfc = None
        if files_summary:
            fs = next((f for f in files_summary if f.get("account_holder_name")), None)
            if fs:
                statement_owner_name = fs.get("account_holder_name")
                statement_owner_rfc = fs.get("account_holder_rfc")
        for m in all_movements:
            detect_own_account_transfer(m, user_accounts, statement_owner_name, statement_owner_rfc)
        total_ing = sum(m.get("monto_deposito") or 0 for m in all_movements)
        total_gas = sum(m.get("monto_retiro") or 0 for m in all_movements)
        total_ing_impactan = sum(
            (m.get("monto_deposito") or 0) for m in all_movements
            if m.get("impacta_contabilidad", True) and (m.get("tipo_movimiento") or "").upper() == "INGRESO"
        )
        total_gas_impactan = sum(
            (m.get("monto_retiro") or 0) for m in all_movements
            if m.get("impacta_contabilidad", True) and (m.get("tipo_movimiento") or "").upper() == "GASTO"
        )
        count_in = sum(1 for m in all_movements if (m.get("tipo_movimiento") or "").upper() == "INGRESO")
        count_out = sum(1 for m in all_movements if (m.get("tipo_movimiento") or "").upper() == "GASTO")
        count_info = sum(1 for m in all_movements if (m.get("tipo_movimiento") or "").upper() == "INFO")
        count_fin = sum(1 for m in all_movements if m.get("es_movimiento_financiero"))
        low_conf = sum(1 for m in all_movements if int(m.get("confianza_clasificacion") or 0) < 60)
        count_revisar = sum(1 for m in all_movements if m.get("requiere_revision"))
        count_duplicados = sum(1 for m in all_movements if m.get("posible_duplicado"))
        global_summary = {
            "files_processed": len(files_summary),
            "files_with_errors": len(file_errors),
            "total_movements": len(all_movements),
            "total_ingresos": round(total_ing, 2),
            "total_gastos": round(total_gas, 2),
            "total_ingresos_que_impactan": round(total_ing_impactan, 2),
            "total_gastos_que_impactan": round(total_gas_impactan, 2),
            "count_ingreso": count_in,
            "count_gasto": count_out,
            "count_info": count_info,
            "count_financiero": count_fin,
            "count_low_confidence": low_conf,
            "count_requiere_revision": count_revisar,
            "count_duplicados": count_duplicados,
        }
        return JSONResponse({
            "ok": True,
            "movements": all_movements,
            "global_summary": global_summary,
            "files_summary": files_summary,
            "file_errors": file_errors,
            "file_warnings": file_warnings,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        })

    @router.get("/bank/accounts", response_class=JSONResponse)
    def portal_bank_accounts_list(issuer: dict = Depends(get_portal_issuer)):
        """Lista cuentas bancarias del usuario (para detectar cuentas propias)."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        accounts = bank_list_accounts(int(issuer["id"]))
        return JSONResponse({"ok": True, "accounts": accounts})

    @router.get("/bank/accounts/manage", response_class=HTMLResponse)
    def portal_bank_accounts_manage(request: Request, issuer: dict = Depends(get_portal_issuer)):
        """Pantalla simple: Mis cuentas bancarias (config para detectar traspasos propios)."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        accounts = bank_list_all_accounts(issuer_id)
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_bank_accounts.html",
            active_page="bank_accounts",
            title="Mis cuentas bancarias",
            extra={"accounts": accounts or []},
        )

    @router.post("/bank/accounts", response_class=JSONResponse)
    def portal_bank_accounts_create(
        issuer: dict = Depends(get_portal_issuer),
        alias: str = Body(..., embed=True),
        bank_name: str = Body(..., embed=True),
        clabe: Optional[str] = Body(None, embed=True),
        account_last4: Optional[str] = Body(None, embed=True),
        holder_name: Optional[str] = Body(None, embed=True),
        rfc_titular: Optional[str] = Body(None, embed=True),
        is_active: bool = Body(True, embed=True),
    ):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        created = bank_create_account(
            int(issuer["id"]), alias=alias, bank_name=bank_name,
            clabe=clabe, account_last4=account_last4, holder_name=holder_name,
            rfc_titular=rfc_titular, is_active=is_active,
        )
        if created.get("error"):
            raise HTTPException(status_code=500, detail=created["error"])
        return JSONResponse({"ok": True, "account": created})

    @router.put("/bank/accounts/{account_id}", response_class=JSONResponse)
    def portal_bank_accounts_update(
        account_id: int,
        issuer: dict = Depends(get_portal_issuer),
        payload: dict = Body(...),
    ):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        allowed = {"alias", "bank_name", "clabe", "account_last4", "holder_name", "rfc_titular", "is_active"}
        kwargs = {k: v for k, v in payload.items() if k in allowed}
        if "account_last4" in kwargs and kwargs["account_last4"]:
            kwargs["account_last4"] = str(kwargs["account_last4"]).strip()[:4]
        updated = bank_update_account(account_id, int(issuer["id"]), **kwargs)
        if not updated:
            raise HTTPException(status_code=404, detail="Cuenta no encontrada")
        return JSONResponse({"ok": True, "account": updated})

    @router.delete("/bank/accounts/{account_id}", response_class=JSONResponse)
    def portal_bank_accounts_delete(account_id: int, issuer: dict = Depends(get_portal_issuer)):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        deleted = bank_delete_account(account_id, int(issuer["id"]))
        if not deleted:
            raise HTTPException(status_code=404, detail="Cuenta no encontrada")
        return JSONResponse({"ok": True})

    @router.post("/bank/statements/ingest", response_class=JSONResponse)
    async def portal_bank_statements_ingest(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        file: UploadFile = File(...),
        bank_account_id: int = Form(..., description="ID de cuenta bancaria del issuer"),
    ):
        """Ingesta estado de cuenta con validación RFC/cuenta. Fases 2+3."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        fn = (file.filename or "").strip()
        if not fn.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser .pdf")
        size = 0
        chunks = []
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BANK_PDF_SIZE:
                raise HTTPException(status_code=400, detail="El PDF excede el máximo de 15MB.")
            chunks.append(chunk)
        if size <= 0:
            raise HTTPException(status_code=400, detail="El PDF está vacío.")
        pdf_bytes = b"".join(chunks)
        sha = hashlib.sha256(pdf_bytes).hexdigest()
        result = parse_bank_statement_preview(pdf_bytes, file_name=fn, file_index=0)
        if result.get("file_error"):
            return JSONResponse(
                {"ok": False, "rejection_reason": result["file_error"], "status": "parse_error"},
                status_code=400,
            )
        storage_root = get_storage_root(BASE_DIR)
        uploads_rel = os.path.join("uploads", str(issuer_id), "bank")
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        pdf_name = f"{stamp}_{sha[:12]}.pdf"
        pdf_rel_path = os.path.join(uploads_rel, pdf_name)
        pdf_abs_path = safe_join(storage_root, pdf_rel_path)
        ensure_parent_dir(pdf_abs_path)
        with open(pdf_abs_path, "wb") as f:
            f.write(pdf_bytes)
        expected_rfc = (issuer.get("rfc") or "").strip()
        ingest_result = ingest_bank_statement(
            issuer_id=issuer_id,
            bank_account_id=bank_account_id,
            pdf_path=pdf_rel_path,
            pdf_sha256=sha,
            source_file_name=fn,
            preview_result=result,
            expected_issuer_rfc=expected_rfc,
        )
        if not ingest_result.get("ok"):
            return JSONResponse(
                {
                    "ok": False,
                    "rejection_reason": ingest_result.get("rejection_reason", "Error desconocido"),
                    "status": ingest_result.get("status", "error"),
                },
                status_code=400,
            )
        statement_id = ingest_result.get("statement_id")
        movements_count = ingest_result.get("movements_count", 0)
        if statement_id and movements_count > 0 and table_exists(db(), "bank_invoice_matches"):
            try:
                conn = db()
                rows = conn.execute(
                    "SELECT id, deposito, retiro, amount, fecha, descripcion, rfc_encontrado, counterparty_rfc_detected, requires_cfdi FROM bank_movements WHERE issuer_id = ? AND bank_statement_id = ?",
                    (issuer_id, statement_id),
                ).fetchall()
                conn.close()
                for r in rows:
                    r = _db_row_to_dict(r)
                    if int(r.get("requires_cfdi") or 0):
                        mov = r
                        candidates = find_cfdi_candidates(issuer_id, mov, direction="received", limit=5)
                        if candidates:
                            save_suggested_matches(issuer_id, int(r["id"]), candidates, "payment")
            except Exception as e:
                logger.exception("bank ingest: matching post-insert failed: %s", e)
        log_action(request, "bank_statement_ingest", issuer_id=issuer_id, entity_id=statement_id)
        return JSONResponse({
            "ok": True,
            "statement_id": statement_id,
            "movements_count": movements_count,
            "inserted_count": ingest_result.get("inserted_count", movements_count),
            "duplicate_movements_count": ingest_result.get("duplicate_movements_count", 0),
            "duplicate": ingest_result.get("duplicate", False),
        })

    @router.post("/bank/matches/{match_id}/confirm", response_class=JSONResponse)
    def portal_bank_match_confirm(match_id: int, issuer: dict = Depends(get_portal_issuer)):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        ok = match_confirm(match_id, int(issuer["id"]))
        if not ok:
            raise HTTPException(status_code=404, detail="Match no encontrado")
        return JSONResponse({"ok": True, "status": "confirmed"})

    @router.post("/bank/matches/{match_id}/reject", response_class=JSONResponse)
    def portal_bank_match_reject(match_id: int, issuer: dict = Depends(get_portal_issuer)):
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        ok = match_reject(match_id, int(issuer["id"]))
        if not ok:
            raise HTTPException(status_code=404, detail="Match no encontrado")
        return JSONResponse({"ok": True, "status": "rejected"})

    @router.patch("/bank/movements/{movement_id}", response_class=JSONResponse)
    async def portal_bank_movement_update(
        movement_id: int,
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
    ):
        """Actualiza descripción y/o categoría de un movimiento (mismo comportamiento que en convertir edo. de cuenta)."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        try:
            body = await request.json() if request.headers.get("content-type", "").strip().startswith("application/json") else {}
        except Exception:
            body = {}
        descripcion = body.get("descripcion")
        categoria = body.get("categoria")
        if descripcion is None and categoria is None:
            return JSONResponse({"ok": True, "updated": False})
        conn = db()
        try:
            row = conn.execute(
                "SELECT id FROM bank_movements WHERE id = ? AND issuer_id = ?",
                (movement_id, issuer_id),
            ).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Movimiento no encontrado")
            updates = []
            params: list = []
            if descripcion is not None:
                updates.append("descripcion = ?")
                params.append(str(descripcion).strip() if descripcion else "")
            if categoria is not None:
                updates.append("categoria = ?")
                params.append(str(categoria).strip() if categoria else "")
            if not updates:
                return JSONResponse({"ok": True, "updated": False})
            params.extend([movement_id, issuer_id])
            conn.execute(
                "UPDATE bank_movements SET " + ", ".join(updates) + " WHERE id = ? AND issuer_id = ?",
                params,
            )
            conn.commit()
            return JSONResponse({"ok": True, "updated": True})
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    @router.post("/bank/movements/delete-all", response_class=JSONResponse)
    def portal_bank_movements_delete_all(issuer: dict = Depends(get_portal_issuer)):
        """Borra todos los movimientos bancarios del emisor actual. Requiere confirmación en el cliente."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        conn = db()
        try:
            cur = conn.execute("DELETE FROM bank_movements WHERE issuer_id = ?", (issuer_id,))
            deleted = cur.rowcount
            conn.commit()
            return JSONResponse({"ok": True, "deleted": deleted})
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    @router.post("/bank/pdf-to-excel/upload")
    async def portal_bank_pdf_to_excel_upload(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        file: UploadFile = File(...),
        preview: Optional[str] = Form(None),
    ):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")

        # Validaciones básicas
        filename = (file.filename or "").strip()
        name_l = filename.lower()
        if not name_l.endswith(".pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser .pdf")
        content_type = (file.content_type or "").lower().strip()
        if content_type and content_type not in ("application/pdf", "application/x-pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser un PDF válido (MIME application/pdf)")

        # Leer PDF en memoria
        sha = hashlib.sha256()
        size = 0
        chunks: list[bytes] = []
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_BANK_PDF_SIZE:
                raise HTTPException(status_code=400, detail="El PDF excede el máximo de 15MB.")
            sha.update(chunk)
            chunks.append(chunk)
        if size <= 0:
            raise HTTPException(status_code=400, detail="El PDF está vacío.")

        # Vista previa Banorte: solo parsear y devolver HTML (no DB, no XLSX)
        if preview and str(preview).strip().lower() in ("1", "true", "yes"):
            import tempfile
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    for ch in chunks:
                        tmp.write(ch)
                    tmp_path = tmp.name
                result = parse_bank_pdf_to_movements_preview(tmp_path)
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_bank_pdf_to_excel.html",
                active_page="bank_pdf_to_excel",
                title="Convertir Edo. de Cuenta",
                extra={
                    "preview_movements": result.get("movements") or [],
                    "preview_summary": result.get("summary") or {},
                },
            )

        storage_root = get_storage_root(BASE_DIR)
        uploads_rel_dir = os.path.join("uploads", str(issuer_id), "bank")
        exports_rel_dir = os.path.join("exports", str(issuer_id), "bank_statements")

        pdf_sha256 = sha.hexdigest()
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        pdf_name = f"{stamp}_{pdf_sha256[:12]}.pdf"
        pdf_rel_path = os.path.join(uploads_rel_dir, pdf_name)
        pdf_abs_path = safe_join(storage_root, pdf_rel_path)
        ensure_parent_dir(pdf_abs_path)
        with open(pdf_abs_path, "wb") as f:
            for ch in chunks:
                f.write(ch)

        pdf_bytes = b"".join(chunks)
        # Get or create bank_statement (dedupe por mismo PDF). Validar RFC antes de guardar.
        conn = db()
        upload_metadata = {}
        result_preview = parse_bank_statement_preview(pdf_bytes, file_name=filename or "documento.pdf", file_index=0)
        if not result_preview.get("file_error"):
            upload_metadata = extract_statement_metadata(result_preview)
            detected_rfc = (upload_metadata.get("detected_holder_rfc") or "").strip().upper().replace(" ", "")
            expected_rfc = (issuer.get("rfc") or "").strip().upper().replace(" ", "")
            if expected_rfc and detected_rfc and expected_rfc != detected_rfc:
                raise HTTPException(
                    status_code=400,
                    detail="El RFC del estado de cuenta no coincide con el RFC de tu cuenta. No se puede procesar este PDF.",
                )
        try:
            _ensure_bank_statements_table(conn)
            row = conn.execute(
                "SELECT id FROM bank_statements WHERE issuer_id = ? AND source_pdf_sha256 = ? LIMIT 1",
                (issuer_id, pdf_sha256),
            ).fetchone()
            if row:
                statement_id = int(row["id"])
            else:
                conn.execute(
                    """
                    INSERT INTO bank_statements (issuer_id, source_pdf_path, source_pdf_sha256, created_at)
                    VALUES (?, ?, ?, datetime('now'))
                    """,
                    (issuer_id, pdf_rel_path, pdf_sha256),
                )
                statement_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                conn.commit()
            period_month = (upload_metadata.get("period_month") or "")[:7]
            if period_month and has_column(conn, "bank_statements", "period_month"):
                conn.execute(
                    "UPDATE bank_statements SET period_month = ?, bank_name = ?, account_last4 = ? WHERE id = ? AND issuer_id = ?",
                    (period_month, upload_metadata.get("bank_name"), upload_metadata.get("account_last4"), statement_id, issuer_id),
                )
                conn.commit()
        finally:
            conn.close()

        file_id = secrets.token_urlsafe(16)
        xlsx_name = f"{stamp}_{file_id[:10]}.xlsx"
        xlsx_rel_path = os.path.join(exports_rel_dir, xlsx_name)
        xlsx_abs_path = safe_join(storage_root, xlsx_rel_path)

        try:
            meta = convert_pdf_to_xlsx(
                pdf_abs_path,
                xlsx_abs_path,
                issuer_id=issuer_id,
                statement_id=statement_id,
            )
        except Exception as e:
            logger.exception("bank pdf-to-excel: error convirtiendo issuer=%s pdf=%s", issuer_id, pdf_rel_path)
            portal_error_type("parse_fail", log_context={"issuer_id": issuer_id, "pdf": pdf_rel_path})

        meta_for_storage = {k: v for k, v in (meta or {}).items() if k != "transactions"}
        meta_json_str = json.dumps(meta_for_storage, ensure_ascii=False)[:4000]

        conn = db()
        try:
            _ensure_bank_exports_table(conn)
            _ensure_bank_movements_table(conn)
            conn.execute(
                """
                INSERT INTO bank_pdf_exports (issuer_id, file_id, pdf_path, xlsx_path, meta_json)
                VALUES (?, ?, ?, ?, ?)
                """,
                (issuer_id, file_id, pdf_rel_path, xlsx_rel_path, meta_json_str),
            )
            default_period = (upload_metadata.get("period_month") or (meta or {}).get("period_start") or "")[:7]
            mov_has_period = has_column(conn, "bank_movements", "period_month")
            mov_has_hash = has_column(conn, "bank_movements", "movement_hash")
            user_accounts = bank_list_accounts_raw(int(issuer_id)) if int(issuer_id) > 0 else []
            statement_owner_name = (upload_metadata.get("detected_holder_name") or "").strip() or None
            statement_owner_rfc = (upload_metadata.get("detected_holder_rfc") or "").strip() or None
            for t in (meta or {}).get("transactions") or []:
                fecha = (t.get("fecha") or "")[:32]
                descripcion = (t.get("descripcion") or "")[:2000]
                deposito = t.get("deposito")
                retiro = t.get("retiro")
                # Reglas Job3 (MVP): traspasos propios + pagos financieros (sin IA)
                try:
                    desc_norm = (descripcion or "").strip().upper()
                    mov_hint = {
                        "raw_text_original": descripcion,
                        "raw_text_normalized": desc_norm,
                        "referencia": "",
                        "contraparte_nombre": (t.get("contraparte_hint") or "").strip(),
                        "rfc_detectado": (t.get("rfc_encontrado") or "").strip(),
                        "tipo_movimiento": (t.get("tipo") or "").strip().upper(),
                        "monto_deposito": float(deposito or 0),
                        "monto_retiro": float(retiro or 0),
                        "categoria_sugerida": (t.get("categoria") or "").strip(),
                        "warnings": [],
                    }
                    detect_own_account_transfer(mov_hint, user_accounts, statement_owner_name, statement_owner_rfc)
                    if mov_hint.get("es_transferencia_propia_probable"):
                        t["categoria"] = "CUENTA_PROPIA"
                    else:
                        metodo = (t.get("metodo_hint") or "").upper()
                        if (
                            ("PAGO CONCENTRACION" in desc_norm or "PAGO TARJETA" in desc_norm or "TARJETA DE CRED" in desc_norm)
                            or ("TARJETA" in metodo and "PAGO" in desc_norm)
                        ):
                            t["categoria"] = "FINANCIERO_PAGO_TARJETA"
                except Exception:
                    pass
                if mov_has_hash:
                    m_hash = _movement_dedup_hash(issuer_id, fecha, descripcion, deposito, retiro)
                    existing = conn.execute(
                        "SELECT 1 FROM bank_movements WHERE issuer_id = ? AND movement_hash = ? LIMIT 1",
                        (issuer_id, m_hash),
                    ).fetchone()
                    if existing:
                        continue
                # Mes por movimiento = fecha del movimiento (cada fila a su mes), no solo periodo del estado
                row_ym = (fecha[:7] if (fecha and len(fecha) >= 7 and fecha[4] == "-" and fecha[:4].isdigit() and fecha[5:7].isdigit()) else None) or default_period
                if mov_has_period and row_ym:
                    if mov_has_hash:
                        conn.execute(
                            """
                            INSERT INTO bank_movements (issuer_id, statement_file_id, period_month, fecha, descripcion, deposito, retiro, saldo, tipo, categoria, metodo_hint, contraparte_hint, rfc_encontrado, confidence_score, source_page_first, movement_hash)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                issuer_id,
                                file_id,
                                row_ym,
                                fecha,
                                descripcion,
                                deposito,
                                retiro,
                                t.get("saldo"),
                                (t.get("tipo") or "DESCONOCIDO")[:32],
                                (t.get("categoria") or "")[:200],
                                (t.get("metodo_hint") or "")[:64],
                                (t.get("contraparte_hint") or "")[:200],
                                (t.get("rfc_encontrado") or "")[:20],
                                int(t.get("confidence_score") or 0),
                                int(t.get("source_page_first") or 0),
                                m_hash if mov_has_hash else None,
                            ),
                        )
                    else:
                        conn.execute(
                            """
                            INSERT INTO bank_movements (issuer_id, statement_file_id, period_month, fecha, descripcion, deposito, retiro, saldo, tipo, categoria, metodo_hint, contraparte_hint, rfc_encontrado, confidence_score, source_page_first)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                issuer_id,
                                file_id,
                                row_ym,
                                fecha,
                                descripcion,
                                deposito,
                                retiro,
                                t.get("saldo"),
                                (t.get("tipo") or "DESCONOCIDO")[:32],
                                (t.get("categoria") or "")[:200],
                                (t.get("metodo_hint") or "")[:64],
                                (t.get("contraparte_hint") or "")[:200],
                                (t.get("rfc_encontrado") or "")[:20],
                                int(t.get("confidence_score") or 0),
                                int(t.get("source_page_first") or 0),
                            ),
                        )
                else:
                    if mov_has_hash:
                        conn.execute(
                            """
                            INSERT INTO bank_movements (issuer_id, statement_file_id, fecha, descripcion, deposito, retiro, saldo, tipo, categoria, metodo_hint, contraparte_hint, rfc_encontrado, confidence_score, source_page_first, movement_hash)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                issuer_id,
                                file_id,
                                fecha,
                                descripcion,
                                deposito,
                                retiro,
                                t.get("saldo"),
                                (t.get("tipo") or "DESCONOCIDO")[:32],
                                (t.get("categoria") or "")[:200],
                                (t.get("metodo_hint") or "")[:64],
                                (t.get("contraparte_hint") or "")[:200],
                                (t.get("rfc_encontrado") or "")[:20],
                                int(t.get("confidence_score") or 0),
                                int(t.get("source_page_first") or 0),
                                m_hash if mov_has_hash else None,
                            ),
                        )
                    else:
                        conn.execute(
                            """
                            INSERT INTO bank_movements (issuer_id, statement_file_id, fecha, descripcion, deposito, retiro, saldo, tipo, categoria, metodo_hint, contraparte_hint, rfc_encontrado, confidence_score, source_page_first)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                issuer_id,
                                file_id,
                                fecha,
                                descripcion,
                                deposito,
                                retiro,
                                t.get("saldo"),
                                (t.get("tipo") or "DESCONOCIDO")[:32],
                                (t.get("categoria") or "")[:200],
                                (t.get("metodo_hint") or "")[:64],
                                (t.get("contraparte_hint") or "")[:200],
                                (t.get("rfc_encontrado") or "")[:20],
                                int(t.get("confidence_score") or 0),
                                int(t.get("source_page_first") or 0),
                            ),
                        )
            conn.commit()
        finally:
            conn.close()

        log_action(request, "bank_pdf_to_excel", issuer_id=issuer_id, entity_id=file_id[:32])
        # Campos resumidos para UI (además de meta completo)
        try:
            processed_count = int((meta or {}).get("processed_count") or 0)
        except Exception:
            processed_count = 0
        try:
            total_ingresos = float((meta or {}).get("total_ingresos") or 0)
        except Exception:
            total_ingresos = 0.0
        try:
            total_gastos = float((meta or {}).get("total_gastos") or 0)
        except Exception:
            total_gastos = 0.0
        try:
            sin_factura_count = int((meta or {}).get("sin_factura_count") or 0)
        except Exception:
            sin_factura_count = 0
        try:
            movements_count = int((meta or {}).get("movements_count") or 0)
        except Exception:
            movements_count = 0
        try:
            ingresos_total = float((meta or {}).get("ingresos_total") or 0)
        except Exception:
            ingresos_total = 0.0
        try:
            gastos_total = float((meta or {}).get("gastos_total") or 0)
        except Exception:
            gastos_total = 0.0
        try:
            sin_parse_count = int((meta or {}).get("sin_parse_count") or 0)
        except Exception:
            sin_parse_count = 0
        quality = (meta or {}).get("quality") if isinstance((meta or {}).get("quality"), dict) else None
        try:
            low_confidence_count = int((meta or {}).get("low_confidence_count") or 0)
        except Exception:
            low_confidence_count = 0
        return JSONResponse(
            {
                "ok": True,
                "file_id": file_id,
                "statement_id": statement_id,
                "meta": meta,
                "processed_count": processed_count,
                "total_ingresos": total_ingresos,
                "total_gastos": total_gastos,
                "sin_factura_count": sin_factura_count,
                "movements_count": movements_count,
                "ingresos_total": ingresos_total,
                "gastos_total": gastos_total,
                "sin_parse_count": sin_parse_count,
                "low_confidence_count": low_confidence_count,
                "quality": quality,
                "download_url": f"/portal/bank/pdf-to-excel/download/{file_id}",
            }
        )

    def _build_preview_export_xlsx(movements: list[dict[str, Any]]) -> bytes:
        """Genera XLSX en memoria desde lista de movimientos (preview editados). Sin DB."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)
        headers = ["idx", "fecha", "concepto", "descripcion_raw", "deposito", "retiro", "saldo", "direction", "method", "category", "bucket", "deductible_hint", "needs_review", "confidence", "notes"]
        ws_mov = wb.create_sheet("Movimientos", 0)
        for col, h in enumerate(headers, 1):
            cell = ws_mov.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        for i, m in enumerate(movements, start=2):
            concept = (m.get("concept") or m.get("description_short") or "")[:500]
            raw = (m.get("description_raw") or "")[:2000]
            notes = (m.get("notes") or "")[:1000]
            ws_mov.append([
                m.get("idx"),
                m.get("date"),
                concept,
                raw,
                m.get("deposit"),
                m.get("withdraw"),
                m.get("balance"),
                m.get("direction"),
                m.get("method"),
                m.get("category"),
                m.get("bucket"),
                m.get("deductible_hint"),
                m.get("needs_review"),
                m.get("confidence"),
                notes,
            ])
        gastos = [m for m in movements if (m.get("withdraw") or 0) > 0]
        ingresos = [m for m in movements if (m.get("deposit") or 0) > 0]
        ws_g = wb.create_sheet("Gastos")
        ws_g.append(headers)
        for m in gastos:
            concept = (m.get("concept") or m.get("description_short") or "")[:500]
            ws_g.append([m.get("idx"), m.get("date"), concept, m.get("description_raw"), m.get("deposit"), m.get("withdraw"), m.get("balance"), m.get("direction"), m.get("method"), m.get("category"), m.get("bucket"), m.get("deductible_hint"), m.get("needs_review"), m.get("confidence"), m.get("notes")])
        ws_i = wb.create_sheet("Ingresos")
        ws_i.append(headers)
        for m in ingresos:
            concept = (m.get("concept") or m.get("description_short") or "")[:500]
            ws_i.append([m.get("idx"), m.get("date"), concept, m.get("description_raw"), m.get("deposit"), m.get("withdraw"), m.get("balance"), m.get("direction"), m.get("method"), m.get("category"), m.get("bucket"), m.get("deductible_hint"), m.get("needs_review"), m.get("confidence"), m.get("notes")])
        total_dep = sum(float(m.get("deposit") or 0) for m in movements)
        total_wd = sum(float(m.get("withdraw") or 0) for m in movements)
        ws_r = wb.create_sheet("Resumen")
        ws_r.append(["campo", "valor"])
        ws_r.append(["total_depositos", total_dep])
        ws_r.append(["total_retiros", total_wd])
        ws_r.append(["neto", total_dep - total_wd])
        ws_r.append(["count_movimientos", len(movements)])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @router.post("/bank/preview/export")
    def portal_bank_preview_export(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Recibe JSON con movimientos editados, devuelve XLSX. Sin DB."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        movements = body.get("movements") if isinstance(body.get("movements"), list) else []
        if not movements:
            raise HTTPException(status_code=400, detail="No hay movimientos para exportar.")
        try:
            xlsx_bytes = _build_preview_export_xlsx(movements)
        except Exception as e:
            logger.exception("bank preview export: %s", e)
            raise HTTPException(status_code=500, detail="Error al generar el Excel.")
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="movimientos_preview.xlsx"'},
        )

    @router.post("/bank/preview/reclassify")
    def portal_bank_preview_reclassify(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Re-clasifica movimientos con preset (conservative | aggressive). Sin DB."""
        if int(issuer.get("id") or 0) <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        movements = body.get("movements") if isinstance(body.get("movements"), list) else []
        preset = (body.get("preset") or "conservative").strip().lower()
        if preset not in ("conservative", "aggressive"):
            preset = "conservative"
        if not movements:
            return JSONResponse({"movements": [], "preset": preset})
        try:
            out = reclassify_movements(movements, preset=preset)
        except Exception as e:
            logger.exception("bank preview reclassify: %s", e)
            raise HTTPException(status_code=500, detail="Error al re-clasificar.")
        return JSONResponse({"movements": out, "preset": preset})

    @router.post("/bank/preview/commit", response_class=JSONResponse)
    def portal_bank_preview_commit(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        body: dict = Body(...),
    ):
        """Persiste movimientos del preview (editados en frontend) sin re-subir PDF."""
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        bank_account_id = body.get("bank_account_id")
        if bank_account_id is None:
            raise HTTPException(status_code=400, detail="Falta bank_account_id")
        try:
            bank_account_id = int(bank_account_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="bank_account_id debe ser un número")
        files = body.get("files") if isinstance(body.get("files"), list) else []
        if not files:
            raise HTTPException(status_code=400, detail="Falta lista de archivos (files) con file_summary y movements.")
        expected_rfc = (issuer.get("rfc") or "").strip()
        results = []
        total_inserted = 0
        total_duplicate_movements = 0
        last_statement_id = None
        period_month = None
        for item in files:
            file_summary = item.get("file_summary") if isinstance(item.get("file_summary"), dict) else {}
            movements = item.get("movements") if isinstance(item.get("movements"), list) else []
            if not movements:
                continue
            out = commit_preview_to_db(
                issuer_id=issuer_id,
                bank_account_id=bank_account_id,
                file_summary=file_summary,
                movements=movements,
                expected_issuer_rfc=expected_rfc,
            )
            if not out.get("ok"):
                return JSONResponse(
                    {"ok": False, "rejection_reason": out.get("rejection_reason", "Error al guardar"), "status": out.get("status", "error")},
                    status_code=400,
                )
            results.append({
                "statement_id": out.get("statement_id"),
                "inserted_count": out.get("inserted_count", 0),
                "duplicate_statement": out.get("duplicate_statement", False),
                "duplicate_movements_count": out.get("duplicate_movements_count", 0),
            })
            total_inserted += out.get("inserted_count", 0)
            total_duplicate_movements += out.get("duplicate_movements_count", 0)
            if out.get("statement_id"):
                last_statement_id = out["statement_id"]
            if out.get("period_month"):
                period_month = out["period_month"]
        log_action(request, "bank_preview_commit", issuer_id=issuer_id, entity_id=last_statement_id)
        return JSONResponse({
            "ok": True,
            "statement_id": last_statement_id,
            "inserted_count": total_inserted,
            "duplicate_movements_count": total_duplicate_movements,
            "results": results,
            "period_month": period_month,
        })

    @router.get("/bank/pdf-to-excel/download/{file_id}")
    def portal_bank_pdf_to_excel_download(
        request: Request,
        file_id: str,
        issuer: dict = Depends(get_portal_issuer),
    ):
        issuer_id = int(issuer.get("id") or 0)
        fid = (file_id or "").strip()
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        if not fid:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")

        conn = db()
        try:
            _ensure_bank_exports_table(conn)
            row = conn.execute(
                "SELECT xlsx_path FROM bank_pdf_exports WHERE issuer_id = ? AND file_id = ? LIMIT 1",
                (issuer_id, fid),
            ).fetchone()
        finally:
            conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")

        storage_root = get_storage_root(BASE_DIR)
        xlsx_rel_path = (row["xlsx_path"] or "").strip()
        if not xlsx_rel_path:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        try:
            xlsx_abs_path = safe_join(storage_root, xlsx_rel_path)
        except ValueError:
            raise HTTPException(status_code=404, detail="Ruta inválida")
        if not os.path.exists(xlsx_abs_path):
            raise HTTPException(status_code=404, detail="El archivo ya no existe en disco")

        filename = f"estado_cuenta_{fid[:8]}.xlsx"
        file_access_log.log_file_access(
            request=request,
            action="download_bank_xlsx",
            issuer_id=issuer_id,
            user_id=getattr(request.state, "user_id", None),
            file_path=xlsx_rel_path,
            entity="bank_pdf_exports",
            entity_id=fid[:64],
        )
        return FileResponse(
            path=xlsx_abs_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    @router.get("/bank/statements", response_class=HTMLResponse)
    def portal_bank_statements(request: Request, issuer: dict = Depends(get_portal_issuer)):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        statements: list = []
        conn = None
        try:
            conn = db()
            conn.row_factory = lambda cursor, row: dict(zip([c[0] for c in cursor.description], row))
            _ensure_bank_exports_table(conn)
            _ensure_bank_movements_table(conn)
            rows = conn.execute(
                "SELECT file_id, pdf_path, xlsx_path, meta_json, created_at FROM bank_pdf_exports WHERE issuer_id = ? ORDER BY created_at DESC",
                (issuer_id,),
            ).fetchall()
            statements = []
            for r in rows:
                r = _db_row_to_dict(r)
                meta = {}
                if r.get("meta_json"):
                    try:
                        meta = json.loads(r["meta_json"] or "{}")
                    except Exception:
                        pass
                period_start = meta.get("period_start") or ""
                period_end = meta.get("period_end") or ""
                bank_name = meta.get("bank_name") or "—"
                account_last4 = meta.get("account_last4") or "—"
                movements_count = int(meta.get("movements_count") or 0)
                total_gastos = float(meta.get("gastos_total") or meta.get("total_gastos") or 0)
                total_ingresos = float(meta.get("ingresos_total") or meta.get("total_ingresos") or 0)
                period_label = f"{period_start} – {period_end}" if (period_start or period_end) else "—"
                statements.append({
                    "file_id": r["file_id"],
                    "statement_key": r["file_id"],
                    "created_at": r["created_at"] or "",
                    "period_label": period_label,
                    "bank_name": bank_name,
                    "account_last4": account_last4,
                    "movements_count": movements_count,
                    "total_gastos": total_gastos,
                    "total_ingresos": total_ingresos,
                    "source": "export",
                })
            if table_exists(conn, "bank_statements"):
                has_pm = has_column(conn, "bank_statements", "period_month")
                has_tm = has_column(conn, "bank_statements", "total_movements")
                if has_pm and has_tm:
                    st_rows = conn.execute(
                        "SELECT id, period_month, bank_name, account_last4, total_movements, status, created_at FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                else:
                    st_rows = conn.execute(
                        "SELECT id, bank_name, account_last4, period_start, period_end, created_at FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                for r in st_rows:
                    r = _db_row_to_dict(r)
                    if has_pm:
                        pm = r.get("period_month") or ""
                    else:
                        pm = (r.get("period_start") or "")[:7]
                    period_label = pm if pm else ((r.get("created_at") or "")[:7] or "—")
                    statements.append({
                        "file_id": None,
                        "statement_key": f"stmt_{r['id']}",
                        "statement_id": r["id"],
                        "created_at": r.get("created_at") or "",
                        "period_label": period_label,
                        "bank_name": r.get("bank_name") or "—",
                        "account_last4": r.get("account_last4") or "—",
                        "movements_count": int(r.get("total_movements") or 0) if has_tm else 0,
                        "total_gastos": 0,
                        "total_ingresos": 0,
                        "source": "ingest",
                    })
            statements.sort(key=lambda x: (x.get("created_at") or ""), reverse=True)
        except Exception as e:
            logger.warning("portal bank/statements: error cargando lista (%s), mostrando vacío", e)
            statements = []
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_bank_statements.html",
            active_page="bank_statements",
            title="Estados de cuenta",
            statements=statements,
        )

    @router.get("/movimientos", response_class=HTMLResponse)
    @router.get("/bank/movements", response_class=HTMLResponse)
    def portal_bank_movements(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: Optional[str] = Query(None, description="Mes YYYY-MM (selector como emitidas/recibidas)"),
        statement_id: Optional[str] = Query(None, description="Filtrar por estado de cuenta (file_id o stmt_N)"),
        period_month: Optional[str] = Query(None, description="YYYY-MM (legacy, usa ym si no viene)"),
        tipo: Optional[str] = Query(None, description="INGRESO, GASTO, INFO"),
        categoria: Optional[str] = Query(None),
        cfdi_match_status: Optional[str] = Query(None, description="pending, suggested, confirmed, rejected"),
        match_filter: Optional[str] = Query(None, description="none|probable (conciliación)"),
        min_confidence: Optional[int] = Query(None, ge=0, le=100),
        search: Optional[str] = Query(None),
        hide_own_transfers: Optional[int] = Query(None, description="1 para ocultar traspasos propios"),
        hide_financial: Optional[int] = Query(None, description="1 para ocultar pagos/cargos financieros"),
        only_real_expenses: Optional[int] = Query(None, description="1 para solo gastos reales"),
        limit: int = Query(200, ge=1, le=500),
        offset: int = Query(0, ge=0, le=MAX_LIST_OFFSET),
    ):
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        # Mes: prioridad ym (selector) > period_month (legacy) > mes actual
        period_month = (ym or period_month or ym_now()).strip()[:7] or ym_now()
        movements: list = []
        total_count = 0
        sum_ingresos = 0.0
        sum_gastos = 0.0
        statements_opt: list = []
        months_with_movements: list[dict] = []
        conn = None
        try:
            conn = db()
            conn.row_factory = lambda cursor, row: dict(zip([c[0] for c in cursor.description], row))
            _ensure_bank_movements_table(conn)
            _ensure_bank_exports_table(conn)
            has_matches = table_exists(conn, "bank_invoice_matches") and table_exists(conn, "sat_cfdi")

            params: list = [issuer_id]
            where_clauses = ["issuer_id = ?"]

            if statement_id:
                sid = statement_id.strip()
                if sid.startswith("stmt_"):
                    try:
                        bid = int(sid.replace("stmt_", ""))
                        if has_column(conn, "bank_movements", "bank_statement_id"):
                            where_clauses.append("bank_statement_id = ?")
                            params.append(bid)
                        else:
                            where_clauses.append("statement_file_id = ?")
                            params.append(sid)
                    except ValueError:
                        where_clauses.append("statement_file_id = ?")
                        params.append(sid)
                else:
                    if has_column(conn, "bank_movements", "statement_file_id"):
                        where_clauses.append("statement_file_id = ?")
                        params.append(sid)
            if has_column(conn, "bank_movements", "period_month"):
                where_clauses.append("period_month = ?")
                params.append(period_month)
            if tipo:
                where_clauses.append("tipo = ?")
                params.append(tipo.strip().upper())
            if categoria:
                where_clauses.append("categoria = ?")
                params.append(categoria.strip())
            if hide_own_transfers:
                where_clauses.append("COALESCE(categoria,'') != 'CUENTA_PROPIA'")
            if hide_financial:
                where_clauses.append("COALESCE(categoria,'') NOT IN ('FINANCIERO_PAGO_TARJETA','MOVIMIENTO_FINANCIERO','COMISIONES BANCARIAS','COMISIONES_BANCARIAS','COMISION_BANCARIA')")
            if only_real_expenses:
                where_clauses.append(
                    "COALESCE(categoria,'') NOT IN ('CUENTA_PROPIA','FINANCIERO_PAGO_TARJETA','MOVIMIENTO_FINANCIERO','COMISIONES BANCARIAS','COMISIONES_BANCARIAS','COMISION_BANCARIA','TRASPASO_PROPIO')"
                )
            if cfdi_match_status and has_column(conn, "bank_movements", "cfdi_match_status"):
                where_clauses.append("cfdi_match_status = ?")
                params.append(cfdi_match_status.strip().lower())
            if match_filter and has_matches:
                mf = (match_filter or "").strip().lower()
                if mf == "probable":
                    where_clauses.append(
                        """EXISTS (
                             SELECT 1 FROM bank_invoice_matches bim
                             WHERE bim.issuer_id = bank_movements.issuer_id
                               AND bim.bank_movement_id = bank_movements.id
                               AND bim.status IN ('suggested','confirmed')
                               AND COALESCE(bim.score,0) >= 80
                           )"""
                    )
                elif mf == "revisar":
                    where_clauses.append(
                        """EXISTS (
                             SELECT 1 FROM bank_invoice_matches bim
                             WHERE bim.issuer_id = bank_movements.issuer_id
                               AND bim.bank_movement_id = bank_movements.id
                               AND bim.status IN ('suggested','confirmed')
                               AND COALESCE(bim.score,0) BETWEEN 50 AND 79
                           )"""
                    )
                elif mf == "none":
                    where_clauses.append(
                        """NOT EXISTS (
                             SELECT 1 FROM bank_invoice_matches bim
                             WHERE bim.issuer_id = bank_movements.issuer_id
                               AND bim.bank_movement_id = bank_movements.id
                               AND bim.status IN ('suggested','confirmed')
                               AND COALESCE(bim.score,0) >= 50
                           )"""
                    )
            if min_confidence is not None:
                where_clauses.append("confidence_score >= ?")
                params.append(min_confidence)
            if search and search.strip():
                q = f"%{search.strip()}%"
                if has_column(conn, "bank_movements", "raw_description"):
                    where_clauses.append("(descripcion LIKE ? OR contraparte_hint LIKE ? OR raw_description LIKE ?)")
                    params.extend([q, q, q])
                else:
                    where_clauses.append("(descripcion LIKE ? OR contraparte_hint LIKE ?)")
                    params.extend([q, q])

            where_sql = " AND ".join(where_clauses)

            total_count_row = conn.execute(
                f"SELECT COUNT(*) AS c FROM bank_movements WHERE {where_sql}",
                params,
            ).fetchone()
            total_count = int(_db_row_to_dict(total_count_row).get("c", 0) or 0)

            sum_row = conn.execute(
                f"SELECT COALESCE(SUM(deposito), 0) AS ing, COALESCE(SUM(retiro), 0) AS gas FROM bank_movements WHERE {where_sql}",
                params,
            ).fetchone()
            sum_row_d = _db_row_to_dict(sum_row)
            sum_ingresos = float(sum_row_d.get("ing", 0) or 0)
            sum_gastos = float(sum_row_d.get("gas", 0) or 0)

            # Construir SELECT solo con columnas que existan (compatibilidad con distintos esquemas)
            sel = ["id"]
            if has_column(conn, "bank_movements", "statement_file_id"):
                sel.append("statement_file_id")
            elif has_column(conn, "bank_movements", "statement_id"):
                sel.append("statement_id AS statement_file_id")
            sel.append("fecha")
            if has_column(conn, "bank_movements", "descripcion"):
                sel.append("descripcion")
            elif has_column(conn, "bank_movements", "descripcion_norm"):
                sel.append("descripcion_norm AS descripcion")
            else:
                sel.append("descripcion_raw AS descripcion")
            sel.extend(["deposito", "retiro", "saldo", "tipo", "categoria", "metodo_hint", "contraparte_hint"])
            if has_column(conn, "bank_movements", "rfc_encontrado"):
                sel.append("rfc_encontrado")
            elif has_column(conn, "bank_movements", "rfc_detectado"):
                sel.append("rfc_detectado AS rfc_encontrado")
            sel.append("confidence_score")
            if has_column(conn, "bank_movements", "bank_statement_id"):
                sel.append("bank_statement_id")
            if has_column(conn, "bank_movements", "cfdi_match_status"):
                sel.append("cfdi_match_status")
            if has_matches:
                sel.append(
                    """(
                        SELECT sc.uuid
                        FROM bank_invoice_matches bim
                        JOIN sat_cfdi sc ON sc.id = bim.cfdi_id
                        WHERE bim.issuer_id = bank_movements.issuer_id
                          AND bim.bank_movement_id = bank_movements.id
                          AND bim.status IN ('suggested','confirmed')
                        ORDER BY bim.score DESC, bim.id DESC
                        LIMIT 1
                    ) AS probable_cfdi_uuid"""
                )
                sel.append(
                    """(
                        SELECT bim.score
                        FROM bank_invoice_matches bim
                        WHERE bim.issuer_id = bank_movements.issuer_id
                          AND bim.bank_movement_id = bank_movements.id
                          AND bim.status IN ('suggested','confirmed')
                        ORDER BY bim.score DESC, bim.id DESC
                        LIMIT 1
                    ) AS probable_cfdi_score"""
                )
                sel.append(
                    """(
                        SELECT bim.status
                        FROM bank_invoice_matches bim
                        WHERE bim.issuer_id = bank_movements.issuer_id
                          AND bim.bank_movement_id = bank_movements.id
                          AND bim.status IN ('suggested','confirmed')
                        ORDER BY bim.score DESC, bim.id DESC
                        LIMIT 1
                    ) AS probable_cfdi_status"""
                )
            select_cols = ", ".join(sel)
            params_ext = params + [limit, offset]
            movements = conn.execute(
                f"SELECT {select_cols} FROM bank_movements WHERE {where_sql} ORDER BY fecha DESC, id DESC LIMIT ? OFFSET ?",
                params_ext,
            ).fetchall()
            movements = [_db_row_to_dict(r) for r in movements]
            for row in movements:
                row.setdefault("fecha", None)
                row.setdefault("descripcion", None)
                row.setdefault("deposito", None)
                row.setdefault("retiro", None)
                row.setdefault("saldo", None)
                row.setdefault("tipo", None)
                row.setdefault("categoria", None)
                row.setdefault("metodo_hint", None)
                row.setdefault("contraparte_hint", None)
                row.setdefault("rfc_encontrado", None)
                row.setdefault("confidence_score", None)
                row.setdefault("cfdi_match_status", None)
                row.setdefault("bank_statement_id", None)
                row.setdefault("probable_cfdi_uuid", None)
                row.setdefault("probable_cfdi_score", None)
                row.setdefault("probable_cfdi_status", None)
                # Asegurar que montos sean numéricos para el formato en plantilla
                for key in ("deposito", "retiro", "saldo", "confidence_score", "probable_cfdi_score"):
                    if row.get(key) is not None and row[key] != "":
                        try:
                            if key == "confidence_score":
                                row[key] = int(float(row[key]))
                            elif key == "probable_cfdi_score":
                                row[key] = int(float(row[key]))
                            else:
                                row[key] = float(row[key])
                        except (TypeError, ValueError):
                            row[key] = None if key != "confidence_score" else 0
                # Concepto = descripción sin prefijo de fecha (igual que en convertir edo. de cuenta)
                row["concepto"] = _strip_date_from_description(row.get("descripcion")) or (row.get("descripcion") or "").strip()

            statements_opt = []
            for r in conn.execute(
                "SELECT file_id, meta_json, created_at FROM bank_pdf_exports WHERE issuer_id = ? ORDER BY created_at DESC",
                (issuer_id,),
            ).fetchall():
                r = _db_row_to_dict(r)
                meta = {}
                if r.get("meta_json"):
                    try:
                        meta = json.loads(r["meta_json"] or "{}")
                    except Exception:
                        pass
                p_start = meta.get("period_start") or ""
                p_end = meta.get("period_end") or ""
                if p_start or p_end:
                    label = f"{p_start} – {p_end}"
                else:
                    label = (r.get("created_at") or "")[:16] or (r["file_id"][:12] + "…")
                statements_opt.append({"statement_id": r["file_id"], "label": label})
            if table_exists(conn, "bank_statements"):
                has_pm = has_column(conn, "bank_statements", "period_month")
                if has_pm:
                    st_opt_rows = conn.execute(
                        "SELECT id, period_month, total_movements FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                else:
                    st_opt_rows = conn.execute(
                        "SELECT id, period_start FROM bank_statements WHERE issuer_id = ? ORDER BY created_at DESC",
                        (issuer_id,),
                    ).fetchall()
                for r in st_opt_rows:
                    r = _db_row_to_dict(r)
                    if has_pm:
                        pm = r.get("period_month") or ""
                    else:
                        pm = (r.get("period_start") or "")[:7]
                    label = pm if pm else f"Estado #{r['id']}"
                    statements_opt.append({"statement_id": f"stmt_{r['id']}", "label": label})
            # Meses con movimientos (para el selector como emitidas/recibidas)
            if has_column(conn, "bank_movements", "period_month"):
                months_rows = conn.execute(
                    """SELECT period_month AS ym, COUNT(*) AS n FROM bank_movements
                       WHERE issuer_id = ? AND period_month IS NOT NULL AND TRIM(period_month) != ''
                       GROUP BY period_month ORDER BY period_month DESC""",
                    (issuer_id,),
                ).fetchall()
                for r in months_rows:
                    r = _db_row_to_dict(r)
                    ym_val = r.get("ym") or ""
                    if ym_val:
                        months_with_movements.append({"ym": ym_val, "n": int(r.get("n") or 0), "label": ym_to_label(ym_val)})
            else:
                # Sin columna period_month: usar meses de bank_statements si existen
                if table_exists(conn, "bank_statements") and has_column(conn, "bank_statements", "period_month"):
                    months_rows = conn.execute(
                        """SELECT period_month AS ym, COALESCE(SUM(total_movements), 0) AS n FROM bank_statements
                           WHERE issuer_id = ? AND period_month IS NOT NULL AND TRIM(period_month) != ''
                           GROUP BY period_month ORDER BY period_month DESC""",
                        (issuer_id,),
                    ).fetchall()
                    for r in months_rows:
                        r = _db_row_to_dict(r)
                        ym_val = r.get("ym") or ""
                        if ym_val:
                            months_with_movements.append({"ym": ym_val, "n": int(r.get("n") or 0), "label": ym_to_label(ym_val)})
        except Exception as e:
            logger.warning("portal movimientos: error cargando datos (%s), mostrando lista vacía", e)
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

        ym_safe = (period_month or ym_now()).strip()[:7] or ym_now()
        try:
            return _render_portal(
                request,
                issuer=issuer,
                template_name="portal_bank_movements.html",
                active_page="bank_movements",
                title="Movimientos",
                movements=movements,
                total_count=total_count,
                sum_ingresos=sum_ingresos,
                sum_gastos=sum_gastos,
                limit=limit,
                offset=offset,
                statement_id=statement_id or "",
                period_month=ym_safe,
                ym=ym_safe,
                ym_label=ym_to_label(ym_safe),
                prev_ym=shift_ym(ym_safe, -1),
                next_ym=shift_ym(ym_safe, +1),
                months=months_with_movements,
                tipo=tipo or "",
                categoria=categoria or "",
                cfdi_match_status=cfdi_match_status or "",
                match_filter=(match_filter or ""),
                min_confidence=min_confidence,
                search=search or "",
                hide_own_transfers=1 if hide_own_transfers else 0,
                hide_financial=1 if hide_financial else 0,
                only_real_expenses=1 if only_real_expenses else 0,
                statements_opt=statements_opt,
                csrf_token=csrf_service.generate_csrf_token(),
            )
        except Exception as e:
            logger.exception("portal movimientos (render): %s", e)
            raise HTTPException(status_code=500, detail=f"Error al mostrar la página: {e!s}")

    @router.post("/bank/movements/reconcile", response_class=RedirectResponse)
    def portal_bank_movements_reconcile(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        ym: str = Form(...),
        csrf_token: str | None = Form(None),
    ):
        token_val = (csrf_token or request.headers.get("X-CSRF-Token") or "").strip()
        if not csrf_service.verify_csrf_token(token_val):
            raise HTTPException(status_code=403, detail="Token CSRF inválido o expirado")
        if rate_limit_service.is_rate_limited(request, "bank_reconcile"):
            raise HTTPException(status_code=429, detail="Demasiados intentos. Espera un minuto.")
        issuer_id = int(issuer.get("id") or 0)
        from services import bank_cfdi_matching as bank_cfdi_matching_service

        bank_cfdi_matching_service.refresh_suggestions_for_month(issuer_id, ym)
        audit.log(
            action="bank_reconcile_run",
            user_id=getattr(request.state, "user_id", 0) or 0,
            issuer_id=issuer_id,
            request=request,
            entity="bank_movements",
            entity_id=ym,
        )
        log_action(request, "bank_reconcile_run", issuer_id=issuer_id, ym=ym)
        return RedirectResponse(url=f"/portal/bank/movements?ym={ym}", status_code=302)

    @router.post("/notifications/{notification_id}/read", response_class=RedirectResponse)
    def portal_notification_mark_read(
        request: Request,
        notification_id: int,
        issuer: dict = Depends(get_portal_issuer),
        next: str = Form("/portal/home"),
        csrf_token: str | None = Form(None),
    ):
        token_val = (csrf_token or request.headers.get("X-CSRF-Token") or "").strip()
        if not csrf_service.verify_csrf_token(token_val):
            raise HTTPException(status_code=403, detail="Token CSRF inválido o expirado")
        issuer_id = int(issuer.get("id") or 0)
        if issuer_id <= 0:
            raise HTTPException(status_code=401, detail="Sesión inválida")
        from services import notifications as notifications_service

        notifications_service.mark_read(int(issuer_id), int(notification_id))
        return RedirectResponse(url=safe_next_url(next), status_code=302)

    # ---------- SAT credentials (FIEL) upload & validate ----------
    MAX_FIEL_SIZE = 2 * 1024 * 1024  # 2 MB
    ALLOWED_CER = (".cer",)
    ALLOWED_KEY = (".key",)

    @router.get("/config/sat", response_class=HTMLResponse)
    def portal_config_sat(request: Request, issuer: dict = Depends(get_portal_issuer)):
        issuer_id = issuer["id"]
        conn = db()
        try:
            _ensure_sat_credentials_validation_columns(conn)
            row = conn.execute(
                "SELECT fiel_cer_path, fiel_key_path, validation_at, validation_ok, validation_message FROM sat_credentials WHERE issuer_id = ?",
                (issuer_id,),
            ).fetchone()
        finally:
            conn.close()
        cred = dict(row) if row else None
        return _render_portal(
            request,
            issuer=issuer,
            template_name="portal_config_sat.html",
            active_page="config_sat",
            title="FIEL / Credenciales SAT",
            extra={
                "sat_cred": cred,
                "has_fiel": cred is not None,
                "validation_at": cred.get("validation_at") if cred else None,
                "validation_ok": cred.get("validation_ok") if cred else None,
                "validation_message": cred.get("validation_message") if cred else None,
            },
        )

    @router.post("/config/sat")
    async def portal_config_sat_save(
        request: Request,
        issuer: dict = Depends(get_portal_issuer),
        fiel_cer: UploadFile = File(...),
        fiel_key: UploadFile = File(...),
        fiel_password: str = Form(""),
        csrf_token: str | None = Form(None),
    ):
        token_val = (csrf_token or request.headers.get("X-CSRF-Token") or "").strip()
        if not csrf_service.verify_csrf_token(token_val):
            raise HTTPException(status_code=403, detail="Token CSRF inválido o expirado")
        if rate_limit_service.is_rate_limited(request, "upload"):
            raise HTTPException(status_code=429, detail="Demasiados intentos. Espera un minuto.")
        issuer_id = issuer["id"]
        # Validar extensiones
        cer_name = (fiel_cer.filename or "").lower()
        key_name = (fiel_key.filename or "").lower()
        if not any(cer_name.endswith(e) for e in ALLOWED_CER):
            raise HTTPException(status_code=400, detail="El archivo del certificado debe ser .cer")
        if not any(key_name.endswith(e) for e in ALLOWED_KEY):
            raise HTTPException(status_code=400, detail="El archivo de la clave debe ser .key")
        if not (fiel_password and fiel_password.strip()):
            raise HTTPException(status_code=400, detail="La contraseña FIEL es obligatoria")
        cer_body = await fiel_cer.read()
        key_body = await fiel_key.read()
        if len(cer_body) > MAX_FIEL_SIZE or len(key_body) > MAX_FIEL_SIZE:
            raise HTTPException(status_code=400, detail="Cada archivo debe medir como máximo 2 MB")
        cred_dir = _credentials_dir(issuer_id)
        # Cifrado at-rest: guardar solo blobs cifrados (AES-GCM). Nunca persistir .cer/.key en claro.
        cer_enc_path = os.path.join(cred_dir, "fiel.cer.enc")
        key_enc_path = os.path.join(cred_dir, "fiel.key.enc")
        rel_cer = f"storage/credentials/{issuer_id}/fiel.cer.enc"
        rel_key = f"storage/credentials/{issuer_id}/fiel.key.enc"
        from services.crypto_at_rest import encrypt_bytes, encrypt_text

        cer_blob = encrypt_bytes(issuer_id=int(issuer_id), plaintext=cer_body, aad=b"fiel.cer")
        key_blob = encrypt_bytes(issuer_id=int(issuer_id), plaintext=key_body, aad=b"fiel.key")
        with open(cer_enc_path, "wb") as f:
            f.write(cer_blob)
        with open(key_enc_path, "wb") as f:
            f.write(key_blob)
        os.chmod(cer_enc_path, 0o600)
        os.chmod(key_enc_path, 0o600)
        # Guardar contraseña tal cual (sin strip) para evitar que espacios válidos rompan el descifrado,
        # pero cifrada en DB.
        password_plain = fiel_password if fiel_password is not None else ""
        password = encrypt_text(issuer_id=int(issuer_id), plaintext=password_plain)
        conn = db()
        try:
            _ensure_sat_credentials_validation_columns(conn)
            conn.execute(
                """
                INSERT INTO sat_credentials (issuer_id, fiel_cer_path, fiel_key_path, fiel_key_password, updated_at)
                VALUES (?, ?, ?, ?, datetime('now'))
                ON CONFLICT(issuer_id) DO UPDATE SET
                    fiel_cer_path = excluded.fiel_cer_path,
                    fiel_key_path = excluded.fiel_key_path,
                    fiel_key_password = excluded.fiel_key_password,
                    updated_at = datetime('now')
                """,
                (issuer_id, rel_cer, rel_key, password),
            )
            conn.commit()
        finally:
            conn.close()
        uid = getattr(request.state, "user_id", None) or 0
        audit.log(action="credentials_uploaded", user_id=uid, issuer_id=issuer_id, request=request, entity="sat_credentials", entity_id=str(issuer_id))
        log_action(request, "credentials_uploaded", issuer_id=issuer_id)
        # Validación post-upload: ejecutar check_fiel y persistir estado para mostrar en UI
        valid_ok, valid_message = _run_fiel_validation(issuer_id)
        audit.log(action="credentials_validated", user_id=uid, issuer_id=issuer_id, request=request, entity="sat_credentials", entity_id=str(issuer_id), details=f"ok={valid_ok}")
        log_action(request, "credentials_validated", issuer_id=issuer_id)
        if request.headers.get("accept", "").find("application/json") >= 0:
            return JSONResponse({"ok": True, "message": "Credenciales guardadas.", "validation_ok": valid_ok, "validation_message": valid_message})
        return RedirectResponse(url="/portal/config/sat?saved=1", status_code=302)

    @router.post("/config/sat/validate", response_class=JSONResponse)
    def portal_config_sat_validate(request: Request, issuer: dict = Depends(get_portal_issuer)):
        if rate_limit_service.is_rate_limited(request, "validate"):
            return JSONResponse({"ok": False, "message": "Demasiados intentos. Espera un minuto."}, status_code=429)
        issuer_id = issuer["id"]
        ok, message = _run_fiel_validation(issuer_id)
        if not message:
            message = "FIEL válida." if ok else "Error al validar la FIEL."
        uid = getattr(request.state, "user_id", None) or 0
        audit.log(action="credentials_validated", user_id=uid, issuer_id=issuer_id, request=request, entity="sat_credentials", entity_id=str(issuer_id), details=f"ok={ok}")
        log_action(request, "credentials_validated", issuer_id=issuer_id)
        return JSONResponse({"ok": ok, "message": message})

    return router
